import threading
import time
import queue
import os
import base64
import csv
import json
import subprocess
import ssl
import zipfile
import shutil
import gc
import tracemalloc
import traceback
import urllib.request
import urllib.error
import urllib.parse
import textwrap
import unicodedata
import uuid
import sys
import webbrowser
import http.server
import socketserver
import httpx
from datetime import datetime
from collections import defaultdict, deque
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, scrolledtext, messagebox, filedialog, simpledialog
try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

from TikTokLive import TikTokLiveClient
from TikTokLive.events import CommentEvent
try:
    from TikTokLive.events import LiveEndEvent
except Exception:
    LiveEndEvent = None
try:
    from TikTokLive.client.web.web_settings import WebDefaults
except Exception:
    WebDefaults = None

import db
import printer_utils
import license_client
import security_utils


def _default_runtime_dir() -> str:
    if os.name == "nt":
        base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
        root = os.path.join(base, "SenNails")
    else:
        root = os.path.join(os.path.expanduser("~"), ".sen_nails")
    try:
        os.makedirs(root, exist_ok=True)
    except Exception:
        root = os.getcwd()
    return root


RUNTIME_DIR = _default_runtime_dir()
DATA_DIR = os.path.join(RUNTIME_DIR, "prints")
APP_FLAVOR_NAME = os.path.splitext(os.path.basename(sys.argv[0] if sys.argv else "app.py"))[0].lower()
APP_IS_PUBLIC_BUILD = True
APP_IS_SERVER_RELAY_BUILD = "server_relay" in APP_FLAVOR_NAME
APP_IS_ARCHIVE_BUILD = ("archive" in APP_FLAVOR_NAME) or ("_ab" in APP_FLAVOR_NAME)
APP_IS_FORTIFIED_BUILD = "fortified" in APP_FLAVOR_NAME
APP_IS_SECURE_BUILD = ("secure" in APP_FLAVOR_NAME) or APP_IS_FORTIFIED_BUILD
APP_IS_SIGNPOOL_BUILD = ("signpool" in APP_FLAVOR_NAME) or ("sign_pool" in APP_FLAVOR_NAME)
APP_IS_SIGNPOOL_RELAY_BUILD = APP_IS_SERVER_RELAY_BUILD and APP_IS_SIGNPOOL_BUILD
if APP_IS_PUBLIC_BUILD:
    APP_SETTINGS_FILE = "app_settings_public.json"
elif APP_IS_FORTIFIED_BUILD and APP_IS_SERVER_RELAY_BUILD and APP_IS_ARCHIVE_BUILD and APP_IS_SECURE_BUILD and APP_IS_SIGNPOOL_BUILD:
    APP_SETTINGS_FILE = "app_settings_server_relay_archive_secure_signpool_fortified.json"
elif APP_IS_FORTIFIED_BUILD and APP_IS_ARCHIVE_BUILD and APP_IS_SECURE_BUILD and APP_IS_SIGNPOOL_BUILD:
    APP_SETTINGS_FILE = "app_settings_archive_secure_signpool_fortified.json"
elif APP_IS_SERVER_RELAY_BUILD and APP_IS_ARCHIVE_BUILD and APP_IS_SECURE_BUILD and APP_IS_SIGNPOOL_BUILD:
    APP_SETTINGS_FILE = "app_settings_server_relay_archive_secure_signpool.json"
elif APP_IS_SERVER_RELAY_BUILD and APP_IS_ARCHIVE_BUILD and APP_IS_SIGNPOOL_BUILD:
    APP_SETTINGS_FILE = "app_settings_server_relay_archive_signpool.json"
elif APP_IS_SERVER_RELAY_BUILD and APP_IS_SECURE_BUILD and APP_IS_SIGNPOOL_BUILD:
    APP_SETTINGS_FILE = "app_settings_server_relay_secure_signpool.json"
elif APP_IS_ARCHIVE_BUILD and APP_IS_SECURE_BUILD and APP_IS_SIGNPOOL_BUILD:
    APP_SETTINGS_FILE = "app_settings_archive_secure_signpool.json"
elif APP_IS_SERVER_RELAY_BUILD and APP_IS_SIGNPOOL_BUILD:
    APP_SETTINGS_FILE = "app_settings_server_relay_signpool.json"
elif APP_IS_ARCHIVE_BUILD and APP_IS_SIGNPOOL_BUILD:
    APP_SETTINGS_FILE = "app_settings_archive_signpool.json"
elif APP_IS_SIGNPOOL_BUILD:
    APP_SETTINGS_FILE = "app_settings_signpool.json"
elif APP_IS_SERVER_RELAY_BUILD and APP_IS_ARCHIVE_BUILD and APP_IS_SECURE_BUILD:
    APP_SETTINGS_FILE = "app_settings_server_relay_archive_secure.json"
elif APP_IS_SERVER_RELAY_BUILD and APP_IS_SECURE_BUILD:
    APP_SETTINGS_FILE = "app_settings_server_relay_secure.json"
elif APP_IS_ARCHIVE_BUILD and APP_IS_SECURE_BUILD:
    APP_SETTINGS_FILE = "app_settings_archive_secure.json"
elif APP_IS_SERVER_RELAY_BUILD and APP_IS_ARCHIVE_BUILD:
    APP_SETTINGS_FILE = "app_settings_server_relay_archive.json"
elif APP_IS_SERVER_RELAY_BUILD:
    APP_SETTINGS_FILE = "app_settings_server_relay.json"
elif APP_IS_ARCHIVE_BUILD:
    APP_SETTINGS_FILE = "app_settings_archive.json"
else:
    APP_SETTINGS_FILE = "app_settings.json"
APP_SETTINGS_VERSION = 3
CANVAS_PRINT_MARKER = "[[SEN_NAILS_CANVAS]]"
FONT_SCALE_MARKER_PREFIX = "[[SEN_FONT_SCALE="
CANVAS_PAYLOAD_MARKER_PREFIX = "[[SEN_CANVAS_PAYLOAD="
SETTINGS_SAVE_DEBOUNCE_MS = 450
POLL_MAX_BATCH = 120
POLL_MAX_SLICE_MS = 120
DEFAULT_SIGN_API_BASE = "https://tiktok.eulerstream.com"
FIXED_LICENSE_SERVER_URL = ""
FIXED_CLOUD_ADMIN_SERVER_URL = ""
SENSITIVE_SETTINGS_KEYS = ("sign_api_key", "cloud_admin_token", "license_machine_token")
LOCKED_LICENSE_SERVER_TEXT = "已锁定官方授权服务器"
LOCKED_ADMIN_SERVER_TEXT = "已锁定官方管理服务器"
PROXY_ROUTE_MODE_OPTIONS = [
    ("全部直连", "direct"),
    ("全部走代理", "all"),
    ("仅 TikTok 走代理", "tiktok_only"),
    ("仅 Sign API 走代理", "sign_only"),
]
PROXY_ROUTE_MODE_LABEL_TO_VALUE = {label: value for label, value in PROXY_ROUTE_MODE_OPTIONS}
PROXY_ROUTE_MODE_VALUE_TO_LABEL = {value: label for label, value in PROXY_ROUTE_MODE_OPTIONS}
LISTEN_SOURCE_MODE_OPTIONS = [
    ("本机直连", "local"),
    ("服务器中转", "relay"),
]
LISTEN_SOURCE_MODE_LABEL_TO_VALUE = {label: value for label, value in LISTEN_SOURCE_MODE_OPTIONS}
LISTEN_SOURCE_MODE_VALUE_TO_LABEL = {value: label for label, value in LISTEN_SOURCE_MODE_OPTIONS}
DESIGNER_MIN_FONT_SIZE = 2
DESIGNER_MAX_FONT_SIZE = 96
DESIGNER_FONT_SIZE_CHOICES = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 16, 18, 20, 24, 28, 32, 36, 40, 48, 56, 64, 72]

# Print template presets
PRESET_TEMPLATES = {
    "简洁型": {
        "template": "ID:{permanent_id}\n{content}",
        "font_size": 1,
        "alignment": "left",
        "line_spacing": 1,
        "paper_width": 40,
    },
    "标准型": {
        "template": "ID:{permanent_id}\nName:{name}\n{content}",
        "font_size": 1,
        "alignment": "left",
        "line_spacing": 1,
        "paper_width": 40,
    },
    "完整型": {
        "template": "ID:{permanent_id}\nUID:{unique_id}\nName:{name}\nTime:{time}\n{content}",
        "font_size": 1,
        "alignment": "left",
        "line_spacing": 1,
        "paper_width": 40,
    },
    "美化型": {
        "template": "═══════════════\nID:{permanent_id}\n名称:{name}\n时间:{time}\n───────────────\n{content}\n═══════════════",
        "font_size": 1,
        "alignment": "center",
        "line_spacing": 1,
        "paper_width": 40,
    },
}

I18N_TEXTS = {
    "zh": {
        "title": "Sen Nails 专用",
        "lang": "语言",
        "custom_name": "自定义主播名(显示用):",
        "live_url": "直播网址(抖音/TikTok):",
        "proxy_enable": "启用代理",
        "proxy_label": "代理(手填优先, 否则系统代理):",
        "ssl_insecure": "SSL不校验",
        "start_listen": "开始监听",
        "stop_listen": "停止监听",
        "save_config": "保存配置",
        "rule_center": "规则中心",
        "admin_mode": "管理员(免密)",
        "emergency_switch": "应急切换",
        "template_editor": "打印模板编辑器",
        "canvas_designer": "画布模板设计器",
        "var_docs": "变量说明",
        "canvas_enable": "启用画布模板",
    },
    "en": {
        "title": "Sen Nails Pro",
        "lang": "Language",
        "custom_name": "Host Name (display):",
        "live_url": "Live URL (Douyin/TikTok):",
        "proxy_enable": "Enable Proxy",
        "proxy_label": "Proxy (manual first, else system):",
        "ssl_insecure": "Disable SSL Verify",
        "start_listen": "Start",
        "stop_listen": "Stop",
        "save_config": "Save Config",
        "rule_center": "Rules",
        "admin_mode": "Admin (No Password)",
        "emergency_switch": "Emergency",
        "template_editor": "Template Editor",
        "canvas_designer": "Canvas Designer",
        "var_docs": "Variable Guide",
        "canvas_enable": "Enable Canvas Template",
    },
    "ja": {
        "title": "Sen Nails 専用",
        "lang": "言語",
        "custom_name": "配信者名(表示用):",
        "live_url": "配信URL(抖音/TikTok):",
        "proxy_enable": "プロキシ有効",
        "proxy_label": "プロキシ (手動優先、次にシステム):",
        "ssl_insecure": "SSL検証なし",
        "start_listen": "監視開始",
        "stop_listen": "監視停止",
        "save_config": "設定保存",
        "rule_center": "ルール",
        "admin_mode": "管理者(パスワード不要)",
        "emergency_switch": "緊急切替",
        "template_editor": "テンプレート編集",
        "canvas_designer": "キャンバス設計",
        "var_docs": "変数ガイド",
        "canvas_enable": "キャンバステンプレート有効",
    },
}

VAR_DOCS = [
    ("permanent_id", "永久编号", "Permanent ID", "永久ID"),
    ("unique_id", "客户唯一ID", "Customer Unique ID", "顧客固有ID"),
    ("name", "客户昵称", "Customer Name", "顧客名"),
    ("time", "弹幕时间", "Comment Time", "コメント時刻"),
    ("content", "弹幕内容", "Comment Content", "コメント内容"),
    ("custom_name", "主播名称(你设置的)", "Custom Host Name", "配信者名(設定値)"),
    ("room_url", "直播间地址", "Live Room URL", "配信URL"),
    ("app_version", "软件版本号", "App Version", "アプリバージョン"),
    ("today_date", "今天日期", "Today's Date", "本日の日付"),
    ("now_hms", "当前时分秒", "Current Time (H:M:S)", "現在時刻(H:M:S)"),
    ("weekday", "星期几", "Weekday", "曜日"),
    ("msg_length", "弹幕字数", "Message Length", "メッセージ文字数"),
    ("is_numeric", "是否纯数字(1/0)", "Is Numeric (1/0)", "数字のみか(1/0)"),
    ("guest_msg_count", "该客人今天第几条", "Guest Message Count Today", "本日の投稿回数"),
    ("today_guest_rank", "今天第几个客人", "Guest Rank Today", "本日の来訪順"),
    ("source_room", "来源直播间", "Source Room", "配信元ルーム"),
    ("lock_tag", "锁单状态标签", "Lock Status Tag", "ロック状態タグ"),
    ("name_wrapped", "自动换行后的昵称", "Wrapped Name", "折り返し済み名前"),
    ("content_wrapped", "自动换行后的弹幕", "Wrapped Content", "折り返し済み内容"),
]

# Canvas designer unit mapping: store/render in internal units, edit as mm.
CANVAS_UNITS_PER_MM = 5.0
FONT_FAMILY_CHOICES = [
    "TkDefaultFont",
    "Arial",
    "Helvetica",
    "Times New Roman",
    "Courier New",
    "PingFang SC",
    "Hiragino Sans GB",
    "Microsoft YaHei",
    "SimHei",
]


class SafeDict(dict):
    def __missing__(self, key):
        return "{" + str(key) + "}"


def _char_cells(ch: str) -> int:
    try:
        return 2 if unicodedata.east_asian_width(str(ch)) in ("W", "F") else 1
    except Exception:
        return 1


def _text_cells(s: str) -> int:
    return sum(_char_cells(ch) for ch in str(s or ""))


def _slice_by_cells(s: str, max_cells: int) -> tuple[str, str]:
    text = str(s or "")
    if max_cells <= 0:
        return "", text
    used = 0
    out = []
    for idx, ch in enumerate(text):
        cw = _char_cells(ch)
        if used + cw > max_cells:
            return "".join(out), text[idx:]
        out.append(ch)
        used += cw
    return "".join(out), ""


def _wrap_by_cells(s: str, width_cells: int) -> list[str]:
    text = str(s or "")
    width = max(1, int(width_cells))
    if not text:
        return [""]
    lines = []
    rest = text
    while rest:
        part, rest = _slice_by_cells(rest, width)
        if part == "" and rest:
            part, rest = rest[0], rest[1:]
        lines.append(part)
    return lines or [""]


def format_print_output(
    content: str,
    font_size: int = 1,
    alignment: str = "left",
    line_spacing: int = 1,
    paper_width: int = 40,
    margin_top: int = 0,
    margin_bottom: int = 0,
    margin_left: int = 0,
    margin_right: int = 0,
    letter_spacing: int = 0,
    paragraph_spacing: int = 0,
) -> str:
    """Format text for thermal printer with spacing and margin controls.

    font_size: global printer font scale hint; does not duplicate lines
    alignment: "left", "center", "right"
    line_spacing: blank lines between wrapped lines
    paper_width: character width of full line
    """
    width = max(20, int(paper_width))
    ml = max(0, int(margin_left))
    mr = max(0, int(margin_right))
    max_width = max(8, width - ml - mr)
    ls = max(1, int(line_spacing))
    ps = max(0, int(paragraph_spacing))
    mt = max(0, int(margin_top))
    mb = max(0, int(margin_bottom))
    try:
        gap = int(letter_spacing)
    except Exception:
        gap = 0

    try:
        fs_level = max(1, int(font_size))
    except Exception:
        fs_level = 1

    def _scale_text(line: str) -> str:
        # Keep letter spacing controlled only by `letter_spacing`.
        s = str(line or "")
        if gap > 0:
            s = (" " * gap).join(list(s))
        return s

    rendered = []
    source_lines = str(content or "").split("\n")
    for line_idx, raw in enumerate(source_lines):
        expanded = _scale_text(raw)
        wrapped = _wrap_by_cells(expanded, max_width)
        for wi, part in enumerate(wrapped):
            clen = _text_cells(part)
            if alignment == "center":
                pad = max(0, (max_width - clen) // 2)
            elif alignment == "right":
                pad = max(0, max_width - clen)
            else:
                pad = 0
            base_line = (" " * ml) + (" " * pad) + part
            rendered.append(base_line)
            if wi < len(wrapped) - 1:
                for _ in range(ls - 1):
                    rendered.append(" " * ml)
        if line_idx < len(source_lines) - 1:
            for _ in range(ls - 1):
                rendered.append(" " * ml)
            for _ in range(ps):
                rendered.append(" " * ml)

    output = ([""] * mt) + rendered + ([""] * mb)
    return "\n".join(output)


def get_app_version() -> str:
    try:
        out = subprocess.check_output(["git", "rev-parse", "--short", "HEAD"], text=True, stderr=subprocess.DEVNULL).strip()
        return out or "dev"
    except Exception:
        return "dev"


def _safe_float(value, default: float) -> float:
    try:
        return float(str(value).strip())
    except Exception:
        return float(default)


class App:
    def _t(self, key: str) -> str:
        lang = getattr(self, "ui_language", "zh")
        return I18N_TEXTS.get(lang, I18N_TEXTS["zh"]).get(key, key)

    def _set_title(self):
        self.root.title(f"{self._t('title')}  v{self.app_version}")

    def _apply_i18n(self):
        self._set_title()
        if hasattr(self, "_i18n_widgets"):
            for w, key in self._i18n_widgets:
                try:
                    w.config(text=self._t(key))
                except Exception:
                    pass

    def _on_language_change(self, event=None):
        # locked to Chinese in this build
        self.ui_language = "zh"
        self._save_settings()
        self._apply_i18n()

    def __init__(self, root):
        self.root = root
        self.app_version = get_app_version()
        self.settings = self._load_settings()
        self._save_settings_after_id = None
        self._debounce_jobs = {}
        self.ui_language = "zh"
        self._i18n_widgets = []
        self._set_title()
        try:
            geo = str(self.settings.get("main_window_geometry", "")).strip()
            if geo:
                self.root.geometry(geo)
        except Exception:
            pass

        # Top controls
        frm = ttk.Frame(root)
        frm.pack(fill=tk.X, padx=8, pady=6)

        self.lbl_custom_name = ttk.Label(frm, text=self._t("custom_name"))
        self.lbl_custom_name.grid(row=0, column=0)
        self._i18n_widgets.append((self.lbl_custom_name, "custom_name"))
        self.custom_name_var = tk.StringVar(value=self.settings.get("custom_name", "主播"))
        ttk.Entry(frm, textvariable=self.custom_name_var, width=20).grid(row=0, column=1)

        self.lbl_live_url = ttk.Label(frm, text=self._t("live_url"))
        self.lbl_live_url.grid(row=0, column=2)
        self._i18n_widgets.append((self.lbl_live_url, "live_url"))
        self.room_url_var = tk.StringVar(value=str(self.settings.get("room_url", "")))
        ttk.Entry(frm, textvariable=self.room_url_var, width=35).grid(row=0, column=3)
        saved_listen_mode = "local" if APP_IS_PUBLIC_BUILD else ("relay" if APP_IS_SIGNPOOL_RELAY_BUILD else str(self.settings.get("listen_source_mode", "local")).strip().lower())
        if saved_listen_mode not in LISTEN_SOURCE_MODE_VALUE_TO_LABEL:
            saved_listen_mode = "relay" if APP_IS_SIGNPOOL_RELAY_BUILD else "local"
        ttk.Label(frm, text="监听模式:").grid(row=0, column=8, sticky=tk.E)
        self.listen_source_mode_var = tk.StringVar(value=LISTEN_SOURCE_MODE_VALUE_TO_LABEL.get(saved_listen_mode, "本机直连"))
        if APP_IS_PUBLIC_BUILD:
            ttk.Label(frm, text="本机直连（公开版）").grid(row=0, column=9, sticky=tk.W, padx=(4, 0))
            self.listen_source_mode_var.set(LISTEN_SOURCE_MODE_VALUE_TO_LABEL["local"])
        elif APP_IS_SIGNPOOL_RELAY_BUILD:
            ttk.Label(frm, text="服务器中转（已锁定）").grid(row=0, column=9, sticky=tk.W, padx=(4, 0))
            self.listen_source_mode_var.set(LISTEN_SOURCE_MODE_VALUE_TO_LABEL["relay"])
        else:
            self.listen_source_mode_cb = ttk.Combobox(
                frm,
                textvariable=self.listen_source_mode_var,
                values=[label for label, _value in LISTEN_SOURCE_MODE_OPTIONS],
                width=12,
                state="readonly",
            )
            self.listen_source_mode_cb.grid(row=0, column=9, sticky=tk.W, padx=(4, 0))

        # Proxy controls
        self.proxy_enabled = tk.IntVar(value=0 if APP_IS_SIGNPOOL_RELAY_BUILD else (1 if bool(self.settings.get("proxy_enabled", False)) else 0))
        self.chk_proxy_enable = ttk.Checkbutton(frm, text=self._t("proxy_enable"), variable=self.proxy_enabled, command=self._toggle_proxy)
        if not APP_IS_SIGNPOOL_RELAY_BUILD:
            self.chk_proxy_enable.grid(row=1, column=0)
            self._i18n_widgets.append((self.chk_proxy_enable, "proxy_enable"))
        self.lbl_proxy = ttk.Label(frm, text=self._t("proxy_label"))
        if not APP_IS_SIGNPOOL_RELAY_BUILD:
            self.lbl_proxy.grid(row=1, column=1)
            self._i18n_widgets.append((self.lbl_proxy, "proxy_label"))
        self.proxy_var = tk.StringVar(value="" if APP_IS_SIGNPOOL_RELAY_BUILD else str(self.settings.get("proxy", "")))
        self.proxy_entry = ttk.Entry(frm, textvariable=self.proxy_var, width=40)
        if not APP_IS_SIGNPOOL_RELAY_BUILD:
            self.proxy_entry.grid(row=1, column=2, columnspan=2)
        saved_proxy_mode = str(self.settings.get("proxy_route_mode", "")).strip().lower()
        if saved_proxy_mode not in PROXY_ROUTE_MODE_VALUE_TO_LABEL:
            saved_proxy_mode = "all" if bool(self.settings.get("proxy_enabled", False)) else "direct"
        if APP_IS_SIGNPOOL_RELAY_BUILD:
            saved_proxy_mode = "direct"
        self.proxy_route_mode_var = tk.StringVar(value=PROXY_ROUTE_MODE_VALUE_TO_LABEL.get(saved_proxy_mode, "全部直连"))
        if not APP_IS_SIGNPOOL_RELAY_BUILD:
            ttk.Label(frm, text="代理用途:").grid(row=3, column=7, sticky=tk.E)
        self.proxy_mode_cb = ttk.Combobox(
            frm,
            textvariable=self.proxy_route_mode_var,
            values=[label for label, _value in PROXY_ROUTE_MODE_OPTIONS],
            width=16,
            state="readonly",
        )
        if not APP_IS_SIGNPOOL_RELAY_BUILD:
            self.proxy_mode_cb.grid(row=3, column=8, sticky=tk.W)
        else:
            ttk.Label(frm, text="网络路由:").grid(row=1, column=0, sticky=tk.W)
            ttk.Label(frm, text="已由服务端统一管理").grid(row=1, column=1, columnspan=3, sticky=tk.W)
        # Stability-first default: disable SSL verification unless user explicitly turns it off.
        self.ssl_insecure_var = tk.IntVar(value=1 if bool(self.settings.get("ssl_insecure", True)) else 0)
        self.chk_ssl_insecure = ttk.Checkbutton(frm, text=self._t("ssl_insecure"), variable=self.ssl_insecure_var)
        self.chk_ssl_insecure.grid(row=1, column=4, padx=6)
        self._i18n_widgets.append((self.chk_ssl_insecure, "ssl_insecure"))

        self.connect_btn = ttk.Button(frm, text=self._t("start_listen"), command=self.start_listen)
        self.connect_btn.grid(row=0, column=4, padx=6)
        self._i18n_widgets.append((self.connect_btn, "start_listen"))
        self.stop_btn = ttk.Button(frm, text=self._t("stop_listen"), command=self.stop_listen)
        self.stop_btn.grid(row=0, column=5, padx=6)
        self._i18n_widgets.append((self.stop_btn, "stop_listen"))
        self.stop_btn.config(state=tk.DISABLED)
        self.save_conn_btn = ttk.Button(frm, text=self._t("save_config"), command=self.save_connection_settings)
        self.save_conn_btn.grid(row=0, column=7, padx=6)
        self._i18n_widgets.append((self.save_conn_btn, "save_config"))
        self.rule_btn = ttk.Button(frm, text=self._t("rule_center"), command=self.open_rule_center)
        self.rule_btn.grid(row=1, column=5, padx=6)
        self._i18n_widgets.append((self.rule_btn, "rule_center"))
        if not APP_IS_PUBLIC_BUILD:
            self.admin_btn = ttk.Button(frm, text=self._t("admin_mode"), command=self.toggle_admin_mode)
            self.admin_btn.grid(row=0, column=6, padx=6)
            self._i18n_widgets.append((self.admin_btn, "admin_mode"))
        self.emergency_mode_var = tk.StringVar(value="正常")
        self.emergency_cb = ttk.Combobox(frm, textvariable=self.emergency_mode_var, values=["正常", "仅记录不打印", "仅数字打印", "关闭锁单"], width=14, state="readonly")
        self.emergency_cb.grid(row=1, column=6, padx=6)
        self.emergency_btn = ttk.Button(frm, text=self._t("emergency_switch"), command=self.apply_emergency_mode)
        self.emergency_btn.grid(row=1, column=7, padx=6)
        self._i18n_widgets.append((self.emergency_btn, "emergency_switch"))
        self.license_server_var = tk.StringVar(value="" if APP_IS_PUBLIC_BUILD else FIXED_LICENSE_SERVER_URL)
        self.license_key_var = tk.StringVar(value="" if APP_IS_PUBLIC_BUILD else str(self.settings.get("license_key", "")))
        if APP_IS_PUBLIC_BUILD:
            ttk.Label(frm, text="公开版: 无需授权，可直接填写 Sign API 配置").grid(row=2, column=0, columnspan=5, sticky=tk.W)
        else:
            ttk.Label(frm, text="授权服务器:").grid(row=2, column=0, sticky=tk.W)
            ttk.Label(frm, text=LOCKED_LICENSE_SERVER_TEXT).grid(row=2, column=1, columnspan=2, sticky=tk.W)
            ttk.Label(frm, text="授权码:").grid(row=2, column=3, sticky=tk.E)
            ttk.Entry(frm, textvariable=self.license_key_var, width=20).grid(row=2, column=4, sticky=tk.W)
        self.sign_api_base_var = tk.StringVar(value=str(self.settings.get("sign_api_base", DEFAULT_SIGN_API_BASE)).strip() or DEFAULT_SIGN_API_BASE)
        self.sign_api_key_var = tk.StringVar(value=str(self.settings.get("sign_api_key", "")))
        self.use_sign_api_var = tk.IntVar(
            value=1
            if bool(self.settings.get("use_sign_api_key", bool(str(self.settings.get("sign_api_key", "")).strip())))
            else 0
        )
        if APP_IS_SIGNPOOL_RELAY_BUILD:
            ttk.Label(frm, text="签名配置:").grid(row=3, column=0, sticky=tk.W)
            ttk.Label(frm, text="已由服务端签名池统一管理").grid(row=3, column=1, columnspan=5, sticky=tk.W)
            self.sign_api_base_var.set("")
            self.sign_api_key_var.set("")
            self.use_sign_api_var.set(0)
        else:
            ttk.Label(frm, text="Sign API Base:").grid(row=3, column=0, sticky=tk.W)
            ttk.Entry(frm, textvariable=self.sign_api_base_var, width=28).grid(row=3, column=1, columnspan=2, sticky=tk.W)
            ttk.Label(frm, text="Sign API Key:").grid(row=3, column=3, sticky=tk.E)
            ttk.Entry(frm, textvariable=self.sign_api_key_var, width=26, show="*").grid(row=3, column=4, sticky=tk.W)
            ttk.Checkbutton(frm, text="使用 Sign API Key 连接", variable=self.use_sign_api_var, command=self._request_save_settings).grid(row=3, column=5, columnspan=2, padx=6, sticky=tk.W)
        self.license_btn = ttk.Button(frm, text=("配置说明" if APP_IS_PUBLIC_BUILD else "激活授权"), command=self.activate_license)
        self.license_btn.grid(row=2, column=5, padx=6)
        if not APP_IS_PUBLIC_BUILD:
            self.customer_front_btn = ttk.Button(frm, text="客户前端", command=self.open_customer_portal)
            self.customer_front_btn.grid(row=2, column=6, padx=6)
        self.license_state_var = tk.StringVar(value=("公开版: 无需授权" if APP_IS_PUBLIC_BUILD else "授权状态: 未校验"))
        ttk.Label(frm, textvariable=self.license_state_var).grid(row=2, column=7, columnspan=2, sticky=tk.W)
        self.status_var = tk.StringVar(value="状态: 未连接")
        ttk.Label(frm, textvariable=self.status_var).grid(row=4, column=0, columnspan=10, sticky=tk.W, pady=(4, 0))

        # Printers and sizes
        printer_frm = ttk.Frame(root)
        printer_frm.pack(fill=tk.X, padx=8)
        ttk.Button(printer_frm, text="检测打印机", command=self.detect_printers).grid(row=0, column=0)
        self.printer_cb = ttk.Combobox(printer_frm, values=[], width=25)
        self.printer_cb.grid(row=0, column=1)
        self.printer_cb.bind("<<ComboboxSelected>>", lambda _e: self._save_settings())
        self.use_default_printer_var = tk.IntVar(value=1 if bool(self.settings.get("use_default_printer", True)) else 0)
        ttk.Checkbutton(
            printer_frm,
            text="使用系统默认打印机",
            variable=self.use_default_printer_var,
            command=self._save_settings,
        ).grid(row=0, column=5, padx=6, sticky=tk.W)
        ttk.Button(printer_frm, text="测试打印", command=self.test_print_selected_printer).grid(row=0, column=4, padx=6)
        ttk.Button(printer_frm, text="打印校准向导", command=self.open_print_calibration_wizard).grid(row=0, column=8, padx=6, sticky=tk.W)
        # Keep presets for canvas tooling only; printing always uses custom mm below.
        self.paper_sizes = {
            "40×30mm": ("40", "30"),
            "30×40mm": ("30", "40"),
            "100×150mm": ("100", "150"),
            "100×100mm": ("100", "100"),
        }

        # Custom thermal size in mm (width x height)
        ttk.Label(printer_frm, text='自定义纸张(mm W×H):').grid(row=1, column=0)
        self.width_var = tk.StringVar(value=str(self.settings.get("custom_paper_width_mm", "40")))
        self.height_var = tk.StringVar(value=str(self.settings.get("custom_paper_height_mm", "30")))
        ttk.Entry(printer_frm, textvariable=self.width_var, width=6).grid(row=1, column=1)
        ttk.Label(printer_frm, text='×').grid(row=1, column=2)
        ttk.Entry(printer_frm, textvariable=self.height_var, width=6).grid(row=1, column=3)
        ttk.Label(printer_frm, text='校准(字宽/行高/边距 mm):').grid(row=1, column=4, sticky=tk.E, padx=(8, 2))
        self.char_width_mm_var = tk.StringVar(value=str(self.settings.get("char_width_mm", "1.50")))
        self.line_height_mm_var = tk.StringVar(value=str(self.settings.get("line_height_mm", "2.80")))
        self.margin_mm_var = tk.StringVar(value=str(self.settings.get("margin_mm", "1.00")))
        ttk.Entry(printer_frm, textvariable=self.char_width_mm_var, width=6).grid(row=1, column=5, sticky=tk.W)
        ttk.Entry(printer_frm, textvariable=self.line_height_mm_var, width=6).grid(row=1, column=6, sticky=tk.W, padx=(4, 0))
        ttk.Entry(printer_frm, textvariable=self.margin_mm_var, width=6).grid(row=1, column=7, sticky=tk.W, padx=(4, 0))
        ttk.Button(printer_frm, text="校准默认", command=self._reset_print_calibration).grid(row=1, column=8, padx=6, sticky=tk.W)

        # Stream display with Treeview (embedded, not separate window)
        disp_frm = ttk.Frame(root)
        disp_frm.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)
        cols = ('time', 'permanent_id', 'unique_id', 'display_name', 'message', 'guest_msg_count', 'today_guest_rank')
        self.stream_tree = ttk.Treeview(disp_frm, columns=cols, show='headings', selectmode='extended')
        self.stream_tree.heading('time', text='时间')
        self.stream_tree.heading('permanent_id', text='永久ID')
        self.stream_tree.heading('unique_id', text='客户ID')
        self.stream_tree.heading('display_name', text='昵称')
        self.stream_tree.heading('message', text='弹幕')
        self.stream_tree.heading('guest_msg_count', text='客人第几条')
        self.stream_tree.heading('today_guest_rank', text='今日第几位客人')
        self.stream_tree.column('time', width=120)
        self.stream_tree.column('permanent_id', width=80)
        self.stream_tree.column('unique_id', width=140)
        self.stream_tree.column('display_name', width=140)
        self.stream_tree.column('message', width=300)
        self.stream_tree.column('guest_msg_count', width=100)
        self.stream_tree.column('today_guest_rank', width=120)
        self.stream_scrollbar = ttk.Scrollbar(disp_frm, orient=tk.VERTICAL, command=self.stream_tree.yview)
        self.stream_hscrollbar = ttk.Scrollbar(disp_frm, orient=tk.HORIZONTAL, command=self.stream_tree.xview)
        self.stream_tree.configure(yscrollcommand=self.stream_scrollbar.set, xscrollcommand=self.stream_hscrollbar.set)
        self.stream_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.stream_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.stream_hscrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        # right-click context menu with enhanced options
        self.stream_menu = tk.Menu(root, tearoff=0)
        self.stream_menu.add_command(label='全选', command=self._select_all_stream)
        self.stream_menu.add_command(label='复制所选', command=self._copy_stream_sel)
        copy_field_menu = tk.Menu(self.stream_menu, tearoff=0)
        copy_field_menu.add_command(label='只复制弹幕', command=lambda: self._copy_stream_field(4, '弹幕'))
        copy_field_menu.add_command(label='只复制客户ID', command=lambda: self._copy_stream_field(2, '客户ID'))
        copy_field_menu.add_command(label='只复制永久ID', command=lambda: self._copy_stream_field(1, '永久ID'))
        copy_field_menu.add_command(label='只复制昵称', command=lambda: self._copy_stream_field(3, '昵称'))
        self.stream_menu.add_cascade(label='按字段复制', menu=copy_field_menu)
        self.stream_menu.add_command(label='粘贴到输入框', command=self._paste_to_input)
        self.stream_menu.add_separator()
        self.stream_menu.add_command(label='补充打印', command=self._reprint_selected)
        self.stream_menu.add_command(label='拉黑并释放ID', command=self._blacklist_selected)
        self.stream_menu.add_command(label='仅释放永久ID', command=self._release_selected)
        self.stream_menu.add_command(label='清空弹幕流水', command=self.clear_stream_rows)
        def stream_right_click(event):
            try:
                row_id = self.stream_tree.identify_row(event.y)
                if row_id:
                    self.stream_tree.selection_set(row_id)
                self.stream_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.stream_menu.grab_release()
        self.stream_tree.bind('<Button-3>', stream_right_click)
        self.stream_tree.bind('<Button-2>', stream_right_click)          # macOS two-finger click compatibility
        self.stream_tree.bind('<Control-Button-1>', stream_right_click)  # macOS control-click compatibility
        self.stream_tree.bind('<Control-a>', lambda e: self._select_all_stream())
        self.stream_tree.bind('<Control-c>', lambda e: self._copy_stream_sel())
        self.stream_tree.bind('<Control-v>', lambda e: self._paste_to_input())

        # Print template with preset selector
        tpl_frm = ttk.Frame(root)
        tpl_frm.pack(fill=tk.X, padx=8)
        ttk.Label(tpl_frm, text="打印模板:").pack(anchor=tk.W)
        tpl_ctrl = ttk.Frame(tpl_frm)
        tpl_ctrl.pack(fill=tk.X)
        ttk.Label(tpl_ctrl, text="预设:").pack(side=tk.LEFT)
        self.preset_cb = ttk.Combobox(tpl_ctrl, values=list(PRESET_TEMPLATES.keys()), width=12, state="readonly")
        self.preset_cb.pack(side=tk.LEFT, padx=4)
        self.preset_cb.bind("<<ComboboxSelected>>", lambda e: self._load_preset())
        self.btn_tpl_editor = ttk.Button(tpl_ctrl, text=self._t("template_editor"), command=self.open_template_editor)
        self.btn_tpl_editor.pack(side=tk.LEFT)
        self._i18n_widgets.append((self.btn_tpl_editor, "template_editor"))
        self.btn_canvas_designer = ttk.Button(tpl_ctrl, text=self._t("canvas_designer"), command=self.open_canvas_template_designer)
        self.btn_canvas_designer.pack(side=tk.LEFT, padx=4)
        self._i18n_widgets.append((self.btn_canvas_designer, "canvas_designer"))
        ttk.Label(tpl_ctrl, text="模板模式:").pack(side=tk.LEFT, padx=(8, 2))
        self.template = tk.StringVar(value=PRESET_TEMPLATES["完整型"]["template"])
        self.font_size_var = tk.IntVar(value=1)
        self.alignment_var = tk.StringVar(value="left")
        self.line_spacing_var = tk.IntVar(value=1)
        self.paper_width_var = tk.IntVar(value=40)
        saved_mode = str(self.settings.get("template_mode", "")).strip().lower()
        if saved_mode not in ("editor", "designer"):
            saved_mode = "designer" if bool(self.settings.get("canvas_template_enabled", False)) else "editor"
        self.template_mode_var = tk.StringVar(value=saved_mode)
        self.canvas_template_enabled = tk.IntVar(value=1 if saved_mode == "designer" else 0)
        ttk.Radiobutton(tpl_ctrl, text="编辑器", value="editor", variable=self.template_mode_var, command=self._on_template_mode_change).pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(tpl_ctrl, text="设计器", value="designer", variable=self.template_mode_var, command=self._on_template_mode_change).pack(side=tk.LEFT, padx=2)
        self.editor_margin_top_var = tk.IntVar(value=int(self.settings.get("editor_margin_top", 0)))
        self.editor_margin_bottom_var = tk.IntVar(value=int(self.settings.get("editor_margin_bottom", 0)))
        self.editor_margin_left_var = tk.IntVar(value=int(self.settings.get("editor_margin_left", 0)))
        self.editor_margin_right_var = tk.IntVar(value=int(self.settings.get("editor_margin_right", 0)))
        self.editor_letter_spacing_var = tk.IntVar(value=int(self.settings.get("editor_letter_spacing", 0)))
        self.editor_paragraph_spacing_var = tk.IntVar(value=int(self.settings.get("editor_paragraph_spacing", 0)))
        self.editor_font_family_var = tk.StringVar(value=str(self.settings.get("editor_font_family", "TkDefaultFont")))
        self.editor_var_font_scale_map = self.settings.get("editor_var_font_scale_map", {}) if isinstance(self.settings.get("editor_var_font_scale_map", {}), dict) else {}

        # Blacklist management
        bl_frm = ttk.Frame(root)
        bl_frm.pack(fill=tk.X, padx=8, pady=6)
        ttk.Label(bl_frm, text="拉黑用户(unique_id):").grid(row=0, column=0)
        self.bl_var = tk.StringVar()
        ttk.Entry(bl_frm, textvariable=self.bl_var, width=30).grid(row=0, column=1)
        ttk.Button(bl_frm, text="添加", command=self.add_blacklist).grid(row=0, column=2)
        ttk.Button(bl_frm, text="移除", command=self.remove_blacklist).grid(row=0, column=3)
        ttk.Button(bl_frm, text="显示拉黑列表", command=self.show_blacklist).grid(row=0, column=4)
        ttk.Button(bl_frm, text="黑名单管理", command=self.open_blacklist_manager).grid(row=0, column=5)

        # Manager windows
        mgr_frm = ttk.Frame(root)
        mgr_frm.pack(fill=tk.X, padx=8, pady=6)
        ttk.Button(mgr_frm, text="永久编号管理", command=self.open_pid_manager).grid(row=0, column=0)
        ttk.Button(mgr_frm, text="跨房分析系统", command=self.open_overlap_analyzer).grid(row=0, column=1)
        if not APP_IS_PUBLIC_BUILD:
            ttk.Button(mgr_frm, text="云数据中心", command=self.open_cloud_center).grid(row=0, column=2)
        ttk.Button(mgr_frm, text="排行榜", command=self.open_leaderboard).grid(row=0, column=3)
        ttk.Button(mgr_frm, text="录制管理", command=self.open_recording_manager).grid(row=0, column=4)
        ttk.Button(mgr_frm, text="性能看板", command=self.open_performance_dashboard).grid(row=0, column=5)
        ttk.Button(mgr_frm, text="健康检查", command=self.open_health_check_panel).grid(row=0, column=6)
        ttk.Button(mgr_frm, text="客户动销分析", command=self.open_growth_analytics_center).grid(row=0, column=7)
        ttk.Button(mgr_frm, text="经营增强中心", command=self.open_business_ops_center).grid(row=0, column=8)
        ttk.Button(mgr_frm, text="开播前检查", command=self.open_startup_checklist).grid(row=0, column=9)
        ttk.Button(mgr_frm, text="飞书与Web看板", command=self.open_integration_center).grid(row=0, column=10)

        # Export buttons
        export_frm = ttk.Frame(root)
        export_frm.pack(fill=tk.X, padx=8, pady=6)
        ttk.Button(export_frm, text="导出本次弹幕", command=self.export_comments).grid(row=0, column=0)
        ttk.Button(export_frm, text="导出永久编号", command=self.export_permanent_ids).grid(row=0, column=1)
        ttk.Button(export_frm, text="导入永久编号", command=self.import_permanent_ids).grid(row=0, column=2)
        ttk.Button(export_frm, text="软件打印任务", command=self.open_job_manager).grid(row=0, column=3)
        ttk.Button(export_frm, text="系统打印队列", command=self.open_system_print_queue).grid(row=0, column=13)
        ttk.Button(export_frm, text="打印预览", command=self.print_preview_selected).grid(row=0, column=4)
        ttk.Button(export_frm, text="今日统计", command=self.open_today_report).grid(row=0, column=5)
        ttk.Button(export_frm, text="一键备份", command=self.backup_project_state).grid(row=0, column=6)
        ttk.Button(export_frm, text="恢复备份", command=self.restore_project_state).grid(row=0, column=7)
        ttk.Button(export_frm, text="清空弹幕流水", command=self.clear_stream_rows).grid(row=0, column=8)
        ttk.Button(export_frm, text="释放内存", command=self.release_memory).grid(row=0, column=9)
        ttk.Button(export_frm, text="弹幕回放", command=self.open_replay_center).grid(row=0, column=10)
        ttk.Button(export_frm, text="模板版本", command=self.open_template_version_manager).grid(row=0, column=11)
        ttk.Button(export_frm, text="操作日志", command=self.open_audit_log_viewer).grid(row=0, column=12)
        ttk.Button(export_frm, text="导出日报包", command=lambda: self.export_standard_report_pack("daily")).grid(row=0, column=14)
        ttk.Button(export_frm, text="导出周报包", command=lambda: self.export_standard_report_pack("weekly")).grid(row=0, column=15)
        ttk.Button(export_frm, text="导出月报包", command=lambda: self.export_standard_report_pack("monthly")).grid(row=0, column=16)
        ttk.Button(export_frm, text="一键诊断包", command=self.export_diagnostic_pack).grid(row=0, column=17)
        ttk.Button(export_frm, text="模板回归检查", command=self.run_template_regression_check).grid(row=0, column=18)

        # Bottom software status bar
        status_frm = ttk.Frame(root)
        status_frm.pack(fill=tk.X, side=tk.BOTTOM, padx=8, pady=(0, 6))
        self.soft_status_var = tk.StringVar(value="软件状态: 初始化中")
        ttk.Label(status_frm, textvariable=self.soft_status_var, anchor=tk.W).pack(fill=tk.X)

        # Internal
        self.queue = queue.Queue()
        self.reconnect_print_window_seconds = int(self.settings.get("reconnect_print_window_seconds", 20))
        self.print_batch_window_seconds = int(self.settings.get("print_batch_window_seconds", 5))
        self.print_batch_limit = int(self.settings.get("print_batch_limit", 200))
        self.print_worker_count = int(self.settings.get("print_worker_count", 50))
        self.print_retry_limit = int(self.settings.get("print_retry_limit", 3))
        self.auto_wrap_print_enabled = bool(self.settings.get("auto_wrap_print_enabled", True))
        self.auto_wrap_name_width = int(self.settings.get("auto_wrap_name_width", 14))
        self.auto_wrap_content_width = int(self.settings.get("auto_wrap_content_width", 22))
        self.max_stream_rows = int(self.settings.get("max_stream_rows", 3000))
        self.auto_archive_stream = bool(self.settings.get("auto_archive_stream", True))
        self.auto_print_numeric = bool(self.settings.get("auto_print_numeric", True))
        self.keyword_print_enabled = bool(self.settings.get("keyword_print_enabled", False))
        self.keyword_print_list = [x.strip() for x in str(self.settings.get("keyword_print_list", "")).split(",") if x.strip()]
        self.print_min_len = int(self.settings.get("print_min_len", 1))
        self.print_max_len = int(self.settings.get("print_max_len", 100))
        self.lock_order_mode = bool(self.settings.get("lock_order_mode", False))
        self.lock_order_window_seconds = int(self.settings.get("lock_order_window_seconds", 5))
        self._lock_order_open_until = 0.0
        self._lock_order_winner_key = None
        self.today_date = datetime.now().date().isoformat()
        self.today_guest_rank = {}  # {unique_id: rank_in_today}
        self.today_guest_counter = 0
        self.guest_message_counter = {}  # {unique_id: message_count}
        self.overlap_events = deque(maxlen=int(self.settings.get("overlap_max_events", 300000)))  # (ts, room_id, unique_id)
        self.overlap_message_events = deque(maxlen=int(self.settings.get("overlap_message_max_events", 300000)))  # (ts, room_id, unique_id, name, content)
        self.audit_log_path = os.path.join(os.getcwd(), "audit.log")
        self.event_store_path = os.path.join(os.getcwd(), "overlap_events.jsonl")
        self.business_event_log_path = os.path.join(os.getcwd(), "business_events.jsonl")
        self.data_retention_days = int(self.settings.get("data_retention_days", 30))
        self.auto_data_maintenance_enabled = bool(self.settings.get("auto_data_maintenance_enabled", True))
        self.overlap_default_hours = int(self.settings.get("overlap_default_hours", 24))
        self.overlap_include_main = bool(self.settings.get("overlap_include_main", True))
        self.sales_goals = self.settings.get("sales_goals", {
            "effective_guests_per_day": 200,
            "conversion_keyword_hits_per_day": 80,
            "print_success_rate": 0.95,
        })
        self.customer_tags = self.settings.get("customer_tags", {})  # {uid: "tag1,tag2"}
        self.followup_queue = self.settings.get("followup_queue", [])  # list[dict]
        self.campaign_calendar = self.settings.get("campaign_calendar", [])  # list[dict]
        self.inventory_map = self.settings.get("inventory_map", {})  # {keyword: stock}
        self.host_scores = self.settings.get("host_scores", {})  # {host: {"score":x, "samples":n}}
        self.team_roles = self.settings.get("team_roles", {"owner": "", "operator": "", "support": ""})
        self.risk_rules = self.settings.get("risk_rules", {"spam_per_10s": 25, "duplicate_text_hits": 8})
        saved_watch_rooms = self.settings.get("overlap_watch_rooms", [])
        if isinstance(saved_watch_rooms, str):
            saved_watch_rooms = [x.strip() for x in saved_watch_rooms.split(",") if x.strip()]
        self.overlap_watch_rooms = [str(x).strip() for x in (saved_watch_rooms or []) if str(x).strip()]
        self.auto_record_enabled = bool(self.settings.get("auto_record_enabled", False))
        self.auto_record_dir = str(self.settings.get("auto_record_dir", os.path.join(os.getcwd(), "recordings"))).strip()
        self.auto_record_cmd = str(self.settings.get("auto_record_cmd", "streamlink")).strip() or "streamlink"
        self.record_workers = {}  # {uid: {"stop": Event, "thread": Thread, "proc": Popen|None}}
        self._last_overlap_refresh_ts = 0.0
        self.per_printer_limit = int(self.settings.get("per_printer_limit", 1))
        self.selected_printer_cached = str(self.settings.get("selected_printer", "")).strip()
        self.use_default_printer_cached = bool(self.settings.get("use_default_printer", True))
        self.queue_alert_threshold = int(self.settings.get("queue_alert_threshold", 200))
        self.fail_alert_threshold = int(self.settings.get("fail_alert_threshold", 20))
        self.peak_warn_threshold = int(self.settings.get("peak_warn_threshold", max(120, self.queue_alert_threshold)))
        self.peak_critical_threshold = int(self.settings.get("peak_critical_threshold", max(self.peak_warn_threshold + 80, self.queue_alert_threshold + 120)))
        self.peak_duplicate_window_seconds = int(self.settings.get("peak_duplicate_window_seconds", 8))
        self.peak_mode = "normal"
        self.peak_drop_counter = 0
        self.peak_merge_counter = 0
        self._last_peak_eval_ts = 0.0
        self._last_peak_pending = 0
        self.blacklist_keywords = [x.strip() for x in str(self.settings.get("blacklist_keywords", "")).split(",") if x.strip()]
        self.whitelist_keywords = [x.strip() for x in str(self.settings.get("whitelist_keywords", "")).split(",") if x.strip()]
        self.auto_blacklist_rate_limit = int(self.settings.get("auto_blacklist_rate_limit", 25))
        self.is_admin_mode = False
        self._admin_unlock_until = 0.0
        self._admin_session_seconds = 120 if APP_IS_FORTIFIED_BUILD else 300
        stored_admin_password_hash = str(self.settings.get("admin_password_hash", "")).strip()
        legacy_admin_password = str(self.settings.get("admin_password", "")).strip()
        if security_utils.is_password_hash(stored_admin_password_hash):
            self.admin_password_hash = stored_admin_password_hash
        elif legacy_admin_password:
            self.admin_password_hash = security_utils.hash_password(legacy_admin_password)
        elif APP_IS_SECURE_BUILD:
            self.admin_password_hash = ""
        else:
            self.admin_password_hash = security_utils.hash_password("8888")
        self.user_msg_timestamps = defaultdict(deque)
        self.recent_print_keys = {}
        self._pid_users_cache = []
        self._pid_cache_ts = 0.0
        self._sign_rate_limited_until = {}
        self._last_alert_ts = 0.0
        self.license_active = bool(APP_IS_PUBLIC_BUILD)
        self.machine_id = license_client.machine_fingerprint()
        self.license_machine_token = str(self.settings.get("license_machine_token", "")).strip()
        self.local_bypass_enabled = False
        self.local_bypass_machine_id = ""
        self.license_last_ok_ts = int(self.settings.get("license_last_ok_ts", 0))
        self.license_grace_hours = 0 if (APP_IS_SECURE_BUILD or APP_IS_PUBLIC_BUILD) else int(self.settings.get("license_grace_hours", 24))
        self.remote_pid_sync_enabled = (False if APP_IS_PUBLIC_BUILD else bool(self.settings.get("remote_pid_sync_enabled", True)))
        self.cloud_upload_enabled = (False if APP_IS_PUBLIC_BUILD else bool(self.settings.get("cloud_upload_enabled", False)))
        self.cloud_client_server = "" if APP_IS_PUBLIC_BUILD else FIXED_LICENSE_SERVER_URL
        self.cloud_admin_server = "" if APP_IS_PUBLIC_BUILD else FIXED_CLOUD_ADMIN_SERVER_URL
        self.cloud_admin_token = ("" if APP_IS_PUBLIC_BUILD else str(self.settings.get("cloud_admin_token", "")).strip())
        self.canvas_template = self._normalize_canvas_template_to_paper(self.settings.get("canvas_template", {
            "canvas_w": 420,
            "canvas_h": 260,
            "elements": [
                {"field": "permanent_id", "label": "ID:{permanent_id}", "x": 12, "y": 12},
                {"field": "name", "label": "Name:{name}", "x": 12, "y": 38},
                {"field": "time", "label": "Time:{time}", "x": 12, "y": 64},
                {"field": "content", "label": "{content}", "x": 12, "y": 100},
            ],
        }))
        self.template_custom_vars = self.settings.get("template_custom_vars", {})
        self.cloud_event_buffer = deque(maxlen=5000)
        self._cloud_uploading = False
        self._cloud_fail_streak = 0
        self._cloud_last_error = ""
        self._cloud_last_ok_ts = 0.0
        self._cloud_next_delay_sec = 5
        self.cloud_queue_path = os.path.join(RUNTIME_DIR, "cloud_queue.jsonl")
        self.dashboard_recent_comments = deque(maxlen=int(self.settings.get("dashboard_recent_comments_limit", 1500)))
        self.dashboard_port = int(self.settings.get("dashboard_port", 8787))
        self.dashboard_autostart = bool(self.settings.get("dashboard_autostart", False))
        self._dashboard_httpd = None
        self._dashboard_thread = None
        self._dashboard_server_started_ts = 0.0
        self.feishu_enabled = bool(self.settings.get("feishu_enabled", False))
        self.feishu_webhook = str(self.settings.get("feishu_webhook", "")).strip()
        self.feishu_push_print_only = bool(self.settings.get("feishu_push_print_only", True))
        self.feishu_keyword_filter = [x.strip() for x in str(self.settings.get("feishu_keyword_filter", "")).split(",") if x.strip()]
        self.feishu_max_per_minute = int(self.settings.get("feishu_max_per_minute", 25))
        self.feishu_push_queue = queue.Queue(maxsize=1200)
        self._feishu_sent_timestamps = deque(maxlen=240)
        self._feishu_thread = threading.Thread(target=self._feishu_worker_loop, daemon=True)
        self._feishu_thread.start()
        self.listener_workers = {}
        self.analysis_listener_workers = {}
        self.printer_locks = defaultdict(lambda: threading.Semaphore(max(1, self.per_printer_limit)))
        # User cache to avoid repeated DB queries
        self.user_cache = {}  # {unique_id: (permanent_id, name)}
        # Printing core: dispatcher claims time-window batches, workers print concurrently.
        self.print_task_queue = queue.Queue()
        self.print_dispatch_thread = threading.Thread(target=self._print_dispatcher, daemon=True)
        self.print_dispatch_thread.start()
        self.print_workers = []
        for _ in range(self.print_worker_count):
            t = threading.Thread(target=self._print_worker, daemon=True)
            t.start()
            self.print_workers.append(t)

        # Control for reconnect loop
        self._stop_event = threading.Event()
        self._analysis_stop_event = threading.Event()
        self.client = None
        self.listen_thread = None
        self.listening = False
        db.init_db()
        self._setup_input_shortcuts()
        self.custom_name_var.trace_add("write", lambda *_: self._request_save_settings())
        self.room_url_var.trace_add("write", lambda *_: self._request_save_settings())
        self.listen_source_mode_var.trace_add("write", lambda *_: self._request_save_settings())
        self.proxy_var.trace_add("write", lambda *_: self._request_save_settings())
        self.proxy_enabled.trace_add("write", lambda *_: self._request_save_settings())
        self.proxy_route_mode_var.trace_add("write", lambda *_: self._on_proxy_route_mode_change())
        self.ssl_insecure_var.trace_add("write", lambda *_: self._request_save_settings())
        self.width_var.trace_add("write", lambda *_: self._request_save_settings())
        self.height_var.trace_add("write", lambda *_: self._request_save_settings())
        self.char_width_mm_var.trace_add("write", lambda *_: self._request_save_settings())
        self.line_height_mm_var.trace_add("write", lambda *_: self._request_save_settings())
        self.margin_mm_var.trace_add("write", lambda *_: self._request_save_settings())
        self.license_server_var.trace_add("write", lambda *_: self._request_save_settings())
        self.license_key_var.trace_add("write", lambda *_: self._request_save_settings())
        self.sign_api_base_var.trace_add("write", lambda *_: self._request_save_settings())
        self.sign_api_key_var.trace_add("write", lambda *_: self._request_save_settings())
        self.use_sign_api_var.trace_add("write", lambda *_: self._request_save_settings())
        try:
            if not tracemalloc.is_tracing():
                tracemalloc.start()
        except Exception:
            pass
        self._toggle_proxy()
        self._check_license_on_startup()
        self._apply_i18n()
        try:
            self._update_admin_button_state()
        except Exception:
            pass
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # Start polling queue (increased interval to reduce CPU usage)
        self.root.after(200, self._poll)
        self.root.after(5000, self._health_check)
        self.root.after(1200, self._refresh_soft_status)
        self.root.after(120000, self._license_heartbeat_loop)
        self.root.after(5000, self._cloud_upload_tick)
        self.root.after(1000, self._load_cloud_queue_from_disk)
        self._load_overlap_events_from_disk()
        self.root.after(18000, self._auto_data_maintenance)
        self.root.after(6 * 60 * 60 * 1000, self._db_maintenance_tick)
        # Try restore printers on startup for auto-print stability (especially on Windows).
        self.detect_printers(silent=True)
        self.root.after(6000, self._printer_health_probe)
        # Keep a local CSV backup of permanent IDs to avoid accidental loss.
        self._sync_local_permanent_ids_backup("startup")
        self.root.after(120000, self._auto_backup_permanent_ids)
        if self.dashboard_autostart:
            try:
                self._start_web_dashboard_server(int(self.dashboard_port))
            except Exception:
                pass

    def _settings_path(self):
        path = os.path.join(RUNTIME_DIR, APP_SETTINGS_FILE)
        legacy = os.path.join(os.getcwd(), APP_SETTINGS_FILE)
        if (not os.path.exists(path)) and os.path.exists(legacy):
            try:
                shutil.copy2(legacy, path)
            except Exception:
                pass
        return path

    def _decode_settings_payload(self, data: dict) -> dict:
        payload = dict(data or {})
        for key in SENSITIVE_SETTINGS_KEYS:
            if key in payload:
                payload[key] = security_utils.unprotect_secret(payload.get(key, ""))
        return payload

    def _encode_settings_payload(self, data: dict) -> dict:
        payload = dict(data or {})
        if APP_IS_SECURE_BUILD:
            for key in SENSITIVE_SETTINGS_KEYS:
                payload[key] = security_utils.protect_secret(payload.get(key, ""))
        return payload

    def _load_settings(self):
        path = self._settings_path()
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    data = self._decode_settings_payload(data)
                    if int(data.get("settings_version", 1)) < APP_SETTINGS_VERSION:
                        data["settings_version"] = APP_SETTINGS_VERSION
                    return data
            except Exception:
                pass
        bak = path + ".bak"
        if os.path.exists(bak):
            try:
                with open(bak, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    data = self._decode_settings_payload(data)
                    data["settings_version"] = APP_SETTINGS_VERSION
                    return data
            except Exception:
                pass
        return {}

    def _request_save_settings(self):
        """Debounce frequent UI setting writes to reduce IO and UI stutter."""
        try:
            if self._save_settings_after_id is not None:
                self.root.after_cancel(self._save_settings_after_id)
        except Exception:
            pass

        def _do_save():
            self._save_settings_after_id = None
            self._save_settings()

        try:
            self._save_settings_after_id = self.root.after(SETTINGS_SAVE_DEBOUNCE_MS, _do_save)
        except Exception:
            self._save_settings()

    def _debounce_call(self, key: str, delay_ms: int, fn):
        """Run fn after delay; repeated calls with same key collapse into one."""
        try:
            prev = self._debounce_jobs.get(str(key))
            if prev is not None:
                self.root.after_cancel(prev)
        except Exception:
            pass

        def _run():
            self._debounce_jobs.pop(str(key), None)
            try:
                fn()
            except Exception:
                pass

        try:
            jid = self.root.after(max(1, int(delay_ms)), _run)
            self._debounce_jobs[str(key)] = jid
        except Exception:
            _run()

    def _save_settings(self):
        self.settings["settings_version"] = APP_SETTINGS_VERSION
        try:
            self.settings["main_window_geometry"] = str(self.root.winfo_geometry())
        except Exception:
            pass
        self.settings["custom_name"] = self.custom_name_var.get().strip() or "主播"
        self.settings["ui_language"] = "zh"
        self.settings["room_url"] = self.room_url_var.get().strip()
        self.settings["listen_source_mode"] = ("local" if APP_IS_PUBLIC_BUILD else ("relay" if APP_IS_SIGNPOOL_RELAY_BUILD else self._get_listen_source_mode()))
        self.settings["proxy_enabled"] = (False if APP_IS_SIGNPOOL_RELAY_BUILD else bool(self.proxy_enabled.get()))
        self.settings["proxy"] = ("" if APP_IS_SIGNPOOL_RELAY_BUILD else self.proxy_var.get().strip())
        self.settings["proxy_route_mode"] = ("direct" if APP_IS_SIGNPOOL_RELAY_BUILD else self._get_proxy_route_mode())
        self.settings["ssl_insecure"] = bool(self.ssl_insecure_var.get())
        self.settings["selected_printer"] = self.printer_cb.get().strip() if hasattr(self, "printer_cb") else ""
        self.settings["use_default_printer"] = bool(self.use_default_printer_var.get()) if hasattr(self, "use_default_printer_var") else True
        self.settings["custom_paper_width_mm"] = str(self.width_var.get()).strip() if hasattr(self, "width_var") else "40"
        self.settings["custom_paper_height_mm"] = str(self.height_var.get()).strip() if hasattr(self, "height_var") else "30"
        self.settings["char_width_mm"] = str(self.char_width_mm_var.get()).strip() if hasattr(self, "char_width_mm_var") else "1.50"
        self.settings["line_height_mm"] = str(self.line_height_mm_var.get()).strip() if hasattr(self, "line_height_mm_var") else "2.80"
        self.settings["margin_mm"] = str(self.margin_mm_var.get()).strip() if hasattr(self, "margin_mm_var") else "1.00"
        self.selected_printer_cached = str(self.settings.get("selected_printer", "")).strip()
        self.use_default_printer_cached = bool(self.settings.get("use_default_printer", True))
        self.settings["license_server"] = ""
        self.settings["license_key"] = ""
        if APP_IS_SIGNPOOL_RELAY_BUILD:
            self.settings["sign_api_base"] = ""
            self.settings["sign_api_key"] = ""
        else:
            self.settings["sign_api_base"] = self.sign_api_base_var.get().strip() if hasattr(self, "sign_api_base_var") else DEFAULT_SIGN_API_BASE
            self.settings["sign_api_key"] = self.sign_api_key_var.get().strip() if hasattr(self, "sign_api_key_var") else ""
        self.settings["license_machine_token"] = ("" if APP_IS_PUBLIC_BUILD else str(getattr(self, "license_machine_token", "")).strip())
        self.settings["use_sign_api_key"] = (False if APP_IS_SIGNPOOL_RELAY_BUILD else bool(self.use_sign_api_var.get())) if hasattr(self, "use_sign_api_var") else False
        self.settings["license_last_ok_ts"] = int(getattr(self, "license_last_ok_ts", 0))
        self.settings["remote_pid_sync_enabled"] = (False if APP_IS_PUBLIC_BUILD else bool(getattr(self, "remote_pid_sync_enabled", True)))
        self.settings["local_bypass_enabled"] = False
        self.settings["local_bypass_machine_id"] = ""
        self.settings["cloud_upload_enabled"] = False if APP_IS_PUBLIC_BUILD else bool(getattr(self, "cloud_upload_enabled", False))
        self.settings["cloud_client_server"] = "" if APP_IS_PUBLIC_BUILD else str(getattr(self, "cloud_client_server", "")).strip()
        self.settings["cloud_admin_server"] = "" if APP_IS_PUBLIC_BUILD else str(getattr(self, "cloud_admin_server", "")).strip()
        self.settings["cloud_admin_token"] = "" if APP_IS_PUBLIC_BUILD else str(getattr(self, "cloud_admin_token", "")).strip()
        self.settings["dashboard_port"] = int(getattr(self, "dashboard_port", 8787))
        self.settings["dashboard_autostart"] = bool(getattr(self, "dashboard_autostart", False))
        self.settings["dashboard_recent_comments_limit"] = int(getattr(self.dashboard_recent_comments, "maxlen", 1500) or 1500)
        self.settings["feishu_enabled"] = bool(getattr(self, "feishu_enabled", False))
        self.settings["feishu_webhook"] = str(getattr(self, "feishu_webhook", "")).strip()
        self.settings["feishu_push_print_only"] = bool(getattr(self, "feishu_push_print_only", True))
        self.settings["feishu_keyword_filter"] = ",".join(getattr(self, "feishu_keyword_filter", []) or [])
        self.settings["feishu_max_per_minute"] = int(getattr(self, "feishu_max_per_minute", 25))
        self.settings["template_mode"] = self._get_template_mode()
        self.settings["canvas_template_enabled"] = bool(self.canvas_template_enabled.get()) if hasattr(self, "canvas_template_enabled") else False
        self.settings["canvas_template"] = self._normalize_canvas_template_to_paper(self.canvas_template) if hasattr(self, "canvas_template") else {}
        self.settings["template_custom_vars"] = self.template_custom_vars if hasattr(self, "template_custom_vars") else {}
        self.settings["editor_margin_top"] = int(self.editor_margin_top_var.get()) if hasattr(self, "editor_margin_top_var") else 0
        self.settings["editor_margin_bottom"] = int(self.editor_margin_bottom_var.get()) if hasattr(self, "editor_margin_bottom_var") else 0
        self.settings["editor_margin_left"] = int(self.editor_margin_left_var.get()) if hasattr(self, "editor_margin_left_var") else 0
        self.settings["editor_margin_right"] = int(self.editor_margin_right_var.get()) if hasattr(self, "editor_margin_right_var") else 0
        self.settings["editor_letter_spacing"] = int(self.editor_letter_spacing_var.get()) if hasattr(self, "editor_letter_spacing_var") else 0
        self.settings["editor_paragraph_spacing"] = int(self.editor_paragraph_spacing_var.get()) if hasattr(self, "editor_paragraph_spacing_var") else 0
        self.settings["editor_font_family"] = str(self.editor_font_family_var.get()).strip() if hasattr(self, "editor_font_family_var") else "TkDefaultFont"
        self.settings["editor_var_font_scale_map"] = self.editor_var_font_scale_map if isinstance(getattr(self, "editor_var_font_scale_map", {}), dict) else {}
        self.settings["reconnect_print_window_seconds"] = int(self.reconnect_print_window_seconds)
        self.settings["print_batch_window_seconds"] = int(self.print_batch_window_seconds)
        self.settings["print_batch_limit"] = int(self.print_batch_limit)
        self.settings["print_worker_count"] = int(self.print_worker_count)
        self.settings["print_retry_limit"] = int(self.print_retry_limit)
        self.settings["auto_wrap_print_enabled"] = bool(self.auto_wrap_print_enabled)
        self.settings["auto_wrap_name_width"] = int(self.auto_wrap_name_width)
        self.settings["auto_wrap_content_width"] = int(self.auto_wrap_content_width)
        self.settings["max_stream_rows"] = int(self.max_stream_rows)
        self.settings["auto_archive_stream"] = bool(self.auto_archive_stream)
        self.settings["auto_print_numeric"] = bool(self.auto_print_numeric)
        self.settings["keyword_print_enabled"] = bool(self.keyword_print_enabled)
        self.settings["keyword_print_list"] = ",".join(self.keyword_print_list)
        self.settings["print_min_len"] = int(self.print_min_len)
        self.settings["print_max_len"] = int(self.print_max_len)
        self.settings["lock_order_mode"] = bool(self.lock_order_mode)
        self.settings["lock_order_window_seconds"] = int(self.lock_order_window_seconds)
        self.settings["per_printer_limit"] = int(self.per_printer_limit)
        self.settings["queue_alert_threshold"] = int(self.queue_alert_threshold)
        self.settings["fail_alert_threshold"] = int(self.fail_alert_threshold)
        self.settings["peak_warn_threshold"] = int(self.peak_warn_threshold)
        self.settings["peak_critical_threshold"] = int(self.peak_critical_threshold)
        self.settings["peak_duplicate_window_seconds"] = int(self.peak_duplicate_window_seconds)
        self.settings["blacklist_keywords"] = ",".join(self.blacklist_keywords)
        self.settings["whitelist_keywords"] = ",".join(self.whitelist_keywords)
        self.settings["auto_blacklist_rate_limit"] = int(self.auto_blacklist_rate_limit)
        self.settings["admin_password"] = ""
        self.settings["admin_password_hash"] = str(getattr(self, "admin_password_hash", "")).strip()
        self.settings["overlap_max_events"] = int(self.overlap_events.maxlen or 300000)
        self.settings["overlap_message_max_events"] = int(self.overlap_message_events.maxlen or 300000)
        self.settings["overlap_default_hours"] = int(self.overlap_default_hours)
        self.settings["overlap_include_main"] = bool(getattr(self, "overlap_include_main", True))
        self.settings["sales_goals"] = dict(getattr(self, "sales_goals", {}))
        self.settings["customer_tags"] = dict(getattr(self, "customer_tags", {}))
        self.settings["followup_queue"] = list(getattr(self, "followup_queue", []))
        self.settings["campaign_calendar"] = list(getattr(self, "campaign_calendar", []))
        self.settings["inventory_map"] = dict(getattr(self, "inventory_map", {}))
        self.settings["host_scores"] = dict(getattr(self, "host_scores", {}))
        self.settings["team_roles"] = dict(getattr(self, "team_roles", {}))
        self.settings["risk_rules"] = dict(getattr(self, "risk_rules", {}))
        self.settings["data_retention_days"] = int(getattr(self, "data_retention_days", 30))
        self.settings["auto_data_maintenance_enabled"] = bool(getattr(self, "auto_data_maintenance_enabled", True))
        self.settings["overlap_watch_rooms"] = list(self.overlap_watch_rooms)
        self.settings["auto_record_enabled"] = bool(self.auto_record_enabled)
        self.settings["auto_record_dir"] = str(self.auto_record_dir or "").strip()
        self.settings["auto_record_cmd"] = str(self.auto_record_cmd or "streamlink").strip()
        path = self._settings_path()
        try:
            tmp_path = path + ".tmp"
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(self._encode_settings_payload(self.settings), f, ensure_ascii=False, indent=2)
            if os.path.exists(path):
                try:
                    shutil.copy2(path, path + ".bak")
                except Exception:
                    pass
            os.replace(tmp_path, path)
        except Exception:
            pass

    def _on_close(self):
        self._sync_local_permanent_ids_backup("on_close")
        try:
            if self._save_settings_after_id is not None:
                self.root.after_cancel(self._save_settings_after_id)
                self._save_settings_after_id = None
        except Exception:
            pass
        try:
            for _k, _jid in list(self._debounce_jobs.items()):
                self.root.after_cancel(_jid)
            self._debounce_jobs.clear()
        except Exception:
            pass
        self._save_settings()
        try:
            self.stop_listen()
        except Exception:
            pass
        try:
            self.stop_analysis_listen()
        except Exception:
            pass
        try:
            self._stop_web_dashboard_server()
        except Exception:
            pass
        try:
            self.feishu_push_queue.put_nowait(None)
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass

    def _is_input_widget(self, widget):
        return isinstance(widget, (tk.Entry, ttk.Entry, ttk.Combobox, tk.Text))

    def _setup_input_shortcuts(self):
        self.input_menu = tk.Menu(self.root, tearoff=0)
        self.input_menu.add_command(label='全选', command=lambda: self._widget_select_all(self._input_menu_widget))
        self.input_menu.add_command(label='复制', command=lambda: self._widget_copy(self._input_menu_widget))
        self.input_menu.add_command(label='粘贴', command=lambda: self._widget_paste(self._input_menu_widget))
        self.input_menu.add_command(label='剪切', command=lambda: self._widget_cut(self._input_menu_widget))
        self._input_menu_widget = None

        for sequence in ('<Button-3>', '<Button-2>', '<Control-Button-1>'):
            self.root.bind_all(sequence, self._show_input_menu, add='+')
        for cls in ('Entry', 'TEntry', 'TCombobox', 'Text'):
            self.root.bind_class(cls, '<Control-a>', self._input_ctrl_a, add='+')

    def _show_input_menu(self, event):
        widget = event.widget
        if not self._is_input_widget(widget):
            return
        self._input_menu_widget = widget
        try:
            if isinstance(widget, tk.Text):
                widget.mark_set("insert", f"@{event.x},{event.y}")
            widget.focus_set()
            self.input_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.input_menu.grab_release()

    def _input_ctrl_a(self, event):
        widget = event.widget
        if not self._is_input_widget(widget):
            return
        self._widget_select_all(widget)
        return "break"

    def _widget_select_all(self, widget):
        if widget is None:
            return
        try:
            if isinstance(widget, tk.Text):
                widget.tag_add("sel", "1.0", "end-1c")
            else:
                widget.selection_range(0, tk.END)
                widget.icursor(tk.END)
                widget.xview_moveto(1.0)
        except Exception:
            pass

    def _widget_copy(self, widget):
        if widget is None:
            return
        try:
            widget.event_generate("<<Copy>>")
        except Exception:
            pass

    def _widget_paste(self, widget):
        if widget is None:
            return
        try:
            widget.event_generate("<<Paste>>")
        except Exception:
            pass

    def _widget_cut(self, widget):
        if widget is None:
            return
        try:
            widget.event_generate("<<Cut>>")
        except Exception:
            pass

    def _local_permanent_ids_backup_path(self):
        return os.path.join(RUNTIME_DIR, "permanent_ids_local_auto.csv")

    def _sync_local_permanent_ids_backup(self, reason: str = "manual"):
        try:
            users = db.list_users()
            path = self._local_permanent_ids_backup_path()
            tmp = path + ".tmp"
            with open(tmp, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                # A列=永久ID, B列=昵称, C列=客户ID
                w.writerow(["permanent_id", "display_name", "unique_id"])
                for unique_id, display_name, permanent_id in users:
                    w.writerow([permanent_id, display_name or "", unique_id])
            os.replace(tmp, path)
            self._audit("pid_local_backup", f"{reason}|count={len(users)}")
            return path, len(users)
        except Exception:
            return "", 0

    def _snapshot_local_permanent_ids_backup(self, reason: str = "manual"):
        try:
            users = db.list_users()
            if not users:
                return "", 0
            snap_dir = os.path.join(RUNTIME_DIR, "permanent_ids_snapshots")
            os.makedirs(snap_dir, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_reason = "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in str(reason or "manual"))[:40]
            path = os.path.join(snap_dir, f"permanent_ids_{ts}_{safe_reason}.csv")
            with open(path, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(["permanent_id", "display_name", "unique_id"])
                for unique_id, display_name, permanent_id in users:
                    w.writerow([permanent_id, display_name or "", unique_id])
            self._audit("pid_local_snapshot", f"{reason}|count={len(users)}|path={path}")
            return path, len(users)
        except Exception:
            return "", 0

    def _auto_backup_permanent_ids(self):
        try:
            self._sync_local_permanent_ids_backup("timer")
        except Exception:
            pass
        finally:
            try:
                self.root.after(120000, self._auto_backup_permanent_ids)
            except Exception:
                pass

    def _set_status(self, text: str):
        self.status_var.set(f"状态: {text}")

    def _audit(self, event: str, detail: str = ""):
        try:
            row = {
                "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "event": str(event or ""),
                "detail": str(detail or ""),
            }
            with open(self.audit_log_path, "a", encoding="utf-8") as f:
                f.write(json.dumps(row, ensure_ascii=False) + "\n")
        except Exception:
            pass

    def _track_business_event(self, event_type: str, payload: dict = None):
        try:
            row = {
                "ts": time.time(),
                "event_type": str(event_type or ""),
                "schema_version": "v1",
                "payload": payload or {},
            }
            with open(self.business_event_log_path, "a", encoding="utf-8") as f:
                f.write(json.dumps(row, ensure_ascii=False) + "\n")
        except Exception:
            pass

    def _enqueue_feishu_event(self, title: str, text: str):
        if not self.feishu_enabled or (not self.feishu_webhook):
            return
        try:
            body = {"title": str(title or "Sen Nails"), "text": str(text or "")}
            self.feishu_push_queue.put_nowait(body)
        except Exception:
            pass

    def _feishu_should_send_comment(self, content: str, rule_hit: str) -> bool:
        if not self.feishu_enabled:
            return False
        if self.feishu_push_print_only and not str(rule_hit or "").strip():
            return False
        text = str(content or "")
        if self.feishu_keyword_filter:
            low = text.lower()
            if not any(str(k).lower() in low for k in self.feishu_keyword_filter):
                return False
        return True

    def _feishu_worker_loop(self):
        while True:
            try:
                item = self.feishu_push_queue.get()
            except Exception:
                time.sleep(0.2)
                continue
            if item is None:
                break
            try:
                now = time.time()
                cutoff = now - 60.0
                while self._feishu_sent_timestamps and self._feishu_sent_timestamps[0] < cutoff:
                    self._feishu_sent_timestamps.popleft()
                if len(self._feishu_sent_timestamps) >= max(1, int(self.feishu_max_per_minute)):
                    continue
                payload = {
                    "msg_type": "text",
                    "content": {"text": f"{item.get('title', 'Sen Nails')}\n{item.get('text', '')}"},
                }
                req = urllib.request.Request(
                    self.feishu_webhook,
                    data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                    headers={"Content-Type": "application/json; charset=utf-8"},
                    method="POST",
                )
                with urllib.request.urlopen(req, timeout=6) as resp:
                    _ = resp.read()
                self._feishu_sent_timestamps.append(time.time())
            except Exception:
                pass
            finally:
                try:
                    self.feishu_push_queue.task_done()
                except Exception:
                    pass

    def _dashboard_collect_snapshot(self) -> dict:
        summary = db.get_today_print_summary()
        top_guests = sorted(
            ((uid, int(cnt)) for uid, cnt in (self.guest_message_counter or {}).items()),
            key=lambda x: (-x[1], x[0]),
        )[:20]
        room_counter = defaultdict(int)
        for _ts, room_id, _uid, _name, _content, _source in list(self.overlap_message_events)[-3000:]:
            room_counter[str(room_id)] += 1
        rooms = sorted(room_counter.items(), key=lambda x: (-x[1], x[0]))[:20]
        comments = list(self.dashboard_recent_comments)[-200:]
        return {
            "ts": int(time.time()),
            "listening_rooms": sorted(list(self.listener_workers.keys())),
            "analysis_rooms": sorted(list(self.analysis_listener_workers.keys())),
            "print_summary": summary,
            "top_guests": [{"uid": u, "count": c} for u, c in top_guests],
            "top_rooms": [{"room": r, "count": c} for r, c in rooms],
            "recent_comments": comments,
        }

    def _dashboard_html(self) -> str:
        return """<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Sen Nails 数据看板</title>
<style>
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:#f3f6fb;margin:0;padding:18px;color:#1d2433}
.grid{display:grid;grid-template-columns:repeat(4,minmax(140px,1fr));gap:12px}
.card{background:#fff;border:1px solid #d6deee;border-radius:10px;padding:12px}
.num{font-size:28px;font-weight:700}
h1{margin:0 0 12px 0}.row{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:12px}
table{width:100%;border-collapse:collapse;background:#fff}
th,td{border-bottom:1px solid #edf1f8;padding:6px 8px;text-align:left;font-size:12px}
.bar{height:8px;background:#dbe7ff;border-radius:999px;overflow:hidden}.bar>i{display:block;height:100%;background:#3b7cff}
</style></head><body>
<h1>Sen Nails 实时看板</h1>
<div class="grid">
<div class="card"><div>待打印</div><div class="num" id="pending">0</div></div>
<div class="card"><div>打印成功</div><div class="num" id="printed">0</div></div>
<div class="card"><div>打印失败</div><div class="num" id="failed">0</div></div>
<div class="card"><div>监听直播间</div><div class="num" id="rooms">0</div></div>
</div>
<div class="row">
<div class="card"><h3>高频客人 Top20</h3><div id="guests"></div></div>
<div class="card"><h3>房间弹幕 Top20</h3><div id="roombars"></div></div>
</div>
<div class="row">
<div class="card" style="grid-column:1 / span 2"><h3>最近弹幕</h3><table><thead><tr><th>时间</th><th>房间</th><th>UID</th><th>昵称</th><th>内容</th></tr></thead><tbody id="comments"></tbody></table></div>
</div>
<script>
function esc(s){return String(s??'').replace(/[&<>"]/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[m]));}
function barRow(name,val,max){const w=max?Math.round((val/max)*100):0;return `<div style="margin:6px 0"><div style="display:flex;justify-content:space-between"><span>${esc(name)}</span><b>${val}</b></div><div class="bar"><i style="width:${w}%"></i></div></div>`;}
async function tick(){
  const r=await fetch('/api/summary',{cache:'no-store'}); const d=await r.json();
  document.getElementById('pending').textContent=d.print_summary.pending||0;
  document.getElementById('printed').textContent=d.print_summary.printed||0;
  document.getElementById('failed').textContent=d.print_summary.failed||0;
  document.getElementById('rooms').textContent=(d.listening_rooms||[]).length;
  const g=d.top_guests||[]; const gmax=g.length?Math.max(...g.map(x=>x.count)):0;
  document.getElementById('guests').innerHTML=g.map(x=>barRow(x.uid,x.count,gmax)).join('')||'<div>暂无数据</div>';
  const rm=d.top_rooms||[]; const rmax=rm.length?Math.max(...rm.map(x=>x.count)):0;
  document.getElementById('roombars').innerHTML=rm.map(x=>barRow(x.room,x.count,rmax)).join('')||'<div>暂无数据</div>';
  document.getElementById('comments').innerHTML=(d.recent_comments||[]).slice(-80).reverse().map(x=>`<tr><td>${esc(x.time)}</td><td>${esc(x.room)}</td><td>${esc(x.uid)}</td><td>${esc(x.name)}</td><td>${esc(x.content)}</td></tr>`).join('');
}
tick(); setInterval(tick, 3000);
</script></body></html>"""

    def _start_web_dashboard_server(self, port: int = None):
        if self._dashboard_httpd is not None:
            return int(self.dashboard_port)
        p = int(port if port is not None else self.dashboard_port)
        p = max(1025, min(65530, p))
        app = self

        class Handler(http.server.BaseHTTPRequestHandler):
            def do_GET(self):
                if self.path.startswith("/api/summary"):
                    body = json.dumps(app._dashboard_collect_snapshot(), ensure_ascii=False).encode("utf-8")
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.send_header("Cache-Control", "no-store")
                    self.send_header("Content-Length", str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                    return
                if self.path == "/" or self.path.startswith("/index"):
                    body = app._dashboard_html().encode("utf-8")
                    self.send_response(200)
                    self.send_header("Content-Type", "text/html; charset=utf-8")
                    self.send_header("Content-Length", str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                    return
                self.send_response(404)
                self.end_headers()

            def log_message(self, format, *args):
                return

        httpd = http.server.ThreadingHTTPServer(("127.0.0.1", p), Handler)
        self._dashboard_httpd = httpd
        self.dashboard_port = int(p)
        self._dashboard_server_started_ts = time.time()
        self._dashboard_thread = threading.Thread(target=httpd.serve_forever, daemon=True)
        self._dashboard_thread.start()
        self._set_status(f"Web看板已启动: http://127.0.0.1:{p}")
        return p

    def _stop_web_dashboard_server(self):
        if self._dashboard_httpd is None:
            return
        try:
            self._dashboard_httpd.shutdown()
            self._dashboard_httpd.server_close()
        except Exception:
            pass
        self._dashboard_httpd = None
        self._dashboard_thread = None
        self._set_status("Web看板已停止")

    def _open_web_dashboard(self):
        port = self._start_web_dashboard_server(self.dashboard_port)
        url = f"http://127.0.0.1:{int(port)}"
        try:
            webbrowser.open(url)
        except Exception:
            pass

    def _test_feishu_connection(self):
        if not self.feishu_webhook:
            messagebox.showwarning("提示", "请先填写飞书Webhook")
            return
        self._enqueue_feishu_event("Sen Nails 测试消息", f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n状态: 接入正常")
        messagebox.showinfo("提示", "已发送测试消息，请到飞书查看。")

    def open_integration_center(self):
        if hasattr(self, "_integration_win") and self._integration_win.winfo_exists():
            self._integration_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("飞书与Web看板")
        win.geometry("760x460")
        self._integration_win = win

        fs = ttk.LabelFrame(win, text="飞书推送")
        fs.pack(fill=tk.X, padx=8, pady=8)
        feishu_enable_var = tk.IntVar(value=1 if self.feishu_enabled else 0)
        feishu_webhook_var = tk.StringVar(value=self.feishu_webhook)
        feishu_print_only_var = tk.IntVar(value=1 if self.feishu_push_print_only else 0)
        feishu_kw_var = tk.StringVar(value=",".join(self.feishu_keyword_filter))
        feishu_rate_var = tk.IntVar(value=max(1, int(self.feishu_max_per_minute)))
        ttk.Checkbutton(fs, text="启用飞书推送", variable=feishu_enable_var).grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Checkbutton(fs, text="仅推送命中打印规则的弹幕", variable=feishu_print_only_var).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
        ttk.Label(fs, text="Webhook").grid(row=1, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(fs, textvariable=feishu_webhook_var, width=78).grid(row=1, column=1, sticky=tk.W, padx=4, pady=4)
        ttk.Label(fs, text="关键词过滤(逗号分隔，可留空)").grid(row=2, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(fs, textvariable=feishu_kw_var, width=45).grid(row=2, column=1, sticky=tk.W, padx=4, pady=4)
        ttk.Label(fs, text="每分钟上限").grid(row=3, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(fs, textvariable=feishu_rate_var, width=8).grid(row=3, column=1, sticky=tk.W, padx=4, pady=4)

        ws = ttk.LabelFrame(win, text="Web数据看板")
        ws.pack(fill=tk.X, padx=8, pady=8)
        dashboard_port_var = tk.IntVar(value=int(self.dashboard_port))
        dashboard_autostart_var = tk.IntVar(value=1 if self.dashboard_autostart else 0)
        dashboard_status_var = tk.StringVar(value=("运行中" if self._dashboard_httpd is not None else "未启动"))
        ttk.Label(ws, text="端口").grid(row=0, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Entry(ws, textvariable=dashboard_port_var, width=8).grid(row=0, column=1, sticky=tk.W, padx=4, pady=4)
        ttk.Checkbutton(ws, text="启动软件时自动启动看板", variable=dashboard_autostart_var).grid(row=0, column=2, sticky=tk.W, padx=4, pady=4)
        ttk.Label(ws, text="状态").grid(row=1, column=0, sticky=tk.W, padx=4, pady=4)
        ttk.Label(ws, textvariable=dashboard_status_var).grid(row=1, column=1, sticky=tk.W, padx=4, pady=4)

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=8)

        def save_conf():
            self.feishu_enabled = bool(feishu_enable_var.get())
            self.feishu_webhook = feishu_webhook_var.get().strip()
            self.feishu_push_print_only = bool(feishu_print_only_var.get())
            self.feishu_keyword_filter = [x.strip() for x in feishu_kw_var.get().split(",") if x.strip()]
            try:
                self.feishu_max_per_minute = max(1, int(feishu_rate_var.get()))
            except Exception:
                self.feishu_max_per_minute = 25
            try:
                self.dashboard_port = max(1025, min(65530, int(dashboard_port_var.get())))
            except Exception:
                self.dashboard_port = 8787
            self.dashboard_autostart = bool(dashboard_autostart_var.get())
            self._save_settings()
            self._set_status("飞书与Web看板配置已保存")
            messagebox.showinfo("完成", "配置已保存")

        def start_dash():
            try:
                self.dashboard_port = max(1025, min(65530, int(dashboard_port_var.get())))
            except Exception:
                self.dashboard_port = 8787
            try:
                self._start_web_dashboard_server(self.dashboard_port)
                dashboard_status_var.set("运行中")
            except Exception as e:
                messagebox.showerror("错误", f"看板启动失败: {e}")

        def stop_dash():
            self._stop_web_dashboard_server()
            dashboard_status_var.set("未启动")

        def open_dash():
            start_dash()
            self._open_web_dashboard()

        ttk.Button(btn, text="保存配置", command=save_conf).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="测试飞书", command=self._test_feishu_connection).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="启动看板", command=start_dash).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="停止看板", command=stop_dash).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="打开看板", command=open_dash).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="关闭", command=win.destroy).pack(side=tk.LEFT, padx=4)

    def _trim_jsonl_by_ts(self, path: str, cutoff_ts: float, ts_key: str = "ts", keep_last: int = 300000):
        if not os.path.exists(path):
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            kept = []
            for line in lines[-max(keep_last * 2, keep_last):]:
                s = line.strip()
                if not s:
                    continue
                try:
                    row = json.loads(s)
                    ts = float(row.get(ts_key, 0))
                except Exception:
                    continue
                if ts >= cutoff_ts:
                    kept.append(s)
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(kept) + ("\n" if kept else ""))
        except Exception:
            pass

    def _auto_data_maintenance(self):
        try:
            if not bool(self.auto_data_maintenance_enabled):
                return
            cutoff = time.time() - max(1, int(self.data_retention_days)) * 86400
            while self.overlap_message_events and float(self.overlap_message_events[0][0]) < cutoff:
                self.overlap_message_events.popleft()
            while self.overlap_events and float(self.overlap_events[0][0]) < cutoff:
                self.overlap_events.popleft()
            self._trim_jsonl_by_ts(self.event_store_path, cutoff, ts_key="ts", keep_last=300000)
            self._trim_jsonl_by_ts(self.business_event_log_path, cutoff, ts_key="ts", keep_last=400000)
        except Exception:
            pass
        finally:
            try:
                self.root.after(21600000, self._auto_data_maintenance)
            except Exception:
                pass

    def _append_overlap_event_to_disk(self, ts: float, room_id: str, unique_id: str, display_name: str, content: str, source_tag: str = "main"):
        try:
            row = {
                "ts": float(ts),
                "room_id": str(room_id or ""),
                "unique_id": str(unique_id or ""),
                "display_name": str(display_name or ""),
                "content": str(content or ""),
                "source": str(source_tag or "main"),
            }
            with open(self.event_store_path, "a", encoding="utf-8") as f:
                f.write(json.dumps(row, ensure_ascii=False) + "\n")
        except Exception:
            pass

    def _load_overlap_events_from_disk(self):
        path = self.event_store_path
        if not os.path.exists(path):
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                lines = f.readlines()[-200000:]
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                try:
                    row = json.loads(line)
                    self.overlap_message_events.append((
                        float(row.get("ts", time.time())),
                        str(row.get("room_id", "")),
                        str(row.get("unique_id", "")),
                        str(row.get("display_name", "")),
                        str(row.get("content", "")),
                        str(row.get("source", "main")),
                    ))
                except Exception:
                    continue
        except Exception:
            pass

    def _set_license_state(self, text: str):
        self.license_state_var.set(f"授权状态: {text}")

    def _license_cache_path(self):
        if APP_IS_SECURE_BUILD and APP_IS_SIGNPOOL_BUILD:
            filename = "license_cache_secure_signpool.json"
        elif APP_IS_SIGNPOOL_BUILD:
            filename = "license_cache_signpool.json"
        else:
            filename = "license_cache_secure.json" if APP_IS_SECURE_BUILD else "license_cache.json"
        path = os.path.join(RUNTIME_DIR, filename)
        legacy = os.path.join(os.getcwd(), filename)
        if (not os.path.exists(path)) and os.path.exists(legacy):
            try:
                shutil.copy2(legacy, path)
            except Exception:
                pass
        return path

    def _license_server_url(self) -> str:
        if APP_IS_PUBLIC_BUILD:
            return ""
        return FIXED_LICENSE_SERVER_URL

    def _remote_pid_server_url(self) -> str:
        return self._license_server_url()

    def _license_key(self) -> str:
        if APP_IS_PUBLIC_BUILD:
            return ""
        return str(self.license_key_var.get().strip() if hasattr(self, "license_key_var") else "")

    def _format_license_days(self, payload: dict | None = None) -> str:
        data = payload or {}
        try:
            total_days = int(data.get("total_days") or 0)
        except Exception:
            total_days = 0
        try:
            remaining_days = int(data.get("remaining_days") or 0)
        except Exception:
            remaining_days = 0
        parts = []
        if total_days > 0:
            parts.append(f"总授权 {total_days} 天")
        if remaining_days >= 0 and (remaining_days > 0 or total_days > 0):
            parts.append(f"剩余 {remaining_days} 天")
        return " | ".join(parts)

    def _normalize_listen_source_mode(self, value: str) -> str:
        raw = str(value or "").strip()
        if raw in LISTEN_SOURCE_MODE_LABEL_TO_VALUE:
            return LISTEN_SOURCE_MODE_LABEL_TO_VALUE[raw]
        raw = raw.lower()
        if raw in LISTEN_SOURCE_MODE_VALUE_TO_LABEL:
            return raw
        return "local"

    def _get_listen_source_mode(self) -> str:
        if APP_IS_PUBLIC_BUILD:
            return "local"
        if APP_IS_SIGNPOOL_RELAY_BUILD:
            return "relay"
        if hasattr(self, "listen_source_mode_var"):
            return self._normalize_listen_source_mode(self.listen_source_mode_var.get())
        return self._normalize_listen_source_mode(self.settings.get("listen_source_mode", "local"))

    def _update_license_machine_token(self, payload: dict | None = None):
        token = str((payload or {}).get("machine_token", "") or "").strip()
        if token and token != getattr(self, "license_machine_token", ""):
            self.license_machine_token = token

    def _perform_license_handshake(self, allow_activate_fallback: bool = True):
        server = self._license_server_url()
        key = self._license_key()
        ok, msg, data = license_client.heartbeat(
            server,
            key,
            self.app_version,
            machine_token=self.license_machine_token,
        )
        if ok:
            self._update_license_machine_token(data)
            return ok, msg, data
        lower_msg = str(msg or "").lower()
        should_activate = allow_activate_fallback and (not self.license_machine_token or "machine token" in lower_msg)
        if should_activate:
            ok, msg, data = license_client.activate(
                server,
                key,
                self.app_version,
                self.custom_name_var.get().strip(),
                machine_token=self.license_machine_token,
            )
            if ok:
                self._update_license_machine_token(data)
        return ok, msg, data

    def _mark_license_ok(self, data: dict | None = None, prefix: str = "已授权"):
        payload = data or {}
        self.license_active = True
        self.license_last_ok_ts = license_client.now_ts()
        self._update_license_machine_token(payload)
        customer_name = str(payload.get("customer_name", "") or "").strip()
        expires_at = str(payload.get("expires_at", "") or "").strip()
        text = prefix
        if customer_name:
            text += f" {customer_name}"
        if expires_at:
            text += f" 到期 {expires_at[:10]}"
        day_text = self._format_license_days(payload)
        if day_text:
            text += f" | {day_text}"
        self._set_license_state(text)
        license_client.save_license_cache(
            self._license_cache_path(),
            {
                "ok": True,
                "checked_at": self.license_last_ok_ts,
                "machine_id": self.machine_id,
                "data": payload,
            },
        )
        self._request_save_settings()

    def _mark_license_failed(self, text: str):
        self.license_active = False
        self._set_license_state(text)
        self._request_save_settings()

    def _license_grace_valid(self, cache: dict | None = None) -> tuple[bool, int]:
        if APP_IS_SECURE_BUILD or int(getattr(self, "license_grace_hours", 0) or 0) <= 0:
            return False, 0
        data = cache or {}
        try:
            checked_at = int(data.get("checked_at") or self.license_last_ok_ts or 0)
        except Exception:
            checked_at = int(self.license_last_ok_ts or 0)
        if checked_at <= 0:
            return False, 0
        if str(data.get("machine_id", self.machine_id) or self.machine_id) != self.machine_id:
            return False, 0
        remain = int((checked_at + max(1, int(self.license_grace_hours)) * 3600) - time.time())
        return remain > 0, max(0, remain)

    def _is_local_bypass(self) -> bool:
        return False

    def enable_local_bypass(self):
        self.local_bypass_enabled = False
        self.local_bypass_machine_id = ""
        self.license_active = False
        self._save_settings()
        messagebox.showwarning("提示", "当前版本已取消本机免授权，必须填写授权码并完成授权。")

    def _refresh_soft_status(self):
        try:
            row_cnt = len(self.stream_tree.get_children()) if hasattr(self, "stream_tree") else 0
            listeners = len(self.listener_workers) if hasattr(self, "listener_workers") else 0
            analysis_listeners = len(self.analysis_listener_workers) if hasattr(self, "analysis_listener_workers") else 0
            mode = "监听中" if self.listening else "未监听"
            mem_txt = self._memory_text()
            peak_mode = str(getattr(self, "peak_mode", "normal"))
            cloud_state = "关"
            if bool(getattr(self, "cloud_upload_enabled", False)):
                if bool(getattr(self, "_cloud_uploading", False)):
                    cloud_state = f"上传中({len(getattr(self, 'cloud_event_buffer', []))})"
                else:
                    cloud_state = f"队列{len(getattr(self, 'cloud_event_buffer', []))} 失败{int(getattr(self, '_cloud_fail_streak', 0))}"
            web_state = f"开:{int(getattr(self, 'dashboard_port', 8787))}" if getattr(self, "_dashboard_httpd", None) is not None else "关"
            feishu_state = "开" if bool(getattr(self, "feishu_enabled", False)) else "关"
            self.soft_status_var.set(
                f"软件状态: {mode} | 主房间:{listeners} | 分析房间:{analysis_listeners} | 弹幕行:{row_cnt} | 队列:{self.queue.qsize()} | 高峰:{peak_mode} 丢弃:{int(getattr(self,'peak_drop_counter',0))} 合并:{int(getattr(self,'peak_merge_counter',0))} | 云:{cloud_state} | 飞书:{feishu_state} | Web:{web_state} | 内存:{mem_txt}"
            )
        except Exception:
            pass
        finally:
            try:
                self.root.after(2000, self._refresh_soft_status)
            except Exception:
                pass

    def _memory_text(self) -> str:
        try:
            current, peak = tracemalloc.get_traced_memory()
            return f"{current / 1024 / 1024:.1f}MB (峰值 {peak / 1024 / 1024:.1f}MB)"
        except Exception:
            return "-"

    def _center_window(self, win, width: int = None, height: int = None):
        try:
            win.update_idletasks()
            w = int(width) if width else max(1, int(win.winfo_width()))
            h = int(height) if height else max(1, int(win.winfo_height()))
            sw = int(win.winfo_screenwidth())
            sh = int(win.winfo_screenheight())
            x = max(0, (sw - w) // 2)
            y = max(0, (sh - h) // 2 - 20)
            win.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            pass

    def save_connection_settings(self):
        self._save_settings()
        self._audit("save_connection_settings", self.room_url_var.get().strip())
        self._set_status("已保存主播/地址/代理配置")
        messagebox.showinfo("保存成功", "已保存主播名、直播地址、代理、SSL与Sign API Base/Key/连接方式配置。")

    def _check_license_on_startup(self):
        if APP_IS_PUBLIC_BUILD:
            self.license_active = True
            self.license_last_ok_ts = int(time.time())
            self._set_license_state("公开版无需授权")
            return
        server = self._license_server_url()
        key = self._license_key()
        if not key:
            self._mark_license_failed("未填写授权码")
            return
        ok, msg, data = self._perform_license_handshake(allow_activate_fallback=True)
        if ok:
            self._mark_license_ok(data, "已授权")
            self._pull_remote_permanent_ids(silent=True)
            return
        cache = license_client.load_license_cache(self._license_cache_path())
        grace_ok, remain = self._license_grace_valid(cache)
        if grace_ok:
            self.license_active = True
            cached_days = self._format_license_days(cache.get("data", {}) if isinstance(cache, dict) else {})
            offline_text = f"离线宽限 {max(1, int(remain / 3600))}h"
            if cached_days:
                offline_text += f" | {cached_days}"
            self._set_license_state(offline_text)
            return
        self._mark_license_failed(f"未授权: {msg}")

    def activate_license(self):
        if APP_IS_PUBLIC_BUILD:
            self.license_active = True
            self._set_license_state("公开版无需授权")
            self._set_status("公开版无需授权，直接填写 Sign API 后开始监听")
            messagebox.showinfo("公开版", "此版本不包含授权与管理员系统。\n请直接填写 Sign API Base / Sign API Key 使用。")
            return
        server = self._license_server_url()
        key = self._license_key()
        if not key:
            self._mark_license_failed("未填写授权码")
            messagebox.showwarning("提示", "请先填写授权码")
            return
        ok, msg, data = license_client.activate(
            server,
            key,
            self.app_version,
            self.custom_name_var.get().strip(),
            machine_token=self.license_machine_token,
        )
        if ok:
            self._update_license_machine_token(data)
        if ok:
            self._mark_license_ok(data, "激活成功")
            self._pull_remote_permanent_ids(silent=True)
            self._set_status("授权激活成功")
            self._save_settings()
            messagebox.showinfo("完成", msg)
            return
        self._mark_license_failed(f"激活失败: {msg}")
        self._set_status(f"授权失败: {msg}")
        messagebox.showerror("授权失败", msg)

    def _ensure_license_active(self, show_dialog: bool = False) -> bool:
        if APP_IS_PUBLIC_BUILD:
            self.license_active = True
            return True
        server = self._license_server_url()
        key = self._license_key()
        if not key:
            self._mark_license_failed("未填写授权码")
            if show_dialog:
                messagebox.showerror("授权失败", "请先填写授权码")
            return False
        if self.license_active:
            return True
        ok, msg, data = self._perform_license_handshake(allow_activate_fallback=True)
        if ok:
            self._mark_license_ok(data, "已授权")
            self._pull_remote_permanent_ids(silent=True)
            return True
        cache = license_client.load_license_cache(self._license_cache_path())
        grace_ok, remain = self._license_grace_valid(cache)
        if grace_ok:
            self.license_active = True
            cached_days = self._format_license_days(cache.get("data", {}) if isinstance(cache, dict) else {})
            offline_text = f"离线宽限 {max(1, int(remain / 3600))}h"
            if cached_days:
                offline_text += f" | {cached_days}"
            self._set_license_state(offline_text)
            return True
        self._mark_license_failed(f"未授权: {msg}")
        if show_dialog:
            messagebox.showerror("授权失败", msg)
        return False

    def _license_heartbeat_loop(self):
        if APP_IS_PUBLIC_BUILD:
            return
        try:
            server = self._license_server_url()
            key = self._license_key()
            if not key:
                self._mark_license_failed("未填写授权码")
                if self.listening:
                    self.stop_listen()
                    self._set_status("未填写授权码，已停止监听")
            else:
                ok, msg, data = self._perform_license_handshake(allow_activate_fallback=True)
                if ok:
                    self._mark_license_ok(data, "已授权")
                else:
                    cache = license_client.load_license_cache(self._license_cache_path())
                    grace_ok, remain = self._license_grace_valid(cache)
                    if grace_ok:
                        self.license_active = True
                        cached_days = self._format_license_days(cache.get("data", {}) if isinstance(cache, dict) else {})
                        offline_text = f"离线宽限 {max(1, int(remain / 3600))}h"
                        if cached_days:
                            offline_text += f" | {cached_days}"
                        self._set_license_state(offline_text)
                    else:
                        self._mark_license_failed(f"未授权: {msg}")
                        if self.listening:
                            self.stop_listen()
                            self._set_status(f"授权失效，已停止监听: {msg}")
        except Exception:
            pass
        finally:
            try:
                self.root.after(120000, self._license_heartbeat_loop)
            except Exception:
                pass

    def _pull_remote_permanent_ids(self, silent: bool = False) -> bool:
        server = self._remote_pid_server_url()
        key = self._license_key()
        if not self.remote_pid_sync_enabled or not server or not key:
            return False
        ok, msg, data = license_client.pull_remote_permanent_ids(
            server,
            key,
            self.app_version,
            machine_token=self.license_machine_token,
        )
        if not ok:
            if not silent:
                messagebox.showerror("同步失败", msg)
            return False
        remote_items = list(data.get("items", []) or [])
        if not remote_items:
            restore_ok, restore_msg = self._recover_remote_permanent_ids_if_server_empty()
            if restore_msg:
                self._set_status(restore_msg)
            if not silent and restore_msg:
                messagebox.showinfo("完成", restore_msg)
            return False if restore_ok is False else True
        imported = 0
        updated = 0
        conflicts = 0
        for item in remote_items:
            try:
                pid = int(item.get("permanent_id", 0))
            except Exception:
                pid = 0
            success, status = db.upsert_user_fixed_permanent_id(
                str(item.get("unique_id", "")).strip(),
                str(item.get("display_name", "")).strip(),
                pid,
            )
            if success and status == "inserted":
                imported += 1
            elif success:
                updated += 1
            else:
                conflicts += 1
        self.user_cache = {}
        self._sync_local_permanent_ids_backup("remote_pull")
        if not silent:
            messagebox.showinfo("完成", f"服务器永久编号同步完成\n新增: {imported}\n更新: {updated}\n冲突: {conflicts}")
        return True

    def _recover_remote_permanent_ids_if_server_empty(self) -> tuple[bool | None, str]:
        local_items = db.list_users_dicts()
        local_count = len(local_items)
        if local_count <= 0:
            return None, "服务器永久编号为空，本地当前也没有可恢复的数据"
        if self._mirror_remote_permanent_ids(silent=True):
            return True, f"检测到服务器永久编号为空，已用本地 {local_count} 条记录自动恢复服务器备份"
        return False, f"检测到服务器永久编号为空，但本地 {local_count} 条记录回灌失败，请稍后点击“推送本地编号”重试"

    def _push_remote_permanent_ids(self, silent: bool = False) -> bool:
        server = self._remote_pid_server_url()
        key = self._license_key()
        if not self.remote_pid_sync_enabled or not server or not key:
            if not silent:
                messagebox.showwarning("提示", "请先启用远程永久编号同步，并填写授权码")
            return False
        items = db.list_users_dicts()
        ok, msg, data = license_client.push_remote_permanent_ids(
            server,
            key,
            items,
            self.app_version,
            machine_token=self.license_machine_token,
        )
        if not ok:
            if not silent:
                messagebox.showerror("同步失败", msg)
            return False
        if not silent:
            messagebox.showinfo(
                "完成",
                f"本地永久编号已推送到服务器\n新增: {int(data.get('imported', 0))}\n更新: {int(data.get('updated', 0))}\n跳过: {int(data.get('skipped', 0))}",
            )
        return True

    def _mirror_remote_permanent_ids(self, silent: bool = False) -> bool:
        server = self._remote_pid_server_url()
        key = self._license_key()
        if not self.remote_pid_sync_enabled or not server or not key:
            if not silent:
                messagebox.showwarning("提示", "请先启用远程永久编号同步，并填写授权码")
            return False
        items = db.list_users_dicts()
        ok, msg, data = license_client.sync_remote_permanent_ids(
            server,
            key,
            items,
            self.app_version,
            machine_token=self.license_machine_token,
        )
        if not ok:
            if not silent:
                messagebox.showerror("同步失败", msg)
            return False
        if not silent:
            messagebox.showinfo(
                "完成",
                f"本地永久编号已镜像同步到服务器\n新增: {int(data.get('imported', 0))}\n更新: {int(data.get('updated', 0))}\n删除: {int(data.get('deleted', 0))}\n跳过: {int(data.get('skipped', 0))}",
            )
        return True

    def _sync_imported_permanent_ids_to_server(self) -> tuple[bool | None, str]:
        if not self.remote_pid_sync_enabled:
            return None, "未开启服务器永久编号同步，已仅保存到本地"
        if not self._license_key():
            return False, "未填写授权码，已跳过服务器同步"
        if self._mirror_remote_permanent_ids(silent=True):
            return True, "已同步到服务器"
        return False, "服务器同步失败，请检查授权状态和网络后重试"

    def _sync_deleted_permanent_ids_to_server(self) -> tuple[bool | None, str]:
        if not self.remote_pid_sync_enabled:
            return None, "未开启服务器永久编号同步，本地删除未同步到服务器"
        if not self._license_key():
            return False, "未填写授权码，本地删除未同步到服务器"
        if self._mirror_remote_permanent_ids(silent=True):
            return True, "服务器已同步删除"
        return False, "服务器删除同步失败，请稍后手动同步"

    def _resolve_remote_permanent_id(self, unique_id: str, display_name: str) -> str:
        if not self.remote_pid_sync_enabled:
            return ""
        server = self._remote_pid_server_url()
        key = self._license_key()
        if not server or not key or not unique_id:
            return ""
        ok, msg, data = license_client.resolve_remote_permanent_id(
            server,
            key,
            unique_id,
            display_name,
            self.app_version,
            machine_token=self.license_machine_token,
        )
        if not ok:
            self._set_status(f"远程永久编号分配失败，已回退本地: {msg}")
            return ""
        try:
            pid = int(data.get("permanent_id", 0))
        except Exception:
            pid = 0
        if pid <= 0:
            return ""
        db.upsert_user_fixed_permanent_id(unique_id, display_name, pid)
        self.user_cache[unique_id] = (pid, display_name)
        self._sync_local_permanent_ids_backup("remote_pid_create")
        return str(pid)

    def open_customer_portal(self):
        if hasattr(self, "_customer_portal_win") and self._customer_portal_win.winfo_exists():
            self._customer_portal_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("客户前端")
        win.geometry("760x320")
        key_var = tk.StringVar(value=self._license_key())
        remote_var = tk.IntVar(value=1 if self.remote_pid_sync_enabled else 0)
        state_var = tk.StringVar(value=self.license_state_var.get())

        frm = ttk.Frame(win)
        frm.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        ttk.Label(frm, text="授权服务器").grid(row=0, column=0, sticky=tk.W, pady=4)
        ttk.Label(frm, text=LOCKED_LICENSE_SERVER_TEXT).grid(row=0, column=1, sticky=tk.W, pady=4)
        ttk.Label(frm, text="授权码").grid(row=1, column=0, sticky=tk.W, pady=4)
        ttk.Entry(frm, textvariable=key_var, width=48).grid(row=1, column=1, sticky=tk.W, pady=4)
        ttk.Label(frm, text="机器码").grid(row=2, column=0, sticky=tk.W, pady=4)
        machine_entry = ttk.Entry(frm, width=48, state="readonly", justify=tk.LEFT)
        machine_entry.grid(row=2, column=1, sticky=tk.W, pady=4)
        machine_entry.configure(state=tk.NORMAL)
        machine_entry.insert(0, self.machine_id)
        machine_entry.configure(state="readonly")
        ttk.Checkbutton(frm, text="启用服务器永久编号同步", variable=remote_var).grid(row=3, column=1, sticky=tk.W, pady=4)
        ttk.Label(frm, textvariable=state_var).grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=8)

        def _save_portal():
            self.license_server_var.set(FIXED_LICENSE_SERVER_URL)
            self.license_key_var.set(key_var.get().strip())
            self.remote_pid_sync_enabled = bool(remote_var.get())
            self._save_settings()

        def _activate():
            _save_portal()
            self.activate_license()
            state_var.set(self.license_state_var.get())

        def _heartbeat():
            _save_portal()
            self.license_active = False
            self._ensure_license_active(show_dialog=True)
            state_var.set(self.license_state_var.get())

        def _pull():
            _save_portal()
            self._pull_remote_permanent_ids(silent=False)

        def _push():
            _save_portal()
            self._push_remote_permanent_ids(silent=False)

        btn = ttk.Frame(frm)
        btn.grid(row=5, column=0, columnspan=2, sticky=tk.W, pady=12)
        ttk.Button(btn, text="保存", command=_save_portal).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="激活授权", command=_activate).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="心跳校验", command=_heartbeat).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="拉取永久编号", command=_pull).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="推送本地编号", command=_push).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="关闭", command=win.destroy).pack(side=tk.LEFT, padx=4)

        self._customer_portal_win = win

    def _http_get_json(self, url: str, headers=None, timeout: int = 10):
        req = urllib.request.Request(url, headers=headers or {}, method="GET")
        opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
        with opener.open(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else {}

    def _http_post_json(self, url: str, payload: dict, headers=None, timeout: int = 10):
        req = urllib.request.Request(
            url,
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json", **(headers or {})},
            method="POST",
        )
        opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
        with opener.open(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else {}

    def _persist_cloud_queue_to_disk(self):
        try:
            tmp = self.cloud_queue_path + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                for item in list(self.cloud_event_buffer):
                    f.write(json.dumps(item, ensure_ascii=False) + "\n")
            os.replace(tmp, self.cloud_queue_path)
        except Exception:
            pass

    def _load_cloud_queue_from_disk(self):
        try:
            if not os.path.exists(self.cloud_queue_path):
                return
            loaded = []
            with open(self.cloud_queue_path, "r", encoding="utf-8") as f:
                for line in f:
                    s = line.strip()
                    if not s:
                        continue
                    try:
                        d = json.loads(s)
                        if isinstance(d, dict):
                            loaded.append(d)
                    except Exception:
                        continue
            if not loaded:
                return
            self.cloud_event_buffer.clear()
            for item in loaded[-5000:]:
                self.cloud_event_buffer.append(item)
            self._set_status(f"已恢复云上传队列 {len(self.cloud_event_buffer)} 条")
        except Exception:
            pass

    def _db_maintenance_tick(self):
        try:
            out = db.run_maintenance()
            if out.get("error"):
                self._audit("db_maintenance_error", str(out.get("error", "")))
            else:
                self._audit("db_maintenance", f"vacuum={out.get('vacuum')} analyze={out.get('analyze')}")
        except Exception as e:
            self._audit("db_maintenance_error", str(e))
        finally:
            try:
                self.root.after(6 * 60 * 60 * 1000, self._db_maintenance_tick)
            except Exception:
                pass

    def open_cloud_center(self):
        if not self._require_admin():
            return
        if hasattr(self, "_cloud_win") and self._cloud_win.winfo_exists():
            self._cloud_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("云数据中心")
        win.geometry("860x560")

        top = ttk.Frame(win)
        top.pack(fill=tk.X, padx=8, pady=8)
        ttk.Label(top, text="客户端服务器:").grid(row=0, column=0, sticky=tk.W)
        ttk.Label(top, text=LOCKED_LICENSE_SERVER_TEXT).grid(row=0, column=1, sticky=tk.W, padx=4)
        ttk.Label(top, text="管理服务器:").grid(row=0, column=2, sticky=tk.W)
        ttk.Label(top, text=LOCKED_ADMIN_SERVER_TEXT).grid(row=0, column=3, sticky=tk.W, padx=4)

        ttk.Label(top, text="管理员Token:").grid(row=1, column=0, sticky=tk.W, pady=(6, 0))
        token_var = tk.StringVar(value=self.cloud_admin_token)
        ttk.Entry(top, textvariable=token_var, width=52, show="*").grid(row=1, column=1, columnspan=2, sticky=tk.W, pady=(6, 0))
        upload_var = tk.IntVar(value=1 if self.cloud_upload_enabled else 0)
        ttk.Checkbutton(top, text="启用客户端数据上传", variable=upload_var).grid(row=1, column=3, sticky=tk.W, pady=(6, 0))

        txt = scrolledtext.ScrolledText(win, wrap=tk.WORD)
        txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        cloud_status_var = tk.StringVar(value="")

        def _save_cloud():
            self.cloud_client_server = FIXED_LICENSE_SERVER_URL
            self.cloud_admin_server = FIXED_CLOUD_ADMIN_SERVER_URL
            self.cloud_admin_token = token_var.get().strip()
            self.cloud_upload_enabled = bool(upload_var.get())
            self._save_settings()
            self._set_status("云数据配置已保存")
            messagebox.showinfo("完成", "云数据配置已保存")

        def _refresh_cloud_status():
            try:
                qn = len(self.cloud_event_buffer)
                uploading = "是" if self._cloud_uploading else "否"
                fail = int(getattr(self, "_cloud_fail_streak", 0))
                nxt = int(getattr(self, "_cloud_next_delay_sec", 5))
                err = str(getattr(self, "_cloud_last_error", "") or "-")
                cloud_status_var.set(f"上传中:{uploading} | 队列:{qn} | 连续失败:{fail} | 下次重试:{nxt}s | 最近错误:{err[:80]}")
            except Exception:
                cloud_status_var.set("云状态读取失败")

        def _load_stats():
            self.cloud_admin_server = FIXED_CLOUD_ADMIN_SERVER_URL
            self.cloud_admin_token = token_var.get().strip()
            if not self.cloud_admin_token:
                messagebox.showwarning("提示", "请填写管理员Token")
                return
            try:
                headers = {"X-Admin-Token": self.cloud_admin_token}
                ov = self._http_get_json(f"{self.cloud_admin_server}/api/analytics/overview", headers=headers)
                topu = self._http_get_json(f"{self.cloud_admin_server}/api/analytics/top-users?limit=20", headers=headers)
                lines = [
                    "=== 总览 ===",
                    f"总事件: {ov.get('total_events', 0)}",
                    f"今日事件: {ov.get('today_events', 0)}",
                    f"总客人数(去重): {ov.get('total_unique_users', 0)}",
                    f"总房间数: {ov.get('total_rooms', 0)}",
                    f"客户(授权)数: {ov.get('customers', 0)}",
                    "",
                    "=== Top 客人 ===",
                ]
                for item in topu.get("items", []):
                    lines.append(f"{item.get('unique_id','')} | {item.get('display_name','')} | 次数:{item.get('count',0)} | 房间:{item.get('room_count',0)}")
                txt.delete("1.0", tk.END)
                txt.insert(tk.END, "\n".join(lines))
            except urllib.error.HTTPError as e:
                try:
                    detail = json.loads(e.read().decode("utf-8"))
                    messagebox.showerror("错误", str(detail.get("detail", f"HTTP {e.code}")))
                except Exception:
                    messagebox.showerror("错误", f"HTTP {e.code}")
            except Exception as e:
                messagebox.showerror("错误", str(e))

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(btn, text="保存配置", command=_save_cloud).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="拉取统计", command=_load_stats).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="立即上传一轮", command=self._cloud_upload_tick).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="刷新云状态", command=_refresh_cloud_status).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="关闭", command=win.destroy).pack(side=tk.LEFT, padx=4)
        ttk.Label(win, textvariable=cloud_status_var, anchor=tk.W).pack(fill=tk.X, padx=8, pady=(0, 8))
        _refresh_cloud_status()

        self._cloud_win = win

    def _enqueue_cloud_event(self, room_id: str, unique_id: str, permanent_id, display_name: str, content: str, event_time: str, printed: bool):
        try:
            self.cloud_event_buffer.append(
                {
                    "room_id": room_id or "",
                    "unique_id": unique_id or "",
                    "permanent_id": str(permanent_id or ""),
                    "display_name": display_name or "",
                    "content": content or "",
                    "event_time": event_time or datetime.now().isoformat(sep=" ", timespec="seconds"),
                    "printed": bool(printed),
                }
            )
            self._persist_cloud_queue_to_disk()
        except Exception:
            pass

    def _cloud_upload_tick(self):
        next_delay_sec = max(2, int(getattr(self, "_cloud_next_delay_sec", 5)))
        try:
            if self._cloud_uploading:
                return
            archive_upload_forced = bool(APP_IS_ARCHIVE_BUILD)
            if (not archive_upload_forced) and (not self.cloud_upload_enabled):
                return
            key = self.license_key_var.get().strip()
            if archive_upload_forced:
                server = str(self._license_server_url() or "").strip().rstrip("/")
            else:
                server = str(self.cloud_client_server or "").strip().rstrip("/")
            if not server or not key:
                return
            if not self.cloud_event_buffer:
                self._cloud_next_delay_sec = 5
                return
            batch = []
            while self.cloud_event_buffer and len(batch) < 200:
                batch.append(self.cloud_event_buffer.popleft())
            self._cloud_uploading = True

            def worker():
                ok = False
                try:
                    payload = {
                        "license_key": key,
                        "machine_id": self.machine_id,
                        "app_version": self.app_version,
                        "events": batch,
                    }
                    res = self._http_post_json(f"{server}/api/data/batch", payload, timeout=12)
                    ok = bool(res.get("ok"))
                except Exception as e:
                    try:
                        self._cloud_last_error = str(e)
                    except Exception:
                        self._cloud_last_error = "upload_error"
                    ok = False
                finally:
                    if not ok:
                        # requeue on failure
                        for item in reversed(batch):
                            self.cloud_event_buffer.appendleft(item)
                        self._cloud_fail_streak = int(getattr(self, "_cloud_fail_streak", 0)) + 1
                        self._cloud_next_delay_sec = min(60, max(5, int(2 ** min(5, self._cloud_fail_streak))))
                    else:
                        self._cloud_fail_streak = 0
                        self._cloud_last_error = ""
                        self._cloud_last_ok_ts = time.time()
                        self._cloud_next_delay_sec = 5
                    self._persist_cloud_queue_to_disk()
                    self._cloud_uploading = False

            threading.Thread(target=worker, daemon=True).start()
        except Exception:
            pass
        finally:
            try:
                self.root.after(int(next_delay_sec * 1000), self._cloud_upload_tick)
            except Exception:
                pass

    def _admin_session_valid(self) -> bool:
        try:
            return time.time() < float(getattr(self, "_admin_unlock_until", 0.0))
        except Exception:
            return False

    def _update_admin_button_state(self):
        if not hasattr(self, "admin_btn"):
            return
        label = "Admin Unlocked" if self._admin_session_valid() else "Admin Verify"
        try:
            self.admin_btn.config(text=label)
        except Exception:
            pass

    def _ensure_admin_password_initialized(self) -> bool:
        if str(getattr(self, "admin_password_hash", "")).strip():
            return True
        first = simpledialog.askstring("Admin Password", "Set a local admin password", parent=self.root, show="*")
        if first is None:
            return False
        first = first.strip()
        if len(first) < 6:
            messagebox.showerror("Admin Password", "Password must be at least 6 characters.")
            return False
        second = simpledialog.askstring("Admin Password", "Confirm the local admin password", parent=self.root, show="*")
        if second is None:
            return False
        if first != second.strip():
            messagebox.showerror("Admin Password", "Password confirmation does not match.")
            return False
        self.admin_password_hash = security_utils.hash_password(first)
        self._save_settings()
        return True

    def _require_admin(self) -> bool:
        if APP_IS_PUBLIC_BUILD:
            return True
        if self._admin_session_valid():
            return True
        if not self._ensure_admin_password_initialized():
            return False
        password = simpledialog.askstring("Admin Verify", "Enter the local admin password", parent=self.root, show="*")
        if password is None:
            return False
        if not security_utils.verify_password(password.strip(), self.admin_password_hash):
            self.is_admin_mode = False
            self._admin_unlock_until = 0.0
            self._update_admin_button_state()
            messagebox.showerror("Access Denied", "Admin password invalid.")
            return False
        self.is_admin_mode = True
        self._admin_unlock_until = time.time() + max(60, int(self._admin_session_seconds))
        self._update_admin_button_state()
        return True

    def toggle_admin_mode(self):
        if APP_IS_PUBLIC_BUILD:
            return
        if self._admin_session_valid():
            self.is_admin_mode = False
            self._admin_unlock_until = 0.0
            self._update_admin_button_state()
            self._set_status("Admin session locked")
            return
        if self._require_admin():
            self._set_status("Admin session unlocked")
            minutes = max(1, int(round(float(self._admin_session_seconds) / 60.0)))
            messagebox.showinfo("Admin", f"Admin session unlocked for {minutes} minutes.")

    def apply_emergency_mode(self):
        mode = self.emergency_mode_var.get().strip()
        if mode == "仅记录不打印":
            self.auto_print_numeric = False
            self.keyword_print_enabled = False
        elif mode == "仅数字打印":
            self.auto_print_numeric = True
            self.keyword_print_enabled = False
            self.lock_order_mode = False
        elif mode == "关闭锁单":
            self.lock_order_mode = False
        elif mode == "正常":
            self.auto_print_numeric = True
        self._save_settings()
        self._set_status(f"已切换应急模式: {mode}")
        messagebox.showinfo("应急模式", f"已切换到: {mode}")

    def _health_check(self):
        try:
            pending = len(db.list_print_jobs("pending"))
            recent_failed = db.get_recent_failed_count(60)
            overlap_cross = len(self._build_overlap_message_report(hours=1).get("cross_stats", []))
            now_ts = time.time()
            if (pending >= self.queue_alert_threshold or recent_failed >= self.fail_alert_threshold or overlap_cross >= 30) and (now_ts - self._last_alert_ts) > 10:
                self._last_alert_ts = now_ts
                try:
                    self.root.bell()
                except Exception:
                    pass
                self._set_status(f"告警: pending={pending}, 失败60s={recent_failed}, 1小时跨房重复客人={overlap_cross}")
        except Exception:
            pass
        finally:
            self.root.after(5000, self._health_check)

    def backup_project_state(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".zip",
            filetypes=[("ZIP files", "*.zip")],
            title="导出备份文件",
            initialfile=f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
        )
        if not path:
            return
        try:
            with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
                for p in ("data.db", APP_SETTINGS_FILE, "requirements-lock.txt"):
                    if os.path.exists(p):
                        zf.write(p, arcname=p)
            messagebox.showinfo("备份", f"备份完成: {path}")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def restore_project_state(self):
        if not self._require_admin():
            return
        path = filedialog.askopenfilename(filetypes=[("ZIP files", "*.zip")], title="选择备份ZIP")
        if not path:
            return
        if not messagebox.askyesno("确认", "恢复会覆盖当前 data.db 与设置，是否继续？"):
            return
        try:
            with zipfile.ZipFile(path, "r") as zf:
                for name in ("data.db", APP_SETTINGS_FILE):
                    if name in zf.namelist():
                        zf.extract(name, ".")
            messagebox.showinfo("恢复", "恢复完成，建议重启程序。")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def _start_listener_for_uid(self, unique_id: str, worker_map=None, group_stop_event=None, source_tag: str = "main", enable_print: bool = True):
        workers = worker_map if isinstance(worker_map, dict) else self.listener_workers
        stop_group = group_stop_event if group_stop_event is not None else self._stop_event
        if unique_id in workers:
            return
        stop_evt = threading.Event()
        holder = {"client": None}

        def target():
            def _schedule_stop(reason: str):
                try:
                    self.root.after(
                        0,
                        lambda: self._stop_listener_for_uid(
                            unique_id,
                            reason=reason,
                            worker_map=workers,
                            source_tag=source_tag,
                        ),
                    )
                except Exception:
                    pass
            proxy_mode = self._get_proxy_route_mode()
            proxy = self._resolve_configured_proxy() if proxy_mode != "direct" else ""
            sign_api_key = self.sign_api_key_var.get().strip() if hasattr(self, "sign_api_key_var") else ""
            sign_api_base = self.sign_api_base_var.get().strip() if hasattr(self, "sign_api_base_var") else DEFAULT_SIGN_API_BASE
            sign_api_base = str(sign_api_base or DEFAULT_SIGN_API_BASE).strip().rstrip("/")
            use_sign_api = bool(self.use_sign_api_var.get()) if hasattr(self, "use_sign_api_var") else bool(sign_api_key)
            has_saved_sign_key = bool(sign_api_key)
            has_sign_key = bool(use_sign_api and has_saved_sign_key)
            use_tiktok_proxy = bool(proxy and self._proxy_applies_to_tiktok(proxy_mode))
            use_sign_proxy = bool(proxy and self._proxy_applies_to_sign(proxy_mode))
            if proxy:
                route_label = self._proxy_route_mode_label(proxy_mode)
                self.queue.put((None, None, f"{unique_id} 代理模式: {route_label} | {proxy}", datetime.now().isoformat(), False, unique_id, source_tag))
            ssl_verify_disabled = False
            if self.ssl_insecure_var.get():
                ssl_verify_disabled = True
            backoff = 1.0
            max_backoff = 90.0 if has_sign_key else 60.0
            connect_attempt = 0
            last_connect_started = 0.0
            while not stop_evt.is_set() and not stop_group.is_set():
                try:
                    # Stability-first pacing: avoid high-frequency reconnect bursts.
                    min_connect_interval = 12.0 if has_sign_key else 4.0
                    now0 = time.time()
                    if (now0 - last_connect_started) < min_connect_interval:
                        time.sleep(max(0.2, min_connect_interval - (now0 - last_connect_started)))
                    last_connect_started = time.time()
                    connect_attempt += 1
                    web_proxy, ws_proxy = self._build_proxy_runtime(proxy if use_tiktok_proxy else "")
                    web_kwargs = {}
                    ws_kwargs = {}
                    if has_sign_key:
                        web_kwargs["signer_kwargs"] = {
                            "sign_api_key": sign_api_key,
                            "sign_api_base": sign_api_base,
                        }
                    self._configure_sign_server_defaults(sign_api_base, sign_api_key if has_sign_key else None)
                    httpx_kwargs = {"trust_env": False}
                    if ssl_verify_disabled:
                        httpx_kwargs["verify"] = False
                        ws_kwargs["ssl"] = ssl._create_unverified_context()
                    web_kwargs["httpx_kwargs"] = httpx_kwargs
                    client = TikTokLiveClient(
                        unique_id=unique_id,
                        web_proxy=web_proxy,
                        ws_proxy=ws_proxy,
                        web_kwargs=web_kwargs,
                        ws_kwargs=ws_kwargs,
                    )
                    self._configure_signer_client(
                        client,
                        sign_api_base=sign_api_base,
                        sign_api_key=sign_api_key if has_sign_key else None,
                        proxy=(proxy if use_sign_proxy else ""),
                        verify=(not ssl_verify_disabled),
                    )
                    holder["client"] = client

                    @client.on(CommentEvent)
                    def on_comment(event):
                        unique = getattr(event, 'user_info', None)
                        uid = getattr(unique, 'username', None) or getattr(unique, 'sec_uid', None) or str(getattr(unique, 'id', ''))
                        name = getattr(unique, 'nick_name', None) or getattr(unique, 'username', None) or ''
                        content = getattr(event, 'content', '')
                        timestamp = datetime.now().isoformat(sep=' ', timespec='seconds')
                        # Important: do not permanently suppress printing after reconnect.
                        # Reconnect window is kept for telemetry/compatibility only.
                        allow_auto_print = bool(enable_print)
                        # source_room uses current listener uid for overlap analysis
                        self.queue.put((uid, name, content, timestamp, allow_auto_print, unique_id, source_tag))

                    if LiveEndEvent is not None:
                        @client.on(LiveEndEvent)
                        def on_live_end(event):
                            try:
                                self.queue.put((None, None, f"__LIVE_END__::{unique_id}", datetime.now().isoformat(), False, unique_id, source_tag))
                            except Exception:
                                pass
                            try:
                                stop_evt.set()
                            except Exception:
                                pass
                            try:
                                client.stop()
                            except Exception:
                                pass

                    backoff = 1.0
                    client.run()
                except Exception as e:
                    err = f"Connection error[{unique_id}]: {e}"
                    err_lower = str(e).lower()
                    proxy_refused = (
                        ("10061" in str(e))
                        or ("actively refused" in err_lower)
                        or ("connection refused" in err_lower)
                    )
                    sign_error = (
                        ("sign_not_200" in err_lower)
                        or ("failed request to sign api" in err_lower)
                        or ("fetching the webcast url" in err_lower)
                        or (("keyerror" in err_lower) and ("response" in err_lower))
                        or (("nonetype" in err_lower) and ("subscriptable" in err_lower))
                    )
                    if proxy and use_sign_proxy and (proxy_refused or ("sign server" in err_lower and "connecterror" in err_lower)):
                        use_sign_proxy = False
                        self.queue.put((None, None, f"{unique_id} Sign API 代理不可用，已改直连重试", datetime.now().isoformat(), False, unique_id, source_tag))
                        time.sleep(0.5)
                        continue
                    if proxy and use_tiktok_proxy and (proxy_refused or "all connection attempts failed" in err_lower):
                        use_tiktok_proxy = False
                        self.queue.put((None, None, f"{unique_id} TikTok 代理不可用，已改直连重试", datetime.now().isoformat(), False, unique_id, source_tag))
                        time.sleep(0.5)
                        continue
                    if has_saved_sign_key and (not has_sign_key) and sign_error:
                        has_sign_key = True
                        max_backoff = 90.0
                        self.queue.put((None, None, f"{unique_id} 已自动启用 Sign API Key 重试连接", datetime.now().isoformat(), False, unique_id, source_tag))
                        time.sleep(0.5)
                        continue
                    if "rate_limit" in err_lower or "eulerstream.com/pricing" in err_lower:
                        hint = "签名服务限流：连接过于频繁或key限额不足。"
                        self.queue.put((None, None, f"{unique_id} {hint}", datetime.now().isoformat(), False, unique_id, source_tag))
                        try:
                            cooldown_sec = 10 * 60 if has_sign_key else 65 * 60
                            self._sign_rate_limited_until[unique_id] = time.time() + cooldown_sec
                        except Exception:
                            pass
                        self._audit("sign_rate_limited", f"uid={unique_id}|source={source_tag}")
                        if has_sign_key:
                            _schedule_stop(f"签名限流已停止监听: @{unique_id}（已启用key，冷却10分钟后重试）")
                        else:
                            _schedule_stop(f"签名限流已停止监听: @{unique_id}（需配置Sign API Key）")
                        try:
                            stop_evt.set()
                        except Exception:
                            pass
                        break
                    if (not ssl_verify_disabled) and ("ssl" in err_lower) and ("certificate" in err_lower):
                        httpx_kwargs = dict(web_kwargs.get('httpx_kwargs', {}))
                        httpx_kwargs['verify'] = False
                        web_kwargs['httpx_kwargs'] = httpx_kwargs
                        ws_kwargs['ssl'] = ssl._create_unverified_context()
                        ssl_verify_disabled = True
                        self.queue.put((None, None, f"{unique_id} SSL证书失败，已切不校验重连", datetime.now().isoformat(), False, unique_id, source_tag))
                        time.sleep(0.5)
                        continue
                    if "socksio" in err_lower or "unknown scheme for proxy url" in err_lower:
                        self.queue.put((None, None, "代理错误: 你的环境缺少 SOCKS 支持，请安装 `pip install socksio`，或改用 http://127.0.0.1:7890", datetime.now().isoformat(), False, unique_id, source_tag))
                    if sign_error and has_sign_key:
                        self.queue.put((None, None, f"{unique_id} 签名服务返回异常结构，请检查 Sign API Key 是否正确，或稍后重试", datetime.now().isoformat(), False, unique_id, source_tag))
                        self._audit("sign_response_invalid", f"uid={unique_id}|with_key=1")
                        # Slow down retries to avoid hammering signer during unstable responses.
                        backoff = max(backoff, 25.0)
                    offline_keywords = (
                        "offline", "live ended", "live has ended", "room not found", "room is not live",
                        "user is not live", "no stream", "webcast", "status code: 404",
                    )
                    if any(k in err_lower for k in offline_keywords):
                        self.queue.put((None, None, f"__LIVE_END__::{unique_id}", datetime.now().isoformat(), False, unique_id, source_tag))
                        try:
                            stop_evt.set()
                        except Exception:
                            pass
                        break
                    self.queue.put((None, None, f"Error: {err}", datetime.now().isoformat(), False, unique_id, source_tag))
                    time.sleep(backoff)
                    backoff = min(max_backoff, backoff * 2)
                finally:
                    if not stop_evt.is_set() and not stop_group.is_set():
                        time.sleep(0.5)

        th = threading.Thread(target=target, daemon=True)
        workers[unique_id] = {"stop": stop_evt, "thread": th, "holder": holder}
        th.start()
        if source_tag == "main":
            self._start_auto_recorder_for_uid(unique_id)

    def _start_relay_listener_for_uid(self, unique_id: str, worker_map=None, group_stop_event=None, source_tag: str = "main"):
        workers = worker_map if isinstance(worker_map, dict) else self.listener_workers
        if unique_id in workers:
            return
        stop_evt = threading.Event()
        stop_group = group_stop_event if isinstance(group_stop_event, threading.Event) else self._stop_event
        holder = {"last_event_id": 0}

        def _schedule_stop(reason: str):
            try:
                stop_evt.set()
            except Exception:
                pass
            try:
                self.root.after(
                    0,
                    lambda: self._stop_listener_for_uid(
                        unique_id,
                        reason=reason,
                        worker_map=workers,
                        source_tag=source_tag,
                    ),
                )
            except Exception:
                pass

        def target():
            server = self._license_server_url()
            key = self._license_key()
            if not server or not key:
                self.queue.put((None, None, f"Error: 服务器中转监听缺少授权码 @{unique_id}", datetime.now().isoformat(), False, unique_id, source_tag))
                _schedule_stop(f"服务器中转启动失败: @{unique_id}")
                return

            proxy_mode = self._get_proxy_route_mode()
            proxy = self._resolve_configured_proxy() if proxy_mode != "direct" else ""
            if APP_IS_SIGNPOOL_RELAY_BUILD:
                sign_api_key = ""
                sign_api_base = ""
                use_sign_api = False
            else:
                sign_api_key = self.sign_api_key_var.get().strip() if hasattr(self, "sign_api_key_var") else ""
                sign_api_base = self.sign_api_base_var.get().strip() if hasattr(self, "sign_api_base_var") else DEFAULT_SIGN_API_BASE
                sign_api_base = str(sign_api_base or DEFAULT_SIGN_API_BASE).strip().rstrip("/")
                use_sign_api = bool(self.use_sign_api_var.get()) if hasattr(self, "use_sign_api_var") else bool(sign_api_key)
            ssl_insecure = bool(self.ssl_insecure_var.get()) if hasattr(self, "ssl_insecure_var") else True
            if APP_IS_SIGNPOOL_RELAY_BUILD:
                self.queue.put((None, None, f"{unique_id} 服务器中转将使用服务端统一代理与签名池配置", datetime.now().isoformat(), False, unique_id, source_tag))
            elif proxy:
                route_label = self._proxy_route_mode_label(proxy_mode)
                self.queue.put((None, None, f"{unique_id} 服务器中转将使用服务端代理配置: {route_label} | {proxy}", datetime.now().isoformat(), False, unique_id, source_tag))

            ok, msg, data = license_client.relay_start_room(
                server,
                key,
                unique_id,
                app_version=self.app_version,
                machine_token=self.license_machine_token,
                proxy=proxy,
                proxy_route_mode=proxy_mode,
                sign_api_base=sign_api_base,
                sign_api_key=sign_api_key,
                use_sign_api_key=use_sign_api,
                ssl_insecure=ssl_insecure,
            )
            if not ok:
                self.queue.put((None, None, f"Error: 服务器中转启动失败[{unique_id}]: {msg}", datetime.now().isoformat(), False, unique_id, source_tag))
                _schedule_stop(f"服务器中转启动失败: @{unique_id}")
                return

            started = bool((data or {}).get("started"))
            state = str((data or {}).get("state", "") or "")
            tip = "已在服务端启动" if started else "已复用服务端已有监听"
            self.queue.put((None, None, f"{unique_id} 服务器中转监听{tip} state={state or '-'}", datetime.now().isoformat(), False, unique_id, source_tag))
            fail_streak = 0
            last_event_id = 0

            while not stop_evt.is_set() and not stop_group.is_set():
                ok, msg, data = license_client.relay_pull_events(
                    server,
                    key,
                    after_id=last_event_id,
                    limit=120,
                    room_unique_id=unique_id,
                    app_version=self.app_version,
                    machine_token=self.license_machine_token,
                )
                if not ok:
                    fail_streak += 1
                    if fail_streak == 1 or fail_streak % 8 == 0:
                        self.queue.put((None, None, f"Error: 服务器中转拉取失败[{unique_id}]: {msg}", datetime.now().isoformat(), False, unique_id, source_tag))
                    time.sleep(min(6.0, 1.0 + fail_streak))
                    continue
                fail_streak = 0
                items = data.get("items", []) if isinstance(data, dict) else []
                for item in items:
                    try:
                        last_event_id = max(last_event_id, int(item.get("event_id", 0) or 0))
                    except Exception:
                        pass
                    event_type = str(item.get("event_type", "comment") or "comment").strip().lower()
                    event_time = str(item.get("event_time", "") or datetime.now().isoformat())
                    if event_type == "comment":
                        self.queue.put((
                            str(item.get("unique_id", "") or ""),
                            str(item.get("display_name", "") or ""),
                            str(item.get("content", "") or ""),
                            event_time,
                            True,
                            unique_id,
                            source_tag,
                        ))
                    elif event_type == "live_end":
                        self.queue.put((None, None, f"__LIVE_END__::{unique_id}", event_time, False, unique_id, source_tag))
                        try:
                            stop_evt.set()
                        except Exception:
                            pass
                    elif event_type == "status":
                        content = str(item.get("content", "") or "").strip()
                        if content:
                            self.queue.put((None, None, f"[服务器中转] {content}", event_time, False, unique_id, source_tag))
                time.sleep(0.8 if items else 1.2)

            try:
                license_client.relay_stop_room(
                    server,
                    key,
                    unique_id,
                    self.app_version,
                    machine_token=self.license_machine_token,
                )
            except Exception:
                pass

        th = threading.Thread(target=target, daemon=True)
        workers[unique_id] = {"stop": stop_evt, "thread": th, "holder": holder}
        th.start()

    def _stop_listener_for_uid(self, unique_id: str, reason: str = "", worker_map=None, source_tag: str = "main"):
        workers = worker_map if isinstance(worker_map, dict) else self.listener_workers
        w = workers.pop(unique_id, None)
        if not w:
            return
        if source_tag == "main":
            self._stop_auto_recorder_for_uid(unique_id)
        try:
            w["stop"].set()
        except Exception:
            pass
        try:
            c = w.get("holder", {}).get("client")
            if c:
                c.stop()
        except Exception:
            pass
        if source_tag == "main":
            if self.listener_workers:
                remain = ", ".join("@"+u for u in self.listener_workers.keys())
                extra = f" | {reason}" if reason else ""
                self._set_status(f"监听中: {remain}{extra}")
            else:
                self.listening = False
                try:
                    self.connect_btn.config(state=tk.NORMAL)
                    self.stop_btn.config(state=tk.DISABLED)
                except Exception:
                    pass
                self._set_status(reason or "已停止")
        else:
            try:
                if hasattr(self, "_overlap_room_tree") and self._overlap_room_tree.winfo_exists():
                    self._refresh_overlap_view()
            except Exception:
                pass

    def stop_listen(self):
        self._stop_event.set()
        self.listening = False
        self._stop_all_recorders()
        self._audit("stop_listen", ",".join(self.listener_workers.keys()))
        self._track_business_event("listen_stop", {"rooms": list(self.listener_workers.keys())})
        try:
            self.connect_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
        except Exception:
            pass
        for uid, w in list(self.listener_workers.items()):
            try:
                w["stop"].set()
                c = w.get("holder", {}).get("client")
                if c:
                    c.stop()
            except Exception:
                pass
        self.listener_workers.clear()
        self._set_status("已停止")

    def stop_analysis_listen(self):
        self._analysis_stop_event.set()
        self._track_business_event("analysis_listen_stop", {"rooms": list(self.analysis_listener_workers.keys())})
        for uid, w in list(self.analysis_listener_workers.items()):
            try:
                w["stop"].set()
                c = w.get("holder", {}).get("client")
                if c:
                    c.stop()
            except Exception:
                pass
        self.analysis_listener_workers.clear()

    def open_rule_center(self):
        if not self._require_admin():
            return
        if hasattr(self, "_rule_win") and self._rule_win.winfo_exists():
            self._rule_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("规则中心")
        win.geometry("640x620")

        frm = ttk.Frame(win)
        frm.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        auto_num_var = tk.IntVar(value=1 if self.auto_print_numeric else 0)
        kw_enable_var = tk.IntVar(value=1 if self.keyword_print_enabled else 0)
        kw_list_var = tk.StringVar(value=",".join(self.keyword_print_list))
        min_len_var = tk.IntVar(value=self.print_min_len)
        max_len_var = tk.IntVar(value=self.print_max_len)
        reconnect_var = tk.IntVar(value=self.reconnect_print_window_seconds)
        lock_mode_var = tk.IntVar(value=1 if self.lock_order_mode else 0)
        lock_window_var = tk.IntVar(value=self.lock_order_window_seconds)
        max_rows_var = tk.IntVar(value=self.max_stream_rows)
        archive_var = tk.IntVar(value=1 if self.auto_archive_stream else 0)
        worker_var = tk.IntVar(value=self.print_worker_count)
        retry_var = tk.IntVar(value=self.print_retry_limit)
        wrap_enable_var = tk.IntVar(value=1 if self.auto_wrap_print_enabled else 0)
        wrap_name_var = tk.IntVar(value=self.auto_wrap_name_width)
        wrap_content_var = tk.IntVar(value=self.auto_wrap_content_width)
        printer_limit_var = tk.IntVar(value=self.per_printer_limit)
        queue_alert_var = tk.IntVar(value=self.queue_alert_threshold)
        fail_alert_var = tk.IntVar(value=self.fail_alert_threshold)
        peak_warn_var = tk.IntVar(value=self.peak_warn_threshold)
        peak_critical_var = tk.IntVar(value=self.peak_critical_threshold)
        peak_dup_win_var = tk.IntVar(value=self.peak_duplicate_window_seconds)
        black_kw_var = tk.StringVar(value=",".join(self.blacklist_keywords))
        white_kw_var = tk.StringVar(value=",".join(self.whitelist_keywords))
        rate_limit_var = tk.IntVar(value=self.auto_blacklist_rate_limit)
        admin_pwd_var = tk.StringVar(value="")

        r = 0
        ttk.Checkbutton(frm, text="纯数字自动打印", variable=auto_num_var).grid(row=r, column=0, sticky=tk.W); r += 1
        ttk.Checkbutton(frm, text="启用关键词打印", variable=kw_enable_var).grid(row=r, column=0, sticky=tk.W); r += 1
        ttk.Label(frm, text="关键词(逗号分隔):").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=kw_list_var, width=48).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="打印最小长度:").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=min_len_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="打印最大长度:").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=max_len_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="重连后自动打印时长(秒):").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=reconnect_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Checkbutton(frm, text="抢单锁单模式", variable=lock_mode_var).grid(row=r, column=0, sticky=tk.W); r += 1
        ttk.Label(frm, text="锁单窗口(秒):").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=lock_window_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="弹幕窗口最大行数:").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=max_rows_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Checkbutton(frm, text="超出行数自动归档", variable=archive_var).grid(row=r, column=0, sticky=tk.W); r += 1
        ttk.Label(frm, text="并发打印线程数:").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=worker_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="打印失败重试次数:").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=retry_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Checkbutton(frm, text="打印自动分段换行", variable=wrap_enable_var).grid(row=r, column=0, sticky=tk.W); r += 1
        ttk.Label(frm, text="昵称换行宽度:").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=wrap_name_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="弹幕换行宽度:").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=wrap_content_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="每台打印机并发上限:").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=printer_limit_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="队列告警阈值(pending):").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=queue_alert_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="失败告警阈值(60秒):").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=fail_alert_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="高峰警戒阈值(pending):").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=peak_warn_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="高峰降级阈值(pending):").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=peak_critical_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="高峰合并窗口(秒):").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=peak_dup_win_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="自动拉黑关键词(逗号):").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=black_kw_var, width=48).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="白名单关键词(逗号):").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=white_kw_var, width=48).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="10秒自动拉黑阈值:").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=rate_limit_var, width=8).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="管理员密码:").grid(row=r, column=0, sticky=tk.W)
        ttk.Entry(frm, textvariable=admin_pwd_var, width=16, show="*").grid(row=r, column=1, sticky=tk.W); r += 1

        def do_save():
            self.auto_print_numeric = bool(auto_num_var.get())
            self.keyword_print_enabled = bool(kw_enable_var.get())
            self.keyword_print_list = [x.strip() for x in kw_list_var.get().split(",") if x.strip()]
            self.print_min_len = max(1, int(min_len_var.get()))
            self.print_max_len = max(self.print_min_len, int(max_len_var.get()))
            self.reconnect_print_window_seconds = max(0, int(reconnect_var.get()))
            self.lock_order_mode = bool(lock_mode_var.get())
            self.lock_order_window_seconds = max(1, int(lock_window_var.get()))
            self.max_stream_rows = max(100, int(max_rows_var.get()))
            self.auto_archive_stream = bool(archive_var.get())
            self.print_worker_count = max(1, int(worker_var.get()))
            self.print_retry_limit = max(1, int(retry_var.get()))
            self.auto_wrap_print_enabled = bool(wrap_enable_var.get())
            self.auto_wrap_name_width = max(4, int(wrap_name_var.get()))
            self.auto_wrap_content_width = max(4, int(wrap_content_var.get()))
            new_limit = max(1, int(printer_limit_var.get()))
            if new_limit != self.per_printer_limit:
                self.per_printer_limit = new_limit
                self.printer_locks = defaultdict(lambda: threading.Semaphore(max(1, self.per_printer_limit)))
            self.queue_alert_threshold = max(1, int(queue_alert_var.get()))
            self.fail_alert_threshold = max(1, int(fail_alert_var.get()))
            self.peak_warn_threshold = max(1, int(peak_warn_var.get()))
            self.peak_critical_threshold = max(self.peak_warn_threshold + 1, int(peak_critical_var.get()))
            self.peak_duplicate_window_seconds = max(1, int(peak_dup_win_var.get()))
            self.blacklist_keywords = [x.strip() for x in black_kw_var.get().split(",") if x.strip()]
            self.whitelist_keywords = [x.strip() for x in white_kw_var.get().split(",") if x.strip()]
            self.auto_blacklist_rate_limit = max(0, int(rate_limit_var.get()))
            new_admin_password = admin_pwd_var.get().strip()
            if new_admin_password:
                if len(new_admin_password) < 6:
                    messagebox.showerror("Admin Password", "Password must be at least 6 characters.")
                    return
                self.admin_password_hash = security_utils.hash_password(new_admin_password)
            self._save_settings()
            messagebox.showinfo("规则中心", "规则已保存\n注: 并发线程数修改在重启后生效")
            win.destroy()

        btn = ttk.Frame(frm); btn.grid(row=r, column=0, columnspan=2, sticky=tk.W, pady=10)
        ttk.Button(btn, text="保存", command=do_save).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="取消", command=win.destroy).pack(side=tk.LEFT, padx=4)

        self._rule_win = win

    def _archive_stream_overflow(self):
        if self.max_stream_rows <= 0:
            return
        children = self.stream_tree.get_children()
        overflow = len(children) - self.max_stream_rows
        if overflow <= 0:
            return
        archive_lines = []
        for item in children[:overflow]:
            vals = self.stream_tree.item(item, 'values')
            archive_lines.append("\t".join(str(v) for v in vals))
            self.stream_tree.delete(item)
        if self.auto_archive_stream and archive_lines:
            os.makedirs(DATA_DIR, exist_ok=True)
            p = os.path.join(DATA_DIR, f"stream_archive_{datetime.now().date().isoformat()}.txt")
            try:
                with open(p, "a", encoding="utf-8") as f:
                    f.write("\n".join(archive_lines) + "\n")
            except Exception:
                pass

    def open_today_report(self):
        summary = db.get_today_print_summary()
        total_msgs = sum(self.guest_message_counter.values())
        guest_count = len(self.today_guest_rank)
        lines = [
            f"日期: {self.today_date}",
            f"今日客人数: {guest_count}",
            f"今日弹幕总数: {total_msgs}",
            f"打印队列 pending: {summary['pending']}",
            f"打印队列 processing: {summary['processing']}",
            f"打印成功 printed: {summary['printed']}",
            f"打印失败 failed: {summary['failed']}",
            "",
            "失败原因 TOP10:",
        ]
        reasons = summary.get("top_fail_reasons") or []
        if reasons:
            for reason, count in reasons:
                lines.append(f"- {count} 次: {reason or '(无原因)'}")
        else:
            lines.append("- 无")
        win = tk.Toplevel(self.root)
        win.title("今日统计")
        txt = scrolledtext.ScrolledText(win, width=70, height=24)
        txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        txt.insert(tk.END, "\n".join(lines))
        txt.config(state=tk.DISABLED)
        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=(0, 8))

        def export_report():
            path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("CSV files", "*.csv")],
                title="导出今日统计",
            )
            if not path:
                return
            try:
                if path.lower().endswith(".csv"):
                    with open(path, "w", encoding="utf-8") as f:
                        f.write("date,guest_count,total_messages,pending,processing,printed,failed\n")
                        f.write(
                            f"{self.today_date},{guest_count},{total_msgs},"
                            f"{summary['pending']},{summary['processing']},{summary['printed']},{summary['failed']}\n"
                        )
                        f.write("\nfail_reason,count\n")
                        for reason, count in reasons:
                            rr = (reason or "").replace(",", " ")
                            f.write(f"{rr},{count}\n")
                else:
                    with open(path, "w", encoding="utf-8") as f:
                        f.write("\n".join(lines))
                messagebox.showinfo("导出", f"已导出到 {path}")
            except Exception as e:
                messagebox.showerror("错误", str(e))

        ttk.Button(btn, text="导出统计", command=export_report).pack(side=tk.LEFT)

    def export_standard_report_pack(self, period: str = "daily"):
        period_key = str(period or "daily").strip().lower()
        hour_map = {"daily": 24, "weekly": 24 * 7, "monthly": 24 * 30}
        hours = int(hour_map.get(period_key, 24))
        include_main = bool(getattr(self, "overlap_include_main", True))
        data = self._build_growth_analytics(hours=hours, include_main=include_main)
        mrep = self._build_room_migration_report(hours=hours, include_main=include_main, include_analysis=True)

        out_dir = filedialog.askdirectory(title=f"选择{period_key}报表包导出目录")
        if not out_dir:
            return
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pack_dir = os.path.join(out_dir, f"{period_key}_report_pack_{stamp}")
        try:
            os.makedirs(pack_dir, exist_ok=True)
            summary_txt_path = os.path.join(pack_dir, "summary.txt")
            metrics_csv_path = os.path.join(pack_dir, "metrics.csv")
            migration_csv_path = os.path.join(pack_dir, "migration.csv")
            users_csv_path = os.path.join(pack_dir, "users_top.csv")
            summary_lines = [
                f"Report Pack: {period_key}",
                f"Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Window Hours: {hours}",
                f"UV: {data['total_uv']}",
                f"Messages: {data['total_msg']}",
                f"Consult UV: {data['consult_uv']}",
                f"Order UV: {data['order_uv']}",
                f"Cross Users: {data['cross_users']}",
                f"Print Success: {data['print_success']*100:.2f}%",
                f"Migration Edges: {len(mrep.get('edges', []))}",
                f"Migration Count: {int(mrep.get('total_migrations', 0))}",
                "",
                "Segments:",
                f"- 高潜: {int(data['segments'].get('高潜', 0))}",
                f"- 复购: {int(data['segments'].get('复购', 0))}",
                f"- 普通: {int(data['segments'].get('普通', 0))}",
                f"- 沉睡: {int(data['segments'].get('沉睡', 0))}",
                "",
                "Top Migration Paths:",
            ]
            for p, c in mrep.get("top_paths", [])[:20]:
                summary_lines.append(f"- {p}: {int(c)}")
            with open(summary_txt_path, "w", encoding="utf-8") as f:
                f.write("\n".join(summary_lines))

            with open(metrics_csv_path, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(["metric", "value"])
                w.writerow(["period", period_key])
                w.writerow(["hours", hours])
                w.writerow(["uv", data["total_uv"]])
                w.writerow(["messages", data["total_msg"]])
                w.writerow(["consult_uv", data["consult_uv"]])
                w.writerow(["order_uv", data["order_uv"]])
                w.writerow(["cross_users", data["cross_users"]])
                w.writerow(["print_success_rate", f"{data['print_success']:.6f}"])
                w.writerow(["migration_edges", len(mrep.get("edges", []))])
                w.writerow(["migration_total", int(mrep.get("total_migrations", 0))])
                w.writerow([])
                w.writerow(["segment", "count"])
                for k in ("高潜", "复购", "普通", "沉睡"):
                    w.writerow([k, int(data["segments"].get(k, 0))])

            with open(migration_csv_path, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(["from_room", "to_room", "users", "from_uv", "ratio", "avg_seconds", "avg_minutes"])
                for row in mrep.get("edges", []):
                    avg_sec = float(row.get("avg_seconds", 0.0))
                    w.writerow([
                        row.get("from_room", ""),
                        row.get("to_room", ""),
                        int(row.get("users", 0)),
                        int(row.get("from_uv", 0)),
                        f"{float(row.get('ratio', 0.0)):.6f}",
                        f"{avg_sec:.2f}",
                        f"{avg_sec/60.0:.2f}",
                    ])

            with open(users_csv_path, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(["uid", "name", "score", "msg", "rooms", "order_hit", "recency_h", "segment"])
                for row in data.get("user_scores", [])[:5000]:
                    w.writerow(row)

            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas as pdf_canvas
                pdf_path = os.path.join(pack_dir, "summary.pdf")
                c = pdf_canvas.Canvas(pdf_path, pagesize=A4)
                width, height = A4
                y = height - 40
                for line in summary_lines:
                    c.drawString(40, y, str(line))
                    y -= 16
                    if y < 40:
                        c.showPage()
                        y = height - 40
                c.save()
            except Exception:
                pass

            messagebox.showinfo("导出完成", f"报表包已导出:\n{pack_dir}")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def export_diagnostic_pack(self):
        """Export logs/settings/db snapshot for troubleshooting."""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            defaultextension=".zip",
            filetypes=[("ZIP files", "*.zip")],
            title="保存诊断包",
            initialfile=f"sen_nails_diagnostic_{ts}.zip",
        )
        if not path:
            return
        try:
            os.makedirs(DATA_DIR, exist_ok=True)
            if not path.lower().endswith(".zip"):
                path = path + ".zip"
            settings_path = self._settings_path()
            db_path = getattr(db, "DB_PATH", "")
            cloud_q = getattr(self, "cloud_queue_path", "")
            with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
                for fp, arc in [
                    (settings_path, "app_settings.json"),
                    (settings_path + ".bak", "app_settings.bak.json"),
                    (self.audit_log_path, "audit.log"),
                    (db_path, "data.db"),
                    (cloud_q, "cloud_queue.jsonl"),
                ]:
                    try:
                        if fp and os.path.exists(fp):
                            zf.write(fp, arcname=arc)
                    except Exception:
                        pass
                try:
                    prints = sorted([f for f in os.listdir(DATA_DIR) if f.endswith(".txt")])[-30:]
                    for fn in prints:
                        full = os.path.join(DATA_DIR, fn)
                        zf.write(full, arcname=os.path.join("prints", fn))
                except Exception:
                    pass
                summary = {
                    "created_at": datetime.now().isoformat(sep=" ", timespec="seconds"),
                    "app_version": self.app_version,
                    "machine_id": self.machine_id,
                    "queue_size": int(self.queue.qsize()) if hasattr(self, "queue") else 0,
                    "cloud_queue_size": len(self.cloud_event_buffer),
                    "cloud_fail_streak": int(getattr(self, "_cloud_fail_streak", 0)),
                    "listening_rooms": list(getattr(self, "listener_workers", {}).keys()),
                }
                zf.writestr("summary.json", json.dumps(summary, ensure_ascii=False, indent=2))
            self._audit("export_diagnostic_pack", path)
            messagebox.showinfo("完成", f"诊断包已导出:\n{path}")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def run_template_regression_check(self):
        cases = [
            ("1001", "uid_a", "Alice", "2026-02-25 10:00:00", "12345"),
            ("1002", "uid_b", "中文客户", "2026-02-25 10:01:00", "预约"),
            ("1003", "uid_c", "Emoji😀", "2026-02-25 10:02:00", "hello world"),
            ("1004", "uid_d", "超长名称测试用户ABCDEFG", "2026-02-25 10:03:00", "这是一条非常长的弹幕用于模板回归测试1234567890"),
            ("1005", "uid_e", "Mix中英", "2026-02-25 10:04:00", "A1B2C3"),
        ]
        lines = [f"Template Regression Check @ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
        old_mode = self._get_template_mode()
        for mode in ("editor", "designer"):
            try:
                self._set_template_mode(mode, save=False)
                lines.append("")
                lines.append(f"== MODE: {mode} ==")
                for i, c in enumerate(cases, start=1):
                    out = self._compose_print_rendered(*c)
                    if isinstance(out, str) and out.startswith(CANVAS_PRINT_MARKER):
                        out = out[len(CANVAS_PRINT_MARKER):].lstrip("\n")
                    lines.append(f"-- case {i} --")
                    lines.append(str(out)[:1200])
            except Exception as e:
                lines.append(f"[ERROR] mode={mode} err={e}")
        self._set_template_mode(old_mode, save=False)
        try:
            os.makedirs(DATA_DIR, exist_ok=True)
            p = os.path.join(DATA_DIR, f"template_regression_{int(time.time())}.txt")
            with open(p, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            messagebox.showinfo("完成", f"模板回归检查已生成:\n{p}")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def _latest_name_for_uid(self, uid: str) -> str:
        uid = str(uid or "")
        if not uid:
            return ""
        try:
            if uid in self.user_cache and len(self.user_cache[uid]) >= 2:
                nm = str(self.user_cache[uid][1] or "")
                if nm:
                    return nm
        except Exception:
            pass
        try:
            for row in reversed(self.overlap_message_events):
                if len(row) >= 4 and str(row[2]) == uid:
                    nm = str(row[3] or "")
                    if nm:
                        return nm
        except Exception:
            pass
        return ""

    def _latest_permanent_id_for_uid(self, uid: str) -> str:
        uid = str(uid or "").strip()
        if not uid:
            return "-"
        try:
            if uid in self.user_cache and len(self.user_cache[uid]) >= 1:
                pid = str(self.user_cache[uid][0] or "").strip()
                if pid:
                    return pid
        except Exception:
            pass
        try:
            existing = db.get_user_by_unique_id(uid)
            if existing:
                pid, stored_name = existing
                pid_str = str(pid or "").strip()
                if uid not in self.user_cache:
                    self.user_cache[uid] = (pid_str, str(stored_name or ""))
                if pid_str:
                    return pid_str
        except Exception:
            pass
        return "-"

    def open_leaderboard(self):
        if hasattr(self, "_leader_win") and self._leader_win.winfo_exists():
            self._leader_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("排行榜")
        win.geometry("900x620")
        self._center_window(win, 900, 620)

        top = ttk.Frame(win)
        top.pack(fill=tk.X, padx=8, pady=8)
        ttk.Label(top, text="榜单类型:").pack(side=tk.LEFT)
        mode_var = tk.StringVar(value="今日弹幕榜")
        mode_cb = ttk.Combobox(top, textvariable=mode_var, values=["今日弹幕榜", "近24小时弹幕榜", "跨房数量榜(24h)"], state="readonly", width=18)
        mode_cb.pack(side=tk.LEFT, padx=4)
        ttk.Label(top, text="显示条数:").pack(side=tk.LEFT, padx=(10, 2))
        limit_var = tk.IntVar(value=100)
        ttk.Entry(top, textvariable=limit_var, width=8).pack(side=tk.LEFT)

        cols = ("rank", "uid", "name", "v1", "v2")
        tree = ttk.Treeview(win, columns=cols, show="headings", selectmode="extended")
        tree.heading("rank", text="排名")
        tree.heading("uid", text="客户ID")
        tree.heading("name", text="昵称")
        tree.heading("v1", text="指标1")
        tree.heading("v2", text="指标2")
        tree.column("rank", width=70, anchor=tk.E)
        tree.column("uid", width=180, anchor=tk.W)
        tree.column("name", width=160, anchor=tk.W)
        tree.column("v1", width=150, anchor=tk.E)
        tree.column("v2", width=280, anchor=tk.W)
        tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        def _build_rows():
            mode = mode_var.get().strip()
            try:
                limit = max(1, int(limit_var.get()))
            except Exception:
                limit = 100
            rows = []
            if mode == "今日弹幕榜":
                for uid, cnt in self.guest_message_counter.items():
                    name = self._latest_name_for_uid(uid)
                    rank_today = self.today_guest_rank.get(uid, "-")
                    rows.append((uid, name, int(cnt), f"今日客人序号:{rank_today}"))
                rows.sort(key=lambda x: (-int(x[2]), str(x[0])))
                return rows[:limit], "弹幕数", "附加"
            if mode == "近24小时弹幕榜":
                cutoff = time.time() - 24 * 3600
                counter = defaultdict(int)
                room_counter = defaultdict(set)
                for row in self.overlap_message_events:
                    if len(row) < 5:
                        continue
                    ts = float(row[0])
                    if ts < cutoff:
                        continue
                    uid = str(row[2])
                    room_id = str(row[1])
                    counter[uid] += 1
                    room_counter[uid].add(room_id)
                for uid, cnt in counter.items():
                    name = self._latest_name_for_uid(uid)
                    rows.append((uid, name, int(cnt), f"跨房数:{len(room_counter.get(uid, set()))}"))
                rows.sort(key=lambda x: (-int(x[2]), str(x[0])))
                return rows[:limit], "24h弹幕数", "附加"
            include_main = bool(getattr(self, "overlap_include_main", True))
            rep = self._build_overlap_message_report(hours=24, include_main=include_main, include_analysis=True)
            for r in rep.get("cross_stats", [])[:limit]:
                rows.append((
                    str(r.get("unique_id", "")),
                    str(r.get("display_name", "")),
                    int(r.get("room_count", 0)),
                    f"总弹幕:{int(r.get('total_messages', 0))} | {str(r.get('room_details', ''))}",
                ))
            rows.sort(key=lambda x: (-int(x[2]), str(x[0])))
            return rows, "跨房数", "附加"

        def refresh():
            for i in tree.get_children():
                tree.delete(i)
            rows, h1, h2 = _build_rows()
            tree.heading("v1", text=h1)
            tree.heading("v2", text=h2)
            for i, (uid, name, v1, v2) in enumerate(rows, start=1):
                tree.insert("", tk.END, values=(i, uid, name, v1, v2))

        def export_csv():
            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")],
                title="导出排行榜",
                initialfile=f"leaderboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            )
            if not path:
                return
            rows, h1, h2 = _build_rows()
            try:
                with open(path, "w", encoding="utf-8", newline="") as f:
                    w = csv.writer(f)
                    w.writerow(["rank", "unique_id", "name", h1, h2])
                    for i, (uid, name, v1, v2) in enumerate(rows, start=1):
                        w.writerow([i, uid, name, v1, v2])
                messagebox.showinfo("导出", f"已导出到 {path}")
            except Exception as e:
                messagebox.showerror("错误", str(e))

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(btn, text="刷新", command=refresh).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="导出CSV", command=export_csv).pack(side=tk.LEFT, padx=4)
        mode_cb.bind("<<ComboboxSelected>>", lambda e: refresh())
        refresh()
        self._leader_win = win

    def _build_growth_analytics(self, hours: int = 24, include_main: bool = True):
        now_ts = time.time()
        cutoff = now_ts - max(1, int(hours)) * 3600
        order_kws = ("下单", "拍", "要", "买", "order", "buy", "预约", "预定")
        consult_kws = ("多少", "价格", "在哪", "怎么", "几天", "优惠", "活动", "shipping")
        by_uid = {}
        by_kw = defaultdict(int)
        hour_stat = defaultdict(lambda: {"uv": set(), "msg": 0, "order_hit": 0})
        room_user = defaultdict(set)
        valid_events = []
        for row in self.overlap_message_events:
            if len(row) >= 6:
                ts, room_id, uid, name, content, source_tag = row
            else:
                ts, room_id, uid, name, content = row[:5]
                source_tag = "main"
            if ts < cutoff:
                continue
            if source_tag == "main" and not include_main:
                continue
            uid = str(uid)
            content = str(content or "")
            name = str(name or "")
            room_id = str(room_id or "")
            if not uid:
                continue
            valid_events.append((float(ts), room_id, uid, name, content))
            one = by_uid.setdefault(uid, {"name": name, "msg": 0, "rooms": set(), "last_ts": 0.0, "order_hit": 0, "consult_hit": 0})
            if name:
                one["name"] = name
            one["msg"] += 1
            one["rooms"].add(room_id)
            one["last_ts"] = max(float(one["last_ts"]), float(ts))
            lower = content.lower()
            hit_order = 0
            for kw in order_kws:
                if kw in lower:
                    by_kw[kw] += 1
                    hit_order = 1
            for kw in consult_kws:
                if kw in lower:
                    one["consult_hit"] += 1
            one["order_hit"] += hit_order
            hh = time.strftime("%H", time.localtime(float(ts)))
            hour_stat[hh]["uv"].add(uid)
            hour_stat[hh]["msg"] += 1
            hour_stat[hh]["order_hit"] += hit_order
            room_user[room_id].add(uid)

        users = list(by_uid.items())
        segments = {"高潜": 0, "复购": 0, "沉睡": 0, "普通": 0}
        user_scores = []
        for uid, d in users:
            recency_h = max(0.0, (now_ts - float(d["last_ts"])) / 3600.0) if d["last_ts"] else 9999
            room_cnt = len(d["rooms"])
            score = d["msg"] + d["order_hit"] * 3 + d["consult_hit"] + room_cnt * 2 - int(min(24, recency_h))
            if d["order_hit"] >= 3 or score >= 20:
                seg = "高潜"
            elif d["msg"] >= 8 and recency_h <= 72:
                seg = "复购"
            elif recency_h > 120:
                seg = "沉睡"
            else:
                seg = "普通"
            segments[seg] += 1
            user_scores.append((uid, d.get("name", ""), score, d["msg"], room_cnt, d["order_hit"], recency_h, seg))

        user_scores.sort(key=lambda x: (-x[2], x[0]))
        total_uv = len(by_uid)
        total_msg = sum(d["msg"] for d in by_uid.values())
        consult_uv = sum(1 for _uid, d in by_uid.items() if d["consult_hit"] > 0)
        order_uv = sum(1 for _uid, d in by_uid.items() if d["order_hit"] > 0)
        summary = db.get_today_print_summary()
        printed = int(summary.get("printed", 0))
        failed = int(summary.get("failed", 0))
        total_jobs = printed + failed
        print_success = (printed / total_jobs) if total_jobs else 0.0
        cross_users = sum(1 for _uid, d in by_uid.items() if len(d["rooms"]) >= 2)
        top_kw = sorted(by_kw.items(), key=lambda x: (-x[1], x[0]))[:20]
        hot_hours = []
        for hh in sorted(hour_stat.keys()):
            st = hour_stat[hh]
            hot_hours.append((hh, len(st["uv"]), int(st["msg"]), int(st["order_hit"])))
        anomalies = []
        if total_uv >= 100 and order_uv / max(1, total_uv) < 0.08:
            anomalies.append("流量高但下单关键词占比偏低")
        if failed >= max(20, printed // 3):
            anomalies.append("打印失败偏高，可能影响动销履约")
        if cross_users >= 80:
            anomalies.append("跨房重叠用户较高，需做竞争房间策略")
        suggestions = [
            "高潜客人优先单独跟进，建议设置专属话术模板",
            "对沉睡客人按高活跃时段做二次触达",
            "对高贡献关键词增加固定展示位",
        ]
        if anomalies:
            suggestions.append("优先处理异常项以提升整体转化效率")
        return {
            "hours": int(hours),
            "events": valid_events,
            "total_uv": total_uv,
            "total_msg": total_msg,
            "consult_uv": consult_uv,
            "order_uv": order_uv,
            "cross_users": cross_users,
            "print_success": print_success,
            "segments": segments,
            "user_scores": user_scores,
            "top_kw": top_kw,
            "hot_hours": hot_hours,
            "anomalies": anomalies,
            "suggestions": suggestions,
        }

    def open_growth_analytics_center(self):
        if hasattr(self, "_growth_win") and self._growth_win.winfo_exists():
            self._growth_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("客户动销分析中心")
        win.geometry("1200x760")
        self._center_window(win, 1200, 760)

        top = ttk.Frame(win)
        top.pack(fill=tk.X, padx=8, pady=8)
        ttk.Label(top, text="分析窗口(小时):").pack(side=tk.LEFT)
        hours_var = tk.IntVar(value=24)
        ttk.Entry(top, textvariable=hours_var, width=8).pack(side=tk.LEFT, padx=4)
        include_main_var = tk.IntVar(value=1 if bool(self.overlap_include_main) else 0)
        ttk.Checkbutton(top, text="包含主监控数据", variable=include_main_var).pack(side=tk.LEFT, padx=8)
        ttk.Label(top, text="日报模板:").pack(side=tk.LEFT, padx=(10, 2))
        report_role_var = tk.StringVar(value="老板版")
        ttk.Combobox(top, textvariable=report_role_var, values=["老板版", "运营版", "客服版"], width=8, state="readonly").pack(side=tk.LEFT)

        nb = ttk.Notebook(win)
        nb.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        tab_overview = ttk.Frame(nb)
        tab_segment = ttk.Frame(nb)
        tab_kw = ttk.Frame(nb)
        tab_timeline = ttk.Frame(nb)
        tab_goal = ttk.Frame(nb)
        tab_report = ttk.Frame(nb)
        nb.add(tab_overview, text="总览/漏斗")
        nb.add(tab_segment, text="客户分层")
        nb.add(tab_kw, text="关键词/时段")
        nb.add(tab_timeline, text="客户旅程")
        nb.add(tab_goal, text="目标管理")
        nb.add(tab_report, text="复盘报告")

        overview_txt = scrolledtext.ScrolledText(tab_overview)
        overview_txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        seg_tree = ttk.Treeview(tab_segment, columns=("uid", "name", "score", "msg", "rooms", "order", "recency", "seg"), show="headings")
        for k, t, w in (
            ("uid", "客户ID", 160), ("name", "昵称", 120), ("score", "价值分", 80), ("msg", "弹幕数", 80),
            ("rooms", "跨房数", 80), ("order", "下单词命中", 100), ("recency", "最近活跃(h)", 100), ("seg", "分层", 80),
        ):
            seg_tree.heading(k, text=t)
            seg_tree.column(k, width=w, anchor=tk.W if k in ("uid", "name", "seg") else tk.E)
        seg_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        kw_txt = scrolledtext.ScrolledText(tab_kw)
        kw_txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        tl_top = ttk.Frame(tab_timeline)
        tl_top.pack(fill=tk.X, padx=8, pady=8)
        ttk.Label(tl_top, text="客户ID:").pack(side=tk.LEFT)
        timeline_uid_var = tk.StringVar(value="")
        ttk.Entry(tl_top, textvariable=timeline_uid_var, width=28).pack(side=tk.LEFT, padx=4)
        timeline_txt = scrolledtext.ScrolledText(tab_timeline)
        timeline_txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        goal_top = ttk.Frame(tab_goal)
        goal_top.pack(fill=tk.X, padx=8, pady=8)
        g1 = tk.IntVar(value=int(self.sales_goals.get("effective_guests_per_day", 200)))
        g2 = tk.IntVar(value=int(self.sales_goals.get("conversion_keyword_hits_per_day", 80)))
        g3 = tk.DoubleVar(value=float(self.sales_goals.get("print_success_rate", 0.95)))
        ttk.Label(goal_top, text="目标-有效客人数/天").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(goal_top, textvariable=g1, width=10).grid(row=0, column=1, sticky=tk.W, padx=4)
        ttk.Label(goal_top, text="目标-下单关键词命中/天").grid(row=1, column=0, sticky=tk.W)
        ttk.Entry(goal_top, textvariable=g2, width=10).grid(row=1, column=1, sticky=tk.W, padx=4)
        ttk.Label(goal_top, text="目标-打印成功率").grid(row=2, column=0, sticky=tk.W)
        ttk.Entry(goal_top, textvariable=g3, width=10).grid(row=2, column=1, sticky=tk.W, padx=4)
        goal_txt = scrolledtext.ScrolledText(tab_goal, height=18)
        goal_txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        report_txt = scrolledtext.ScrolledText(tab_report)
        report_txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        state = {"data": None}

        def refresh_all():
            try:
                h = max(1, int(hours_var.get()))
            except Exception:
                h = 24
            include_main = bool(include_main_var.get())
            self.overlap_include_main = include_main
            data = self._build_growth_analytics(hours=h, include_main=include_main)
            state["data"] = data

            funnel_conv = (data["order_uv"] / max(1, data["total_uv"])) * 100.0
            lines = [
                f"分析窗口: 最近 {data['hours']} 小时",
                f"总UV: {data['total_uv']}",
                f"总弹幕: {data['total_msg']}",
                f"咨询UV: {data['consult_uv']}",
                f"下单关键词UV: {data['order_uv']}",
                f"跨房重叠客人: {data['cross_users']}",
                f"下单关键词转化率(粗略): {funnel_conv:.2f}%",
                f"打印成功率: {data['print_success']*100:.2f}%",
                "",
                "客户分层:",
            ]
            for k in ("高潜", "复购", "普通", "沉睡"):
                lines.append(f"- {k}: {data['segments'].get(k, 0)}")
            lines.append("")
            lines.append("异常检测:")
            if data["anomalies"]:
                for a in data["anomalies"]:
                    lines.append(f"- {a}")
            else:
                lines.append("- 暂无明显异常")
            lines.append("")
            lines.append("策略建议:")
            for s in data["suggestions"]:
                lines.append(f"- {s}")
            overview_txt.config(state=tk.NORMAL)
            overview_txt.delete("1.0", tk.END)
            overview_txt.insert(tk.END, "\n".join(lines))
            overview_txt.config(state=tk.DISABLED)

            for i in seg_tree.get_children():
                seg_tree.delete(i)
            for uid, name, score, msg, rooms, order_hit, recency_h, seg in data["user_scores"][:1000]:
                seg_tree.insert("", tk.END, values=(uid, name, score, msg, rooms, order_hit, f"{recency_h:.1f}", seg))

            kw_lines = ["关键词贡献TOP:", ]
            for kw, cnt in data["top_kw"]:
                kw_lines.append(f"- {kw}: {cnt}")
            kw_lines.append("")
            kw_lines.append("时段热力(小时 | UV | 弹幕 | 下单词命中):")
            for hh, uv, msg, oh in data["hot_hours"]:
                kw_lines.append(f"- {hh}: {uv} | {msg} | {oh}")
            kw_txt.config(state=tk.NORMAL)
            kw_txt.delete("1.0", tk.END)
            kw_txt.insert(tk.END, "\n".join(kw_lines))
            kw_txt.config(state=tk.DISABLED)

            goal_report = []
            goal_report.append(f"有效客人数: {data['total_uv']} / 目标 {int(g1.get())}")
            goal_report.append(f"下单关键词UV: {data['order_uv']} / 目标 {int(g2.get())}")
            goal_report.append(f"打印成功率: {data['print_success']*100:.2f}% / 目标 {float(g3.get())*100:.2f}%")
            goal_report.append("")
            goal_report.append("达成判断:")
            goal_report.append(f"- 有效客人数: {'达成' if data['total_uv'] >= int(g1.get()) else '未达成'}")
            goal_report.append(f"- 下单关键词UV: {'达成' if data['order_uv'] >= int(g2.get()) else '未达成'}")
            goal_report.append(f"- 打印成功率: {'达成' if data['print_success'] >= float(g3.get()) else '未达成'}")
            goal_txt.config(state=tk.NORMAL)
            goal_txt.delete("1.0", tk.END)
            goal_txt.insert(tk.END, "\n".join(goal_report))
            goal_txt.config(state=tk.DISABLED)

            report_lines = [
                f"【动销复盘报告】{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"窗口: 最近{data['hours']}小时",
                f"UV={data['total_uv']} 弹幕={data['total_msg']} 咨询UV={data['consult_uv']} 下单关键词UV={data['order_uv']}",
                f"跨房重叠客人={data['cross_users']} 打印成功率={data['print_success']*100:.2f}%",
                "分层: " + " / ".join([f"{k}:{data['segments'].get(k, 0)}" for k in ('高潜', '复购', '普通', '沉睡')]),
                "异常: " + ("; ".join(data["anomalies"]) if data["anomalies"] else "无"),
                "建议:",
            ]
            for s in data["suggestions"]:
                report_lines.append(f"- {s}")
            report_txt.config(state=tk.NORMAL)
            report_txt.delete("1.0", tk.END)
            report_txt.insert(tk.END, "\n".join(report_lines))
            report_txt.config(state=tk.DISABLED)

        def show_timeline():
            uid = timeline_uid_var.get().strip()
            data = state.get("data") or {}
            events = data.get("events", [])
            rows = [x for x in events if str(x[2]) == uid] if uid else []
            lines = [f"客户ID: {uid}", f"事件数: {len(rows)}", ""]
            for ts, room_id, _uid, name, content in rows[-500:]:
                lines.append(f"{time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(ts))} | @{room_id} | {name}: {content}")
            timeline_txt.config(state=tk.NORMAL)
            timeline_txt.delete("1.0", tk.END)
            timeline_txt.insert(tk.END, "\n".join(lines))
            timeline_txt.config(state=tk.DISABLED)

        def save_goals():
            self.sales_goals = {
                "effective_guests_per_day": int(g1.get()),
                "conversion_keyword_hits_per_day": int(g2.get()),
                "print_success_rate": float(g3.get()),
            }
            self._save_settings()
            messagebox.showinfo("保存", "目标已保存")
            refresh_all()

        def export_report():
            path = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("CSV files", "*.csv")],
                title="导出动销分析报告",
                initialfile=f"growth_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            )
            if not path:
                return
            try:
                data = state.get("data") or self._build_growth_analytics(hours=max(1, int(hours_var.get())), include_main=bool(include_main_var.get()))
                if path.lower().endswith(".csv"):
                    with open(path, "w", encoding="utf-8", newline="") as f:
                        w = csv.writer(f)
                        w.writerow(["metric", "value"])
                        w.writerow(["hours", data["hours"]])
                        w.writerow(["total_uv", data["total_uv"]])
                        w.writerow(["total_msg", data["total_msg"]])
                        w.writerow(["consult_uv", data["consult_uv"]])
                        w.writerow(["order_uv", data["order_uv"]])
                        w.writerow(["cross_users", data["cross_users"]])
                        w.writerow(["print_success_rate", f"{data['print_success']:.4f}"])
                        w.writerow([])
                        w.writerow(["uid", "name", "score", "msg", "rooms", "order_hit", "recency_h", "segment"])
                        for row in data["user_scores"][:2000]:
                            w.writerow(row)
                else:
                    txt = report_txt.get("1.0", tk.END).strip()
                    with open(path, "w", encoding="utf-8") as f:
                        f.write(txt)
                messagebox.showinfo("导出", f"已导出到 {path}")
            except Exception as e:
                messagebox.showerror("错误", str(e))

        op = ttk.Frame(top)
        op.pack(side=tk.RIGHT)
        ttk.Button(op, text="刷新", command=refresh_all).pack(side=tk.LEFT, padx=4)
        ttk.Button(op, text="保存目标", command=save_goals).pack(side=tk.LEFT, padx=4)
        ttk.Button(op, text="导出报告", command=export_report).pack(side=tk.LEFT, padx=4)

        tl_btn = ttk.Frame(tab_timeline)
        tl_btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(tl_btn, text="查询旅程", command=show_timeline).pack(side=tk.LEFT, padx=4)
        ttk.Button(tl_btn, text="使用分层选中ID", command=lambda: timeline_uid_var.set((seg_tree.item(seg_tree.selection()[0], "values")[0] if seg_tree.selection() else ""))).pack(side=tk.LEFT, padx=4)
        refresh_all()
        self._growth_win = win

    def open_business_ops_center(self):
        if hasattr(self, "_bizops_win") and self._bizops_win.winfo_exists():
            self._bizops_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("经营增强中心")
        win.geometry("1260x800")
        self._center_window(win, 1260, 800)

        top = ttk.Frame(win)
        top.pack(fill=tk.X, padx=8, pady=8)
        ttk.Label(top, text="分析窗口(小时):").pack(side=tk.LEFT)
        hours_var = tk.IntVar(value=24)
        ttk.Entry(top, textvariable=hours_var, width=8).pack(side=tk.LEFT, padx=4)
        include_main_var = tk.IntVar(value=1 if bool(self.overlap_include_main) else 0)
        ttk.Checkbutton(top, text="包含主监控数据", variable=include_main_var).pack(side=tk.LEFT, padx=8)
        ttk.Label(top, text="导出模板:").pack(side=tk.LEFT)
        report_role_var = tk.StringVar(value="老板版")
        ttk.Combobox(top, textvariable=report_role_var, values=["老板版", "运营版", "客服版"], width=8, state="readonly").pack(side=tk.LEFT, padx=4)
        ttk.Label(top, text="保留天数:").pack(side=tk.LEFT)
        retention_var = tk.IntVar(value=max(1, int(self.data_retention_days)))
        ttk.Entry(top, textvariable=retention_var, width=6).pack(side=tk.LEFT, padx=4)
        auto_maint_var = tk.IntVar(value=1 if bool(self.auto_data_maintenance_enabled) else 0)
        ttk.Checkbutton(top, text="自动维护", variable=auto_maint_var).pack(side=tk.LEFT, padx=6)

        nb = ttk.Notebook(win)
        nb.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        tabs = {}
        for key, title in [
            ("dashboard", "经营首页"),
            ("tags", "客户标签"),
            ("follow", "私域跟进"),
            ("campaign", "活动日历"),
            ("inventory", "库存联动"),
            ("host", "主播评分"),
            ("risk", "风控中心"),
            ("team", "团队权限"),
            ("strategy", "自动策略"),
            ("predict", "预测面板"),
            ("events", "运营日志"),
        ]:
            frm = ttk.Frame(nb)
            nb.add(frm, text=title)
            tabs[key] = frm

        dashboard_txt = scrolledtext.ScrolledText(tabs["dashboard"])
        dashboard_txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        tag_top = ttk.Frame(tabs["tags"])
        tag_top.pack(fill=tk.X, padx=8, pady=8)
        ttk.Label(tag_top, text="客户ID").pack(side=tk.LEFT)
        tag_uid_var = tk.StringVar(value="")
        ttk.Entry(tag_top, textvariable=tag_uid_var, width=20).pack(side=tk.LEFT, padx=4)
        ttk.Label(tag_top, text="标签(逗号分隔)").pack(side=tk.LEFT)
        tag_val_var = tk.StringVar(value="")
        ttk.Entry(tag_top, textvariable=tag_val_var, width=28).pack(side=tk.LEFT, padx=4)
        tag_tree = ttk.Treeview(tabs["tags"], columns=("uid", "name", "tags", "score"), show="headings")
        for k, t, w in (("uid", "客户ID", 180), ("name", "昵称", 140), ("tags", "标签", 260), ("score", "价值分", 90)):
            tag_tree.heading(k, text=t)
            tag_tree.column(k, width=w, anchor=tk.W if k != "score" else tk.E)
        tag_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        follow_top = ttk.Frame(tabs["follow"])
        follow_top.pack(fill=tk.X, padx=8, pady=8)
        fu_uid_var = tk.StringVar()
        fu_note_var = tk.StringVar()
        fu_next_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Label(follow_top, text="客户ID").pack(side=tk.LEFT)
        ttk.Entry(follow_top, textvariable=fu_uid_var, width=16).pack(side=tk.LEFT, padx=4)
        ttk.Label(follow_top, text="备注").pack(side=tk.LEFT)
        ttk.Entry(follow_top, textvariable=fu_note_var, width=30).pack(side=tk.LEFT, padx=4)
        ttk.Label(follow_top, text="下次日期").pack(side=tk.LEFT)
        ttk.Entry(follow_top, textvariable=fu_next_var, width=12).pack(side=tk.LEFT, padx=4)
        follow_tree = ttk.Treeview(tabs["follow"], columns=("uid", "note", "next", "status"), show="headings")
        for k, t, w in (("uid", "客户ID", 180), ("note", "备注", 420), ("next", "下次跟进", 130), ("status", "状态", 100)):
            follow_tree.heading(k, text=t)
            follow_tree.column(k, width=w, anchor=tk.W)
        follow_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        camp_top = ttk.Frame(tabs["campaign"])
        camp_top.pack(fill=tk.X, padx=8, pady=8)
        camp_name_var = tk.StringVar()
        camp_range_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        camp_note_var = tk.StringVar()
        ttk.Label(camp_top, text="活动名").pack(side=tk.LEFT)
        ttk.Entry(camp_top, textvariable=camp_name_var, width=18).pack(side=tk.LEFT, padx=4)
        ttk.Label(camp_top, text="日期").pack(side=tk.LEFT)
        ttk.Entry(camp_top, textvariable=camp_range_var, width=12).pack(side=tk.LEFT, padx=4)
        ttk.Label(camp_top, text="策略").pack(side=tk.LEFT)
        ttk.Entry(camp_top, textvariable=camp_note_var, width=36).pack(side=tk.LEFT, padx=4)
        camp_tree = ttk.Treeview(tabs["campaign"], columns=("name", "date", "note"), show="headings")
        for k, t, w in (("name", "活动名", 220), ("date", "日期", 120), ("note", "策略/结果", 620)):
            camp_tree.heading(k, text=t)
            camp_tree.column(k, width=w, anchor=tk.W)
        camp_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        inv_top = ttk.Frame(tabs["inventory"])
        inv_top.pack(fill=tk.X, padx=8, pady=8)
        inv_kw_var = tk.StringVar()
        inv_stock_var = tk.IntVar(value=100)
        ttk.Label(inv_top, text="关键词").pack(side=tk.LEFT)
        ttk.Entry(inv_top, textvariable=inv_kw_var, width=18).pack(side=tk.LEFT, padx=4)
        ttk.Label(inv_top, text="库存").pack(side=tk.LEFT)
        ttk.Entry(inv_top, textvariable=inv_stock_var, width=8).pack(side=tk.LEFT, padx=4)
        inv_tree = ttk.Treeview(tabs["inventory"], columns=("kw", "stock", "hits", "status"), show="headings")
        for k, t, w in (("kw", "关键词", 220), ("stock", "库存", 100), ("hits", "近24h命中", 140), ("status", "建议", 420)):
            inv_tree.heading(k, text=t)
            inv_tree.column(k, width=w, anchor=tk.W if k in ("kw", "status") else tk.E)
        inv_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        host_top = ttk.Frame(tabs["host"])
        host_top.pack(fill=tk.X, padx=8, pady=8)
        host_name_var = tk.StringVar(value=self.custom_name_var.get().strip() or "主播")
        host_score_var = tk.DoubleVar(value=5.0)
        ttk.Label(host_top, text="主播").pack(side=tk.LEFT)
        ttk.Entry(host_top, textvariable=host_name_var, width=18).pack(side=tk.LEFT, padx=4)
        ttk.Label(host_top, text="评分(0-10)").pack(side=tk.LEFT)
        ttk.Entry(host_top, textvariable=host_score_var, width=8).pack(side=tk.LEFT, padx=4)
        host_tree = ttk.Treeview(tabs["host"], columns=("host", "score", "samples"), show="headings")
        for k, t, w in (("host", "主播", 220), ("score", "平均分", 120), ("samples", "样本数", 120)):
            host_tree.heading(k, text=t)
            host_tree.column(k, width=w, anchor=tk.W if k == "host" else tk.E)
        host_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        risk_txt = scrolledtext.ScrolledText(tabs["risk"])
        risk_txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        team_top = ttk.Frame(tabs["team"])
        team_top.pack(fill=tk.X, padx=8, pady=8)
        owner_var = tk.StringVar(value=str(self.team_roles.get("owner", "")))
        op_var = tk.StringVar(value=str(self.team_roles.get("operator", "")))
        sup_var = tk.StringVar(value=str(self.team_roles.get("support", "")))
        ttk.Label(team_top, text="老板").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(team_top, textvariable=owner_var, width=24).grid(row=0, column=1, sticky=tk.W, padx=4)
        ttk.Label(team_top, text="运营").grid(row=1, column=0, sticky=tk.W)
        ttk.Entry(team_top, textvariable=op_var, width=24).grid(row=1, column=1, sticky=tk.W, padx=4)
        ttk.Label(team_top, text="客服").grid(row=2, column=0, sticky=tk.W)
        ttk.Entry(team_top, textvariable=sup_var, width=24).grid(row=2, column=1, sticky=tk.W, padx=4)
        team_txt = scrolledtext.ScrolledText(tabs["team"], height=18)
        team_txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        strategy_txt = scrolledtext.ScrolledText(tabs["strategy"])
        strategy_txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        predict_txt = scrolledtext.ScrolledText(tabs["predict"])
        predict_txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        event_tree = ttk.Treeview(tabs["events"], columns=("time", "etype", "payload"), show="headings")
        event_tree.heading("time", text="时间")
        event_tree.heading("etype", text="事件")
        event_tree.heading("payload", text="明细")
        event_tree.column("time", width=160, anchor=tk.W)
        event_tree.column("etype", width=180, anchor=tk.W)
        event_tree.column("payload", width=860, anchor=tk.W)
        event_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        state = {"data": None}

        def refresh_all():
            try:
                h = max(1, int(hours_var.get()))
            except Exception:
                h = 24
            include_main = bool(include_main_var.get())
            self.overlap_include_main = include_main
            data = self._build_growth_analytics(hours=h, include_main=include_main)
            state["data"] = data

            goals = self.sales_goals or {}
            g_uv = int(goals.get("effective_guests_per_day", 200))
            g_ord = int(goals.get("conversion_keyword_hits_per_day", 80))
            g_ps = float(goals.get("print_success_rate", 0.95))
            gaps = [
                f"有效客数差额: {max(0, g_uv - int(data['total_uv']))}",
                f"下单关键词UV差额: {max(0, g_ord - int(data['order_uv']))}",
                f"打印成功率差额: {max(0.0, g_ps - float(data['print_success']))*100:.2f}%",
            ]
            dash = [
                f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"窗口: 最近{data['hours']}小时",
                f"UV={data['total_uv']} 弹幕={data['total_msg']} 咨询UV={data['consult_uv']} 下单关键词UV={data['order_uv']}",
                f"跨房重叠={data['cross_users']} 打印成功率={data['print_success']*100:.2f}%",
                f"分层: 高潜{data['segments'].get('高潜',0)} / 复购{data['segments'].get('复购',0)} / 普通{data['segments'].get('普通',0)} / 沉睡{data['segments'].get('沉睡',0)}",
                "",
                "今日目标差额:",
            ] + [f"- {x}" for x in gaps]
            dashboard_txt.config(state=tk.NORMAL)
            dashboard_txt.delete("1.0", tk.END)
            dashboard_txt.insert(tk.END, "\n".join(dash))
            dashboard_txt.config(state=tk.DISABLED)

            for i in tag_tree.get_children():
                tag_tree.delete(i)
            for uid, name, score, _msg, _rooms, _order, _recency, _seg in data["user_scores"][:800]:
                tag_tree.insert("", tk.END, values=(uid, name, self.customer_tags.get(uid, ""), score))

            for i in follow_tree.get_children():
                follow_tree.delete(i)
            for row in self.followup_queue[-2000:]:
                follow_tree.insert("", tk.END, values=(row.get("uid", ""), row.get("note", ""), row.get("next", ""), row.get("status", "todo")))

            for i in camp_tree.get_children():
                camp_tree.delete(i)
            for row in self.campaign_calendar[-1000:]:
                camp_tree.insert("", tk.END, values=(row.get("name", ""), row.get("date", ""), row.get("note", "")))

            kw_hits = dict(data.get("top_kw", []))
            for i in inv_tree.get_children():
                inv_tree.delete(i)
            for kw, stock in sorted((self.inventory_map or {}).items(), key=lambda x: str(x[0])):
                hits = int(kw_hits.get(str(kw), 0))
                st = "补货" if int(stock) <= max(20, hits * 2) else "正常"
                inv_tree.insert("", tk.END, values=(kw, int(stock), hits, st))

            for i in host_tree.get_children():
                host_tree.delete(i)
            hs = sorted((self.host_scores or {}).items(), key=lambda x: -float(x[1].get("score", 0)))
            for hname, d in hs:
                host_tree.insert("", tk.END, values=(hname, f"{float(d.get('score', 0)):.2f}", int(d.get("samples", 0))))

            spam_limit = int(self.risk_rules.get("spam_per_10s", 25))
            dup_limit = int(self.risk_rules.get("duplicate_text_hits", 8))
            hot_spam = []
            for uid, dq in self.user_msg_timestamps.items():
                hot_spam.append((uid, len(dq)))
            hot_spam.sort(key=lambda x: -x[1])
            text_counter = defaultdict(int)
            for ev in data.get("events", [])[-8000:]:
                text_counter[str(ev[4]).strip().lower()] += 1
            dup_hits = sorted([(t, c) for t, c in text_counter.items() if t and c >= dup_limit], key=lambda x: -x[1])[:50]
            risk_lines = [
                f"风控阈值: 10秒频次>={spam_limit}, 重复文本命中>={dup_limit}",
                "高频发言用户TOP:",
            ]
            for uid, cnt in hot_spam[:30]:
                flag = " [风险]" if cnt >= spam_limit else ""
                risk_lines.append(f"- {uid}: {cnt}{flag}")
            risk_lines.append("")
            risk_lines.append("高重复文本TOP:")
            for txt, cnt in dup_hits:
                risk_lines.append(f"- {cnt}次: {txt[:80]}")
            if not dup_hits:
                risk_lines.append("- 无")
            risk_txt.config(state=tk.NORMAL)
            risk_txt.delete("1.0", tk.END)
            risk_txt.insert(tk.END, "\n".join(risk_lines))
            risk_txt.config(state=tk.DISABLED)

            team_lines = [
                f"当前角色: 老板={owner_var.get().strip() or '-'} | 运营={op_var.get().strip() or '-'} | 客服={sup_var.get().strip() or '-'}",
                "",
                "权限建议:",
                "- 老板: 查看全部报表 + 目标配置 + 导出",
                "- 运营: 标签/策略/活动编辑 + 报表查看",
                "- 客服: 跟进队列 + 客户画像查看",
                "",
                "说明: 当前为本地单机版，先做角色配置留档，不阻断功能。",
            ]
            team_txt.config(state=tk.NORMAL)
            team_txt.delete("1.0", tk.END)
            team_txt.insert(tk.END, "\n".join(team_lines))
            team_txt.config(state=tk.DISABLED)

            auto_strategy = []
            if data["segments"].get("高潜", 0) > 0:
                auto_strategy.append("优先处理高潜客人：安排客服跟进 + 设置专属标签。")
            if data["order_uv"] / max(1, data["total_uv"]) < 0.1:
                auto_strategy.append("下单词转化偏低：建议切换强促销话术并缩短引导路径。")
            if data["cross_users"] > 50:
                auto_strategy.append("跨房重叠高：建议做竞房差异化优惠与专属套餐。")
            if data["print_success"] < 0.9:
                auto_strategy.append("履约成功率偏低：优先处理打印失败和队列积压。")
            if not auto_strategy:
                auto_strategy.append("核心指标稳定，建议维持当前策略并小幅A/B测试。")
            strategy_txt.config(state=tk.NORMAL)
            strategy_txt.delete("1.0", tk.END)
            strategy_txt.insert(tk.END, "\n".join([f"- {x}" for x in auto_strategy]))
            strategy_txt.config(state=tk.DISABLED)

            hour_map = {h: (uv, msg, oh) for h, uv, msg, oh in data.get("hot_hours", [])}
            hours_sorted = sorted(hour_map.keys())
            recent = [hour_map[h][1] for h in hours_sorted[-3:]] if hours_sorted else []
            recent_uv = [hour_map[h][0] for h in hours_sorted[-3:]] if hours_sorted else []
            recent_order = [hour_map[h][2] for h in hours_sorted[-3:]] if hours_sorted else []
            pred_msg = (sum(recent) / len(recent)) if recent else 0.0
            pred_uv = (sum(recent_uv) / len(recent_uv)) if recent_uv else 0.0
            pred_order = (sum(recent_order) / len(recent_order)) if recent_order else 0.0
            pred_lines = [
                "未来1小时预测(简单移动平均):",
                f"- 预计弹幕量: {pred_msg:.1f}",
                f"- 预计UV: {pred_uv:.1f}",
                f"- 预计下单关键词命中: {pred_order:.1f}",
                "",
                "运营动作建议:",
                f"- 若预计下单词<{max(5.0, pred_uv*0.1):.1f}，提前上促销话术",
                f"- 若预计弹幕>{max(50.0, pred_msg):.1f}，提前检查打印队列与客服排班",
            ]
            predict_txt.config(state=tk.NORMAL)
            predict_txt.delete("1.0", tk.END)
            predict_txt.insert(tk.END, "\n".join(pred_lines))
            predict_txt.config(state=tk.DISABLED)
            refresh_event_log()

        def apply_ops_settings():
            try:
                self.data_retention_days = max(1, int(retention_var.get()))
            except Exception:
                self.data_retention_days = 30
                retention_var.set(30)
            self.auto_data_maintenance_enabled = bool(auto_maint_var.get())
            self._save_settings()

        def refresh_event_log():
            for i in event_tree.get_children():
                event_tree.delete(i)
            path = self.business_event_log_path
            if not os.path.exists(path):
                return
            try:
                with open(path, "r", encoding="utf-8") as f:
                    lines = f.readlines()[-1200:]
                for line in reversed(lines):
                    s = line.strip()
                    if not s:
                        continue
                    try:
                        row = json.loads(s)
                    except Exception:
                        continue
                    ts = float(row.get("ts", time.time()))
                    tm = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
                    etype = str(row.get("event_type", ""))
                    payload = row.get("payload", {})
                    payload_txt = json.dumps(payload, ensure_ascii=False)[:500]
                    event_tree.insert("", tk.END, values=(tm, etype, payload_txt))
            except Exception:
                pass

        def export_event_log():
            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("JSONL files", "*.jsonl"), ("Text files", "*.txt")],
                title="导出运营日志",
                initialfile=f"business_events_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            )
            if not path:
                return
            ext = os.path.splitext(path)[1].lower()
            try:
                if ext == ".jsonl":
                    with open(self.business_event_log_path, "r", encoding="utf-8") as src, open(path, "w", encoding="utf-8") as dst:
                        dst.write(src.read())
                else:
                    with open(path, "w", encoding="utf-8", newline="") as f:
                        w = csv.writer(f)
                        w.writerow(["time", "event_type", "payload"])
                        for iid in event_tree.get_children():
                            w.writerow(event_tree.item(iid, "values"))
                messagebox.showinfo("导出", f"已导出到 {path}")
            except Exception as e:
                messagebox.showerror("错误", str(e))

        def clear_event_log():
            if not messagebox.askyesno("确认", "确认清空运营日志文件？"):
                return
            try:
                with open(self.business_event_log_path, "w", encoding="utf-8"):
                    pass
            except Exception as e:
                messagebox.showerror("错误", str(e))
                return
            refresh_event_log()
            self._set_status("已清空运营日志")

        def save_tags():
            uid = tag_uid_var.get().strip()
            if not uid:
                return
            self.customer_tags[uid] = tag_val_var.get().strip()
            self._save_settings()
            refresh_all()

        def on_pick_tag(event=None):
            sel = tag_tree.selection()
            if not sel:
                return
            vals = tag_tree.item(sel[0], "values")
            if not vals:
                return
            tag_uid_var.set(str(vals[0]))
            tag_val_var.set(str(vals[2]) if len(vals) > 2 else "")

        def add_followup():
            uid = fu_uid_var.get().strip()
            if not uid:
                return
            self.followup_queue.append({
                "uid": uid,
                "note": fu_note_var.get().strip(),
                "next": fu_next_var.get().strip(),
                "status": "todo",
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            self._save_settings()
            refresh_all()

        def mark_followup_done():
            sel = follow_tree.selection()
            if not sel:
                return
            vals = follow_tree.item(sel[0], "values")
            uid = str(vals[0]) if vals else ""
            nxt = str(vals[2]) if len(vals) > 2 else ""
            for row in reversed(self.followup_queue):
                if str(row.get("uid")) == uid and str(row.get("next")) == nxt and str(row.get("status", "todo")) != "done":
                    row["status"] = "done"
                    break
            self._save_settings()
            refresh_all()

        def add_campaign():
            name = camp_name_var.get().strip()
            if not name:
                return
            self.campaign_calendar.append({
                "name": name,
                "date": camp_range_var.get().strip(),
                "note": camp_note_var.get().strip(),
            })
            self._save_settings()
            refresh_all()

        def save_inventory():
            kw = inv_kw_var.get().strip()
            if not kw:
                return
            self.inventory_map[kw] = int(inv_stock_var.get())
            self._save_settings()
            refresh_all()

        def add_host_score():
            hn = host_name_var.get().strip()
            if not hn:
                return
            s = max(0.0, min(10.0, float(host_score_var.get())))
            old = self.host_scores.get(hn, {"score": 0.0, "samples": 0})
            n = int(old.get("samples", 0)) + 1
            new_score = ((float(old.get("score", 0.0)) * (n - 1)) + s) / n
            self.host_scores[hn] = {"score": float(new_score), "samples": n}
            self._save_settings()
            refresh_all()

        def save_team():
            self.team_roles = {
                "owner": owner_var.get().strip(),
                "operator": op_var.get().strip(),
                "support": sup_var.get().strip(),
            }
            self._save_settings()
            refresh_all()

        def export_daily_report():
            data = state.get("data")
            if not data:
                refresh_all()
                data = state.get("data")
            if not data:
                messagebox.showwarning("提示", "暂无可导出数据")
                return
            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("PDF files", "*.pdf"), ("Text files", "*.txt")],
                title="导出经营日报",
                initialfile=f"biz_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            )
            if not path:
                return
            ext = os.path.splitext(path)[1].lower()
            try:
                if ext == ".csv":
                    with open(path, "w", encoding="utf-8", newline="") as f:
                        w = csv.writer(f)
                        w.writerow(["metric", "value"])
                        w.writerow(["generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                        w.writerow(["hours", data["hours"]])
                        w.writerow(["total_uv", data["total_uv"]])
                        w.writerow(["total_msg", data["total_msg"]])
                        w.writerow(["consult_uv", data["consult_uv"]])
                        w.writerow(["order_uv", data["order_uv"]])
                        w.writerow(["cross_users", data["cross_users"]])
                        w.writerow(["print_success_rate", f"{data['print_success']:.4f}"])
                        w.writerow([])
                        w.writerow(["segment", "count"])
                        for k in ("高潜", "复购", "普通", "沉睡"):
                            w.writerow([k, int(data["segments"].get(k, 0))])
                        w.writerow([])
                        w.writerow(["top_keywords", "hits"])
                        for kw, hits in data.get("top_kw", []):
                            w.writerow([kw, hits])
                        w.writerow([])
                        w.writerow(["uid", "name", "score", "msg", "rooms", "order_hit", "recency_h", "segment"])
                        for row in data.get("user_scores", [])[:2000]:
                            w.writerow(row)
                elif ext == ".pdf":
                    try:
                        from reportlab.lib.pagesizes import A4
                        from reportlab.pdfgen import canvas as pdf_canvas
                        c = pdf_canvas.Canvas(path, pagesize=A4)
                        width, height = A4
                        y = height - 40
                        lines = [
                            "Business Daily Report",
                            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                            f"Window(hours): {data['hours']}",
                            f"UV: {data['total_uv']}  Msg: {data['total_msg']}  ConsultUV: {data['consult_uv']}",
                            f"OrderUV: {data['order_uv']}  CrossUsers: {data['cross_users']}",
                            f"PrintSuccess: {data['print_success']*100:.2f}%",
                            "Segments:",
                        ]
                        for k in ("高潜", "复购", "普通", "沉睡"):
                            lines.append(f"  {k}: {int(data['segments'].get(k, 0))}")
                        lines.append("Top Keywords:")
                        for kw, hits in data.get("top_kw", [])[:20]:
                            lines.append(f"  {kw}: {hits}")
                        for line in lines:
                            c.drawString(40, y, str(line))
                            y -= 16
                            if y < 40:
                                c.showPage()
                                y = height - 40
                        c.save()
                    except Exception:
                        messagebox.showwarning("提示", "PDF导出需要 reportlab，已改为TXT导出。")
                        txt_path = os.path.splitext(path)[0] + ".txt"
                        with open(txt_path, "w", encoding="utf-8") as f:
                            f.write(dashboard_txt.get("1.0", tk.END).strip())
                        path = txt_path
                else:
                    role = report_role_var.get().strip()
                    if role == "运营版":
                        body = strategy_txt.get("1.0", tk.END).strip() + "\n\n" + predict_txt.get("1.0", tk.END).strip()
                    elif role == "客服版":
                        body = "【客服跟进视图】\n\n" + "\n".join(
                            f"{r.get('uid','')} | {r.get('next','')} | {r.get('status','todo')} | {r.get('note','')}"
                            for r in self.followup_queue[-500:]
                        )
                    else:
                        body = dashboard_txt.get("1.0", tk.END).strip()
                    with open(path, "w", encoding="utf-8") as f:
                        f.write(body)
                messagebox.showinfo("导出", f"已导出到 {path}")
            except Exception as e:
                messagebox.showerror("错误", str(e))

        btn = ttk.Frame(top)
        btn.pack(side=tk.RIGHT)
        ttk.Button(btn, text="刷新", command=refresh_all).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="保存设置", command=apply_ops_settings).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="导出日报", command=export_daily_report).pack(side=tk.LEFT, padx=4)

        tag_btn = ttk.Frame(tabs["tags"])
        tag_btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(tag_btn, text="保存标签", command=save_tags).pack(side=tk.LEFT, padx=4)
        tag_tree.bind("<<TreeviewSelect>>", on_pick_tag)

        follow_btn = ttk.Frame(tabs["follow"])
        follow_btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(follow_btn, text="加入跟进", command=add_followup).pack(side=tk.LEFT, padx=4)
        ttk.Button(follow_btn, text="标记完成", command=mark_followup_done).pack(side=tk.LEFT, padx=4)

        camp_btn = ttk.Frame(tabs["campaign"])
        camp_btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(camp_btn, text="添加活动", command=add_campaign).pack(side=tk.LEFT, padx=4)

        inv_btn = ttk.Frame(tabs["inventory"])
        inv_btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(inv_btn, text="保存库存", command=save_inventory).pack(side=tk.LEFT, padx=4)

        host_btn = ttk.Frame(tabs["host"])
        host_btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(host_btn, text="添加评分", command=add_host_score).pack(side=tk.LEFT, padx=4)

        team_btn = ttk.Frame(tabs["team"])
        team_btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(team_btn, text="保存角色", command=save_team).pack(side=tk.LEFT, padx=4)
        event_btn = ttk.Frame(tabs["events"])
        event_btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(event_btn, text="刷新日志", command=refresh_event_log).pack(side=tk.LEFT, padx=4)
        ttk.Button(event_btn, text="导出日志", command=export_event_log).pack(side=tk.LEFT, padx=4)
        ttk.Button(event_btn, text="清空日志", command=clear_event_log).pack(side=tk.LEFT, padx=4)

        refresh_all()
        self._track_business_event("open_business_ops_center", {"hours": int(hours_var.get()), "include_main": bool(include_main_var.get())})
        self._bizops_win = win

    def _parse_room_inputs(self, raw_text: str):
        text = str(raw_text or "").strip()
        if not text:
            return []
        parts = [x.strip() for x in text.replace("\n", ",").split(",") if x.strip()]
        out = []
        seen = set()
        for raw in parts:
            uid = ""
            if "tiktok.com/@" in raw and "@" in raw:
                try:
                    uid = raw.split("@", 1)[1].split("/")[0].split("?")[0].strip()
                except Exception:
                    uid = ""
            elif raw.startswith("@"):
                uid = raw[1:].strip()
            else:
                uid = raw.strip()
            uid = uid.replace(" ", "")
            if uid and uid not in seen:
                out.append(uid)
                seen.add(uid)
        return out

    def _record_overlap_message_event(self, room_id: str, unique_id: str, display_name: str, content: str, when_ts: float, source_tag: str = "main"):
        if not room_id or not unique_id:
            return
        try:
            self.overlap_message_events.append((
                float(when_ts),
                str(room_id),
                str(unique_id),
                str(display_name or ""),
                str(content or ""),
                str(source_tag or "main"),
            ))
            self._append_overlap_event_to_disk(when_ts, room_id, unique_id, display_name, content, source_tag=source_tag)
            self._track_business_event(
                "comment_collected",
                {
                    "room_id": str(room_id or ""),
                    "unique_id": str(unique_id or ""),
                    "source": str(source_tag or "main"),
                    "content_len": len(str(content or "")),
                    "is_numeric": bool(str(content or "").isdigit()),
                },
            )
        except Exception:
            pass

    def _start_auto_recorder_for_uid(self, unique_id: str):
        if not self.auto_record_enabled:
            return
        uid = str(unique_id or "").strip()
        if not uid:
            return
        if uid in self.record_workers:
            return
        stop_evt = threading.Event()
        holder = {"proc": None}

        def target():
            missing_tool_logged = False
            while not stop_evt.is_set() and not self._stop_event.is_set():
                out_dir = self.auto_record_dir.strip() if self.auto_record_dir else os.path.join(os.getcwd(), "recordings")
                try:
                    os.makedirs(out_dir, exist_ok=True)
                except Exception:
                    out_dir = os.getcwd()
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                out_file = os.path.join(out_dir, f"{uid}_{stamp}.mp4")
                cmd = [self.auto_record_cmd or "streamlink", f"https://www.tiktok.com/@{uid}/live", "best", "-o", out_file]
                proxy = ""
                if self._proxy_applies_to_tiktok():
                    proxy = self._resolve_configured_proxy()
                if proxy:
                    cmd += ["--http-proxy", proxy]
                env = os.environ.copy()
                if self.ssl_insecure_var.get():
                    env["PYTHONHTTPSVERIFY"] = "0"
                try:
                    proc = subprocess.Popen(
                        cmd,
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        env=env,
                    )
                except FileNotFoundError:
                    if not missing_tool_logged:
                        self.queue.put((None, None, f"[录制] 未找到命令: {self.auto_record_cmd}，请安装 streamlink 或修改录制命令", datetime.now().isoformat(), False))
                        missing_tool_logged = True
                    break
                except Exception as e:
                    self.queue.put((None, None, f"[录制] 启动失败 @{uid}: {e}", datetime.now().isoformat(), False))
                    time.sleep(8)
                    continue
                holder["proc"] = proc
                self.queue.put((None, None, f"[录制] 已启动 @{uid}", datetime.now().isoformat(), False))
                while proc.poll() is None and not stop_evt.is_set() and not self._stop_event.is_set():
                    time.sleep(1.0)
                if proc.poll() is None:
                    try:
                        proc.terminate()
                    except Exception:
                        pass
                    time.sleep(1.0)
                    if proc.poll() is None:
                        try:
                            proc.kill()
                        except Exception:
                            pass
                holder["proc"] = None
                if stop_evt.is_set() or self._stop_event.is_set():
                    break
                time.sleep(5.0)
            self.queue.put((None, None, f"[录制] 已停止 @{uid}", datetime.now().isoformat(), False))

        th = threading.Thread(target=target, daemon=True)
        self.record_workers[uid] = {"stop": stop_evt, "thread": th, "holder": holder}
        th.start()

    def _stop_auto_recorder_for_uid(self, unique_id: str):
        uid = str(unique_id or "").strip()
        if not uid:
            return
        worker = self.record_workers.pop(uid, None)
        if not worker:
            return
        try:
            worker["stop"].set()
        except Exception:
            pass
        try:
            proc = worker.get("holder", {}).get("proc")
            if proc and proc.poll() is None:
                proc.terminate()
        except Exception:
            pass

    def _stop_all_recorders(self):
        for uid in list(self.record_workers.keys()):
            self._stop_auto_recorder_for_uid(uid)

    def _build_overlap_message_report(self, hours: int = 24, include_main: bool = True, include_analysis: bool = True):
        now_ts = time.time()
        cutoff = now_ts - (max(1, int(hours)) * 3600)
        room_user_counter = defaultdict(lambda: defaultdict(int))
        user_name_latest = {}
        message_rows = []
        for row in self.overlap_message_events:
            if len(row) >= 6:
                ts, room_id, unique_id, display_name, content, source_tag = row
            else:
                ts, room_id, unique_id, display_name, content = row
                source_tag = "main"
            if (source_tag == "main" and not include_main) or (source_tag == "analysis" and not include_analysis):
                continue
            if ts < cutoff:
                continue
            room_user_counter[room_id][unique_id] += 1
            if display_name:
                user_name_latest[unique_id] = display_name
            message_rows.append((ts, room_id, unique_id, display_name, content))
        cross_stats = []
        user_rooms = defaultdict(set)
        total_msgs_by_user = defaultdict(int)
        for room_id, uc in room_user_counter.items():
            for uid, cnt in uc.items():
                user_rooms[uid].add(room_id)
                total_msgs_by_user[uid] += cnt
        for uid, rooms in user_rooms.items():
            if len(rooms) < 2:
                continue
            room_msg_detail = []
            for r in sorted(rooms):
                room_msg_detail.append(f"{r}:{room_user_counter[r].get(uid, 0)}")
            cross_stats.append({
                "permanent_id": self._latest_permanent_id_for_uid(uid),
                "unique_id": uid,
                "display_name": user_name_latest.get(uid, ""),
                "room_count": len(rooms),
                "total_messages": total_msgs_by_user.get(uid, 0),
                "room_details": " | ".join(room_msg_detail),
            })
        cross_stats.sort(key=lambda x: (-int(x["room_count"]), -int(x["total_messages"]), str(x["unique_id"])))
        message_rows.sort(key=lambda x: x[0], reverse=True)
        overlap_uids = {str(x.get("unique_id", "")) for x in cross_stats}
        overlap_message_rows = [x for x in message_rows if str(x[2]) in overlap_uids]
        return {
            "hours": hours,
            "cross_stats": cross_stats,
            "message_rows": message_rows,
            "overlap_message_rows": overlap_message_rows,
        }

    def _build_room_migration_report(self, hours: int = 24, include_main: bool = True, include_analysis: bool = True, max_gap_minutes: int = 120):
        now_ts = time.time()
        cutoff = now_ts - (max(1, int(hours)) * 3600)
        max_gap = max(1, int(max_gap_minutes)) * 60
        by_uid = defaultdict(list)
        for row in self.overlap_message_events:
            if len(row) >= 6:
                ts, room_id, unique_id, _display_name, _content, source_tag = row
            else:
                ts, room_id, unique_id, _display_name, _content = row[:5]
                source_tag = "main"
            if (source_tag == "main" and not include_main) or (source_tag == "analysis" and not include_analysis):
                continue
            if float(ts) < cutoff:
                continue
            rid = str(room_id or "").strip()
            uid = str(unique_id or "").strip()
            if not rid or not uid:
                continue
            by_uid[uid].append((float(ts), rid))

        edge_counter = defaultdict(int)         # (from_room, to_room) -> users
        edge_gap_sum = defaultdict(float)       # (from_room, to_room) -> seconds
        edge_gap_cnt = defaultdict(int)
        room_users = defaultdict(set)
        path_counter = defaultdict(int)         # "a -> b -> c" -> users

        for uid, events in by_uid.items():
            events.sort(key=lambda x: x[0])
            compressed = []
            last_room = ""
            for ts, rid in events:
                room_users[rid].add(uid)
                if rid == last_room:
                    continue
                compressed.append((ts, rid))
                last_room = rid
            if len(compressed) >= 2:
                path_rooms = [x[1] for x in compressed]
                path_counter[" -> ".join(path_rooms)] += 1
            for i in range(1, len(compressed)):
                prev_ts, prev_room = compressed[i - 1]
                cur_ts, cur_room = compressed[i]
                if cur_room == prev_room:
                    continue
                gap = max(0.0, float(cur_ts) - float(prev_ts))
                if gap > max_gap:
                    continue
                key = (prev_room, cur_room)
                edge_counter[key] += 1
                edge_gap_sum[key] += gap
                edge_gap_cnt[key] += 1

        total_migrations = sum(edge_counter.values())
        edges = []
        for (a, b), cnt in edge_counter.items():
            avg_gap = (edge_gap_sum[(a, b)] / edge_gap_cnt[(a, b)]) if edge_gap_cnt[(a, b)] else 0.0
            uv_from = len(room_users.get(a, set()))
            ratio = (cnt / max(1, uv_from)) if uv_from else 0.0
            edges.append({
                "from_room": a,
                "to_room": b,
                "users": int(cnt),
                "from_uv": int(uv_from),
                "ratio": float(ratio),
                "avg_seconds": float(avg_gap),
            })
        edges.sort(key=lambda x: (-int(x["users"]), -float(x["ratio"]), str(x["from_room"]), str(x["to_room"])))
        top_paths = sorted(path_counter.items(), key=lambda x: (-int(x[1]), x[0]))[:100]
        return {
            "hours": int(hours),
            "max_gap_minutes": int(max_gap_minutes),
            "total_migrations": int(total_migrations),
            "edges": edges,
            "top_paths": top_paths,
        }

    def _record_overlap_event(self, room_id: str, unique_id: str, when_ts: float, source_tag: str = "main"):
        if not room_id or not unique_id:
            return
        try:
            self.overlap_events.append((float(when_ts), str(room_id), str(unique_id), str(source_tag or "main")))
        except Exception:
            pass

    def _build_overlap_report(self, hours: int = 24, include_main: bool = True, include_analysis: bool = True):
        now_ts = time.time()
        cutoff = now_ts - (max(1, int(hours)) * 3600)
        room_users = {}
        user_rooms = {}
        for row in self.overlap_events:
            if len(row) >= 4:
                ts, room_id, unique_id, source_tag = row
            else:
                ts, room_id, unique_id = row
                source_tag = "main"
            if (source_tag == "main" and not include_main) or (source_tag == "analysis" and not include_analysis):
                continue
            if ts < cutoff:
                continue
            room_users.setdefault(room_id, set()).add(unique_id)
            user_rooms.setdefault(unique_id, set()).add(room_id)
        rooms = sorted(room_users.keys())
        overlap_counts = {}
        overlap_jaccard = {}
        for a in rooms:
            for b in rooms:
                sa = room_users.get(a, set())
                sb = room_users.get(b, set())
                inter = len(sa.intersection(sb))
                uni = len(sa.union(sb))
                overlap_counts[(a, b)] = inter
                overlap_jaccard[(a, b)] = (inter / uni) if uni else 0.0
        cross_users = sorted(
            [(uid, len(rs)) for uid, rs in user_rooms.items() if len(rs) >= 2],
            key=lambda x: (-x[1], x[0]),
        )
        return {
            "hours": hours,
            "rooms": rooms,
            "room_users": room_users,
            "overlap_counts": overlap_counts,
            "overlap_jaccard": overlap_jaccard,
            "cross_users": cross_users,
        }

    def _render_overlap_text(self, report: dict) -> str:
        rooms = report["rooms"]
        lines = [f"=== 跨房分析报告（最近 {report['hours']} 小时） ==="]
        if not rooms:
            lines.append("暂无数据（请先监听并等待弹幕事件）")
            return "\n".join(lines)
        lines.append("")
        lines.append("房间独立客人数(UV):")
        for r in rooms:
            lines.append(f"- {r}: {len(report['room_users'].get(r, set()))}")
        lines.append("")
        lines.append("重叠人数矩阵:")
        header = "room".ljust(20) + "".join(x[:12].ljust(14) for x in rooms)
        lines.append(header)
        for a in rooms:
            row = a[:18].ljust(20)
            for b in rooms:
                row += str(report["overlap_counts"][(a, b)]).ljust(14)
            lines.append(row)
        lines.append("")
        lines.append("Jaccard重叠率矩阵:")
        lines.append(header)
        for a in rooms:
            row = a[:18].ljust(20)
            for b in rooms:
                row += f"{report['overlap_jaccard'][(a, b)]:.3f}".ljust(14)
            lines.append(row)
        lines.append("")
        lines.append("跨房高频客人(>=2房):")
        for uid, cnt in report["cross_users"][:50]:
            lines.append(f"- {uid}: {cnt} rooms")
        return "\n".join(lines)

    def _refresh_overlap_view(self):
        try:
            hours = max(1, int(self._overlap_hours_var.get()))
        except Exception:
            hours = self.overlap_default_hours
        include_main = bool(getattr(self, "overlap_include_main", True))
        report = self._build_overlap_report(hours=hours, include_main=include_main, include_analysis=True)
        msg_report = self._build_overlap_message_report(hours=hours, include_main=include_main, include_analysis=True)
        migration_report = self._build_room_migration_report(hours=hours, include_main=include_main, include_analysis=True)
        content = self._render_overlap_text(report)
        content += (
            f"\n\n弹幕统计: 全部 {len(msg_report.get('message_rows', []))} 条，"
            f"相同客人弹幕 {len(msg_report.get('overlap_message_rows', []))} 条"
        )
        content += (
            f"\n迁移路径: 边数 {len(migration_report.get('edges', []))}，"
            f"总迁移次数 {int(migration_report.get('total_migrations', 0))}"
        )
        if hasattr(self, "_overlap_txt") and self._overlap_txt.winfo_exists():
            self._overlap_txt.config(state=tk.NORMAL)
            self._overlap_txt.delete("1.0", tk.END)
            self._overlap_txt.insert(tk.END, content)
            self._overlap_txt.config(state=tk.DISABLED)
        self._refresh_overlap_data_views(report, msg_report)

    def _refresh_overlap_data_views(self, report: dict = None, msg_report: dict = None):
        try:
            hours = max(1, int(self._overlap_hours_var.get()))
        except Exception:
            hours = self.overlap_default_hours
        include_main = bool(getattr(self, "overlap_include_main", True))
        if report is None:
            report = self._build_overlap_report(hours=hours, include_main=include_main, include_analysis=True)
        if msg_report is None:
            msg_report = self._build_overlap_message_report(hours=hours, include_main=include_main, include_analysis=True)
        room_filter = ""
        uid_filter = ""
        kw_filter = ""
        try:
            room_filter = (self._overlap_filter_room_var.get().strip() if hasattr(self, "_overlap_filter_room_var") else "").lower()
            uid_filter = (self._overlap_filter_uid_var.get().strip() if hasattr(self, "_overlap_filter_uid_var") else "").lower()
            kw_filter = (self._overlap_filter_kw_var.get().strip() if hasattr(self, "_overlap_filter_kw_var") else "").lower()
        except Exception:
            pass
        if hasattr(self, "_overlap_user_tree") and self._overlap_user_tree.winfo_exists():
            tree = self._overlap_user_tree
            for r in tree.get_children():
                tree.delete(r)
            for row in msg_report.get("cross_stats", [])[:1000]:
                uid = str(row.get("unique_id", ""))
                name = str(row.get("display_name", ""))
                details = str(row.get("room_details", ""))
                if uid_filter and uid_filter not in uid.lower() and uid_filter not in name.lower():
                    continue
                if room_filter and room_filter not in details.lower():
                    continue
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        row.get("permanent_id", "-"),
                        uid,
                        name,
                        row.get("room_count", 0),
                        row.get("total_messages", 0),
                        details,
                    ),
                )
        if hasattr(self, "_overlap_msg_tree") and self._overlap_msg_tree.winfo_exists():
            tree = self._overlap_msg_tree
            for r in tree.get_children():
                tree.delete(r)
            for ts, room_id, unique_id, display_name, content in msg_report.get("overlap_message_rows", [])[:5000]:
                if room_filter and room_filter not in str(room_id).lower():
                    continue
                if uid_filter and uid_filter not in str(unique_id).lower() and uid_filter not in str(display_name).lower():
                    continue
                if kw_filter and kw_filter not in str(content).lower():
                    continue
                t = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))
                tree.insert("", tk.END, values=(t, room_id, unique_id, display_name, content))
        if hasattr(self, "_overlap_room_tree") and self._overlap_room_tree.winfo_exists():
            room_msg_counts = defaultdict(int)
            room_uv = {}
            for ts, room_id, unique_id, _display_name, _content in msg_report.get("message_rows", []):
                room_msg_counts[room_id] += 1
            for rid in report.get("rooms", []):
                room_uv[rid] = len(report.get("room_users", {}).get(rid, set()))
            tree = self._overlap_room_tree
            keep_selected = list(tree.selection())
            for r in tree.get_children():
                tree.delete(r)
            for uid in self.overlap_watch_rooms:
                status = "监听中" if uid in self.analysis_listener_workers else "未监听"
                tree.insert(
                    "",
                    tk.END,
                    iid=uid,
                    values=(uid, status, int(room_uv.get(uid, 0)), int(room_msg_counts.get(uid, 0))),
                )
            for iid in keep_selected:
                if tree.exists(iid):
                    tree.selection_add(iid)
        if hasattr(self, "_overlap_path_tree") and self._overlap_path_tree.winfo_exists():
            mrep = self._build_room_migration_report(hours=hours, include_main=include_main, include_analysis=True)
            tree = self._overlap_path_tree
            for r in tree.get_children():
                tree.delete(r)
            for row in mrep.get("edges", [])[:1000]:
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        row.get("from_room", ""),
                        row.get("to_room", ""),
                        int(row.get("users", 0)),
                        f"{float(row.get('ratio', 0.0))*100:.2f}%",
                        f"{float(row.get('avg_seconds', 0.0))/60.0:.1f}",
                    ),
                )

    def _export_overlap_csv(self):
        try:
            hours = max(1, int(self._overlap_hours_var.get()))
        except Exception:
            hours = self.overlap_default_hours
        include_main = bool(getattr(self, "overlap_include_main", True))
        report = self._build_overlap_report(hours=hours, include_main=include_main, include_analysis=True)
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            title="导出跨房分析",
            initialfile=f"overlap_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(["room_id", "uv"])
                for r in report["rooms"]:
                    w.writerow([r, len(report["room_users"].get(r, set()))])
                w.writerow([])
                w.writerow(["overlap_matrix"])
                header = ["room"] + report["rooms"]
                w.writerow(header)
                for a in report["rooms"]:
                    row = [a] + [report["overlap_counts"][(a, b)] for b in report["rooms"]]
                    w.writerow(row)
                w.writerow([])
                w.writerow(["cross_room_users", "room_count"])
                for uid, cnt in report["cross_users"]:
                    w.writerow([uid, cnt])
            messagebox.showinfo("导出", f"已导出到 {path}")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def _export_overlap_migration_csv(self):
        try:
            hours = max(1, int(self._overlap_hours_var.get()))
        except Exception:
            hours = self.overlap_default_hours
        include_main = bool(getattr(self, "overlap_include_main", True))
        mrep = self._build_room_migration_report(hours=hours, include_main=include_main, include_analysis=True)
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            title="导出迁移路径",
            initialfile=f"room_migration_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(["from_room", "to_room", "users", "from_uv", "ratio", "avg_seconds", "avg_minutes"])
                for row in mrep.get("edges", []):
                    avg_sec = float(row.get("avg_seconds", 0.0))
                    w.writerow([
                        row.get("from_room", ""),
                        row.get("to_room", ""),
                        int(row.get("users", 0)),
                        int(row.get("from_uv", 0)),
                        f"{float(row.get('ratio', 0.0)):.6f}",
                        f"{avg_sec:.2f}",
                        f"{avg_sec/60.0:.2f}",
                    ])
                w.writerow([])
                w.writerow(["top_paths"])
                w.writerow(["path", "users"])
                for p, c in mrep.get("top_paths", []):
                    w.writerow([p, c])
            messagebox.showinfo("导出", f"已导出到 {path}")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def _export_overlap_detail_csv(self):
        try:
            hours = max(1, int(self._overlap_hours_var.get()))
        except Exception:
            hours = self.overlap_default_hours
        include_main = bool(getattr(self, "overlap_include_main", True))
        msg_report = self._build_overlap_message_report(hours=hours, include_main=include_main, include_analysis=True)
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            title="导出跨房明细",
            initialfile=f"overlap_detail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerow(["cross_users"])
                w.writerow(["permanent_id", "unique_id", "display_name", "room_count", "total_messages", "room_details"])
                for row in msg_report.get("cross_stats", []):
                    w.writerow([
                        row.get("permanent_id", "-"),
                        row.get("unique_id", ""),
                        row.get("display_name", ""),
                        row.get("room_count", 0),
                        row.get("total_messages", 0),
                        row.get("room_details", ""),
                    ])
                w.writerow([])
                w.writerow(["overlap_messages"])
                w.writerow(["time", "room_id", "unique_id", "display_name", "content"])
                for ts, room_id, unique_id, display_name, content in msg_report.get("overlap_message_rows", []):
                    t = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))
                    w.writerow([t, room_id, unique_id, display_name, content])
                w.writerow([])
                w.writerow(["all_messages"])
                w.writerow(["time", "room_id", "unique_id", "display_name", "content"])
                for ts, room_id, unique_id, display_name, content in msg_report.get("message_rows", []):
                    t = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))
                    w.writerow([t, room_id, unique_id, display_name, content])
            messagebox.showinfo("导出", f"已导出到 {path}")
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def open_overlap_analyzer(self):
        if hasattr(self, "_overlap_win") and self._overlap_win.winfo_exists():
            self._overlap_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("跨房分析系统")
        win.geometry("1260x760")
        if not self.overlap_watch_rooms:
            self.overlap_watch_rooms = self._parse_room_inputs(self.room_url_var.get().strip())

        top = ttk.Frame(win)
        top.pack(fill=tk.X, padx=8, pady=8)
        ttk.Label(top, text="新增监控直播间(支持 @id / url / id，逗号分隔):").grid(row=0, column=0, sticky=tk.W)
        self._overlap_rooms_var = tk.StringVar(value=",".join(self.overlap_watch_rooms))
        ttk.Entry(top, textvariable=self._overlap_rooms_var, width=80).grid(row=0, column=1, columnspan=4, sticky=tk.W)
        ttk.Label(top, text="分析窗口(小时):").grid(row=1, column=0, sticky=tk.W)
        self._overlap_hours_var = tk.StringVar(value=str(self.overlap_default_hours))
        ttk.Entry(top, textvariable=self._overlap_hours_var, width=8).grid(row=1, column=1, sticky=tk.W)
        ttk.Label(top, text="左侧可选中房间执行开始/停止").grid(row=1, column=2, columnspan=3, sticky=tk.W)
        ttk.Label(top, text="筛选房间:").grid(row=3, column=0, sticky=tk.W)
        self._overlap_filter_room_var = tk.StringVar(value="")
        ttk.Entry(top, textvariable=self._overlap_filter_room_var, width=14).grid(row=3, column=1, sticky=tk.W)
        ttk.Label(top, text="筛选客人:").grid(row=3, column=2, sticky=tk.E)
        self._overlap_filter_uid_var = tk.StringVar(value="")
        ttk.Entry(top, textvariable=self._overlap_filter_uid_var, width=14).grid(row=3, column=3, sticky=tk.W)
        ttk.Label(top, text="关键词:").grid(row=3, column=4, sticky=tk.E)
        self._overlap_filter_kw_var = tk.StringVar(value="")
        ttk.Entry(top, textvariable=self._overlap_filter_kw_var, width=14).grid(row=3, column=5, sticky=tk.W)
        include_main_var = tk.IntVar(value=1 if bool(self.overlap_include_main) else 0)
        ttk.Checkbutton(top, text="分析时包含主监控打印数据", variable=include_main_var).grid(row=4, column=0, columnspan=3, sticky=tk.W)
        auto_rec_var = tk.IntVar(value=1 if self.auto_record_enabled else 0)
        rec_dir_var = tk.StringVar(value=self.auto_record_dir)
        rec_cmd_var = tk.StringVar(value=self.auto_record_cmd)
        ttk.Checkbutton(top, text="自动录制(开播录制/下播停止)", variable=auto_rec_var).grid(row=2, column=0, sticky=tk.W)
        ttk.Label(top, text="录制目录:").grid(row=2, column=1, sticky=tk.E)
        ttk.Entry(top, textvariable=rec_dir_var, width=34).grid(row=2, column=2, sticky=tk.W)
        ttk.Label(top, text="录制命令:").grid(row=2, column=3, sticky=tk.E)
        ttk.Entry(top, textvariable=rec_cmd_var, width=14).grid(row=2, column=4, sticky=tk.W)

        # Safe defaults to avoid unbound closure when UI build is interrupted.
        add_rooms_from_input = lambda: None
        remove_selected_rooms = lambda: None
        clear_rooms = lambda: None
        start_selected = lambda: None
        stop_selected = lambda: None
        start_all = lambda: None
        stop_all = lambda: None
        sync_to_main = lambda: None
        save_overlap_settings = lambda: None

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=4)
        ttk.Button(btn, text="添加到监控清单", command=lambda: add_rooms_from_input()).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="删除选中房间", command=lambda: remove_selected_rooms()).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="清空监控清单", command=lambda: clear_rooms()).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="开始选中", command=lambda: start_selected()).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="结束选中", command=lambda: stop_selected()).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="开始全部", command=lambda: start_all()).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="结束全部", command=lambda: stop_all()).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="同步清单到主监听", command=lambda: sync_to_main()).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="保存分析设置", command=lambda: save_overlap_settings()).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="刷新分析", command=self._refresh_overlap_view).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="导出CSV", command=self._export_overlap_csv).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="导出明细CSV", command=self._export_overlap_detail_csv).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="导出迁移CSV", command=self._export_overlap_migration_csv).pack(side=tk.LEFT, padx=4)

        main = ttk.Panedwindow(win, orient=tk.HORIZONTAL)
        main.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        left = ttk.Frame(main, width=340)
        right = ttk.Panedwindow(main, orient=tk.VERTICAL)
        main.add(left, weight=1)
        main.add(right, weight=3)

        room_box = ttk.LabelFrame(left, text="监控房间")
        room_box.pack(fill=tk.BOTH, expand=True)
        room_wrap = ttk.Frame(room_box)
        room_wrap.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        room_wrap.rowconfigure(0, weight=1)
        room_wrap.columnconfigure(0, weight=1)
        room_tree = ttk.Treeview(room_wrap, columns=("uid", "status", "uv", "msg"), show="headings", selectmode="extended")
        room_tree.heading("uid", text="主播ID")
        room_tree.heading("status", text="状态")
        room_tree.heading("uv", text="UV")
        room_tree.heading("msg", text="弹幕数")
        room_tree.column("uid", width=140, anchor=tk.W)
        room_tree.column("status", width=80, anchor=tk.CENTER)
        room_tree.column("uv", width=60, anchor=tk.E)
        room_tree.column("msg", width=80, anchor=tk.E)
        room_vbar = ttk.Scrollbar(room_wrap, orient=tk.VERTICAL, command=room_tree.yview)
        room_hbar = ttk.Scrollbar(room_wrap, orient=tk.HORIZONTAL, command=room_tree.xview)
        room_tree.configure(yscrollcommand=room_vbar.set, xscrollcommand=room_hbar.set)
        room_tree.grid(row=0, column=0, sticky="nsew")
        room_vbar.grid(row=0, column=1, sticky="ns")
        room_hbar.grid(row=1, column=0, sticky="ew")
        self._overlap_room_tree = room_tree

        report_box = ttk.LabelFrame(right, text="跨房报告")
        path_box = ttk.LabelFrame(right, text="迁移路径(A→B)")
        user_box = ttk.LabelFrame(right, text="相同客人统计(跨房>=2)")
        msg_box = ttk.LabelFrame(right, text="弹幕明细(最近窗口)")
        right.add(report_box, weight=2)
        right.add(path_box, weight=1)
        right.add(user_box, weight=2)
        right.add(msg_box, weight=3)

        txt = scrolledtext.ScrolledText(report_box, width=110, height=12)
        txt.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        path_wrap = ttk.Frame(path_box)
        path_wrap.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        path_wrap.rowconfigure(0, weight=1)
        path_wrap.columnconfigure(0, weight=1)
        path_tree = ttk.Treeview(path_wrap, columns=("from", "to", "users", "ratio", "avgm"), show="headings", selectmode="extended")
        path_tree.heading("from", text="来源房间")
        path_tree.heading("to", text="去向房间")
        path_tree.heading("users", text="迁移人数")
        path_tree.heading("ratio", text="转移率")
        path_tree.heading("avgm", text="平均迁移(分钟)")
        path_tree.column("from", width=130, anchor=tk.W)
        path_tree.column("to", width=130, anchor=tk.W)
        path_tree.column("users", width=90, anchor=tk.E)
        path_tree.column("ratio", width=90, anchor=tk.E)
        path_tree.column("avgm", width=120, anchor=tk.E)
        path_vbar = ttk.Scrollbar(path_wrap, orient=tk.VERTICAL, command=path_tree.yview)
        path_hbar = ttk.Scrollbar(path_wrap, orient=tk.HORIZONTAL, command=path_tree.xview)
        path_tree.configure(yscrollcommand=path_vbar.set, xscrollcommand=path_hbar.set)
        path_tree.grid(row=0, column=0, sticky="nsew")
        path_vbar.grid(row=0, column=1, sticky="ns")
        path_hbar.grid(row=1, column=0, sticky="ew")
        user_wrap = ttk.Frame(user_box)
        user_wrap.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        user_wrap.rowconfigure(0, weight=1)
        user_wrap.columnconfigure(0, weight=1)
        user_tree = ttk.Treeview(user_wrap, columns=("pid", "uid", "name", "rooms", "msgs", "details"), show="headings", selectmode="extended")
        user_tree.heading("pid", text="永久ID")
        user_tree.heading("uid", text="客户ID")
        user_tree.heading("name", text="昵称")
        user_tree.heading("rooms", text="房间数")
        user_tree.heading("msgs", text="总弹幕")
        user_tree.heading("details", text="房间弹幕分布")
        user_tree.column("pid", width=100, anchor=tk.W)
        user_tree.column("uid", width=160, anchor=tk.W)
        user_tree.column("name", width=120, anchor=tk.W)
        user_tree.column("rooms", width=80, anchor=tk.E)
        user_tree.column("msgs", width=80, anchor=tk.E)
        user_tree.column("details", width=380, anchor=tk.W)
        user_vbar = ttk.Scrollbar(user_wrap, orient=tk.VERTICAL, command=user_tree.yview)
        user_hbar = ttk.Scrollbar(user_wrap, orient=tk.HORIZONTAL, command=user_tree.xview)
        user_tree.configure(yscrollcommand=user_vbar.set, xscrollcommand=user_hbar.set)
        user_tree.grid(row=0, column=0, sticky="nsew")
        user_vbar.grid(row=0, column=1, sticky="ns")
        user_hbar.grid(row=1, column=0, sticky="ew")
        user_tree.bind("<Double-1>", lambda e: self.open_user_profile_from_overlap())
        msg_wrap = ttk.Frame(msg_box)
        msg_wrap.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        msg_wrap.rowconfigure(0, weight=1)
        msg_wrap.columnconfigure(0, weight=1)
        msg_tree = ttk.Treeview(msg_wrap, columns=("t", "room", "uid", "name", "content"), show="headings", selectmode="extended")
        msg_tree.heading("t", text="时间")
        msg_tree.heading("room", text="房间")
        msg_tree.heading("uid", text="客户ID")
        msg_tree.heading("name", text="昵称")
        msg_tree.heading("content", text="弹幕")
        msg_tree.column("t", width=140, anchor=tk.W)
        msg_tree.column("room", width=120, anchor=tk.W)
        msg_tree.column("uid", width=150, anchor=tk.W)
        msg_tree.column("name", width=110, anchor=tk.W)
        msg_tree.column("content", width=460, anchor=tk.W)
        msg_vbar = ttk.Scrollbar(msg_wrap, orient=tk.VERTICAL, command=msg_tree.yview)
        msg_hbar = ttk.Scrollbar(msg_wrap, orient=tk.HORIZONTAL, command=msg_tree.xview)
        msg_tree.configure(yscrollcommand=msg_vbar.set, xscrollcommand=msg_hbar.set)
        msg_tree.grid(row=0, column=0, sticky="nsew")
        msg_vbar.grid(row=0, column=1, sticky="ns")
        msg_hbar.grid(row=1, column=0, sticky="ew")
        self._overlap_user_tree = user_tree
        self._overlap_msg_tree = msg_tree
        self._overlap_path_tree = path_tree
        self._overlap_filter_room_var.trace_add("write", lambda *_: self._debounce_call("overlap_filter_refresh", 220, self._refresh_overlap_data_views))
        self._overlap_filter_uid_var.trace_add("write", lambda *_: self._debounce_call("overlap_filter_refresh", 220, self._refresh_overlap_data_views))
        self._overlap_filter_kw_var.trace_add("write", lambda *_: self._debounce_call("overlap_filter_refresh", 220, self._refresh_overlap_data_views))
        include_main_var.trace_add(
            "write",
            lambda *_: (
                setattr(self, "overlap_include_main", bool(include_main_var.get())),
                self._request_save_settings(),
                self._debounce_call("overlap_view_refresh", 120, self._refresh_overlap_view),
            ),
        )

        def add_rooms_from_input():
            uids = self._parse_room_inputs(self._overlap_rooms_var.get())
            if not uids:
                return
            for uid in uids:
                if uid not in self.overlap_watch_rooms:
                    self.overlap_watch_rooms.append(uid)
            self.overlap_include_main = bool(include_main_var.get())
            self.auto_record_enabled = bool(auto_rec_var.get())
            self.auto_record_dir = rec_dir_var.get().strip() or os.path.join(os.getcwd(), "recordings")
            self.auto_record_cmd = rec_cmd_var.get().strip() or "streamlink"
            self._overlap_rooms_var.set(",".join(self.overlap_watch_rooms))
            self._save_settings()
            self._refresh_overlap_view()

        def selected_uids():
            sel = room_tree.selection()
            out = []
            for iid in sel:
                vals = room_tree.item(iid, "values")
                uid = vals[0] if vals else iid
                if uid:
                    out.append(str(uid))
            return out

        def remove_selected_rooms():
            sel = set(selected_uids())
            if not sel:
                return
            self.overlap_watch_rooms = [u for u in self.overlap_watch_rooms if u not in sel]
            self._overlap_rooms_var.set(",".join(self.overlap_watch_rooms))
            self._save_settings()
            self._refresh_overlap_view()

        def clear_rooms():
            self.overlap_watch_rooms = []
            self._overlap_rooms_var.set("")
            self._save_settings()
            self._refresh_overlap_view()

        def _start_uids(uids):
            if not uids:
                return
            self.overlap_include_main = bool(include_main_var.get())
            self.auto_record_enabled = bool(auto_rec_var.get())
            self.auto_record_dir = rec_dir_var.get().strip() or os.path.join(os.getcwd(), "recordings")
            self.auto_record_cmd = rec_cmd_var.get().strip() or "streamlink"
            self._analysis_stop_event.clear()
            for uid in uids:
                self._start_listener_for_uid(
                    uid,
                    worker_map=self.analysis_listener_workers,
                    group_stop_event=self._analysis_stop_event,
                    source_tag="analysis",
                    enable_print=False,
                )
            self._set_status(f"分析监听中: {', '.join('@'+u for u in self.analysis_listener_workers.keys())}")
            self._refresh_overlap_view()

        def _stop_uids(uids):
            if not uids:
                return
            for uid in uids:
                self._stop_listener_for_uid(uid, reason=f"分析手动停止: @{uid}", worker_map=self.analysis_listener_workers, source_tag="analysis")
            self._refresh_overlap_view()

        def start_selected():
            _start_uids(selected_uids())

        def stop_selected():
            _stop_uids(selected_uids())

        def start_all():
            _start_uids(list(self.overlap_watch_rooms))

        def stop_all():
            _stop_uids(list(self.overlap_watch_rooms))

        def sync_to_main():
            joined = ",".join(f"https://www.tiktok.com/@{u}" for u in self.overlap_watch_rooms)
            self.room_url_var.set(joined)

        def save_overlap_settings():
            self.overlap_include_main = bool(include_main_var.get())
            self.auto_record_enabled = bool(auto_rec_var.get())
            self.auto_record_dir = rec_dir_var.get().strip() or os.path.join(os.getcwd(), "recordings")
            self.auto_record_cmd = rec_cmd_var.get().strip() or "streamlink"
            try:
                self.overlap_default_hours = max(1, int(self._overlap_hours_var.get()))
            except Exception:
                pass
            current_rooms = self._parse_room_inputs(self._overlap_rooms_var.get())
            if current_rooms:
                self.overlap_watch_rooms = current_rooms
            self._save_settings()
            self._set_status("已保存跨房分析设置")
            messagebox.showinfo("保存成功", "跨房分析设置已保存")

        def _copy_text(text: str):
            try:
                self.root.clipboard_clear()
                self.root.clipboard_append(str(text or ""))
            except Exception:
                pass

        def _on_room_right_click(event):
            row_id = room_tree.identify_row(event.y)
            if row_id:
                room_tree.selection_set(row_id)
            menu = tk.Menu(win, tearoff=0)
            menu.add_command(label="开始选中房间", command=start_selected)
            menu.add_command(label="结束选中房间", command=stop_selected)
            menu.add_command(label="删除选中房间", command=remove_selected_rooms)
            menu.add_separator()
            menu.add_command(label="复制房间ID", command=lambda: _copy_text(",".join(selected_uids())))
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        def _selected_user_id():
            sel = user_tree.selection()
            if not sel:
                return ""
            vals = user_tree.item(sel[0], "values")
            return str(vals[1]) if len(vals) > 1 else ""

        def _selected_user_name():
            sel = user_tree.selection()
            if not sel:
                return ""
            vals = user_tree.item(sel[0], "values")
            return str(vals[2]) if len(vals) > 2 else ""

        def _blacklist_selected_user():
            uid = _selected_user_id()
            if not uid:
                return
            db.add_blacklist(uid)
            self._audit("blacklist_add_overlap_user", uid)
            messagebox.showinfo("完成", f"已拉黑用户: {uid}")

        def _on_user_right_click(event):
            row_id = user_tree.identify_row(event.y)
            if row_id:
                user_tree.selection_set(row_id)
            menu = tk.Menu(win, tearoff=0)
            menu.add_command(label="查看客户画像", command=self.open_user_profile_from_overlap)
            menu.add_command(label="拉黑该客户", command=_blacklist_selected_user)
            menu.add_separator()
            menu.add_command(label="复制客户ID", command=lambda: _copy_text(_selected_user_id()))
            menu.add_command(label="复制昵称", command=lambda: _copy_text(_selected_user_name()))
            menu.add_command(label="按该客户筛选", command=lambda: self._overlap_filter_uid_var.set(_selected_user_id()))
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        def _selected_msg_values():
            sel = msg_tree.selection()
            if not sel:
                return ()
            vals = msg_tree.item(sel[0], "values")
            return vals if vals else ()

        def _selected_path_values():
            sel = path_tree.selection()
            if not sel:
                return ()
            vals = path_tree.item(sel[0], "values")
            return vals if vals else ()

        def _on_msg_right_click(event):
            row_id = msg_tree.identify_row(event.y)
            if row_id:
                msg_tree.selection_set(row_id)
            menu = tk.Menu(win, tearoff=0)
            menu.add_command(label="复制弹幕", command=lambda: _copy_text((_selected_msg_values()[4] if len(_selected_msg_values()) > 4 else "")))
            menu.add_command(label="复制客户ID", command=lambda: _copy_text((_selected_msg_values()[2] if len(_selected_msg_values()) > 2 else "")))
            menu.add_command(label="按该客户筛选", command=lambda: self._overlap_filter_uid_var.set((_selected_msg_values()[2] if len(_selected_msg_values()) > 2 else "")))
            menu.add_command(label="按该房间筛选", command=lambda: self._overlap_filter_room_var.set((_selected_msg_values()[1] if len(_selected_msg_values()) > 1 else "")))
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        def _on_path_right_click(event):
            row_id = path_tree.identify_row(event.y)
            if row_id:
                path_tree.selection_set(row_id)
            menu = tk.Menu(win, tearoff=0)
            menu.add_command(label="复制来源房间", command=lambda: _copy_text((_selected_path_values()[0] if len(_selected_path_values()) > 0 else "")))
            menu.add_command(label="复制去向房间", command=lambda: _copy_text((_selected_path_values()[1] if len(_selected_path_values()) > 1 else "")))
            menu.add_command(label="按来源房间筛选", command=lambda: self._overlap_filter_room_var.set((_selected_path_values()[0] if len(_selected_path_values()) > 0 else "")))
            menu.add_command(label="按去向房间筛选", command=lambda: self._overlap_filter_room_var.set((_selected_path_values()[1] if len(_selected_path_values()) > 1 else "")))
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        room_tree.bind("<Button-3>", _on_room_right_click)
        room_tree.bind("<Button-2>", _on_room_right_click)
        room_tree.bind("<Control-Button-1>", _on_room_right_click)
        user_tree.bind("<Button-3>", _on_user_right_click)
        user_tree.bind("<Button-2>", _on_user_right_click)
        user_tree.bind("<Control-Button-1>", _on_user_right_click)
        msg_tree.bind("<Button-3>", _on_msg_right_click)
        msg_tree.bind("<Button-2>", _on_msg_right_click)
        msg_tree.bind("<Control-Button-1>", _on_msg_right_click)
        path_tree.bind("<Button-3>", _on_path_right_click)
        path_tree.bind("<Button-2>", _on_path_right_click)
        path_tree.bind("<Control-Button-1>", _on_path_right_click)

        def _on_overlap_close():
            self.overlap_include_main = bool(include_main_var.get())
            self.auto_record_enabled = bool(auto_rec_var.get())
            self.auto_record_dir = rec_dir_var.get().strip() or os.path.join(os.getcwd(), "recordings")
            self.auto_record_cmd = rec_cmd_var.get().strip() or "streamlink"
            self._save_settings()
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", _on_overlap_close)
        self._overlap_txt = txt
        self._overlap_win = win
        self._refresh_overlap_view()
        self._center_window(win, 1260, 760)

    def open_user_profile_from_overlap(self):
        if not hasattr(self, "_overlap_user_tree") or not self._overlap_user_tree.winfo_exists():
            return
        sel = self._overlap_user_tree.selection()
        if not sel:
            return
        vals = self._overlap_user_tree.item(sel[0], "values")
        if not vals:
            return
        pid = str(vals[0]) if len(vals) > 0 else "-"
        uid = str(vals[1]) if len(vals) > 1 else ""
        name = str(vals[2]) if len(vals) > 2 else ""
        room_count = int(vals[3]) if len(vals) > 3 else 0
        msg_count = int(vals[4]) if len(vals) > 4 else 0
        events = [x for x in self.overlap_message_events if str(x[2]) == uid]
        events.sort(key=lambda x: x[0])
        first_ts = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(events[0][0])) if events else "-"
        last_ts = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(events[-1][0])) if events else "-"
        by_hour = defaultdict(int)
        by_kw = defaultdict(int)
        for row in events:
            ts = row[0]
            content = row[4] if len(row) > 4 else ""
            hh = time.strftime("%H", time.localtime(ts))
            by_hour[hh] += 1
            for kw in ("1", "2", "3", "4", "5", "预定", "预约", "buy", "order"):
                if kw in str(content):
                    by_kw[kw] += 1
        hot_hours = ", ".join(f"{k}:{v}" for k, v in sorted(by_hour.items(), key=lambda x: -x[1])[:5]) or "-"
        hot_kw = ", ".join(f"{k}:{v}" for k, v in sorted(by_kw.items(), key=lambda x: -x[1])[:8]) or "-"
        lines = [
            f"永久ID: {pid}",
            f"客户ID: {uid}",
            f"昵称: {name}",
            f"跨房数量: {room_count}",
            f"总弹幕数: {msg_count}",
            f"首次出现: {first_ts}",
            f"最近出现: {last_ts}",
            f"高活跃时段: {hot_hours}",
            f"常见关键词: {hot_kw}",
        ]
        win = tk.Toplevel(self.root)
        win.title("客户画像卡")
        txt = scrolledtext.ScrolledText(win, width=70, height=18)
        txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        txt.insert(tk.END, "\n".join(lines))
        txt.config(state=tk.DISABLED)

    def open_replay_center(self):
        if hasattr(self, "_replay_win") and self._replay_win.winfo_exists():
            self._replay_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("弹幕回放")
        win.geometry("980x620")
        ctrl = ttk.Frame(win)
        ctrl.pack(fill=tk.X, padx=8, pady=8)
        ttk.Label(ctrl, text="回放房间(可空):").pack(side=tk.LEFT)
        room_var = tk.StringVar(value="")
        ttk.Entry(ctrl, textvariable=room_var, width=14).pack(side=tk.LEFT, padx=4)
        ttk.Label(ctrl, text="回放客人(可空):").pack(side=tk.LEFT)
        uid_var = tk.StringVar(value="")
        ttk.Entry(ctrl, textvariable=uid_var, width=14).pack(side=tk.LEFT, padx=4)
        speed_var = tk.DoubleVar(value=1.0)
        ttk.Label(ctrl, text="速度").pack(side=tk.LEFT)
        ttk.Scale(ctrl, from_=0.2, to=4.0, variable=speed_var, orient=tk.HORIZONTAL, length=140).pack(side=tk.LEFT, padx=4)
        tree = ttk.Treeview(win, columns=("t", "room", "uid", "name", "content"), show="headings")
        for k, w in (("t", 140), ("room", 120), ("uid", 150), ("name", 110), ("content", 430)):
            tree.heading(k, text=k)
            tree.column(k, width=w, anchor=tk.W)
        tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        state = {"rows": [], "idx": 0, "playing": False}

        def load_rows():
            r = room_var.get().strip().lower()
            u = uid_var.get().strip().lower()
            rows = []
            for row in list(self.overlap_message_events):
                ts = row[0]
                room_id = row[1] if len(row) > 1 else ""
                unique_id = row[2] if len(row) > 2 else ""
                display_name = row[3] if len(row) > 3 else ""
                content = row[4] if len(row) > 4 else ""
                if r and r not in str(room_id).lower():
                    continue
                if u and u not in str(unique_id).lower() and u not in str(display_name).lower():
                    continue
                rows.append((ts, room_id, unique_id, display_name, content))
            rows.sort(key=lambda x: x[0])
            state["rows"] = rows
            state["idx"] = 0
            for x in tree.get_children():
                tree.delete(x)

        def tick():
            if not state["playing"]:
                return
            rows = state["rows"]
            if state["idx"] >= len(rows):
                state["playing"] = False
                return
            ts, room_id, unique_id, display_name, content = rows[state["idx"]]
            t = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))
            tree.insert("", tk.END, values=(t, room_id, unique_id, display_name, content))
            if tree.get_children():
                tree.see(tree.get_children()[-1])
            state["idx"] += 1
            delay = max(30, int(300 / max(0.2, float(speed_var.get()))))
            win.after(delay, tick)

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(btn, text="加载数据", command=load_rows).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="开始回放", command=lambda: (state.update({"playing": True}), tick())).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="暂停", command=lambda: state.update({"playing": False})).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="清空画面", command=lambda: [tree.delete(x) for x in tree.get_children()]).pack(side=tk.LEFT, padx=4)
        self._replay_win = win

    def open_template_version_manager(self):
        if hasattr(self, "_tplver_win") and self._tplver_win.winfo_exists():
            self._tplver_win.lift()
            return
        versions = self.settings.get("template_versions", [])
        if not isinstance(versions, list):
            versions = []
        win = tk.Toplevel(self.root)
        win.title("模板版本管理")
        win.geometry("820x520")
        cols = ("name", "time", "mode")
        tree = ttk.Treeview(win, columns=cols, show="headings", selectmode="browse")
        tree.heading("name", text="版本名")
        tree.heading("time", text="时间")
        tree.heading("mode", text="模板类型")
        tree.column("name", width=220, anchor=tk.W)
        tree.column("time", width=180, anchor=tk.W)
        tree.column("mode", width=120, anchor=tk.W)
        tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        name_var = tk.StringVar(value=f"版本_{datetime.now().strftime('%m%d_%H%M')}")
        frm = ttk.Frame(win)
        frm.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Entry(frm, textvariable=name_var, width=28).pack(side=tk.LEFT)

        def refresh():
            for i in tree.get_children():
                tree.delete(i)
            for i, row in enumerate(versions):
                tree.insert("", tk.END, iid=str(i), values=(row.get("name", ""), row.get("saved_at", ""), row.get("mode", "")))

        def save_new():
            mode = self._get_template_mode()
            row = {
                "name": name_var.get().strip() or f"版本_{datetime.now().strftime('%m%d_%H%M')}",
                "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "mode": mode,
                "template": self.template.get(),
                "canvas_template": self._normalize_canvas_template_to_paper(self.canvas_template),
                "canvas_enabled": int(self.canvas_template_enabled.get()),
            }
            versions.append(row)
            self.settings["template_versions"] = versions[-200:]
            self._save_settings()
            refresh()
            self._audit("template_version_save", row["name"])

        def restore_one():
            sel = tree.selection()
            if not sel:
                return
            row = versions[int(sel[0])]
            self.template.set(str(row.get("template", self.template.get())))
            self.canvas_template = self._normalize_canvas_template_to_paper(row.get("canvas_template", self.canvas_template))
            restore_mode = str(row.get("mode", "")).strip().lower()
            if restore_mode not in ("editor", "designer"):
                restore_mode = "designer" if int(row.get("canvas_enabled", 0)) else "editor"
            self._set_template_mode(restore_mode, save=False)
            self._save_settings()
            self._audit("template_version_restore", row.get("name", ""))
            messagebox.showinfo("恢复", "模板版本已恢复")

        def delete_one():
            sel = tree.selection()
            if not sel:
                return
            idx = int(sel[0])
            row = versions.pop(idx)
            self.settings["template_versions"] = versions
            self._save_settings()
            refresh()
            self._audit("template_version_delete", row.get("name", ""))

        ttk.Button(frm, text="保存当前版本", command=save_new).pack(side=tk.LEFT, padx=4)
        ttk.Button(frm, text="恢复选中版本", command=restore_one).pack(side=tk.LEFT, padx=4)
        ttk.Button(frm, text="删除选中版本", command=delete_one).pack(side=tk.LEFT, padx=4)
        refresh()
        self._tplver_win = win

    def open_recording_manager(self):
        if hasattr(self, "_rec_win") and self._rec_win.winfo_exists():
            self._rec_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("录制管理")
        win.geometry("920x580")
        path_var = tk.StringVar(value=self.auto_record_dir)
        enabled_var = tk.IntVar(value=1 if self.auto_record_enabled else 0)
        cmd_var = tk.StringVar(value=self.auto_record_cmd)
        top = ttk.Frame(win)
        top.pack(fill=tk.X, padx=8, pady=8)
        ttk.Checkbutton(top, text="启用自动录制", variable=enabled_var).pack(side=tk.LEFT)
        ttk.Label(top, text="目录").pack(side=tk.LEFT, padx=(10, 2))
        ttk.Entry(top, textvariable=path_var, width=34).pack(side=tk.LEFT)
        ttk.Label(top, text="命令").pack(side=tk.LEFT, padx=(10, 2))
        ttk.Entry(top, textvariable=cmd_var, width=14).pack(side=tk.LEFT)
        cols = ("uid", "status")
        active_tree = ttk.Treeview(win, columns=cols, show="headings")
        active_tree.heading("uid", text="房间")
        active_tree.heading("status", text="状态")
        active_tree.column("uid", width=160, anchor=tk.W)
        active_tree.column("status", width=120, anchor=tk.W)
        active_tree.pack(fill=tk.X, padx=8, pady=4)
        file_tree = ttk.Treeview(win, columns=("name", "size", "mtime"), show="headings")
        for k, t, w in (("name", "文件", 500), ("size", "大小MB", 100), ("mtime", "修改时间", 180)):
            file_tree.heading(k, text=t)
            file_tree.column(k, width=w, anchor=tk.W)
        file_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        def refresh():
            self.auto_record_enabled = bool(enabled_var.get())
            self.auto_record_dir = path_var.get().strip() or os.path.join(os.getcwd(), "recordings")
            self.auto_record_cmd = cmd_var.get().strip() or "streamlink"
            self._save_settings()
            for x in active_tree.get_children():
                active_tree.delete(x)
            for uid in sorted(self.listener_workers.keys()):
                st = "录制中" if uid in self.record_workers else "待机"
                active_tree.insert("", tk.END, values=(uid, st))
            for x in file_tree.get_children():
                file_tree.delete(x)
            d = self.auto_record_dir
            if os.path.isdir(d):
                files = []
                for n in os.listdir(d):
                    p = os.path.join(d, n)
                    if os.path.isfile(p):
                        try:
                            st = os.stat(p)
                            files.append((n, st.st_size, st.st_mtime))
                        except Exception:
                            pass
                files.sort(key=lambda x: x[2], reverse=True)
                for n, s, mt in files[:1000]:
                    file_tree.insert("", tk.END, values=(n, f"{s/1024/1024:.2f}", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(mt))))

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(btn, text="刷新", command=refresh).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="应用设置", command=refresh).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="停止全部录制", command=self._stop_all_recorders).pack(side=tk.LEFT, padx=4)
        refresh()
        self._rec_win = win

    def open_audit_log_viewer(self):
        if hasattr(self, "_audit_win") and self._audit_win.winfo_exists():
            self._audit_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("操作日志")
        win.geometry("900x560")
        txt = scrolledtext.ScrolledText(win)
        txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        try:
            if os.path.exists(self.audit_log_path):
                with open(self.audit_log_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()[-3000:]
                txt.insert(tk.END, "".join(lines))
        except Exception as e:
            txt.insert(tk.END, f"读取失败: {e}")
        txt.config(state=tk.DISABLED)
        self._audit_win = win

    def open_health_check_panel(self):
        checks = []
        try:
            checks.append(("Python", "OK"))
            checks.append(("代理", self._proxy_route_mode_label(self._get_proxy_route_mode())))
            checks.append(("SSL校验", "关闭(稳定优先)" if self.ssl_insecure_var.get() else "开启"))
            checks.append(("监听房间", str(len(self.listener_workers))))
            checks.append(("打印队列pending", str(len(db.list_print_jobs("pending")))))
            checks.append(("最近60秒失败", str(db.get_recent_failed_count(60))))
            rec_dir = self.auto_record_dir if self.auto_record_dir else os.path.join(os.getcwd(), "recordings")
            free_mb = "-"
            try:
                du = shutil.disk_usage(rec_dir if os.path.exists(rec_dir) else os.getcwd())
                free_mb = f"{du.free/1024/1024:.0f}"
            except Exception:
                pass
            checks.append(("录制目录可用空间MB", free_mb))
        except Exception as e:
            checks.append(("HealthError", str(e)))
        win = tk.Toplevel(self.root)
        win.title("一键健康检查")
        tree = ttk.Treeview(win, columns=("k", "v"), show="headings")
        tree.heading("k", text="项目")
        tree.heading("v", text="结果")
        tree.column("k", width=220, anchor=tk.W)
        tree.column("v", width=520, anchor=tk.W)
        tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        for k, v in checks:
            tree.insert("", tk.END, values=(k, v))

    def open_startup_checklist(self):
        if hasattr(self, "_sop_win") and self._sop_win.winfo_exists():
            self._sop_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("开播前检查(SOP)")
        win.geometry("760x520")
        self._center_window(win, 760, 520)
        checks = [
            ("直播间地址已填", bool(self.room_url_var.get().strip())),
            ("打印机已选择", bool(self.printer_cb.get().strip())),
            ("代理配置合理", (self._get_proxy_route_mode() == "direct") or bool(self.proxy_var.get().strip()) or bool(self._resolve_configured_proxy())),
            ("SSL稳定模式", bool(self.ssl_insecure_var.get())),
            ("模板已启用", bool(self.canvas_template_enabled.get()) or bool(self.template.get().strip())),
            ("自动录制配置", (not self.auto_record_enabled) or bool(self.auto_record_cmd.strip())),
            ("数据保留天数>=7", int(self.data_retention_days) >= 7),
        ]
        tree = ttk.Treeview(win, columns=("item", "status"), show="headings")
        tree.heading("item", text="检查项")
        tree.heading("status", text="状态")
        tree.column("item", width=520, anchor=tk.W)
        tree.column("status", width=140, anchor=tk.W)
        tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        for k, ok in checks:
            tree.insert("", tk.END, values=(k, "通过" if ok else "待处理"))
        txt = scrolledtext.ScrolledText(win, height=8)
        txt.pack(fill=tk.BOTH, expand=False, padx=8, pady=(0, 8))
        tips = [
            "建议顺序: 检查网络 -> 监听测试 -> 打印测试 -> 录制测试 -> 正式开播。",
            "如果出现SSL错误，优先保持“SSL不校验”开启（稳定优先）。",
            "开播前先清理上次异常任务，避免队列污染。",
        ]
        txt.insert(tk.END, "\n".join(tips))
        txt.config(state=tk.DISABLED)
        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(btn, text="保存当前配置", command=self.save_connection_settings).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="打开健康检查", command=self.open_health_check_panel).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="关闭", command=win.destroy).pack(side=tk.LEFT, padx=4)
        self._track_business_event("open_startup_checklist", {"ok_count": sum(1 for _, x in checks if x), "total": len(checks)})
        self._sop_win = win

    def open_performance_dashboard(self):
        if hasattr(self, "_perf_win") and self._perf_win.winfo_exists():
            self._perf_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("性能看板")
        win.geometry("760x420")
        txt = scrolledtext.ScrolledText(win)
        txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        def tick():
            if not win.winfo_exists():
                return
            try:
                mem = self._memory_text()
                listeners = len(self.listener_workers)
                recorders = len(self.record_workers)
                q = self.queue.qsize()
                rows = len(self.stream_tree.get_children())
                pending = len(db.list_print_jobs("pending"))
                failed60 = db.get_recent_failed_count(60)
                lines = [
                    f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    f"监听房间: {listeners}",
                    f"录制任务: {recorders}",
                    f"内存: {mem}",
                    f"消息队列: {q}",
                    f"弹幕行数: {rows}",
                    f"打印pending: {pending}",
                    f"最近60秒失败: {failed60}",
                ]
                txt.config(state=tk.NORMAL)
                txt.delete("1.0", tk.END)
                txt.insert(tk.END, "\n".join(lines))
                txt.config(state=tk.DISABLED)
            except Exception:
                pass
            win.after(1000, tick)

        tick()
        self._perf_win = win
    def _load_preset(self):
        preset_name = self.preset_cb.get()
        if preset_name in PRESET_TEMPLATES:
            preset = PRESET_TEMPLATES[preset_name]
            self.template.set(preset["template"])
            self.font_size_var.set(preset["font_size"])
            self.alignment_var.set(preset["alignment"])
            self.line_spacing_var.set(preset["line_spacing"])
            self.paper_width_var.set(preset["paper_width"])
            self._set_template_mode("editor", save=False)

    def _load_paper_preset(self, event=None):
        # Printing uses custom paper size only; preset entry removed.
        return

    def _print_calibration(self):
        try:
            char_width_mm = max(0.8, float(str(self.char_width_mm_var.get()).strip()))
        except Exception:
            char_width_mm = 1.5
        try:
            line_height_mm = max(1.0, float(str(self.line_height_mm_var.get()).strip()))
        except Exception:
            line_height_mm = 2.8
        try:
            margin_mm = max(0.2, float(str(self.margin_mm_var.get()).strip()))
        except Exception:
            margin_mm = 1.0
        return char_width_mm, line_height_mm, margin_mm

    def _reset_print_calibration(self):
        try:
            self.char_width_mm_var.set("1.50")
            self.line_height_mm_var.set("2.80")
            self.margin_mm_var.set("1.00")
            self._save_settings()
            self._set_status("打印校准已恢复默认")
        except Exception:
            pass

    def _resolve_active_printer(self) -> str:
        printer = ""
        try:
            if bool(self.use_default_printer_var.get()):
                printer = printer_utils.get_default_printer().strip()
            else:
                printer = self.printer_cb.get().strip()
        except Exception:
            printer = ""
        if not printer:
            try:
                printer = self.printer_cb.get().strip()
            except Exception:
                printer = ""
        return str(printer or "").strip()

    def _build_calibration_sheet_text(self, round_no: int) -> str:
        sample_30 = "123456789012345678901234567890"
        lines = [
            "Sen Nails 打印校准页",
            f"轮次: {int(round_no)} / 3",
            "测量说明:",
            "1) 用尺测量 A30 行竖线 |...| 之间的宽度(mm)",
            "2) 用尺测量 B10 第一行到第十行总高度(mm)",
            "3) 用尺测量纸左边到首字符的空白边距(mm)",
            "",
            f"A30|{sample_30}|",
            "B10|01|校准高度",
            "B10|02|校准高度",
            "B10|03|校准高度",
            "B10|04|校准高度",
            "B10|05|校准高度",
            "B10|06|校准高度",
            "B10|07|校准高度",
            "B10|08|校准高度",
            "B10|09|校准高度",
            "B10|10|校准高度",
            "",
            "END",
        ]
        return "\n".join(lines) + "\n"

    def _print_calibration_sheet(self, round_no: int, char_width_mm: float, line_height_mm: float, margin_mm: float):
        printer = self._resolve_active_printer()
        if not printer:
            return False, "未找到可用打印机", ""
        try:
            width_mm = int(float(self.width_var.get().strip()))
            height_mm = int(float(self.height_var.get().strip()))
        except Exception:
            width_mm, height_mm = 40, 30
        os.makedirs(DATA_DIR, exist_ok=True)
        path = os.path.join(DATA_DIR, f"calibration_round_{int(round_no)}_{int(time.time())}.txt")
        text = self._build_calibration_sheet_text(round_no)
        printer_utils.print_to_file(text, path)
        ok, detail = printer_utils.send_to_printer_debug(
            printer,
            path,
            width_mm=width_mm,
            height_mm=height_mm,
            canvas_mode=False,
            char_width_mm=float(char_width_mm),
            line_height_mm=float(line_height_mm),
            margin_mm=float(margin_mm),
        )
        return bool(ok), str(detail), path

    def open_print_calibration_wizard(self):
        if hasattr(self, "_calib_win") and self._calib_win.winfo_exists():
            self._calib_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("打印校准向导（3轮）")
        win.geometry("760x560")

        top = ttk.LabelFrame(win, text="当前参数")
        top.pack(fill=tk.X, padx=8, pady=8)
        round_var = tk.IntVar(value=1)
        char_var = tk.StringVar(value=str(self.char_width_mm_var.get()).strip() or "1.50")
        line_var = tk.StringVar(value=str(self.line_height_mm_var.get()).strip() or "2.80")
        margin_var = tk.StringVar(value=str(self.margin_mm_var.get()).strip() or "1.00")
        ttk.Label(top, text="轮次").grid(row=0, column=0, padx=4, pady=4, sticky=tk.W)
        ttk.Label(top, textvariable=round_var).grid(row=0, column=1, padx=4, pady=4, sticky=tk.W)
        ttk.Label(top, text="字宽mm").grid(row=0, column=2, padx=4, pady=4, sticky=tk.W)
        ttk.Entry(top, textvariable=char_var, width=8).grid(row=0, column=3, padx=4, pady=4, sticky=tk.W)
        ttk.Label(top, text="行高mm").grid(row=0, column=4, padx=4, pady=4, sticky=tk.W)
        ttk.Entry(top, textvariable=line_var, width=8).grid(row=0, column=5, padx=4, pady=4, sticky=tk.W)
        ttk.Label(top, text="边距mm").grid(row=0, column=6, padx=4, pady=4, sticky=tk.W)
        ttk.Entry(top, textvariable=margin_var, width=8).grid(row=0, column=7, padx=4, pady=4, sticky=tk.W)

        measure = ttk.LabelFrame(win, text="测量输入（打印后填写）")
        measure.pack(fill=tk.X, padx=8, pady=(0, 8))
        width30_var = tk.StringVar(value="")
        height10_var = tk.StringVar(value="")
        left_margin_var = tk.StringVar(value="")
        ttk.Label(measure, text="A30 宽度(mm)").grid(row=0, column=0, padx=4, pady=4, sticky=tk.W)
        ttk.Entry(measure, textvariable=width30_var, width=12).grid(row=0, column=1, padx=4, pady=4, sticky=tk.W)
        ttk.Label(measure, text="B10 高度(mm)").grid(row=0, column=2, padx=4, pady=4, sticky=tk.W)
        ttk.Entry(measure, textvariable=height10_var, width=12).grid(row=0, column=3, padx=4, pady=4, sticky=tk.W)
        ttk.Label(measure, text="左边距(mm)").grid(row=0, column=4, padx=4, pady=4, sticky=tk.W)
        ttk.Entry(measure, textvariable=left_margin_var, width=12).grid(row=0, column=5, padx=4, pady=4, sticky=tk.W)

        rec = ttk.LabelFrame(win, text="自动推荐")
        rec.pack(fill=tk.X, padx=8, pady=(0, 8))
        rec_char_var = tk.StringVar(value="-")
        rec_line_var = tk.StringVar(value="-")
        rec_margin_var = tk.StringVar(value="-")
        ttk.Label(rec, text="推荐字宽mm").grid(row=0, column=0, padx=4, pady=4, sticky=tk.W)
        ttk.Label(rec, textvariable=rec_char_var).grid(row=0, column=1, padx=4, pady=4, sticky=tk.W)
        ttk.Label(rec, text="推荐行高mm").grid(row=0, column=2, padx=4, pady=4, sticky=tk.W)
        ttk.Label(rec, textvariable=rec_line_var).grid(row=0, column=3, padx=4, pady=4, sticky=tk.W)
        ttk.Label(rec, text="推荐边距mm").grid(row=0, column=4, padx=4, pady=4, sticky=tk.W)
        ttk.Label(rec, textvariable=rec_margin_var).grid(row=0, column=5, padx=4, pady=4, sticky=tk.W)

        log_frm = ttk.LabelFrame(win, text="校准记录")
        log_frm.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        log_txt = scrolledtext.ScrolledText(log_frm, height=12)
        log_txt.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        log_txt.insert(tk.END, "流程: 打印校准页 -> 填测量值 -> 自动计算 -> 应用并下一轮\n")

        calc_cache = {"char": None, "line": None, "margin": None}

        def append_log(msg: str):
            log_txt.insert(tk.END, str(msg).rstrip() + "\n")
            log_txt.see(tk.END)

        def _parse_float(v: str, name: str) -> float:
            try:
                return float(str(v).strip())
            except Exception:
                raise ValueError(f"{name} 不是有效数字")

        def print_round():
            try:
                cw = max(0.8, _parse_float(char_var.get(), "字宽mm"))
                lh = max(1.0, _parse_float(line_var.get(), "行高mm"))
                mg = max(0.2, _parse_float(margin_var.get(), "边距mm"))
            except Exception as e:
                messagebox.showwarning("提示", str(e))
                return
            ok, detail, path = self._print_calibration_sheet(round_var.get(), cw, lh, mg)
            append_log(f"打印轮次{round_var.get()} -> {'成功' if ok else '失败'} | {detail} | {path}")
            if not ok:
                messagebox.showerror("打印失败", detail)

        def auto_calc():
            try:
                w30 = _parse_float(width30_var.get(), "A30宽度")
                h10 = _parse_float(height10_var.get(), "B10高度")
                lm = _parse_float(left_margin_var.get(), "左边距")
                if w30 <= 0 or h10 <= 0 or lm < 0:
                    raise ValueError("测量值必须大于0（左边距可为0）")
            except Exception as e:
                messagebox.showwarning("提示", str(e))
                return
            rec_char = max(0.8, min(3.5, w30 / 30.0))
            rec_line = max(1.0, min(8.0, h10 / 10.0))
            rec_margin = max(0.2, min(8.0, lm))
            calc_cache["char"] = rec_char
            calc_cache["line"] = rec_line
            calc_cache["margin"] = rec_margin
            rec_char_var.set(f"{rec_char:.3f}")
            rec_line_var.set(f"{rec_line:.3f}")
            rec_margin_var.set(f"{rec_margin:.3f}")
            append_log(
                f"轮次{round_var.get()}推荐 -> 字宽={rec_char:.3f} 行高={rec_line:.3f} 边距={rec_margin:.3f}"
            )

        def apply_recommend():
            if calc_cache["char"] is None:
                messagebox.showwarning("提示", "请先点击“自动计算推荐”")
                return
            char_var.set(f"{float(calc_cache['char']):.3f}")
            line_var.set(f"{float(calc_cache['line']):.3f}")
            margin_var.set(f"{float(calc_cache['margin']):.3f}")
            self.char_width_mm_var.set(char_var.get())
            self.line_height_mm_var.set(line_var.get())
            self.margin_mm_var.set(margin_var.get())
            self._save_settings()
            append_log("已应用推荐参数到主界面并保存")

        def next_round():
            cur = int(round_var.get())
            if cur >= 3:
                messagebox.showinfo("完成", "已完成3轮校准，可直接使用当前参数")
                return
            round_var.set(cur + 1)
            append_log(f"进入第{round_var.get()}轮")

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(btn, text="1) 打印本轮校准页", command=print_round).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="2) 自动计算推荐", command=auto_calc).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="3) 应用推荐参数", command=apply_recommend).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="下一轮", command=next_round).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="关闭", command=win.destroy).pack(side=tk.RIGHT, padx=4)

        self._calib_win = win

    def _get_template_mode(self) -> str:
        mode = str(self.template_mode_var.get()).strip().lower() if hasattr(self, "template_mode_var") else ""
        if mode in ("editor", "designer"):
            return mode
        return "designer" if bool(self.canvas_template_enabled.get()) else "editor"

    def _set_template_mode(self, mode: str, save: bool = True):
        target = "designer" if str(mode).strip().lower() == "designer" else "editor"
        if hasattr(self, "template_mode_var"):
            self.template_mode_var.set(target)
        if hasattr(self, "canvas_template_enabled"):
            self.canvas_template_enabled.set(1 if target == "designer" else 0)
        if save:
            self._save_settings()

    def _on_template_mode_change(self):
        self._set_template_mode(self._get_template_mode(), save=True)

    def open_template_editor(self):
        if hasattr(self, '_tpl_win') and self._tpl_win.winfo_exists():
            self._tpl_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title('打印模板编辑器')
        win.geometry('600x600')

        # Template text
        ttk.Label(win, text='模板文本 (可用: {permanent_id}, {unique_id}, {name}, {time}, {content}):').pack(anchor=tk.W, padx=6, pady=6)
        tpl_text = scrolledtext.ScrolledText(win, height=10)
        tpl_text.pack(fill=tk.BOTH, expand=True, padx=6, pady=2)
        tpl_text.insert(tk.END, self.template.get())

        # Formatting controls
        ctrl_frm = ttk.LabelFrame(win, text='格式化选项')
        ctrl_frm.pack(fill=tk.X, padx=6, pady=6)

        # Font size
        sz_frm = ttk.Frame(ctrl_frm)
        sz_frm.pack(fill=tk.X, padx=6, pady=4)
        ttk.Label(sz_frm, text='字体大小:').pack(side=tk.LEFT)
        self.tpl_font_size = tk.Scale(sz_frm, from_=1, to=3, orient=tk.HORIZONTAL, variable=self.font_size_var, length=200)
        self.tpl_font_size.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(sz_frm, text='(1=正常, 2=大, 3=超大)').pack(side=tk.LEFT)

        ff_frm = ttk.Frame(ctrl_frm)
        ff_frm.pack(fill=tk.X, padx=6, pady=4)
        ttk.Label(ff_frm, text='字体族:').pack(side=tk.LEFT)
        ttk.Combobox(ff_frm, textvariable=self.editor_font_family_var, values=FONT_FAMILY_CHOICES, width=20).pack(side=tk.LEFT, padx=4)
        ttk.Button(ff_frm, text='变量字号', command=lambda: open_var_scale_manager()).pack(side=tk.LEFT, padx=6)

        # Alignment
        al_frm = ttk.Frame(ctrl_frm)
        al_frm.pack(fill=tk.X, padx=6, pady=4)
        ttk.Label(al_frm, text='对齐:').pack(side=tk.LEFT)
        for align in ['left', 'center', 'right']:
            ttk.Radiobutton(al_frm, text=align, variable=self.alignment_var, value=align).pack(side=tk.LEFT, padx=4)

        # Line spacing
        sp_frm = ttk.Frame(ctrl_frm)
        sp_frm.pack(fill=tk.X, padx=6, pady=4)
        ttk.Label(sp_frm, text='行间距:').pack(side=tk.LEFT)
        self.tpl_line_sp = tk.Scale(sp_frm, from_=1, to=3, orient=tk.HORIZONTAL, variable=self.line_spacing_var, length=200)
        self.tpl_line_sp.pack(side=tk.LEFT, fill=tk.X, expand=True)

        ext_sp_frm = ttk.Frame(ctrl_frm)
        ext_sp_frm.pack(fill=tk.X, padx=6, pady=4)
        ttk.Label(ext_sp_frm, text='字间距:').pack(side=tk.LEFT)
        tk.Spinbox(ext_sp_frm, from_=-6, to=6, width=6, textvariable=self.editor_letter_spacing_var).pack(side=tk.LEFT, padx=4)
        ttk.Label(ext_sp_frm, text='段距:').pack(side=tk.LEFT, padx=(10, 0))
        tk.Spinbox(ext_sp_frm, from_=0, to=6, width=6, textvariable=self.editor_paragraph_spacing_var).pack(side=tk.LEFT, padx=4)

        # Paper width
        pw_frm = ttk.Frame(ctrl_frm)
        pw_frm.pack(fill=tk.X, padx=6, pady=4)
        ttk.Label(pw_frm, text='纸张宽度(字符数):').pack(side=tk.LEFT)
        self.tpl_paper_w = tk.Scale(pw_frm, from_=20, to=80, orient=tk.HORIZONTAL, variable=self.paper_width_var, length=200)
        self.tpl_paper_w.pack(side=tk.LEFT, fill=tk.X, expand=True)

        mg_frm = ttk.Frame(ctrl_frm)
        mg_frm.pack(fill=tk.X, padx=6, pady=4)
        ttk.Label(mg_frm, text='上边距').pack(side=tk.LEFT)
        tk.Spinbox(mg_frm, from_=0, to=12, width=5, textvariable=self.editor_margin_top_var).pack(side=tk.LEFT, padx=(4, 10))
        ttk.Label(mg_frm, text='下边距').pack(side=tk.LEFT)
        tk.Spinbox(mg_frm, from_=0, to=12, width=5, textvariable=self.editor_margin_bottom_var).pack(side=tk.LEFT, padx=(4, 10))
        ttk.Label(mg_frm, text='左边距').pack(side=tk.LEFT)
        tk.Spinbox(mg_frm, from_=0, to=20, width=5, textvariable=self.editor_margin_left_var).pack(side=tk.LEFT, padx=(4, 10))
        ttk.Label(mg_frm, text='右边距').pack(side=tk.LEFT)
        tk.Spinbox(mg_frm, from_=0, to=20, width=5, textvariable=self.editor_margin_right_var).pack(side=tk.LEFT, padx=4)

        # Preview
        ttk.Label(win, text='预览:').pack(anchor=tk.W, padx=6, pady=(6, 2))
        prev_text = scrolledtext.ScrolledText(win, height=8)
        prev_text.pack(fill=tk.BOTH, expand=True, padx=6, pady=2)

        def open_var_scale_manager():
            if hasattr(self, "_tpl_var_scale_win") and self._tpl_var_scale_win.winfo_exists():
                self._tpl_var_scale_win.lift()
                return
            vw = tk.Toplevel(win)
            vw.title("编辑器变量字号")
            vw.geometry("560x420")
            self._tpl_var_scale_win = vw
            frm = ttk.Frame(vw)
            frm.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
            top = ttk.Frame(frm)
            top.pack(fill=tk.X)
            key_var = tk.StringVar(value="")
            level_var = tk.IntVar(value=1)
            keys = [k for k, _zh, _en, _ja in VAR_DOCS] + sorted((self.template_custom_vars or {}).keys())
            ttk.Label(top, text="变量名").pack(side=tk.LEFT)
            ttk.Combobox(top, textvariable=key_var, values=sorted(set(keys)), width=20).pack(side=tk.LEFT, padx=4)
            ttk.Label(top, text="字号级别").pack(side=tk.LEFT, padx=(8, 2))
            ttk.Combobox(top, textvariable=level_var, values=[1, 2, 3], width=6, state="readonly").pack(side=tk.LEFT)
            ttk.Label(top, text="1=正常 2=放大 3=更大").pack(side=tk.LEFT, padx=6)

            tree = ttk.Treeview(frm, columns=("k", "lv"), show="headings")
            tree.heading("k", text="变量")
            tree.heading("lv", text="级别")
            tree.column("k", width=240, anchor=tk.W)
            tree.column("lv", width=90, anchor=tk.CENTER)
            tree.pack(fill=tk.BOTH, expand=True, pady=8)

            def refresh_tree():
                for r in tree.get_children():
                    tree.delete(r)
                mp = self.editor_var_font_scale_map if isinstance(self.editor_var_font_scale_map, dict) else {}
                for k, lv in sorted(mp.items(), key=lambda x: str(x[0])):
                    tree.insert("", tk.END, values=(str(k), int(lv)))

            def save_rule():
                k = key_var.get().strip().strip("{}")
                if not k:
                    messagebox.showwarning("提示", "变量名不能为空")
                    return
                lv = max(1, min(3, int(level_var.get())))
                if not isinstance(self.editor_var_font_scale_map, dict):
                    self.editor_var_font_scale_map = {}
                self.editor_var_font_scale_map[k] = int(lv)
                self._save_settings()
                refresh_tree()
                update_preview()

            def del_rule():
                sel = tree.selection()
                if not sel:
                    return
                vals = tree.item(sel[0], "values")
                if not vals:
                    return
                k = str(vals[0]).strip()
                if isinstance(self.editor_var_font_scale_map, dict):
                    self.editor_var_font_scale_map.pop(k, None)
                self._save_settings()
                refresh_tree()
                update_preview()

            def on_pick(_evt=None):
                sel = tree.selection()
                if not sel:
                    return
                vals = tree.item(sel[0], "values")
                if len(vals) >= 2:
                    key_var.set(str(vals[0]))
                    try:
                        level_var.set(int(vals[1]))
                    except Exception:
                        level_var.set(1)

            btn = ttk.Frame(frm)
            btn.pack(fill=tk.X)
            ttk.Button(btn, text="新增/更新", command=save_rule).pack(side=tk.LEFT, padx=4)
            ttk.Button(btn, text="删除", command=del_rule).pack(side=tk.LEFT, padx=4)
            ttk.Button(btn, text="关闭", command=vw.destroy).pack(side=tk.LEFT, padx=4)
            tree.bind("<<TreeviewSelect>>", on_pick)
            refresh_tree()

        def update_preview():
            tpl = tpl_text.get('1.0', tk.END).strip()
            vm = self._build_render_value_map(
                permanent_id='1',
                unique_id='test_uid',
                name='TestUser',
                timestamp='2026-02-18 12:00:00',
                content='测试内容',
                extra_vars={"guest_msg_count": "3", "today_guest_rank": "5"},
            )
            vm = self._apply_editor_var_font_scales(vm)
            sample = self._safe_format(tpl, vm)
            formatted = format_print_output(
                sample,
                self.font_size_var.get(),
                self.alignment_var.get(),
                self.line_spacing_var.get(),
                self.paper_width_var.get(),
                margin_top=self.editor_margin_top_var.get(),
                margin_bottom=self.editor_margin_bottom_var.get(),
                margin_left=self.editor_margin_left_var.get(),
                margin_right=self.editor_margin_right_var.get(),
                letter_spacing=self.editor_letter_spacing_var.get(),
                paragraph_spacing=self.editor_paragraph_spacing_var.get(),
            )
            prev_text.config(state=tk.NORMAL)
            prev_text.delete('1.0', tk.END)
            prev_text.insert(tk.END, formatted)
            try:
                # Match print behavior: global font size also scales print font itself.
                preview_pt = max(1, min(30, int(float(self.font_size_var.get()))))
                prev_text.configure(font=(str(self.editor_font_family_var.get()).strip() or "Consolas", preview_pt))
            except Exception:
                pass
            prev_text.config(state=tk.DISABLED)

        # Save and preview
        btn_frm = ttk.Frame(win)
        btn_frm.pack(fill=tk.X, padx=6, pady=6)
        ttk.Button(btn_frm, text='更新预览', command=update_preview).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frm, text='保存', command=lambda: (self.template.set(tpl_text.get('1.0', tk.END).strip()), self._set_template_mode("editor", save=False), self._save_settings(), win.destroy())).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frm, text='取消', command=win.destroy).pack(side=tk.LEFT, padx=2)

        self._tpl_win = win
        update_preview()  # show initial preview

    def _build_render_value_map(self, permanent_id, unique_id, name, timestamp, content, extra_vars=None):
        now = datetime.now()
        value_map = {
            "permanent_id": str(permanent_id),
            "unique_id": str(unique_id),
            "name": str(name or ""),
            "time": str(timestamp),
            "content": str(content or ""),
            "custom_name": self.custom_name_var.get().strip() if hasattr(self, "custom_name_var") else "",
            "room_url": self.room_url_var.get().strip() if hasattr(self, "room_url_var") else "",
            "app_version": str(self.app_version),
            "today_date": now.strftime("%Y-%m-%d"),
            "now_hms": now.strftime("%H:%M:%S"),
            "now_ts": str(int(time.time())),
            "weekday": ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][now.weekday()],
            "msg_length": str(len(str(content or ""))),
            "is_numeric": "1" if str(content or "").isdigit() else "0",
        }
        if isinstance(self.template_custom_vars, dict):
            for k, v in self.template_custom_vars.items():
                kk = str(k).strip()
                if kk:
                    value_map[kk] = str(v)
        if isinstance(extra_vars, dict):
            for k, v in extra_vars.items():
                kk = str(k).strip()
                if kk:
                    value_map[kk] = str(v)
        return value_map

    def _editor_scale_text_by_level(self, text: str, level: int) -> str:
        s = str(text or "")
        try:
            lv = int(level)
        except Exception:
            lv = 1
        lv = max(1, min(3, lv))
        if lv <= 1:
            return s
        sep = " " * (lv - 1)
        out_lines = []
        for line in s.splitlines() or [s]:
            out_lines.append(sep.join(list(str(line))))
        return "\n".join(out_lines)

    def _apply_editor_var_font_scales(self, value_map: dict) -> dict:
        out = dict(value_map or {})
        scale_map = self.editor_var_font_scale_map if isinstance(getattr(self, "editor_var_font_scale_map", {}), dict) else {}
        if not scale_map:
            return out
        for k, lv in scale_map.items():
            key = str(k).strip()
            if not key or key not in out:
                continue
            out[key] = self._editor_scale_text_by_level(out.get(key, ""), lv)
        return out

    def _designer_font_scale_hint(self, tpl: dict) -> float:
        return 1.0

    def _get_current_paper_size_mm(self) -> tuple[float, float]:
        width_src = self.width_var.get() if hasattr(self, "width_var") else self.settings.get("custom_paper_width_mm", "40")
        height_src = self.height_var.get() if hasattr(self, "height_var") else self.settings.get("custom_paper_height_mm", "30")
        width_mm = max(20.0, _safe_float(width_src, 40.0))
        height_mm = max(20.0, _safe_float(height_src, 30.0))
        return float(width_mm), float(height_mm)

    def _current_paper_canvas_units(self) -> tuple[int, int]:
        width_mm, height_mm = self._get_current_paper_size_mm()
        return (
            max(100, int(round(width_mm * CANVAS_UNITS_PER_MM))),
            max(100, int(round(height_mm * CANVAS_UNITS_PER_MM))),
        )

    def _get_print_layout_metrics(self) -> dict:
        width_mm, height_mm = self._get_current_paper_size_mm()
        char_width_mm, line_height_mm, margin_mm = self._print_calibration()
        printable_w_mm = max(8.0, float(width_mm) - (2.0 * float(margin_mm)))
        printable_h_mm = max(8.0, float(height_mm) - (2.0 * float(margin_mm)))
        chars = max(10, min(180, int(printable_w_mm / max(0.8, float(char_width_mm)))))
        rows = max(8, min(220, int(printable_h_mm / max(1.0, float(line_height_mm)))))
        return {
            "width_mm": float(width_mm),
            "height_mm": float(height_mm),
            "char_width_mm": float(char_width_mm),
            "line_height_mm": float(line_height_mm),
            "margin_mm": float(margin_mm),
            "printable_w_mm": float(printable_w_mm),
            "printable_h_mm": float(printable_h_mm),
            "chars": int(chars),
            "rows": int(rows),
        }

    def _normalize_canvas_template_to_paper(self, tpl=None) -> dict:
        raw = dict(tpl or {})
        target_w, target_h = self._current_paper_canvas_units()
        paper_w_mm, paper_h_mm = self._get_current_paper_size_mm()
        source_w = max(100, int(raw.get("canvas_w", target_w)))
        source_h = max(100, int(raw.get("canvas_h", target_h)))
        src_elements = [dict(x) for x in (raw.get("elements") or [])]
        need_scale = (source_w != target_w) or (source_h != target_h)
        scale_x = (float(target_w) / float(source_w)) if need_scale and source_w > 0 else 1.0
        scale_y = (float(target_h) / float(source_h)) if need_scale and source_h > 0 else 1.0
        elements = []
        for elem in src_elements:
            item = dict(elem)
            if need_scale:
                item["x"] = int(round(max(0, float(item.get("x", 0))) * scale_x))
                item["y"] = int(round(max(0, float(item.get("y", 0))) * scale_y))
                item["w"] = int(round(max(40, float(item.get("w", 160))) * scale_x))
                item["h"] = int(round(max(20, float(item.get("h", 34))) * scale_y))
            item["x"] = max(0, min(target_w - 40, int(item.get("x", 0))))
            item["y"] = max(0, min(target_h - 20, int(item.get("y", 0))))
            item["w"] = max(40, min(target_w, int(item.get("w", 160))))
            item["h"] = max(20, min(target_h, int(item.get("h", 34))))
            item["x"] = max(0, min(target_w - item["w"], int(item.get("x", 0))))
            item["y"] = max(0, min(target_h - item["h"], int(item.get("y", 0))))
            elements.append(item)
        return {
            "canvas_w": int(target_w),
            "canvas_h": int(target_h),
            "paper_mm_w": round(float(paper_w_mm), 2),
            "paper_mm_h": round(float(paper_h_mm), 2),
            "font_scale": float(raw.get("font_scale", 1.0)),
            "elements": elements,
        }

    def _safe_format(self, tpl_text: str, value_map: dict) -> str:
        try:
            return str(tpl_text).format_map(SafeDict(value_map))
        except Exception:
            return str(tpl_text)

    def _auto_wrap_value(self, text: str, width: int) -> str:
        t = str(text or "")
        w = max(4, int(width))
        if not t:
            return t
        lines = []
        for raw in t.splitlines() or [t]:
            # keep words when possible; fallback to hard break for long CJK strings
            part = textwrap.fill(raw, width=w, break_long_words=True, break_on_hyphens=False)
            lines.append(part)
        return "\n".join(lines)

    def open_custom_var_manager(self):
        if hasattr(self, "_var_win") and self._var_win.winfo_exists():
            self._var_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("变量管理")
        win.geometry("620x440")

        top = ttk.Frame(win)
        top.pack(fill=tk.X, padx=8, pady=8)
        ttk.Label(top, text="变量名:").grid(row=0, column=0, sticky=tk.W)
        key_var = tk.StringVar()
        ttk.Entry(top, textvariable=key_var, width=22).grid(row=0, column=1, sticky=tk.W, padx=4)
        ttk.Label(top, text="变量值:").grid(row=0, column=2, sticky=tk.W)
        val_var = tk.StringVar()
        ttk.Entry(top, textvariable=val_var, width=36).grid(row=0, column=3, sticky=tk.W, padx=4)

        cols = ("key", "value")
        tree = ttk.Treeview(win, columns=cols, show="headings", selectmode="browse")
        tree.heading("key", text="变量名")
        tree.heading("value", text="变量值")
        tree.column("key", width=180)
        tree.column("value", width=380)
        tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        builtins = (
            "内置变量: {permanent_id} {unique_id} {name} {time} {content} {custom_name} "
            "{room_url} {app_version} {today_date} {now_hms} {weekday} {msg_length} {is_numeric} "
            "{guest_msg_count} {today_guest_rank} {source_room}"
        )
        ttk.Label(win, text=builtins, foreground="#333").pack(anchor=tk.W, padx=8, pady=(0, 6))

        def refresh():
            for r in tree.get_children():
                tree.delete(r)
            for k, v in sorted((self.template_custom_vars or {}).items(), key=lambda x: str(x[0])):
                tree.insert("", tk.END, values=(k, v))

        def save_var():
            k = key_var.get().strip()
            if not k:
                messagebox.showwarning("提示", "变量名不能为空")
                return
            self.template_custom_vars[k] = val_var.get()
            self._save_settings()
            refresh()

        def del_var():
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            k = vals[0] if vals else ""
            if not k:
                return
            self.template_custom_vars.pop(str(k), None)
            self._save_settings()
            refresh()

        def on_pick(event=None):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            if len(vals) >= 2:
                key_var.set(str(vals[0]))
                val_var.set(str(vals[1]))

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(btn, text="新增/更新", command=save_var).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="删除", command=del_var).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="关闭", command=win.destroy).pack(side=tk.LEFT, padx=4)
        tree.bind("<<TreeviewSelect>>", on_pick)

        refresh()
        self._var_win = win

    def open_variable_docs(self):
        # kept for backward compatibility; variable docs are integrated into canvas designer.
        self.open_canvas_template_designer()

    def _canvas_element_font_family(self, elem: dict) -> str:
        family = str((elem or {}).get("font_family", "TkDefaultFont") or "TkDefaultFont").strip()
        return family if family else "TkDefaultFont"

    def _canvas_element_font_size(self, elem: dict) -> int:
        try:
            if "font_size" in (elem or {}):
                return max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, int(elem.get("font_size", 12))))
            return max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, int(round(float(elem.get("scale", 1.0)) * 12))))
        except Exception:
            return 12

    def _resolve_canvas_element_text(self, elem: dict, value_map: dict) -> str:
        field = str((elem or {}).get("field", "literal"))
        raw_label = str((elem or {}).get("label", ""))
        if field == "literal":
            text = raw_label
        elif field == "custom":
            custom_key = str((elem or {}).get("custom_key", "")).strip()
            if raw_label and "{" in raw_label and "}" in raw_label:
                text = self._safe_format(raw_label, value_map)
            else:
                text = value_map.get(custom_key, "")
        else:
            if raw_label and "{" in raw_label and "}" in raw_label:
                text = self._safe_format(raw_label, value_map)
            else:
                text = value_map.get(field, "")
        return str(text or "")

    def _build_canvas_render_payload(self, permanent_id, unique_id, name, timestamp, content, extra_vars=None, tpl_override=None) -> dict:
        base_tpl = tpl_override if isinstance(tpl_override, dict) else (self.canvas_template if isinstance(self.canvas_template, dict) else {})
        tpl = self._normalize_canvas_template_to_paper(base_tpl)
        canvas_w = max(100, int(tpl.get("canvas_w", 420)))
        canvas_h = max(100, int(tpl.get("canvas_h", 260)))
        metrics = self._get_print_layout_metrics()
        value_map = self._build_render_value_map(permanent_id, unique_id, name, timestamp, content, extra_vars=extra_vars)
        elements = []
        for elem in sorted((tpl.get("elements") or []), key=lambda x: (int(x.get("y", 0)), int(x.get("x", 0)))):
            text = self._resolve_canvas_element_text(elem, value_map)
            if not text:
                continue
            item = dict(elem)
            item["font_family"] = self._canvas_element_font_family(item)
            item["font_size"] = self._canvas_element_font_size(item)
            item["bold"] = 1 if int(item.get("bold", 0)) else 0
            item["letter_spacing"] = int(item.get("letter_spacing", 0))
            item["paragraph_spacing"] = int(item.get("paragraph_spacing", 0))
            item["render_text"] = text
            elements.append(item)
        return {
            "canvas_w": canvas_w,
            "canvas_h": canvas_h,
            "paper_mm_w": float(metrics["width_mm"]),
            "paper_mm_h": float(metrics["height_mm"]),
            "margin_mm": float(metrics["margin_mm"]),
            "elements": elements,
        }

    def _render_canvas_text(self, permanent_id, unique_id, name, timestamp, content, extra_vars=None, tpl_override=None) -> str:
        base_tpl = tpl_override if isinstance(tpl_override, dict) else (self.canvas_template if isinstance(self.canvas_template, dict) else {})
        tpl = self._normalize_canvas_template_to_paper(base_tpl)
        elements = tpl.get("elements") or []
        if not elements:
            return f"ID:{permanent_id}\n{name}\n{content}"
        canvas_w = max(100, int(tpl.get("canvas_w", 420)))
        canvas_h = max(100, int(tpl.get("canvas_h", 260)))
        layout = self._get_print_layout_metrics()
        chars = int(layout["chars"])
        rows = int(layout["rows"])
        lines = [list(" " * chars) for _ in range(rows)]
        value_map = self._build_render_value_map(permanent_id, unique_id, name, timestamp, content, extra_vars=extra_vars)

        def _write_line_cells(dst_row: list, start_col: int, text: str, max_cols: int):
            if not dst_row:
                return
            col = max(0, int(start_col))
            left = max(0, int(max_cols))
            if left <= 0 or col >= len(dst_row):
                return
            remaining = str(text or "")
            while remaining and left > 0 and col < len(dst_row):
                ch = remaining[0]
                remaining = remaining[1:]
                cw = _char_cells(ch)
                if cw > left:
                    break
                dst_row[col] = ch
                if cw == 2 and (col + 1) < len(dst_row):
                    dst_row[col + 1] = " "
                col += cw
                left -= cw

        def _scale_single_line_text(s: str, letter_spacing_val: int) -> str:
            try:
                gap = max(-6, min(8, int(letter_spacing_val)))
            except Exception:
                gap = 0
            if gap <= 0:
                return s
            sep = " " * gap
            return sep.join(list(str(s)))

        for e in sorted(elements, key=lambda x: (int(x.get("y", 0)), int(x.get("x", 0)))):
            text = self._resolve_canvas_element_text(e, value_map)
            if not text:
                continue
            x = max(0, int(e.get("x", 0)))
            y = max(0, int(e.get("y", 0)))
            elem_w = max(40, int(e.get("w", 160)))
            elem_h = max(20, int(e.get("h", 34)))
            base_col = min(chars - 1, int((x / canvas_w) * chars))
            row = min(rows - 1, int((y / canvas_h) * rows))
            elem_rows = max(1, int((elem_h / canvas_h) * rows))
            align = str(e.get("align", "left")).lower()
            valign = str(e.get("valign", "top")).lower()
            letter_spacing_val = int(e.get("letter_spacing", 0))
            paragraph_spacing_val = max(0, int(e.get("paragraph_spacing", 0)))
            parts = str(text).splitlines() or [str(text)]
            elem_cols = max(1, int((elem_w / canvas_w) * chars))
            effective_rows = max(1, min(elem_rows, rows - row))
            line_parts = []
            for part in parts:
                scaled = _scale_single_line_text(str(part), letter_spacing_val)
                wrapped = _wrap_by_cells(scaled, max(1, elem_cols))
                line_parts.extend(wrapped)
            required_rows = 1 + max(0, len(line_parts) - 1) * (1 + paragraph_spacing_val)
            block_rows = min(effective_rows, required_rows)
            if valign in ("middle", "center"):
                row_start = row + max(0, (effective_rows - block_rows) // 2)
            elif valign == "bottom":
                row_start = row + max(0, effective_rows - block_rows)
            else:
                row_start = row
            for li, part in enumerate(line_parts):
                p, _rest = _slice_by_cells(str(part), elem_cols)
                p_cells = _text_cells(p)
                if align == "center":
                    col = base_col + max(0, (elem_cols - p_cells) // 2)
                elif align == "right":
                    col = base_col + max(0, elem_cols - p_cells)
                else:
                    col = base_col
                col = min(chars - 1, max(0, col))
                rr = row_start + (li * (1 + paragraph_spacing_val))
                if rr >= rows or rr >= (row + effective_rows):
                    break
                _write_line_cells(lines[rr], col, p, min(elem_cols, max(0, chars - col)))
        out_lines = ["".join(r).rstrip() for r in lines]
        while out_lines and out_lines[-1] == "":
            out_lines.pop()
        if not out_lines:
            return f"ID:{permanent_id}\n{name}\n{content}"
        return "\n".join(out_lines)

    def _compose_print_rendered(self, permanent_id, unique_id, name, timestamp, content, extra_vars=None) -> str:
        wrapped_name = self._auto_wrap_value(name, self.auto_wrap_name_width) if self.auto_wrap_print_enabled else str(name or "")
        wrapped_content = self._auto_wrap_value(content, self.auto_wrap_content_width) if self.auto_wrap_print_enabled else str(content or "")
        merged_vars = dict(extra_vars or {})
        merged_vars["name_wrapped"] = wrapped_name
        merged_vars["content_wrapped"] = wrapped_content
        if self._get_template_mode() == "designer":
            tpl = self._normalize_canvas_template_to_paper(self.canvas_template if isinstance(self.canvas_template, dict) else {})
            canvas_payload = self._build_canvas_render_payload(
                permanent_id,
                unique_id,
                wrapped_name,
                timestamp,
                wrapped_content,
                extra_vars=merged_vars,
                tpl_override=tpl,
            )
            body = self._render_canvas_text(permanent_id, unique_id, wrapped_name, timestamp, wrapped_content, extra_vars=merged_vars, tpl_override=tpl)
            scale_hint = self._designer_font_scale_hint(tpl)
            payload_json = json.dumps(canvas_payload, ensure_ascii=False, separators=(",", ":"))
            payload_b64 = base64.b64encode(payload_json.encode("utf-8")).decode("ascii")
            return CANVAS_PRINT_MARKER + "\n" + f"{CANVAS_PAYLOAD_MARKER_PREFIX}{payload_b64}]]\n" + f"{FONT_SCALE_MARKER_PREFIX}{scale_hint:.2f}]]\n" + body
        else:
            tpl = self.template.get()
            value_map = self._build_render_value_map(permanent_id, unique_id, wrapped_name, timestamp, wrapped_content, extra_vars=merged_vars)
            value_map = self._apply_editor_var_font_scales(value_map)
            body = self._safe_format(tpl, value_map)
            rendered = format_print_output(
                body,
                self.font_size_var.get(),
                self.alignment_var.get(),
                self.line_spacing_var.get(),
                self.paper_width_var.get(),
                margin_top=self.editor_margin_top_var.get(),
                margin_bottom=self.editor_margin_bottom_var.get(),
                margin_left=self.editor_margin_left_var.get(),
                margin_right=self.editor_margin_right_var.get(),
                letter_spacing=self.editor_letter_spacing_var.get(),
                paragraph_spacing=self.editor_paragraph_spacing_var.get(),
            )
            return f"{FONT_SCALE_MARKER_PREFIX}{max(1, min(3, int(self.font_size_var.get()))):.2f}]]\n" + rendered

    def open_canvas_template_designer(self):
        if hasattr(self, "_canvas_tpl_win") and self._canvas_tpl_win.winfo_exists():
            self._canvas_tpl_win.lift()
            return

        win = tk.Toplevel(self.root)
        win.title("画布模板设计器（含变量管理）")
        win.geometry("1260x760")

        root_pane = ttk.Panedwindow(win, orient=tk.HORIZONTAL)
        root_pane.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        left = ttk.Frame(root_pane, width=360)
        right = ttk.Frame(root_pane)
        root_pane.add(left, weight=1)
        root_pane.add(right, weight=3)

        template_data = self._normalize_canvas_template_to_paper(self.canvas_template if isinstance(self.canvas_template, dict) else {})
        state = {
            "canvas_w": int(template_data.get("canvas_w", 420)),
            "canvas_h": int(template_data.get("canvas_h", 260)),
            "font_scale": float(template_data.get("font_scale", 1.0)),
            "elements": [dict(x) for x in (template_data.get("elements") or [])],
            "selected_idx": None,
            "selected_ids": set(),
            "drag_start": (0, 0),
            "drag_mode": "move",
            "drag_handle": "",
            "marquee_start": (0, 0),
            "marquee_end": (0, 0),
            "clipboard": [],
            "history": [],
            "future": [],
            "history_lock": False,
            "form_sync_lock": False,
            "inline_edit_idx": None,
            "inline_edit_panel": None,
            "inline_edit_text_var": None,
            "inline_edit_scale_var": None,
            "inline_edit_bold_var": None,
            "inline_edit_original": ("", 12, 0),
            "inline_edit_history_pushed": False,
            "dragging": False,
            "preview_after_id": None,
            "guides": [],
        }
        state["canvas_w"] = max(100, int(state["canvas_w"]))
        state["canvas_h"] = max(100, int(state["canvas_h"]))

        def _to_mm(units: int) -> float:
            return round(max(1.0, float(units) / CANVAS_UNITS_PER_MM), 2)

        def _to_units(mm_value) -> int:
            try:
                mm = float(mm_value)
            except Exception:
                mm = 20.0
            return max(100, int(round(max(1.0, mm) * CANVAS_UNITS_PER_MM)))

        def _paper_default_canvas_units() -> tuple[int, int]:
            width_mm, height_mm = self._get_current_paper_size_mm()
            return _to_units(width_mm), _to_units(height_mm)

        left_scroll_wrap = ttk.Frame(left)
        left_scroll_wrap.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        left_canvas = tk.Canvas(left_scroll_wrap, highlightthickness=0)
        left_scroll = ttk.Scrollbar(left_scroll_wrap, orient=tk.VERTICAL, command=left_canvas.yview)
        left_scroll_x = ttk.Scrollbar(left_scroll_wrap, orient=tk.HORIZONTAL, command=left_canvas.xview)
        left_canvas.configure(yscrollcommand=left_scroll.set, xscrollcommand=left_scroll_x.set)
        left_canvas.grid(row=0, column=0, sticky="nsew")
        left_scroll.grid(row=0, column=1, sticky="ns")
        left_scroll_x.grid(row=1, column=0, sticky="ew")
        left_scroll_wrap.rowconfigure(0, weight=1)
        left_scroll_wrap.columnconfigure(0, weight=1)

        left_pane = ttk.Panedwindow(left_canvas, orient=tk.VERTICAL)
        left_window_id = left_canvas.create_window((0, 0), window=left_pane, anchor="nw")

        def _refresh_left_scrollregion(event=None):
            try:
                left_canvas.configure(scrollregion=left_canvas.bbox("all"))
            except Exception:
                pass

        left_pane.bind("<Configure>", _refresh_left_scrollregion)
        left_canvas.bind("<Configure>", _refresh_left_scrollregion)

        def _on_left_mousewheel(event):
            delta = 0
            if hasattr(event, "delta") and event.delta:
                delta = -1 if event.delta > 0 else 1
            elif getattr(event, "num", None) == 4:
                delta = -1
            elif getattr(event, "num", None) == 5:
                delta = 1
            if delta != 0:
                left_canvas.yview_scroll(delta, "units")
                return "break"

        left_canvas.bind("<MouseWheel>", _on_left_mousewheel)
        left_canvas.bind("<Button-4>", _on_left_mousewheel)
        left_canvas.bind("<Button-5>", _on_left_mousewheel)

        cfg = ttk.LabelFrame(left_pane, text="画布")
        ttk.Label(cfg, text="宽(mm)").grid(row=0, column=0, padx=4, pady=4, sticky=tk.W)
        canvas_w_var = tk.DoubleVar(value=_to_mm(state["canvas_w"]))
        ttk.Entry(cfg, textvariable=canvas_w_var, width=8).grid(row=0, column=1, padx=4, pady=4, sticky=tk.W)
        ttk.Label(cfg, text="高(mm)").grid(row=0, column=2, padx=4, pady=4, sticky=tk.W)
        canvas_h_var = tk.DoubleVar(value=_to_mm(state["canvas_h"]))
        ttk.Entry(cfg, textvariable=canvas_h_var, width=8).grid(row=0, column=3, padx=4, pady=4, sticky=tk.W)
        ttk.Button(cfg, text="按打印尺寸", command=lambda: apply_canvas_from_paper()).grid(row=0, column=4, padx=4, pady=4, sticky=tk.W)
        ttk.Label(cfg, text="尺寸预设").grid(row=0, column=5, padx=4, pady=4, sticky=tk.W)
        canvas_preset_var = tk.StringVar(value="")
        canvas_preset_cb = ttk.Combobox(cfg, textvariable=canvas_preset_var, values=list(getattr(self, "paper_sizes", {}).keys()), width=12, state="readonly")
        canvas_preset_cb.grid(row=0, column=6, padx=4, pady=4, sticky=tk.W)
        ttk.Label(cfg, text="缩放").grid(row=1, column=0, padx=4, pady=4, sticky=tk.W)
        zoom_var = tk.IntVar(value=100)
        ttk.Scale(cfg, from_=50, to=300, variable=zoom_var, orient=tk.HORIZONTAL, length=220).grid(row=1, column=1, columnspan=3, padx=4, pady=4, sticky=tk.W)
        ttk.Label(cfg, text="字号模式").grid(row=2, column=0, padx=4, pady=4, sticky=tk.W)
        global_scale_var = tk.DoubleVar(value=1.0)
        ttk.Label(cfg, text="按元素单独调整").grid(row=2, column=1, columnspan=3, padx=4, pady=4, sticky=tk.W)
        snap_var = tk.IntVar(value=0)
        ttk.Checkbutton(cfg, text="自动对齐吸附", variable=snap_var).grid(row=3, column=0, columnspan=2, padx=4, pady=2, sticky=tk.W)
        perf_mode_var = tk.IntVar(value=1)
        ttk.Checkbutton(cfg, text="性能模式(编辑更流畅)", variable=perf_mode_var).grid(row=3, column=2, columnspan=2, padx=4, pady=2, sticky=tk.W)

        var_box = ttk.LabelFrame(left_pane, text="变量管理")
        var_key = tk.StringVar()
        var_val = tk.StringVar()
        ttk.Label(var_box, text="变量名").grid(row=0, column=0, padx=4, pady=4, sticky=tk.W)
        ttk.Entry(var_box, textvariable=var_key, width=11).grid(row=0, column=1, padx=4, pady=4, sticky=tk.W)
        ttk.Label(var_box, text="变量值").grid(row=0, column=2, padx=4, pady=4, sticky=tk.W)
        ttk.Entry(var_box, textvariable=var_val, width=12).grid(row=0, column=3, padx=4, pady=4, sticky=tk.W)
        var_tree = ttk.Treeview(var_box, columns=("k", "v"), show="headings", height=6)
        var_tree.heading("k", text="键")
        var_tree.heading("v", text="值")
        var_tree.column("k", width=95)
        var_tree.column("v", width=130)
        var_tree.grid(row=1, column=0, columnspan=4, padx=4, pady=4, sticky="nsew")
        var_box.rowconfigure(1, weight=1)
        ttk.Label(var_box, text="内置变量说明(双击可自动新增到画布):", foreground="#333").grid(row=2, column=0, columnspan=4, sticky=tk.W, padx=4, pady=(2, 0))
        var_docs_tree = ttk.Treeview(var_box, columns=("ph", "desc"), show="headings", height=6)
        var_docs_tree.heading("ph", text="占位符")
        var_docs_tree.heading("desc", text="说明")
        var_docs_tree.column("ph", width=120, anchor=tk.W)
        var_docs_tree.column("desc", width=180, anchor=tk.W)
        var_docs_tree.grid(row=3, column=0, columnspan=4, padx=4, pady=4, sticky="nsew")
        var_box.rowconfigure(3, weight=1)
        for k, zh, _en, _ja in VAR_DOCS:
            var_docs_tree.insert("", tk.END, iid=str(k), values=("{"+str(k)+"}", zh))

        elem = ttk.LabelFrame(left_pane, text="元素")
        field_var = tk.StringVar(value="content")
        custom_key_var = tk.StringVar(value="")
        label_var = tk.StringVar(value="")
        align_var = tk.StringVar(value="left")
        valign_var = tk.StringVar(value="top")
        font_family_var = tk.StringVar(value="TkDefaultFont")
        font_size_var = tk.IntVar(value=12)
        letter_spacing_var = tk.IntVar(value=0)
        paragraph_spacing_var = tk.IntVar(value=0)
        bold_var = tk.IntVar(value=0)
        pos_x_var = tk.IntVar(value=0)
        pos_y_var = tk.IntVar(value=0)
        elem_w_var = tk.IntVar(value=160)
        elem_h_var = tk.IntVar(value=34)
        ttk.Label(elem, text="字段").grid(row=0, column=0, padx=4, pady=3, sticky=tk.W)
        ttk.Combobox(elem, textvariable=field_var, values=["permanent_id", "unique_id", "name", "time", "content", "custom", "literal"], state="readonly", width=12).grid(row=0, column=1, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="自定义键").grid(row=1, column=0, padx=4, pady=3, sticky=tk.W)
        custom_key_cb = ttk.Combobox(elem, textvariable=custom_key_var, values=sorted((self.template_custom_vars or {}).keys()), width=12)
        custom_key_cb.grid(row=1, column=1, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="文本").grid(row=2, column=0, padx=4, pady=3, sticky=tk.W)
        ttk.Entry(elem, textvariable=label_var, width=20).grid(row=2, column=1, columnspan=3, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="水平对齐").grid(row=3, column=0, padx=4, pady=3, sticky=tk.W)
        ttk.Combobox(elem, textvariable=align_var, values=["left", "center", "right"], state="readonly", width=10).grid(row=3, column=1, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="垂直对齐").grid(row=3, column=2, padx=4, pady=3, sticky=tk.W)
        ttk.Combobox(elem, textvariable=valign_var, values=["top", "middle", "bottom"], state="readonly", width=10).grid(row=3, column=3, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="字体").grid(row=4, column=0, padx=4, pady=3, sticky=tk.W)
        ttk.Combobox(elem, textvariable=font_family_var, values=FONT_FAMILY_CHOICES, width=12).grid(row=4, column=1, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="字号(pt)").grid(row=4, column=2, padx=4, pady=3, sticky=tk.W)
        font_ctl = ttk.Frame(elem)
        font_ctl.grid(row=4, column=3, padx=4, pady=3, sticky=tk.W)
        font_minus_btn = ttk.Button(font_ctl, text="A-", width=3)
        font_minus_btn.pack(side=tk.LEFT, padx=(0, 2))
        font_size_cb = ttk.Combobox(
            font_ctl,
            textvariable=font_size_var,
            values=DESIGNER_FONT_SIZE_CHOICES,
            width=5,
        )
        font_size_cb.pack(side=tk.LEFT)
        font_plus_btn = ttk.Button(font_ctl, text="A+", width=3)
        font_plus_btn.pack(side=tk.LEFT, padx=(2, 0))
        ttk.Checkbutton(font_ctl, text="加粗", variable=bold_var).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(elem, text="字间距").grid(row=5, column=0, padx=4, pady=3, sticky=tk.W)
        tk.Spinbox(elem, from_=-6, to=8, textvariable=letter_spacing_var, width=6).grid(row=5, column=1, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="段落间距").grid(row=5, column=2, padx=4, pady=3, sticky=tk.W)
        tk.Spinbox(elem, from_=0, to=6, textvariable=paragraph_spacing_var, width=6).grid(row=5, column=3, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="X 位置").grid(row=6, column=0, padx=4, pady=3, sticky=tk.W)
        pos_x_scale = ttk.Scale(elem, from_=0, to=state["canvas_w"], variable=pos_x_var, orient=tk.HORIZONTAL, length=180)
        pos_x_scale.grid(row=6, column=1, columnspan=3, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="Y 位置").grid(row=7, column=0, padx=4, pady=3, sticky=tk.W)
        pos_y_scale = ttk.Scale(elem, from_=0, to=state["canvas_h"], variable=pos_y_var, orient=tk.HORIZONTAL, length=180)
        pos_y_scale.grid(row=7, column=1, columnspan=3, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="宽度").grid(row=8, column=0, padx=4, pady=3, sticky=tk.W)
        elem_w_scale = ttk.Scale(elem, from_=40, to=state["canvas_w"], variable=elem_w_var, orient=tk.HORIZONTAL, length=180)
        elem_w_scale.grid(row=8, column=1, columnspan=3, padx=4, pady=3, sticky=tk.W)
        ttk.Label(elem, text="高度").grid(row=9, column=0, padx=4, pady=3, sticky=tk.W)
        elem_h_scale = ttk.Scale(elem, from_=20, to=240, variable=elem_h_var, orient=tk.HORIZONTAL, length=180)
        elem_h_scale.grid(row=9, column=1, columnspan=3, padx=4, pady=3, sticky=tk.W)
        elem_list = tk.Listbox(elem, height=10, selectmode=tk.EXTENDED, exportselection=False)
        elem_list.grid(row=10, column=0, columnspan=4, padx=4, pady=4, sticky="nsew")
        elem.rowconfigure(10, weight=1)

        left_pane.add(cfg, weight=1)
        left_pane.add(var_box, weight=2)
        left_pane.add(elem, weight=4)

        right_pane = ttk.Panedwindow(right, orient=tk.VERTICAL)
        right_pane.pack(fill=tk.BOTH, expand=True)
        editor_frame = ttk.Frame(right_pane)
        preview_frame = ttk.LabelFrame(right_pane, text="实时打印预览")
        right_pane.add(editor_frame, weight=3)
        right_pane.add(preview_frame, weight=2)

        cv_wrap = ttk.Frame(editor_frame)
        cv_wrap.pack(fill=tk.BOTH, expand=True)
        vbar = ttk.Scrollbar(cv_wrap, orient=tk.VERTICAL)
        hbar = ttk.Scrollbar(cv_wrap, orient=tk.HORIZONTAL)
        canvas = tk.Canvas(cv_wrap, bg="#f7f7f7", highlightthickness=1, highlightbackground="#888")
        canvas.configure(yscrollcommand=vbar.set, xscrollcommand=hbar.set)
        vbar.configure(command=canvas.yview)
        hbar.configure(command=canvas.xview)
        cv_wrap.rowconfigure(0, weight=1)
        cv_wrap.columnconfigure(0, weight=1)
        canvas.grid(row=0, column=0, sticky="nsew")
        vbar.grid(row=0, column=1, sticky="ns")
        hbar.grid(row=1, column=0, sticky="ew")

        pv_wrap = ttk.Frame(preview_frame)
        pv_wrap.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        pv_vbar = ttk.Scrollbar(pv_wrap, orient=tk.VERTICAL)
        pv_hbar = ttk.Scrollbar(pv_wrap, orient=tk.HORIZONTAL)
        preview_canvas = tk.Canvas(pv_wrap, bg="#ffffff", highlightthickness=1, highlightbackground="#b5b5b5")
        preview_canvas.configure(yscrollcommand=pv_vbar.set, xscrollcommand=pv_hbar.set)
        pv_vbar.configure(command=preview_canvas.yview)
        pv_hbar.configure(command=preview_canvas.xview)
        pv_wrap.rowconfigure(0, weight=1)
        pv_wrap.columnconfigure(0, weight=1)
        preview_canvas.grid(row=0, column=0, sticky="nsew")
        pv_vbar.grid(row=0, column=1, sticky="ns")
        pv_hbar.grid(row=1, column=0, sticky="ew")
        font_obj_cache = {}

        def _get_cached_font(family: str, size: int, weight: str):
            key = (family, int(size), str(weight))
            fo = font_obj_cache.get(key)
            if fo is None:
                try:
                    fo = tkfont.Font(family=family, size=int(size), weight=str(weight))
                except Exception:
                    fo = tkfont.Font(family="TkDefaultFont", size=int(size), weight=str(weight))
                font_obj_cache[key] = fo
            return fo

        def refresh_var_tree():
            for r in var_tree.get_children():
                var_tree.delete(r)
            for k, v in sorted((self.template_custom_vars or {}).items(), key=lambda x: str(x[0])):
                var_tree.insert("", tk.END, values=(k, v))
            custom_key_cb.configure(values=sorted((self.template_custom_vars or {}).keys()))

        def _snapshot_state():
            return {
                "canvas_w": int(state["canvas_w"]),
                "canvas_h": int(state["canvas_h"]),
                "font_scale": float(global_scale_var.get()),
                "elements": [dict(x) for x in state["elements"]],
            }

        def _restore_snapshot(snap):
            state["history_lock"] = True
            try:
                state["canvas_w"] = int(snap.get("canvas_w", state["canvas_w"]))
                state["canvas_h"] = int(snap.get("canvas_h", state["canvas_h"]))
                state["elements"] = [dict(x) for x in (snap.get("elements") or [])]
                canvas_w_var.set(_to_mm(state["canvas_w"]))
                canvas_h_var.set(_to_mm(state["canvas_h"]))
                global_scale_var.set(float(snap.get("font_scale", global_scale_var.get())))
                pos_x_scale.configure(to=state["canvas_w"])
                pos_y_scale.configure(to=state["canvas_h"])
                elem_w_scale.configure(to=state["canvas_w"])
                _apply_selection([], primary=None, update_form=False)
                refresh_elem_list()
            finally:
                state["history_lock"] = False
            draw()

        def _push_history(clear_future=True):
            if state.get("history_lock"):
                return
            snap = _snapshot_state()
            hist = state["history"]
            if hist and hist[-1] == snap:
                return
            hist.append(snap)
            if len(hist) > 120:
                del hist[0]
            if clear_future:
                state["future"] = []

        def undo_action(event=None):
            if event is not None and _is_input_widget(getattr(event, "widget", None)):
                return
            if len(state["history"]) <= 1:
                return "break"
            current = state["history"].pop()
            state["future"].append(current)
            prev = state["history"][-1]
            _restore_snapshot(prev)
            return "break"

        def redo_action(event=None):
            if event is not None and _is_input_widget(getattr(event, "widget", None)):
                return
            if not state["future"]:
                return "break"
            snap = state["future"].pop()
            _restore_snapshot(snap)
            _push_history(clear_future=False)
            return "break"

        def _selected_indices() -> list[int]:
            sel = sorted(int(i) for i in (state.get("selected_ids") or set()) if 0 <= int(i) < len(state["elements"]))
            if sel:
                return sel
            idx = state.get("selected_idx")
            if idx is None or idx < 0 or idx >= len(state["elements"]):
                return []
            return [int(idx)]

        def _elem_font_size(e) -> int:
            try:
                if "font_size" in e:
                    return max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, int(e.get("font_size", 12))))
                return max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, int(round(float(e.get("scale", 1.0)) * 12))))
            except Exception:
                return 12

        def _elem_font_family(e) -> str:
            fam = str(e.get("font_family", "TkDefaultFont") or "TkDefaultFont").strip()
            return fam if fam else "TkDefaultFont"

        def _set_form_from_element(e):
            state["form_sync_lock"] = True
            try:
                field_var.set(str(e.get("field", "content")))
                custom_key_var.set(str(e.get("custom_key", "")))
                label_var.set(str(e.get("label", "")))
                align_var.set(str(e.get("align", "left")))
                valign_var.set(str(e.get("valign", "top")))
                font_family_var.set(_elem_font_family(e))
                try:
                    font_size_var.set(_elem_font_size(e))
                except Exception:
                    font_size_var.set(12)
                letter_spacing_var.set(int(e.get("letter_spacing", 0)))
                paragraph_spacing_var.set(int(e.get("paragraph_spacing", 0)))
                bold_var.set(1 if int(e.get("bold", 0)) else 0)
                pos_x_var.set(int(e.get("x", 0)))
                pos_y_var.set(int(e.get("y", 0)))
                elem_w_var.set(int(e.get("w", 160)))
                elem_h_var.set(int(e.get("h", 34)))
            finally:
                state["form_sync_lock"] = False

        def _apply_selection(indices, primary=None, update_form=True):
            valid = sorted(set(int(i) for i in indices if 0 <= int(i) < len(state["elements"])))
            state["selected_ids"] = set(valid)
            if primary is None:
                primary = valid[0] if valid else None
            if primary is not None and primary not in state["selected_ids"]:
                state["selected_ids"].add(int(primary))
            state["selected_idx"] = int(primary) if primary is not None else None
            elem_list.selection_clear(0, tk.END)
            for i in sorted(state["selected_ids"]):
                elem_list.selection_set(i)
            if update_form and state["selected_idx"] is not None and 0 <= state["selected_idx"] < len(state["elements"]):
                e = state["elements"][state["selected_idx"]]
                _set_form_from_element(e)

        def refresh_elem_list():
            keep = set(int(i) for i in (state.get("selected_ids") or set()))
            elem_list.delete(0, tk.END)
            for i, e in enumerate(state["elements"]):
                elem_list.insert(
                    tk.END,
                    f"{i+1}. {e.get('field','')}:{e.get('custom_key','')} @({int(e.get('x',0))},{int(e.get('y',0))}) "
                    f"{e.get('align','left')}/{e.get('valign','top')} 字体{_elem_font_family(e)} 字号{_elem_font_size(e)} 间距{int(e.get('letter_spacing',0))} 段距{int(e.get('paragraph_spacing',0))} 加粗{int(e.get('bold',0))} 尺寸{int(e.get('w',160))}x{int(e.get('h',34))}",
                )
            if keep:
                _apply_selection(sorted(i for i in keep if i < len(state["elements"])), primary=state.get("selected_idx"), update_form=False)

        def _preview_text_for_element(e, value_map) -> str:
            field = str(e.get("field", "literal"))
            raw_label = str(e.get("label", ""))
            if field == "literal":
                text = raw_label
            elif field == "custom":
                custom_key = str(e.get("custom_key", "")).strip()
                if raw_label and "{" in raw_label and "}" in raw_label:
                    text = self._safe_format(raw_label, value_map)
                else:
                    text = value_map.get(custom_key, "")
            else:
                if raw_label and "{" in raw_label and "}" in raw_label:
                    text = self._safe_format(raw_label, value_map)
                else:
                    text = value_map.get(field, "")
            return str(text or "")

        def _preview_apply_letter_spacing(text: str, spacing: int) -> str:
            try:
                gap = max(-6, min(8, int(spacing)))
            except Exception:
                gap = 0
            if gap <= 0:
                return str(text or "")
            return (" " * gap).join(list(str(text or "")))

        def _wrap_preview_line(text: str, font_obj, max_width_px: int) -> list[str]:
            raw = str(text or "")
            if max_width_px <= 4:
                return [raw] if raw else [""]
            if not raw:
                return [""]
            out = []
            remaining = raw
            while remaining:
                width = 0
                cut = 0
                last_space = -1
                for idx, ch in enumerate(remaining):
                    ch_width = max(1, int(font_obj.measure(ch)))
                    if cut > 0 and (width + ch_width) > max_width_px:
                        break
                    width += ch_width
                    cut = idx + 1
                    if ch.isspace():
                        last_space = cut
                if cut <= 0:
                    cut = 1
                if 0 < last_space < cut:
                    part = remaining[:last_space].rstrip()
                    remaining = remaining[last_space:].lstrip()
                else:
                    part = remaining[:cut]
                    remaining = remaining[cut:]
                out.append(part or " ")
            return out or [""]

        def _wrap_preview_text(text: str, font_obj, max_width_px: int, letter_spacing_val: int, paragraph_spacing_val: int) -> list[str]:
            parts = str(text or "").splitlines() or [""]
            line_parts = []
            for part in parts:
                spaced = _preview_apply_letter_spacing(part, letter_spacing_val)
                line_parts.extend(_wrap_preview_line(spaced, font_obj, max_width_px))
            if not line_parts:
                return [""]
            out = []
            gap_rows = max(0, int(paragraph_spacing_val))
            for idx, line in enumerate(line_parts):
                out.append(line)
                if idx != (len(line_parts) - 1) and gap_rows:
                    out.extend([""] * gap_rows)
            return out

        def sample_render_preview():
            preview_canvas.delete("all")
            tpl_preview = self._normalize_canvas_template_to_paper({
                "canvas_w": int(state["canvas_w"]),
                "canvas_h": int(state["canvas_h"]),
                "font_scale": float(global_scale_var.get()),
                "elements": [dict(x) for x in state["elements"]],
            })
            metrics = self._get_print_layout_metrics()
            width_mm = float(metrics["width_mm"])
            height_mm = float(metrics["height_mm"])
            margin_mm = float(metrics["margin_mm"])
            try:
                avail_w = max(320, int(preview_canvas.winfo_width()) - 24)
            except Exception:
                avail_w = 420
            try:
                avail_h = max(240, int(preview_canvas.winfo_height()) - 24)
            except Exception:
                avail_h = 320
            px_per_mm = max(4.0, min(12.0, min(float(avail_w) / max(1.0, width_mm), float(avail_h) / max(1.0, height_mm))))
            paper_w_px = max(160, int(round(width_mm * px_per_mm)))
            paper_h_px = max(120, int(round(height_mm * px_per_mm)))
            offset_x = 10
            offset_y = 10
            margin_px = max(4, int(round(margin_mm * px_per_mm)))
            preview_canvas.configure(scrollregion=(0, 0, paper_w_px + 20, paper_h_px + 20))
            preview_canvas.create_rectangle(offset_x, offset_y, offset_x + paper_w_px, offset_y + paper_h_px, outline="#7d7d7d", fill="#ffffff")
            preview_canvas.create_rectangle(
                offset_x + margin_px,
                offset_y + margin_px,
                offset_x + paper_w_px - margin_px,
                offset_y + paper_h_px - margin_px,
                outline="#d6d6d6",
                dash=(3, 2),
            )
            preview_canvas.create_text(
                offset_x + 8,
                offset_y - 2,
                text=f"{int(round(width_mm))}x{int(round(height_mm))}mm",
                anchor=tk.SW,
                fill="#555",
                font=_get_cached_font("TkDefaultFont", 9, "normal"),
            )
            elements = tpl_preview.get("elements") or []
            if not elements:
                rendered = self._render_canvas_text(
                    permanent_id="1001",
                    unique_id="u_demo_001",
                    name="DemoUser",
                    timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    content="这是实时预览内容 12345",
                    extra_vars={"guest_msg_count": 3, "today_guest_rank": 12, "source_room": "demo_room"},
                    tpl_override=tpl_preview,
                )
                line_px = max(10, int(round(max(20, paper_h_px - (margin_px * 2)) / max(1, int(metrics["rows"])))))
                mono_font = _get_cached_font("Consolas", max(DESIGNER_MIN_FONT_SIZE, int(round(line_px * 0.72))), "normal")
                y = offset_y + margin_px
                for line in rendered.splitlines() or [""]:
                    preview_canvas.create_text(
                        offset_x + margin_px,
                        y,
                        text=str(line),
                        anchor=tk.NW,
                        fill="#111",
                        font=mono_font,
                    )
                    y += line_px
                return

            canvas_w_preview = max(100, int(tpl_preview.get("canvas_w", state["canvas_w"])))
            canvas_h_preview = max(100, int(tpl_preview.get("canvas_h", state["canvas_h"])))
            scale_x = float(paper_w_px) / float(canvas_w_preview)
            scale_y = float(paper_h_px) / float(canvas_h_preview)
            base_scale = max(0.35, min(scale_x, scale_y))
            value_map = self._build_render_value_map(
                "1001",
                "u_demo_001",
                "DemoUser",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "这是实时预览内容 12345",
                extra_vars={"guest_msg_count": 3, "today_guest_rank": 12, "source_room": "demo_room"},
            )
            for e in sorted(elements, key=lambda x: (int(x.get("y", 0)), int(x.get("x", 0)))):
                text = _preview_text_for_element(e, value_map)
                if not text:
                    continue
                x = offset_x + int(round(int(e.get("x", 0)) * scale_x))
                y = offset_y + int(round(int(e.get("y", 0)) * scale_y))
                w = max(12, int(round(int(e.get("w", 160)) * scale_x)))
                h = max(12, int(round(int(e.get("h", 34)) * scale_y)))
                inset_x = max(2, int(round(4 * base_scale)))
                inset_y = max(2, int(round(3 * base_scale)))
                font_px = max(DESIGNER_MIN_FONT_SIZE, int(round(_elem_font_size(e) * base_scale)))
                font_obj = _get_cached_font(
                    _elem_font_family(e),
                    font_px,
                    "bold" if int(e.get("bold", 0)) else "normal",
                )
                box_width = max(8, w - (inset_x * 2))
                line_height_px = max(font_obj.metrics("linespace"), font_px + max(2, int(round(base_scale * 2))))
                wrapped_lines = _wrap_preview_text(
                    text,
                    font_obj,
                    box_width,
                    int(e.get("letter_spacing", 0)),
                    int(e.get("paragraph_spacing", 0)),
                )
                total_height = max(line_height_px, len(wrapped_lines) * line_height_px)
                valign = str(e.get("valign", "top")).lower()
                if valign in ("middle", "center"):
                    yy = y + max(0, (h - total_height) // 2)
                elif valign == "bottom":
                    yy = y + max(inset_y, h - total_height - inset_y)
                else:
                    yy = y + inset_y
                align = str(e.get("align", "left")).lower()
                for line in wrapped_lines:
                    if yy > (y + h - line_height_px):
                        break
                    text_width = max(0, int(font_obj.measure(line)))
                    if align == "center":
                        tx = x + max(0, (w - text_width) // 2)
                    elif align == "right":
                        tx = x + max(inset_x, w - inset_x - text_width)
                    else:
                        tx = x + inset_x
                    preview_canvas.create_text(
                        tx,
                        yy,
                        text=str(line),
                        anchor=tk.NW,
                        fill="#111",
                        font=font_obj,
                    )
                    yy += line_height_px

        def _schedule_preview():
            if state.get("preview_after_id"):
                try:
                    win.after_cancel(state["preview_after_id"])
                except Exception:
                    pass
                state["preview_after_id"] = None
            delay = 120 if perf_mode_var.get() else 0
            if delay <= 0:
                sample_render_preview()
                return
            def _run_preview():
                state["preview_after_id"] = None
                sample_render_preview()
            state["preview_after_id"] = win.after(delay, _run_preview)

        def draw(force_preview=False):
            canvas.delete("all")
            zoom = max(0.5, min(3.0, float(zoom_var.get()) / 100.0))
            cw = int(state["canvas_w"] * zoom)
            ch = int(state["canvas_h"] * zoom)
            canvas.create_rectangle(2, 2, cw - 2, ch - 2, outline="#888")
            canvas.configure(scrollregion=(0, 0, cw + 40, ch + 40))
            for i, e in enumerate(state["elements"]):
                ex, ey, ew, eh = _elem_box(i)
                x = int(ex * zoom)
                y = int(ey * zoom)
                w = int(ew * zoom)
                h = int(eh * zoom)
                txt = str(e.get("label", "")) or ("{" + str(e.get("field", "text")) + "}")
                base_fs = _elem_font_size(e)
                draw_fs = max(DESIGNER_MIN_FONT_SIZE, int(base_fs))
                draw_weight = "bold" if int(e.get("bold", 0)) else "normal"
                draw_family = _elem_font_family(e)
                fill = "#d1e9ff" if i in state["selected_ids"] else "#e8edf3"
                canvas.create_rectangle(x, y, x + w, y + h, fill=fill, outline="#5f7389")
                align = str(e.get("align", "left")).lower()
                valign = str(e.get("valign", "top")).lower()
                if align == "center":
                    tx = x + w // 2
                elif align == "right":
                    tx = x + w - 6
                else:
                    tx = x + 6
                if valign in ("middle", "center"):
                    ty = y + h // 2
                    if align == "center":
                        anchor = tk.CENTER
                    elif align == "right":
                        anchor = tk.E
                    else:
                        anchor = tk.W
                elif valign == "bottom":
                    ty = y + h - 6
                    if align == "center":
                        anchor = tk.S
                    elif align == "right":
                        anchor = tk.SE
                    else:
                        anchor = tk.SW
                else:
                    ty = y + 6
                    if align == "center":
                        anchor = tk.N
                    elif align == "right":
                        anchor = tk.NE
                    else:
                        anchor = tk.NW
                canvas.create_text(tx, ty, text=txt, anchor=anchor, fill="#223", font=_get_cached_font(draw_family, draw_fs, draw_weight))
                if state["selected_idx"] == i:
                    hs = 12
                    handles = {
                        "nw": (x, y),
                        "n": (x + w // 2, y),
                        "ne": (x + w, y),
                        "e": (x + w, y + h // 2),
                        "se": (x + w, y + h),
                        "s": (x + w // 2, y + h),
                        "sw": (x, y + h),
                        "w": (x, y + h // 2),
                    }
                    for hx, hy in handles.values():
                        canvas.create_rectangle(hx - hs, hy - hs, hx + hs, hy + hs, fill="#2563eb", outline="#1d4ed8")
            if state.get("drag_mode") == "marquee":
                sx, sy = state.get("marquee_start", (0, 0))
                ex, ey = state.get("marquee_end", (0, 0))
                canvas.create_rectangle(int(sx * zoom), int(sy * zoom), int(ex * zoom), int(ey * zoom), outline="#2563eb", dash=(4, 2))
            for g in state.get("guides", []):
                axis, pos = g
                if axis == "x":
                    p = int(pos * zoom)
                    canvas.create_line(p, 0, p, ch, fill="#ef4444", dash=(4, 2))
                else:
                    p = int(pos * zoom)
                    canvas.create_line(0, p, cw, p, fill="#ef4444", dash=(4, 2))
            if state.get("inline_edit_idx") is not None and state["inline_edit_idx"] >= len(state["elements"]):
                _close_inline_editor(restore=False)
            _place_inline_editor()
            if force_preview or (not perf_mode_var.get()) or (not state.get("dragging")):
                _schedule_preview()

        def apply_canvas_size():
            try:
                _push_history()
                old_w = max(100, int(state["canvas_w"]))
                old_h = max(100, int(state["canvas_h"]))
                new_w = _to_units(canvas_w_var.get())
                new_h = _to_units(canvas_h_var.get())
                scale_x = float(new_w) / float(old_w) if old_w > 0 else 1.0
                scale_y = float(new_h) / float(old_h) if old_h > 0 else 1.0
                state["canvas_w"] = int(new_w)
                state["canvas_h"] = int(new_h)
                if old_w != new_w or old_h != new_h:
                    for e in state["elements"]:
                        e["x"] = int(round(max(0, float(e.get("x", 0))) * scale_x))
                        e["y"] = int(round(max(0, float(e.get("y", 0))) * scale_y))
                        e["w"] = int(round(max(40, float(e.get("w", 160))) * scale_x))
                        e["h"] = int(round(max(20, float(e.get("h", 34))) * scale_y))
                        e["w"] = max(40, min(state["canvas_w"], int(e.get("w", 160))))
                        e["h"] = max(20, min(state["canvas_h"], int(e.get("h", 34))))
                        e["x"] = max(0, min(state["canvas_w"] - e["w"], int(e.get("x", 0))))
                        e["y"] = max(0, min(state["canvas_h"] - e["h"], int(e.get("y", 0))))
                pos_x_scale.configure(to=state["canvas_w"])
                pos_y_scale.configure(to=state["canvas_h"])
                elem_w_scale.configure(to=state["canvas_w"])
                refresh_elem_list()
                if state["selected_idx"] is not None and 0 <= state["selected_idx"] < len(state["elements"]):
                    _set_form_from_element(state["elements"][state["selected_idx"]])
                draw()
            except Exception:
                messagebox.showwarning("提示", "画布尺寸必须是数字(mm)")

        def apply_canvas_from_paper():
            w_units, h_units = _paper_default_canvas_units()
            canvas_w_var.set(_to_mm(w_units))
            canvas_h_var.set(_to_mm(h_units))
            apply_canvas_size()

        def apply_canvas_from_preset(event=None):
            name = str(canvas_preset_var.get()).strip()
            if not name:
                return
            presets = getattr(self, "paper_sizes", {})
            if name not in presets:
                return
            w_mm, h_mm = presets.get(name, ("40", "30"))
            try:
                canvas_w_var.set(float(w_mm))
                canvas_h_var.set(float(h_mm))
            except Exception:
                return
            apply_canvas_size()

        def add_or_update_var():
            k = var_key.get().strip()
            if not k:
                return
            self.template_custom_vars[k] = var_val.get()
            refresh_var_tree()
            sample_render_preview()

        def delete_var():
            sel = var_tree.selection()
            if not sel:
                return
            vals = var_tree.item(sel[0], "values")
            k = vals[0] if vals else ""
            if k:
                self.template_custom_vars.pop(str(k), None)
            refresh_var_tree()
            sample_render_preview()

        def on_var_pick(event=None):
            sel = var_tree.selection()
            if not sel:
                return
            vals = var_tree.item(sel[0], "values")
            if len(vals) >= 2:
                var_key.set(str(vals[0]))
                var_val.set(str(vals[1]))

        def _build_default_elem_for_key(key: str) -> dict:
            key = str(key).strip()
            field_key_map = {"permanent_id", "unique_id", "name", "time", "content"}
            if key in field_key_map:
                fld = key
                ckey = ""
            elif key in (self.template_custom_vars or {}):
                fld = "custom"
                ckey = key
            else:
                fld = "literal"
                ckey = ""
            return {
                "field": fld,
                "custom_key": ckey,
                "label": "{" + key + "}",
                "x": 20,
                "y": 20 + 30 * len(state["elements"]),
                "align": "left",
                "valign": "top",
                "font_family": str(font_family_var.get()).strip() or "TkDefaultFont",
                "font_size": max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, int(font_size_var.get()))),
                "scale": max(0.6, min(3.0, float(font_size_var.get()) / 12.0)),
                "letter_spacing": max(-6, min(8, int(letter_spacing_var.get()))),
                "paragraph_spacing": max(0, min(6, int(paragraph_spacing_var.get()))),
                "bold": 1 if int(bold_var.get()) else 0,
                "w": max(40, min(state["canvas_w"], int(elem_w_var.get()))),
                "h": max(20, min(state["canvas_h"], int(elem_h_var.get()))),
            }

        def add_custom_var_at_canvas(event=None):
            zoom = max(0.5, min(3.0, float(zoom_var.get()) / 100.0))
            cx = int(canvas.canvasx(getattr(event, "x", 0)) / zoom)
            cy = int(canvas.canvasy(getattr(event, "y", 0)) / zoom)
            var_key_name = simpledialog.askstring("新建变量", "请输入变量名(英文/数字/下划线):", parent=win)
            if not var_key_name:
                return
            key = str(var_key_name).strip()
            if not key:
                return
            if key.startswith("{") and key.endswith("}"):
                key = key[1:-1].strip()
            if not key:
                return
            default_text = "{" + key + "}"
            label_text = simpledialog.askstring("显示文本", "请输入显示文本(可用占位符，如 {my_var}):", initialvalue=default_text, parent=win)
            if label_text is None:
                return
            if key not in (self.template_custom_vars or {}):
                self.template_custom_vars[key] = ""
            e = {
                "field": "custom",
                "custom_key": key,
                "label": str(label_text).strip() or default_text,
                "x": max(0, min(state["canvas_w"] - 40, cx)),
                "y": max(0, min(state["canvas_h"] - 20, cy)),
                "align": "left",
                "valign": "top",
                "font_family": "TkDefaultFont",
                "font_size": 12,
                "scale": 1.0,
                "letter_spacing": 0,
                "paragraph_spacing": 0,
                "bold": 0,
                "w": max(40, min(state["canvas_w"], 180)),
                "h": max(20, min(state["canvas_h"], 34)),
            }
            _push_history()
            state["elements"].append(e)
            refresh_var_tree()
            refresh_elem_list()
            _apply_selection([len(state["elements"]) - 1], primary=len(state["elements"]) - 1)
            draw()

        def add_element():
            _push_history()
            fld = field_var.get().strip() or "content"
            ckey = custom_key_var.get().strip()
            lb = label_var.get().strip()
            if not lb:
                if fld == "custom":
                    lb = "{" + (ckey or "my_var") + "}"
                elif fld == "literal":
                    lb = "自定义文本"
                elif fld == "content":
                    lb = "{content}"
                else:
                    lb = f"{fld}:{{{fld}}}"
            state["elements"].append(
                {
                    "field": fld,
                    "custom_key": ckey,
                    "label": lb,
                    "x": 20,
                    "y": 20 + 30 * len(state["elements"]),
                    "align": align_var.get().strip() or "left",
                    "valign": valign_var.get().strip() or "top",
                    "font_family": str(font_family_var.get()).strip() or "TkDefaultFont",
                    "font_size": max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, int(font_size_var.get()))),
                    "scale": max(0.6, min(3.0, float(font_size_var.get()) / 12.0)),
                    "letter_spacing": max(-6, min(8, int(letter_spacing_var.get()))),
                    "paragraph_spacing": max(0, min(6, int(paragraph_spacing_var.get()))),
                    "bold": 1 if int(bold_var.get()) else 0,
                    "w": max(40, min(state["canvas_w"], int(elem_w_var.get()))),
                    "h": max(20, min(state["canvas_h"], int(elem_h_var.get()))),
                }
            )
            refresh_elem_list()
            _apply_selection([len(state["elements"]) - 1], primary=len(state["elements"]) - 1)
            draw()

        def add_doc_var_element(event=None):
            sel = var_docs_tree.selection()
            if not sel:
                return
            key = str(sel[0]).strip()
            if not key:
                return
            field_var.set("content")
            custom_key_var.set("")
            label_var.set("{" + key + "}")
            _push_history()
            state["elements"].append(_build_default_elem_for_key(key))
            refresh_elem_list()
            _apply_selection([len(state["elements"]) - 1], primary=len(state["elements"]) - 1)
            draw()

        def delete_element():
            idxs = _selected_indices()
            if not idxs:
                return
            _push_history()
            for idx in sorted(idxs, reverse=True):
                state["elements"].pop(int(idx))
            state["selected_idx"] = None
            state["selected_ids"] = set()
            refresh_elem_list()
            draw()

        def copy_selected_elements(event=None):
            if event is not None and _is_input_widget(getattr(event, "widget", None)):
                return
            idxs = _selected_indices()
            if not idxs:
                return "break"
            state["clipboard"] = [dict(state["elements"][i]) for i in idxs]
            return "break"

        def paste_elements(event=None):
            if event is not None and _is_input_widget(getattr(event, "widget", None)):
                return
            clip = state.get("clipboard") or []
            if not clip:
                return "break"
            _push_history()
            new_ids = []
            for src in clip:
                dup = dict(src)
                dup["x"] = max(0, min(state["canvas_w"] - int(dup.get("w", 160)), int(dup.get("x", 0)) + 16))
                dup["y"] = max(0, min(state["canvas_h"] - int(dup.get("h", 34)), int(dup.get("y", 0)) + 16))
                state["elements"].append(dup)
                new_ids.append(len(state["elements"]) - 1)
            refresh_elem_list()
            _apply_selection(new_ids, primary=(new_ids[0] if new_ids else None))
            draw()
            return "break"

        def duplicate_element():
            idxs = _selected_indices()
            if not idxs:
                return
            _push_history()
            new_ids = []
            for idx in idxs:
                src = dict(state["elements"][int(idx)])
                src["x"] = int(src.get("x", 0)) + 12
                src["y"] = int(src.get("y", 0)) + 12
                state["elements"].append(src)
                new_ids.append(len(state["elements"]) - 1)
            refresh_elem_list()
            _apply_selection(new_ids, primary=(new_ids[0] if new_ids else None))
            draw()

        def select_all_elements(event=None):
            if event is not None and _is_input_widget(getattr(event, "widget", None)):
                return
            if not state["elements"]:
                return "break"
            ids = list(range(len(state["elements"])))
            _apply_selection(ids, primary=ids[0])
            draw()
            return "break"

        def clear_selection(event=None):
            _apply_selection([], primary=None, update_form=False)
            draw()
            return "break"

        def move_layer(delta: int):
            idxs = _selected_indices()
            if not idxs:
                return
            _push_history()
            idx = int(idxs[0])
            ni = idx + delta
            if ni < 0 or ni >= len(state["elements"]):
                return
            state["elements"][idx], state["elements"][ni] = state["elements"][ni], state["elements"][idx]
            refresh_elem_list()
            _apply_selection([ni], primary=ni)
            draw()

        def reorder_selection(to_front: bool):
            idxs = _selected_indices()
            if not idxs:
                return
            _push_history()
            selected = [dict(state["elements"][i]) for i in idxs]
            remain = [dict(state["elements"][i]) for i in range(len(state["elements"])) if i not in set(idxs)]
            if to_front:
                state["elements"] = remain + selected
                new_ids = list(range(len(remain), len(remain) + len(selected)))
            else:
                state["elements"] = selected + remain
                new_ids = list(range(0, len(selected)))
            refresh_elem_list()
            _apply_selection(new_ids, primary=(new_ids[0] if new_ids else None))
            draw()

        def on_elem_pick(event=None):
            idxs = elem_list.curselection()
            if not idxs:
                return
            primary = int(idxs[0])
            if state.get("selected_idx") in idxs:
                primary = int(state.get("selected_idx"))
            _apply_selection([int(x) for x in idxs], primary=primary)
            draw()

        def update_element():
            idxs = elem_list.curselection()
            if not idxs:
                return
            _push_history()
            idx = int(idxs[0])
            e = state["elements"][idx]
            e["field"] = field_var.get().strip() or "content"
            e["custom_key"] = custom_key_var.get().strip()
            e["label"] = label_var.get().strip()
            e["align"] = align_var.get().strip() or "left"
            e["valign"] = valign_var.get().strip() or "top"
            e["font_family"] = str(font_family_var.get()).strip() or "TkDefaultFont"
            try:
                fsz = max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, int(font_size_var.get())))
            except Exception:
                fsz = 12
            e["font_size"] = int(fsz)
            e["scale"] = max(0.6, min(3.0, float(fsz) / 12.0))
            e["letter_spacing"] = max(-6, min(8, int(letter_spacing_var.get())))
            e["paragraph_spacing"] = max(0, min(6, int(paragraph_spacing_var.get())))
            e["bold"] = 1 if int(bold_var.get()) else 0
            e["w"] = max(40, min(state["canvas_w"], int(elem_w_var.get())))
            e["h"] = max(20, min(state["canvas_h"], int(elem_h_var.get())))
            refresh_elem_list()
            draw()

        def _elem_box(i):
            e = state["elements"][i]
            txt = str(e.get("label", "")) or ("{" + str(e.get("field", "text")) + "}")
            w = max(40, min(state["canvas_w"], int(e.get("w", max(80, min(300, len(txt) * 8 + 20))))))
            h = max(20, min(state["canvas_h"], int(e.get("h", 34))))
            return int(e.get("x", 0)), int(e.get("y", 0)), w, h

        def _close_inline_editor(restore=False):
            idx = state.get("inline_edit_idx")
            panel = state.get("inline_edit_panel")
            if restore and idx is not None and 0 <= int(idx) < len(state["elements"]):
                old_text, old_font_size, old_bold = state.get("inline_edit_original", ("", 12, 0))
                e = state["elements"][int(idx)]
                e["label"] = str(old_text)
                e["font_size"] = max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, int(old_font_size)))
                e["scale"] = max(0.6, min(3.0, float(e["font_size"]) / 12.0))
                e["bold"] = 1 if int(old_bold) else 0
                _set_form_from_element(e)
                refresh_elem_list()
            try:
                if panel is not None and panel.winfo_exists():
                    panel.destroy()
            except Exception:
                pass
            state["inline_edit_idx"] = None
            state["inline_edit_panel"] = None
            state["inline_edit_text_var"] = None
            state["inline_edit_scale_var"] = None
            state["inline_edit_bold_var"] = None
            state["inline_edit_original"] = ("", 12, 0)
            state["inline_edit_history_pushed"] = False

        def _apply_inline_editor_change(*_):
            idx = state.get("inline_edit_idx")
            if idx is None or idx < 0 or idx >= len(state["elements"]):
                return
            tv = state.get("inline_edit_text_var")
            sv = state.get("inline_edit_scale_var")
            bv = state.get("inline_edit_bold_var")
            if tv is None or sv is None:
                return
            e = state["elements"][idx]
            new_label = str(tv.get())
            try:
                new_font_size = max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, int(float(sv.get()))))
            except Exception:
                new_font_size = _elem_font_size(e)
            new_bold = 1 if int(bv.get()) else 0 if bv is not None else int(e.get("bold", 0))
            changed = (
                (new_label != str(e.get("label", "")))
                or (new_font_size != _elem_font_size(e))
                or (new_bold != int(e.get("bold", 0)))
            )
            if not changed:
                return
            if not state.get("inline_edit_history_pushed"):
                _push_history()
                state["inline_edit_history_pushed"] = True
            e["label"] = new_label
            e["font_size"] = int(new_font_size)
            e["scale"] = max(0.6, min(3.0, float(new_font_size) / 12.0))
            e["bold"] = int(new_bold)
            _set_form_from_element(e)
            refresh_elem_list()
            draw()

        def _place_inline_editor():
            idx = state.get("inline_edit_idx")
            panel = state.get("inline_edit_panel")
            if idx is None or panel is None or idx < 0 or idx >= len(state["elements"]):
                return
            zoom = max(0.5, min(3.0, float(zoom_var.get()) / 100.0))
            ex, ey, ew, _eh = _elem_box(idx)
            px = int(ex * zoom)
            py = int(ey * zoom) - 38
            if py < 4:
                py = int((ey + 28) * zoom)
            pw = min(max(320, int(ew * zoom) + 220), 820)
            view_left = int(canvas.canvasx(0))
            view_right = int(canvas.canvasx(max(1, canvas.winfo_width())))
            view_top = int(canvas.canvasy(0))
            view_bottom = int(canvas.canvasy(max(1, canvas.winfo_height())))
            x = max(view_left + 4, min(px, view_right - pw - 4))
            y = max(view_top + 4, min(py, view_bottom - 40))
            panel.place(x=int(x), y=int(y), width=int(pw))

        def _open_inline_editor(idx: int):
            if idx is None or idx < 0 or idx >= len(state["elements"]):
                return
            _close_inline_editor(restore=False)
            _apply_selection([idx], primary=idx)
            e = state["elements"][idx]
            text_var = tk.StringVar(value=str(e.get("label", "")))
            font_size_local_var = tk.IntVar(value=_elem_font_size(e))
            bold_local_var = tk.IntVar(value=1 if int(e.get("bold", 0)) else 0)
            panel = ttk.Frame(canvas)
            entry = ttk.Entry(panel, textvariable=text_var)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 4), pady=4)
            ttk.Combobox(panel, textvariable=font_size_local_var, values=DESIGNER_FONT_SIZE_CHOICES, width=5).pack(side=tk.LEFT, padx=4, pady=4)
            ttk.Checkbutton(panel, text="加粗", variable=bold_local_var).pack(side=tk.LEFT, padx=(2, 4), pady=4)
            ttk.Button(panel, text="完成", width=5, command=lambda: (_apply_inline_editor_change(), _close_inline_editor(False), draw())).pack(side=tk.LEFT, padx=(2, 6), pady=4)

            def _on_enter(event=None):
                _apply_inline_editor_change()
                _close_inline_editor(False)
                draw()
                return "break"

            def _on_escape(event=None):
                _close_inline_editor(restore=True)
                draw()
                return "break"

            entry.bind("<Return>", _on_enter)
            entry.bind("<Escape>", _on_escape)
            entry.bind("<KeyRelease>", _apply_inline_editor_change)
            panel.bind("<Escape>", _on_escape)
            font_size_local_var.trace_add("write", _apply_inline_editor_change)

            state["inline_edit_idx"] = int(idx)
            state["inline_edit_panel"] = panel
            state["inline_edit_text_var"] = text_var
            state["inline_edit_scale_var"] = font_size_local_var
            state["inline_edit_bold_var"] = bold_local_var
            state["inline_edit_original"] = (str(e.get("label", "")), _elem_font_size(e), int(e.get("bold", 0)))
            state["inline_edit_history_pushed"] = False
            _place_inline_editor()
            entry.focus_set()
            entry.selection_range(0, tk.END)

        def hit_test(px, py):
            zoom = max(0.5, min(3.0, float(zoom_var.get()) / 100.0))
            x0 = int(canvas.canvasx(px) / zoom)
            y0 = int(canvas.canvasy(py) / zoom)
            handle_tol = max(8, int(round(12 / zoom)))

            def _hit_handle_only(ex, ey, ew, eh):
                handles = {
                    "nw": (ex, ey),
                    "n": (ex + ew // 2, ey),
                    "ne": (ex + ew, ey),
                    "e": (ex + ew, ey + eh // 2),
                    "se": (ex + ew, ey + eh),
                    "s": (ex + ew // 2, ey + eh),
                    "sw": (ex, ey + eh),
                    "w": (ex, ey + eh // 2),
                }
                for name, (hx, hy) in handles.items():
                    if abs(x0 - hx) <= handle_tol and abs(y0 - hy) <= handle_tol:
                        return name
                return ""

            # Priority 1: selected element handle -> resize
            sidx = state.get("selected_idx")
            if sidx is not None and 0 <= int(sidx) < len(state["elements"]):
                ex, ey, ew, eh = _elem_box(int(sidx))
                h = _hit_handle_only(ex, ey, ew, eh)
                if h:
                    return int(sidx), "resize", h
                # Priority 2: selected element body -> move
                if ex <= x0 <= ex + ew and ey <= y0 <= ey + eh:
                    return int(sidx), "move", ""

            # Priority 3: other element handle -> resize, body -> move
            for i in range(len(state["elements"]) - 1, -1, -1):
                ex, ey, ew, eh = _elem_box(i)
                if sidx is not None and int(sidx) == i:
                    continue
                h = _hit_handle_only(ex, ey, ew, eh)
                if h:
                    return i, "resize", h
                if ex <= x0 <= ex + ew and ey <= y0 <= ey + eh:
                    return i, "move", ""
            return None, "move", ""

        def _move_targets(indices, dx, dy):
            if not indices:
                return
            min_x = None
            min_y = None
            max_x = None
            max_y = None
            for i in indices:
                ex, ey, ew, eh = _elem_box(i)
                min_x = ex if min_x is None else min(min_x, ex)
                min_y = ey if min_y is None else min(min_y, ey)
                max_x = ex + ew if max_x is None else max(max_x, ex + ew)
                max_y = ey + eh if max_y is None else max(max_y, ey + eh)
            dx = max(-min_x, min(state["canvas_w"] - max_x, dx))
            dy = max(-min_y, min(state["canvas_h"] - max_y, dy))
            for i in indices:
                e = state["elements"][i]
                e["x"] = int(e.get("x", 0)) + dx
                e["y"] = int(e.get("y", 0)) + dy

        def on_press(event):
            try:
                canvas.focus_set()
            except Exception:
                pass
            state["dragging"] = False
            idx, mode, handle_name = hit_test(event.x, event.y)
            inline_idx = state.get("inline_edit_idx")
            if inline_idx is not None and idx != inline_idx:
                _close_inline_editor(restore=False)
            state["drag_start"] = (canvas.canvasx(event.x), canvas.canvasy(event.y))
            x0 = int(canvas.canvasx(event.x) / max(0.5, min(3.0, float(zoom_var.get()) / 100.0)))
            y0 = int(canvas.canvasy(event.y) / max(0.5, min(3.0, float(zoom_var.get()) / 100.0)))
            toggle = bool(getattr(event, "state", 0) & 0x0005)  # shift/ctrl
            if idx is None:
                if not toggle:
                    _apply_selection([], primary=None, update_form=False)
                state["drag_mode"] = "marquee"
                state["drag_handle"] = ""
                state["marquee_start"] = (x0, y0)
                state["marquee_end"] = (x0, y0)
            else:
                if mode == "resize":
                    _push_history()
                    _apply_selection([idx], primary=idx)
                elif toggle:
                    cur = set(state.get("selected_ids") or set())
                    if idx in cur:
                        cur.remove(idx)
                        _apply_selection(sorted(cur), primary=(min(cur) if cur else None), update_form=bool(cur))
                    else:
                        cur.add(idx)
                        _apply_selection(sorted(cur), primary=idx)
                elif idx not in (state.get("selected_ids") or set()):
                    _apply_selection([idx], primary=idx)
                elif mode == "move":
                    _push_history()
                state["drag_mode"] = mode
                state["drag_handle"] = handle_name
            draw()

        def on_double_click(event):
            zoom = max(0.5, min(3.0, float(zoom_var.get()) / 100.0))
            x0 = int(canvas.canvasx(event.x) / zoom)
            y0 = int(canvas.canvasy(event.y) / zoom)
            idx = None
            for i in range(len(state["elements"]) - 1, -1, -1):
                ex, ey, ew, eh = _elem_box(i)
                if ex <= x0 <= ex + ew and ey <= y0 <= ey + eh:
                    idx = i
                    break
            if idx is None:
                return
            _open_inline_editor(int(idx))
            draw()
            return "break"

        def on_drag(event):
            state["dragging"] = True
            sx, sy = state["drag_start"]
            cx, cy = canvas.canvasx(event.x), canvas.canvasy(event.y)
            state["drag_start"] = (cx, cy)
            zoom = max(0.5, min(3.0, float(zoom_var.get()) / 100.0))
            dx, dy = int((cx - sx) / zoom), int((cy - sy) / zoom)
            if state.get("drag_mode") == "marquee":
                state["marquee_end"] = (int(cx / zoom), int(cy / zoom))
                sx0, sy0 = state["marquee_start"]
                ex0, ey0 = state["marquee_end"]
                x1, x2 = sorted((sx0, ex0))
                y1, y2 = sorted((sy0, ey0))
                hit = []
                for i in range(len(state["elements"])):
                    exb, eyb, ewb, ehb = _elem_box(i)
                    if exb <= x2 and (exb + ewb) >= x1 and eyb <= y2 and (eyb + ehb) >= y1:
                        hit.append(i)
                _apply_selection(hit, primary=(hit[-1] if hit else None), update_form=bool(hit))
                draw()
                return
            idx = state["selected_idx"]
            if idx is None:
                return
            e = state["elements"][idx]
            state["guides"] = []
            if state["drag_mode"] == "resize":
                ex, ey, ew, eh = _elem_box(idx)
                nx, ny, nw, nh = ex, ey, ew, eh
                handle_name = str(state.get("drag_handle", "")).lower()
                min_w, min_h = 40, 20
                if "w" in handle_name:
                    nx = ex + dx
                    nw = ew - dx
                if "e" in handle_name:
                    nw = ew + dx
                if "n" in handle_name:
                    ny = ey + dy
                    nh = eh - dy
                if "s" in handle_name:
                    nh = eh + dy

                if nw < min_w:
                    if "w" in handle_name:
                        nx -= (min_w - nw)
                    nw = min_w
                if nh < min_h:
                    if "n" in handle_name:
                        ny -= (min_h - nh)
                    nh = min_h

                nx = max(0, min(state["canvas_w"] - nw, nx))
                ny = max(0, min(state["canvas_h"] - nh, ny))
                nw = max(min_w, min(state["canvas_w"] - nx, nw))
                nh = max(min_h, min(state["canvas_h"] - ny, nh))
                e["x"], e["y"], e["w"], e["h"] = int(nx), int(ny), int(nw), int(nh)
                _set_form_from_element(e)
            else:
                targets = _selected_indices()
                _move_targets(targets, dx, dy)
                if snap_var.get() and len(targets) == 1:
                    tol = 6
                    ex, ey, ew, eh = _elem_box(idx)
                    candidates_x = [0, state["canvas_w"] // 2, state["canvas_w"]]
                    candidates_y = [0, state["canvas_h"] // 2, state["canvas_h"]]
                    for j in range(len(state["elements"])):
                        if j == idx:
                            continue
                        ox, oy, ow, oh = _elem_box(j)
                        candidates_x += [ox, ox + ow // 2, ox + ow]
                        candidates_y += [oy, oy + oh // 2, oy + oh]
                    cur_left, cur_mid_x, cur_right = ex, ex + ew // 2, ex + ew
                    cur_top, cur_mid_y, cur_bot = ey, ey + eh // 2, ey + eh
                    for cx in candidates_x:
                        if abs(cur_left - cx) <= tol:
                            e["x"] += cx - cur_left
                            state["guides"].append(("x", cx))
                            break
                        if abs(cur_mid_x - cx) <= tol:
                            e["x"] += cx - cur_mid_x
                            state["guides"].append(("x", cx))
                            break
                        if abs(cur_right - cx) <= tol:
                            e["x"] += cx - cur_right
                            state["guides"].append(("x", cx))
                            break
                    ex, ey, ew, eh = _elem_box(idx)
                    cur_top, cur_mid_y, cur_bot = ey, ey + eh // 2, ey + eh
                    for cy in candidates_y:
                        if abs(cur_top - cy) <= tol:
                            e["y"] += cy - cur_top
                            state["guides"].append(("y", cy))
                            break
                        if abs(cur_mid_y - cy) <= tol:
                            e["y"] += cy - cur_mid_y
                            state["guides"].append(("y", cy))
                            break
                        if abs(cur_bot - cy) <= tol:
                            e["y"] += cy - cur_bot
                            state["guides"].append(("y", cy))
                            break
            refresh_elem_list()
            if state["selected_idx"] is not None and state["selected_idx"] < len(state["elements"]):
                s = state["elements"][state["selected_idx"]]
                _set_form_from_element(s)
            draw()

        def on_release(event=None):
            if state.get("drag_mode") == "marquee":
                state["marquee_start"] = (0, 0)
                state["marquee_end"] = (0, 0)
            state["drag_mode"] = "move"
            state["drag_handle"] = ""
            state["dragging"] = False
            state["guides"] = []
            draw(force_preview=True)

        def on_wheel(event):
            # Windows/most users expect wheel to scroll canvas.
            # Use Ctrl + wheel for single-element font size scaling.
            delta = 0
            if hasattr(event, "delta") and event.delta:
                delta = 1 if event.delta > 0 else -1
            elif getattr(event, "num", None) == 4:
                delta = 1
            elif getattr(event, "num", None) == 5:
                delta = -1
            if delta == 0:
                return

            ctrl_pressed = bool(getattr(event, "state", 0) & 0x0004)
            if ctrl_pressed:
                idx = state["selected_idx"]
                if idx is None:
                    return
                e = state["elements"][idx]
                next_size = max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, _elem_font_size(e) + delta))
                e["font_size"] = int(next_size)
                e["scale"] = max(0.6, min(3.0, float(next_size) / 12.0))
                font_size_var.set(int(e["font_size"]))
                refresh_elem_list()
                draw()
                return "break"

            # Default: vertical scroll
            canvas.yview_scroll(-1 * delta, "units")
            return "break"

        def nudge_selected_font_size(delta: int):
            idx = state["selected_idx"]
            if idx is None:
                return
            e = state["elements"][idx]
            next_size = max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, _elem_font_size(e) + int(delta)))
            e["font_size"] = int(next_size)
            e["scale"] = max(0.6, min(3.0, float(next_size) / 12.0))
            font_size_var.set(int(e["font_size"]))
            refresh_elem_list()
            draw()

        def on_key(event):
            idx = state["selected_idx"]
            if idx is None:
                return
            step = 10 if (event.state & 0x0001) else 1
            dx = 0
            dy = 0
            if event.keysym == "Left":
                dx = -step
            elif event.keysym == "Right":
                dx = step
            elif event.keysym == "Up":
                dy = -step
            elif event.keysym == "Down":
                dy = step
            if dx != 0 or dy != 0:
                _push_history()
            _move_targets(_selected_indices(), dx, dy)
            refresh_elem_list()
            if state["selected_idx"] is not None and state["selected_idx"] < len(state["elements"]):
                e = state["elements"][state["selected_idx"]]
                _set_form_from_element(e)
            draw()
            return "break"

        def apply_pos_from_sliders(*_):
            if state.get("history_lock") or state.get("form_sync_lock"):
                return
            idx = state["selected_idx"]
            if idx is None:
                return
            _push_history()
            e = state["elements"][idx]
            try:
                _, _, ew, eh = _elem_box(idx)
                e["x"] = max(0, min(state["canvas_w"] - ew, int(float(pos_x_var.get()))))
                e["y"] = max(0, min(state["canvas_h"] - eh, int(float(pos_y_var.get()))))
            except Exception:
                return
            refresh_elem_list()
            draw()

        def apply_size_from_sliders(*_):
            if state.get("history_lock") or state.get("form_sync_lock"):
                return
            idx = state["selected_idx"]
            if idx is None:
                return
            _push_history()
            e = state["elements"][idx]
            try:
                e["w"] = max(40, min(state["canvas_w"], int(float(elem_w_var.get()))))
                e["h"] = max(20, min(state["canvas_h"], int(float(elem_h_var.get()))))
                e["x"] = max(0, min(state["canvas_w"] - e["w"], int(e.get("x", 0))))
                e["y"] = max(0, min(state["canvas_h"] - e["h"], int(e.get("y", 0))))
            except Exception:
                return
            refresh_elem_list()
            draw()

        def apply_font_style_from_inputs(*_):
            if state.get("history_lock") or state.get("form_sync_lock"):
                return
            idx = state["selected_idx"]
            if idx is None:
                return
            e = state["elements"][idx]
            try:
                new_font_size = max(DESIGNER_MIN_FONT_SIZE, min(DESIGNER_MAX_FONT_SIZE, int(float(font_size_var.get()))))
            except Exception:
                return
            new_family = str(font_family_var.get()).strip() or "TkDefaultFont"
            new_letter_spacing = max(-6, min(8, int(letter_spacing_var.get())))
            new_paragraph_spacing = max(0, min(6, int(paragraph_spacing_var.get())))
            new_bold = 1 if int(bold_var.get()) else 0
            if (
                new_font_size == _elem_font_size(e)
                and new_family == _elem_font_family(e)
                and new_letter_spacing == int(e.get("letter_spacing", 0))
                and new_paragraph_spacing == int(e.get("paragraph_spacing", 0))
                and new_bold == int(e.get("bold", 0))
            ):
                return
            _push_history()
            e["font_family"] = str(new_family)
            e["font_size"] = int(new_font_size)
            e["scale"] = max(0.6, min(3.0, float(new_font_size) / 12.0))
            e["letter_spacing"] = int(new_letter_spacing)
            e["paragraph_spacing"] = int(new_paragraph_spacing)
            e["bold"] = int(new_bold)
            refresh_elem_list()
            draw()

        var_btn = ttk.Frame(var_box)
        var_btn.grid(row=4, column=0, columnspan=4, sticky=tk.W, padx=4, pady=4)
        ttk.Button(var_btn, text="新增/更新变量", command=add_or_update_var).pack(side=tk.LEFT, padx=2)
        ttk.Button(var_btn, text="删除变量", command=delete_var).pack(side=tk.LEFT, padx=2)
        ttk.Button(var_btn, text="选中说明自动新增元素", command=add_doc_var_element).pack(side=tk.LEFT, padx=2)

        elem_btn = ttk.Frame(elem)
        elem_btn.grid(row=11, column=0, columnspan=4, sticky=tk.W, padx=4, pady=4)
        ttk.Button(elem_btn, text="添加元素", command=add_element).pack(side=tk.LEFT, padx=2)
        ttk.Button(elem_btn, text="删除元素", command=delete_element).pack(side=tk.LEFT, padx=2)
        ttk.Button(elem_btn, text="复制元素", command=duplicate_element).pack(side=tk.LEFT, padx=2)
        ttk.Button(elem_btn, text="复制所选", command=copy_selected_elements).pack(side=tk.LEFT, padx=2)
        ttk.Button(elem_btn, text="粘贴元素", command=paste_elements).pack(side=tk.LEFT, padx=2)
        ttk.Button(elem_btn, text="更新元素", command=update_element).pack(side=tk.LEFT, padx=2)
        ttk.Button(elem_btn, text="应用画布大小", command=apply_canvas_size).pack(side=tk.LEFT, padx=2)
        ttk.Button(elem_btn, text="上移图层", command=lambda: move_layer(1)).pack(side=tk.LEFT, padx=2)
        ttk.Button(elem_btn, text="下移图层", command=lambda: move_layer(-1)).pack(side=tk.LEFT, padx=2)
        ttk.Button(elem_btn, text="置顶", command=lambda: reorder_selection(True)).pack(side=tk.LEFT, padx=2)
        ttk.Button(elem_btn, text="置底", command=lambda: reorder_selection(False)).pack(side=tk.LEFT, padx=2)
        def align_selected(mode: str):
            indices = _selected_indices()
            if not indices:
                return
            _push_history()
            if len(indices) == 1:
                idx = indices[0]
                ex, ey, ew, eh = _elem_box(idx)
                e = state["elements"][idx]
                if mode == "left":
                    e["x"] = 0
                elif mode == "hcenter":
                    e["x"] = max(0, (state["canvas_w"] - ew) // 2)
                elif mode == "right":
                    e["x"] = max(0, state["canvas_w"] - ew)
                elif mode == "top":
                    e["y"] = 0
                elif mode == "vcenter":
                    e["y"] = max(0, (state["canvas_h"] - eh) // 2)
                elif mode == "bottom":
                    e["y"] = max(0, state["canvas_h"] - eh)
            else:
                anchor_idx = state["selected_idx"] if state["selected_idx"] in indices else indices[0]
                ax, ay, aw, ah = _elem_box(anchor_idx)
                for idx in indices:
                    e = state["elements"][idx]
                    _, _, ew, eh = _elem_box(idx)
                    if mode == "left":
                        e["x"] = ax
                    elif mode == "hcenter":
                        e["x"] = ax + (aw // 2) - (ew // 2)
                    elif mode == "right":
                        e["x"] = ax + aw - ew
                    elif mode == "top":
                        e["y"] = ay
                    elif mode == "vcenter":
                        e["y"] = ay + (ah // 2) - (eh // 2)
                    elif mode == "bottom":
                        e["y"] = ay + ah - eh
            refresh_elem_list()
            draw()

        def distribute_elements(axis: str):
            items = []
            candidates = _selected_indices()
            if len(candidates) < 3:
                candidates = list(range(len(state["elements"])))
            for i in candidates:
                ex, ey, ew, eh = _elem_box(i)
                items.append((i, ex, ey, ew, eh))
            if len(items) < 3:
                return
            _push_history()
            if axis == "x":
                items.sort(key=lambda t: t[1])
                first_x = items[0][1]
                last_end = items[-1][1] + items[-1][3]
                total_w = sum(t[3] for t in items)
                span = last_end - first_x
                gap = (span - total_w) / float(len(items) - 1)
                cur = float(first_x)
                for i, _x, _y, w, _h in items:
                    state["elements"][i]["x"] = int(round(cur))
                    cur += w + gap
            elif axis == "y":
                items.sort(key=lambda t: t[2])
                first_y = items[0][2]
                last_end = items[-1][2] + items[-1][4]
                total_h = sum(t[4] for t in items)
                span = last_end - first_y
                gap = (span - total_h) / float(len(items) - 1)
                cur = float(first_y)
                for i, _x, _y, _w, h in items:
                    state["elements"][i]["y"] = int(round(cur))
                    cur += h + gap
            refresh_elem_list()
            draw()

        def apply_inner_align(h_align: str = None, v_align: str = None):
            indices = _selected_indices()
            if not indices:
                return
            _push_history()
            for idx in indices:
                e = state["elements"][idx]
                if h_align in ("left", "center", "right"):
                    e["align"] = h_align
                if v_align in ("top", "middle", "bottom"):
                    e["valign"] = v_align
            if state["selected_idx"] is not None and state["selected_idx"] < len(state["elements"]):
                _set_form_from_element(state["elements"][state["selected_idx"]])
            refresh_elem_list()
            draw()

        align_btn = ttk.Frame(elem)
        align_btn.grid(row=12, column=0, columnspan=4, sticky=tk.W, padx=4, pady=(0, 4))
        ttk.Button(align_btn, text="左对齐", command=lambda: align_selected("left")).pack(side=tk.LEFT, padx=2)
        ttk.Button(align_btn, text="水平居中", command=lambda: align_selected("hcenter")).pack(side=tk.LEFT, padx=2)
        ttk.Button(align_btn, text="右对齐", command=lambda: align_selected("right")).pack(side=tk.LEFT, padx=2)
        ttk.Button(align_btn, text="上对齐", command=lambda: align_selected("top")).pack(side=tk.LEFT, padx=2)
        ttk.Button(align_btn, text="垂直居中", command=lambda: align_selected("vcenter")).pack(side=tk.LEFT, padx=2)
        ttk.Button(align_btn, text="下对齐", command=lambda: align_selected("bottom")).pack(side=tk.LEFT, padx=2)
        distribute_btn = ttk.Frame(elem)
        distribute_btn.grid(row=13, column=0, columnspan=4, sticky=tk.W, padx=4, pady=(0, 4))
        ttk.Button(distribute_btn, text="全体水平等距", command=lambda: distribute_elements("x")).pack(side=tk.LEFT, padx=2)
        ttk.Button(distribute_btn, text="全体垂直等距", command=lambda: distribute_elements("y")).pack(side=tk.LEFT, padx=2)
        ttk.Button(distribute_btn, text="全选", command=select_all_elements).pack(side=tk.LEFT, padx=2)
        ttk.Button(distribute_btn, text="清空选择", command=clear_selection).pack(side=tk.LEFT, padx=2)

        foot = ttk.Frame(win)
        foot.pack(fill=tk.X, padx=8, pady=(0, 8))

        def save_all():
            self.canvas_template = self._normalize_canvas_template_to_paper({
                "canvas_w": state["canvas_w"],
                "canvas_h": state["canvas_h"],
                "font_scale": float(global_scale_var.get()),
                "elements": state["elements"],
            })
            self._set_template_mode("designer", save=False)
            self._save_settings()
            messagebox.showinfo("保存", "画布模板和变量已保存并启用")

        ttk.Button(foot, text="保存并启用", command=save_all).pack(side=tk.LEFT, padx=4)
        ttk.Button(foot, text="撤销", command=undo_action).pack(side=tk.LEFT, padx=4)
        ttk.Button(foot, text="重做", command=redo_action).pack(side=tk.LEFT, padx=4)
        ttk.Button(foot, text="关闭", command=win.destroy).pack(side=tk.LEFT, padx=4)

        def _is_input_widget(w):
            return isinstance(w, (tk.Entry, tk.Text, ttk.Entry, ttk.Combobox))

        def on_delete_key(event=None):
            if event is not None and isinstance(getattr(event, "widget", None), (tk.Entry, tk.Text, ttk.Entry, ttk.Combobox)):
                return
            delete_element()
            return "break"

        def on_ctrl_d(event=None):
            if event is not None and _is_input_widget(getattr(event, "widget", None)):
                return
            duplicate_element()
            return "break"

        canvas_ctx_menu = tk.Menu(win, tearoff=0)
        canvas_ctx_menu.add_command(label="在此新建变量", command=lambda: add_custom_var_at_canvas())
        canvas_ctx_menu.add_separator()
        canvas_ctx_menu.add_command(label="框内居中(横+竖)", command=lambda: apply_inner_align("center", "middle"))
        align_h_menu = tk.Menu(canvas_ctx_menu, tearoff=0)
        align_h_menu.add_command(label="框内左对齐", command=lambda: apply_inner_align("left", None))
        align_h_menu.add_command(label="框内水平居中", command=lambda: apply_inner_align("center", None))
        align_h_menu.add_command(label="框内右对齐", command=lambda: apply_inner_align("right", None))
        canvas_ctx_menu.add_cascade(label="框内水平对齐", menu=align_h_menu)
        align_v_menu = tk.Menu(canvas_ctx_menu, tearoff=0)
        align_v_menu.add_command(label="框内上对齐", command=lambda: apply_inner_align(None, "top"))
        align_v_menu.add_command(label="框内垂直居中", command=lambda: apply_inner_align(None, "middle"))
        align_v_menu.add_command(label="框内下对齐", command=lambda: apply_inner_align(None, "bottom"))
        canvas_ctx_menu.add_cascade(label="框内垂直对齐", menu=align_v_menu)
        canvas_ctx_menu.add_separator()
        canvas_ctx_menu.add_command(label="复制所选", command=copy_selected_elements)
        canvas_ctx_menu.add_command(label="粘贴元素", command=paste_elements)
        canvas_ctx_menu.add_command(label="删除所选", command=delete_element)

        def on_canvas_context(event):
            idx, _mode, _handle = hit_test(event.x, event.y)
            if idx is not None and idx not in (state.get("selected_ids") or set()):
                _apply_selection([idx], primary=idx)
                draw()
            # remember click point for context "new variable at here"
            state["ctx_x"] = int(getattr(event, "x", 0))
            state["ctx_y"] = int(getattr(event, "y", 0))
            try:
                canvas_ctx_menu.tk_popup(event.x_root, event.y_root)
            finally:
                canvas_ctx_menu.grab_release()
            return "break"

        # route context menu command position to current click point
        def add_custom_var_from_ctx():
            class _E:
                pass
            e = _E()
            e.x = int(state.get("ctx_x", 0))
            e.y = int(state.get("ctx_y", 0))
            add_custom_var_at_canvas(e)

        canvas_ctx_menu.entryconfigure("在此新建变量", command=add_custom_var_from_ctx)

        var_tree.bind("<<TreeviewSelect>>", on_var_pick)
        var_docs_tree.bind("<Double-1>", add_doc_var_element)
        elem_list.bind("<<ListboxSelect>>", on_elem_pick)
        canvas.bind("<Button-1>", on_press)
        canvas.bind("<Double-Button-1>", on_double_click)
        canvas.bind("<Button-3>", on_canvas_context)
        canvas.bind("<Button-2>", on_canvas_context)
        canvas.bind("<B1-Motion>", on_drag)
        canvas.bind("<ButtonRelease-1>", on_release)
        canvas.bind("<MouseWheel>", on_wheel)
        canvas.bind("<Button-4>", lambda e: on_wheel(type("Event", (), {"delta": 120})()))
        canvas.bind("<Button-5>", lambda e: on_wheel(type("Event", (), {"delta": -120})()))
        win.bind("<Left>", on_key)
        win.bind("<Right>", on_key)
        win.bind("<Up>", on_key)
        win.bind("<Down>", on_key)
        win.bind("<Control-a>", select_all_elements)
        win.bind("<Command-a>", select_all_elements)
        win.bind("<Control-c>", copy_selected_elements)
        win.bind("<Command-c>", copy_selected_elements)
        win.bind("<Control-v>", paste_elements)
        win.bind("<Command-v>", paste_elements)
        win.bind("<Control-d>", on_ctrl_d)
        win.bind("<Command-d>", on_ctrl_d)
        win.bind("<Control-z>", undo_action)
        win.bind("<Command-z>", undo_action)
        win.bind("<Control-y>", redo_action)
        win.bind("<Command-y>", redo_action)
        win.bind("<Control-Shift-Z>", redo_action)
        win.bind("<Command-Shift-Z>", redo_action)
        win.bind("<Control-Shift-z>", redo_action)
        win.bind("<Command-Shift-z>", redo_action)
        win.bind("<Delete>", on_delete_key)
        win.bind("<BackSpace>", on_delete_key)
        win.bind("<Escape>", clear_selection)
        canvas_preset_cb.bind("<<ComboboxSelected>>", apply_canvas_from_preset)
        zoom_var.trace_add("write", lambda *_: draw())
        global_scale_var.trace_add("write", lambda *_: draw())
        perf_mode_var.trace_add("write", lambda *_: draw(force_preview=True))
        pos_x_var.trace_add("write", apply_pos_from_sliders)
        pos_y_var.trace_add("write", apply_pos_from_sliders)
        elem_w_var.trace_add("write", apply_size_from_sliders)
        elem_h_var.trace_add("write", apply_size_from_sliders)
        font_size_var.trace_add("write", apply_font_style_from_inputs)
        font_family_var.trace_add("write", apply_font_style_from_inputs)
        letter_spacing_var.trace_add("write", apply_font_style_from_inputs)
        paragraph_spacing_var.trace_add("write", apply_font_style_from_inputs)
        bold_var.trace_add("write", apply_font_style_from_inputs)
        font_minus_btn.configure(command=lambda: nudge_selected_font_size(-1))
        font_plus_btn.configure(command=lambda: nudge_selected_font_size(1))

        refresh_var_tree()
        refresh_elem_list()
        _push_history(clear_future=True)
        draw()
        self._canvas_tpl_win = win
        self._center_window(win, 1260, 760)
    def _toggle_proxy(self):
        try:
            if hasattr(self, "proxy_route_mode_var") and self.proxy_enabled.get():
                current_mode = self._normalize_proxy_route_mode(self.proxy_route_mode_var.get())
                if current_mode == "direct":
                    self.proxy_route_mode_var.set(PROXY_ROUTE_MODE_VALUE_TO_LABEL["all"])
            elif hasattr(self, "proxy_route_mode_var") and not self.proxy_enabled.get():
                self.proxy_route_mode_var.set(PROXY_ROUTE_MODE_VALUE_TO_LABEL["direct"])
            mode = self._get_proxy_route_mode()
            state = tk.NORMAL if mode != "direct" else tk.DISABLED
            self.proxy_entry.configure(state=state)
            if hasattr(self, "proxy_mode_cb"):
                self.proxy_mode_cb.configure(state="readonly")
        except Exception:
            pass
        self._save_settings()

    def _on_proxy_route_mode_change(self):
        try:
            mode = self._normalize_proxy_route_mode(self.proxy_route_mode_var.get() if hasattr(self, "proxy_route_mode_var") else "direct")
            if hasattr(self, "proxy_enabled"):
                desired = 0 if mode == "direct" else 1
                if int(self.proxy_enabled.get()) != desired:
                    self.proxy_enabled.set(desired)
            state = tk.NORMAL if mode != "direct" else tk.DISABLED
            if hasattr(self, "proxy_entry"):
                self.proxy_entry.configure(state=state)
        except Exception:
            pass
        self._request_save_settings()

    def _normalize_proxy_route_mode(self, value: str) -> str:
        raw = str(value or "").strip()
        if raw in PROXY_ROUTE_MODE_LABEL_TO_VALUE:
            return PROXY_ROUTE_MODE_LABEL_TO_VALUE[raw]
        raw = raw.lower()
        if raw in PROXY_ROUTE_MODE_VALUE_TO_LABEL:
            return raw
        return "direct"

    def _get_proxy_route_mode(self) -> str:
        if APP_IS_SIGNPOOL_RELAY_BUILD:
            return "direct"
        if hasattr(self, "proxy_route_mode_var"):
            mode = self._normalize_proxy_route_mode(self.proxy_route_mode_var.get())
        else:
            mode = self._normalize_proxy_route_mode(self.settings.get("proxy_route_mode", ""))
        if mode == "direct" and hasattr(self, "proxy_enabled") and bool(self.proxy_enabled.get()):
            saved = self._normalize_proxy_route_mode(self.settings.get("proxy_route_mode", "all"))
            if saved in ("all", "tiktok_only", "sign_only"):
                return saved
        return mode

    def _proxy_route_mode_label(self, mode: str) -> str:
        return PROXY_ROUTE_MODE_VALUE_TO_LABEL.get(self._normalize_proxy_route_mode(mode), "全部直连")

    def _proxy_applies_to_tiktok(self, mode: str = None) -> bool:
        current = self._normalize_proxy_route_mode(mode or self._get_proxy_route_mode())
        return current in ("all", "tiktok_only")

    def _proxy_applies_to_sign(self, mode: str = None) -> bool:
        current = self._normalize_proxy_route_mode(mode or self._get_proxy_route_mode())
        return current in ("all", "sign_only")

    def _resolve_configured_proxy(self) -> str:
        if APP_IS_SIGNPOOL_RELAY_BUILD:
            return ""
        manual_proxy = self._normalize_proxy_url(self.proxy_var.get() if hasattr(self, "proxy_var") else "")
        if manual_proxy:
            return manual_proxy
        try:
            proxies = urllib.request.getproxies() or {}
            for key in ("https", "http", "all"):
                v = self._normalize_proxy_url(proxies.get(key, ""))
                if v:
                    return v
        except Exception:
            pass
        return ""

    def _normalize_proxy_url(self, proxy_url: str) -> str:
        proxy = str(proxy_url or "").strip()
        if not proxy:
            return ""
        if "://" not in proxy:
            proxy = "http://" + proxy
        return proxy

    def _resolve_system_proxy(self) -> str:
        """Backward-compatible wrapper that returns the active proxy string."""
        mode = self._get_proxy_route_mode()
        if mode == "direct":
            return ""
        return self._resolve_configured_proxy()

    def _build_proxy_runtime(self, proxy_url: str):
        """Build proxy objects for TikTokLiveClient(web_proxy/ws_proxy)."""
        web_proxy = None
        ws_proxy = None
        if not proxy_url:
            return web_proxy, ws_proxy
        p = self._normalize_proxy_url(proxy_url)
        try:
            pobj = httpx.Proxy(p)
            web_proxy = pobj
            ws_proxy = pobj
        except Exception:
            # Fallback to raw URL if parsing fails (library may still accept it)
            web_proxy = p
            ws_proxy = p
        return web_proxy, ws_proxy

    def _configure_sign_server_defaults(self, sign_api_base: str, sign_api_key: str = None):
        base = str(sign_api_base or DEFAULT_SIGN_API_BASE).strip().rstrip("/")
        try:
            if WebDefaults is not None:
                WebDefaults.tiktok_sign_url = base
                WebDefaults.tiktok_sign_api_key = sign_api_key or None
        except Exception:
            pass

    def _configure_signer_client(self, client, sign_api_base: str, sign_api_key: str = None, proxy: str = "", verify: bool = True):
        if client is None:
            return
        try:
            signer = getattr(getattr(client, "web", None), "signer", None)
            if signer is None:
                return
            base = str(sign_api_base or DEFAULT_SIGN_API_BASE).strip().rstrip("/")
            try:
                signer._sign_api_base = base
            except Exception:
                pass
            try:
                signer._sign_api_key = sign_api_key or None
            except Exception:
                pass
            old_client = getattr(signer, "_httpx", None)
            headers = {}
            timeout = httpx.Timeout(20.0, connect=20.0, read=20.0, write=20.0)
            try:
                if old_client is not None:
                    headers = dict(getattr(old_client, "headers", {}) or {})
                    prev_timeout = getattr(old_client, "timeout", None)
                    if prev_timeout is not None:
                        timeout = prev_timeout
            except Exception:
                headers = {}
            if sign_api_key:
                headers["X-Api-Key"] = sign_api_key
            else:
                headers.pop("X-Api-Key", None)
            proxy_url = self._normalize_proxy_url(proxy)
            signer._httpx = httpx.AsyncClient(
                headers=headers,
                verify=bool(verify),
                timeout=timeout,
                trust_env=False,
                proxy=proxy_url or None,
            )
        except Exception:
            pass

    def clear_stream_rows(self):
        children = self.stream_tree.get_children()
        count = len(children)
        if count == 0:
            return
        if not messagebox.askyesno("确认", f"确认清空 {count} 条弹幕流水吗？"):
            return
        self.stream_tree.delete(*children)
        self._audit("clear_stream_rows", f"count={count}")
        self._set_status(f"已清空弹幕流水: {count} 条")

    def release_memory(self):
        if not messagebox.askyesno("确认", "释放内存将清空缓存和待处理队列，是否继续？"):
            return
        dropped = 0
        while True:
            try:
                self.queue.get_nowait()
                dropped += 1
            except queue.Empty:
                break
        self.user_msg_timestamps.clear()
        self.user_cache.clear()
        self.overlap_events.clear()
        self._lock_order_winner_key = None
        self._lock_order_open_until = 0.0
        gc.collect()
        self._audit("release_memory", f"dropped_queue={dropped}")
        self._set_status(f"内存释放完成，清理队列 {dropped} 条")
        self._refresh_soft_status()

    def detect_printers(self, silent: bool = False):
        printers = printer_utils.detect_printers()
        self.printer_cb['values'] = printers
        saved = str(self.settings.get("selected_printer", "")).strip()
        default_printer = printer_utils.get_default_printer()
        target = ""
        if saved and saved in printers:
            target = saved
        elif default_printer and default_printer in printers:
            target = default_printer
        elif printers:
            target = printers[0]
        if target:
            self.printer_cb.set(target)
        self._save_settings()
        if not silent:
            messagebox.showinfo("检测完成", f"发现打印机: {printers}")

    def _printer_health_probe(self):
        try:
            printers = printer_utils.detect_printers()
            default_printer = printer_utils.get_default_printer().strip()
            selected = str(self.printer_cb.get().strip()) if hasattr(self, "printer_cb") else ""
            use_default = bool(self.use_default_printer_var.get()) if hasattr(self, "use_default_printer_var") else True
            target = default_printer if use_default else selected
            if target and target not in printers:
                self._set_status(f"打印机异常: 目标打印机不可用 ({target})")
                self._audit("printer_health_warn", f"target={target}|available={len(printers)}")
        except Exception as e:
            self._audit("printer_health_error", str(e))
        finally:
            try:
                self.root.after(60000, self._printer_health_probe)
            except Exception:
                pass

    def test_print_selected_printer(self):
        if bool(self.use_default_printer_var.get()):
            printer = printer_utils.get_default_printer().strip()
        else:
            printer = self.printer_cb.get().strip()
        if not printer:
            messagebox.showwarning("提示", "未找到可用打印机（默认或手动）")
            return
        try:
            os.makedirs(DATA_DIR, exist_ok=True)
            stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            path = os.path.join(DATA_DIR, f"test_print_{int(time.time())}.txt")
            try:
                width_mm = int(float(self.width_var.get().strip()))
                height_mm = int(float(self.height_var.get().strip()))
            except Exception:
                width_mm, height_mm = 40, 30
            char_width_mm, line_height_mm, margin_mm = self._print_calibration()
            body = (
                "Sen Nails 打印测试\n"
                f"时间: {stamp}\n"
                f"打印机: {printer}\n"
                f"纸张: {width_mm}x{height_mm}mm\n"
                f"校准: 字宽{char_width_mm:.2f} 行高{line_height_mm:.2f} 边距{margin_mm:.2f}mm\n"
                "状态: 如果你看到这张纸，自动打印链路正常。\n"
            )
            printer_utils.print_to_file(body, path)
            ok, detail = printer_utils.send_to_printer_debug(
                printer,
                path,
                width_mm=width_mm,
                height_mm=height_mm,
                char_width_mm=char_width_mm,
                line_height_mm=line_height_mm,
                margin_mm=margin_mm,
            )
            if ok:
                messagebox.showinfo("测试打印", f"已发送到打印机: {printer}")
            else:
                log_path = os.path.join(RUNTIME_DIR, "sen_nails_print_debug.log")
                try:
                    with open(log_path, "a", encoding="utf-8") as f:
                        f.write(f"\n[{stamp}] printer={printer} file={path}\n{detail}\n")
                except Exception:
                    pass
                messagebox.showerror("测试打印", f"发送失败。\n\n{detail}\n\n日志: {log_path}")
        except Exception as e:
            messagebox.showerror("测试打印错误", str(e))

    def open_system_print_queue(self):
        if os.name != "nt":
            messagebox.showinfo("提示", "当前系统仅支持在 Windows 打开系统打印队列。")
            return
        try:
            # Open Windows printers page/queue entry point.
            subprocess.Popen(["cmd", "/c", "start", "", "control", "printers"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            messagebox.showerror("错误", f"打开系统打印队列失败: {e}")

    def add_blacklist(self):
        uid = self.bl_var.get().strip()
        if not uid:
            return
        db.add_blacklist(uid)
        self._audit("blacklist_add", uid)
        messagebox.showinfo("已添加", uid)

    def remove_blacklist(self):
        uid = self.bl_var.get().strip()
        if not uid:
            return
        db.remove_blacklist(uid)
        self._audit("blacklist_remove", uid)
        messagebox.showinfo("已移除", uid)

    def show_blacklist(self):
        self.open_blacklist_manager()

    def open_blacklist_manager(self):
        if hasattr(self, "_bl_win") and self._bl_win.winfo_exists():
            self._bl_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title("黑名单管理")
        win.geometry("620x520")

        top = ttk.Frame(win)
        top.pack(fill=tk.X, padx=8, pady=8)
        ttk.Label(top, text="搜索UID:").pack(side=tk.LEFT)
        q_var = tk.StringVar()
        q_ent = ttk.Entry(top, textvariable=q_var, width=28)
        q_ent.pack(side=tk.LEFT, padx=6)

        tree = ttk.Treeview(win, columns=("uid",), show="headings", selectmode="extended")
        tree.heading("uid", text="Unique ID")
        tree.column("uid", width=560)
        tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        def refresh():
            q = q_var.get().strip().lower()
            for r in tree.get_children():
                tree.delete(r)
            for uid in db.list_blacklist():
                if (not q) or (q in str(uid).lower()):
                    tree.insert("", tk.END, values=(uid,))
            self._set_status(f"黑名单数量: {len(tree.get_children())}")

        def remove_selected():
            sels = tree.selection()
            if not sels:
                return
            for s in sels:
                vals = tree.item(s, "values")
                if vals:
                    db.remove_blacklist(str(vals[0]))
            refresh()

        def import_txt():
            path = filedialog.askopenfilename(filetypes=[("Text/CSV", "*.txt *.csv"), ("All", "*.*")], title="导入黑名单")
            if not path:
                return
            added = 0
            try:
                with open(path, "r", encoding="utf-8-sig") as f:
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        uid = line.split(",")[0].strip()
                        if uid:
                            db.add_blacklist(uid)
                            added += 1
            except Exception as e:
                messagebox.showerror("错误", str(e))
                return
            refresh()
            messagebox.showinfo("导入完成", f"已处理 {added} 条")

        def export_txt():
            path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text", "*.txt"), ("CSV", "*.csv")], title="导出黑名单")
            if not path:
                return
            try:
                with open(path, "w", encoding="utf-8") as f:
                    for uid in db.list_blacklist():
                        f.write(f"{uid}\n")
                messagebox.showinfo("导出", f"已导出到 {path}")
            except Exception as e:
                messagebox.showerror("错误", str(e))

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(btn, text="刷新", command=refresh).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="移除所选", command=remove_selected).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="导入", command=import_txt).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="导出", command=export_txt).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn, text="关闭", command=win.destroy).pack(side=tk.LEFT, padx=4)
        ttk.Button(top, text="搜索", command=refresh).pack(side=tk.LEFT, padx=4)
        ttk.Button(top, text="清空", command=lambda: (q_var.set(""), refresh())).pack(side=tk.LEFT, padx=4)

        refresh()
        self._bl_win = win



    def open_pid_manager(self):
        if not self._require_admin():
            return
        if hasattr(self, '_pid_win') and self._pid_win.winfo_exists():
            self._pid_win.lift()
            return
        win = tk.Toplevel(self.root)
        win.title('永久编号管理')
        # Search box
        search_frm = ttk.Frame(win)
        search_frm.pack(fill=tk.X, padx=6, pady=6)
        ttk.Label(search_frm, text='搜索:').pack(side=tk.LEFT)
        search_var = tk.StringVar(value="")
        search_entry = ttk.Entry(search_frm, textvariable=search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(search_frm, text='查找', command=lambda: self._pid_search(search_var.get())).pack(side=tk.LEFT, padx=4)
        ttk.Button(search_frm, text='清除', command=lambda: (search_var.set(""), self._refresh_pid_tree(tree))).pack(side=tk.LEFT)

        cols = ('permanent_id', 'unique_id', 'display_name')
        tree_wrap = ttk.Frame(win)
        tree_wrap.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))
        tree_wrap.rowconfigure(0, weight=1)
        tree_wrap.columnconfigure(0, weight=1)
        tree = ttk.Treeview(tree_wrap, columns=cols, show='headings')
        tree.heading('permanent_id', text='永久ID')
        tree.heading('unique_id', text='Unique ID')
        tree.heading('display_name', text='名称')
        vbar = ttk.Scrollbar(tree_wrap, orient=tk.VERTICAL, command=tree.yview)
        hbar = ttk.Scrollbar(tree_wrap, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vbar.set, xscrollcommand=hbar.set)
        tree.grid(row=0, column=0, sticky='nsew')
        vbar.grid(row=0, column=1, sticky='ns')
        hbar.grid(row=1, column=0, sticky='ew')

        btn_frm = ttk.Frame(win)
        btn_frm.pack(fill=tk.X)
        ttk.Button(btn_frm, text='刷新', command=lambda: self._refresh_pid_tree(tree)).pack(side=tk.LEFT)
        ttk.Button(btn_frm, text='删除并拉黑', command=lambda: self._pid_delete_and_blacklist(tree)).pack(side=tk.LEFT)
        ttk.Button(btn_frm, text='删除并释放ID', command=lambda: self._pid_delete_release(tree)).pack(side=tk.LEFT)
        ttk.Button(btn_frm, text='全部删除', command=lambda: self._pid_delete_all(tree)).pack(side=tk.LEFT)
        
        # context menu for copying
        pid_menu = tk.Menu(win, tearoff=0)
        pid_menu.add_command(label='全选', command=lambda: self._select_all_pid(tree))
        pid_menu.add_command(label='复制所选', command=lambda: self._copy_pid_selection(tree))
        pid_menu.add_command(label='粘贴导入', command=lambda: self._paste_pid_from_clipboard(tree))
        def pid_right(event):
            try:
                row_id = tree.identify_row(event.y)
                if row_id:
                    tree.selection_set(row_id)
                pid_menu.tk_popup(event.x_root, event.y_root)
            finally:
                pid_menu.grab_release()
        tree.bind('<Button-3>', pid_right)
        tree.bind('<Button-2>', pid_right)
        tree.bind('<Control-Button-1>', pid_right)
        tree.bind('<Control-a>', lambda e: self._select_all_pid(tree))
        tree.bind('<Control-c>', lambda e: self._copy_pid_selection(tree))
        tree.bind('<Control-v>', lambda e: self._paste_pid_from_clipboard(tree))

        self._pid_tree = tree
        self._pid_win = win
        self._refresh_pid_tree(tree)
        search_var.trace_add("write", lambda *_: self._debounce_call("pid_search", 180, lambda: self._pid_search(search_var.get())))

    def _copy_stream_sel(self):
        selected = self.stream_tree.selection()
        if not selected:
            messagebox.showwarning('警告', '请先选择一条弹幕')
            return
        lines = []
        for item in selected:
            vals = self.stream_tree.item(item, 'values')
            if vals:
                lines.append('\t'.join(str(v) for v in vals))
        copied_text = '\n'.join(lines)
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(copied_text)
            messagebox.showinfo('复制', f'已复制 {len(lines)} 条弹幕')
        except Exception as e:
            messagebox.showerror('错误', str(e))

    def _copy_stream_field(self, field_index: int, field_name: str):
        selected = self.stream_tree.selection()
        if not selected:
            messagebox.showwarning('警告', '请先选择一条弹幕')
            return
        lines = []
        for item in selected:
            vals = self.stream_tree.item(item, 'values')
            if len(vals) > field_index:
                lines.append(str(vals[field_index]))
        if not lines:
            messagebox.showwarning('警告', f'未找到可复制的{field_name}')
            return
        copied_text = '\n'.join(lines)
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(copied_text)
            messagebox.showinfo('复制', f'已复制 {len(lines)} 条{field_name}')
        except Exception as e:
            messagebox.showerror('错误', str(e))

    def _select_all_stream(self):
        children = self.stream_tree.get_children()
        if not children:
            return "break"
        self.stream_tree.selection_set(children)
        self.stream_tree.focus(children[-1])
        self.stream_tree.see(children[-1])
        return "break"

    def _paste_to_input(self):
        try:
            text = self.root.clipboard_get().strip()
            if not text:
                messagebox.showwarning('警告', '剪贴板为空')
                return
            first_meta = None
            parsed_lines = []
            for line in text.splitlines():
                cols = line.split('\t')
                if len(cols) >= 5:
                    if first_meta is None:
                        first_meta = cols
                    parsed_lines.append(cols[4])
                else:
                    parsed_lines.append(line)
            paste_seed = '\n'.join(parsed_lines).strip()
            paste_win = tk.Toplevel(self.root)
            paste_win.title('粘贴并补充打印')
            paste_win.geometry('400x150')
            ttk.Label(paste_win, text='修改内容后点击打印:').pack(padx=8, pady=8)
            paste_text = scrolledtext.ScrolledText(paste_win, height=5, width=50)
            paste_text.pack(padx=8, pady=4, fill=tk.BOTH, expand=True)
            paste_text.insert(tk.END, paste_seed)
            
            def do_print():
                content = paste_text.get('1.0', tk.END).strip()
                if not content:
                    messagebox.showwarning('警告', '内容不能为空')
                    return
                sel = self.stream_tree.selection()
                if sel:
                    vals = self.stream_tree.item(sel[0], 'values')
                    unique_id = vals[2] if len(vals) > 2 else 'manual'
                    display_name = vals[3] if len(vals) > 3 else 'manual'
                    permanent_id = vals[1] if len(vals) > 1 else '0'
                    timestamp = vals[0] if len(vals) > 0 else datetime.now().isoformat(sep=' ', timespec='seconds')
                elif first_meta:
                    unique_id = first_meta[2] if len(first_meta) > 2 else 'manual'
                    display_name = first_meta[3] if len(first_meta) > 3 else 'manual'
                    permanent_id = first_meta[1] if len(first_meta) > 1 else '0'
                    timestamp = first_meta[0] if len(first_meta) > 0 else datetime.now().isoformat(sep=' ', timespec='seconds')
                else:
                    unique_id = 'manual'
                    display_name = 'manual'
                    permanent_id = '0'
                    timestamp = datetime.now().isoformat(sep=' ', timespec='seconds')
                rendered = self._compose_print_rendered(permanent_id, unique_id, display_name, timestamp, content)
                try:
                    size_str = f"{self.width_var.get().strip()}x{self.height_var.get().strip()}"
                    trace_id = self._new_trace_id()
                    jid = db.add_print_job(
                        permanent_id, unique_id, display_name, timestamp, content, rendered, ("" if bool(self.use_default_printer_var.get()) else self.printer_cb.get()), size_str,
                        raw_message=content, rule_hit="manual_paste", trace_id=trace_id,
                    )
                    messagebox.showinfo('成功', f'已加入打印队列 (JID:{jid}, Trace:{trace_id})')
                    self.open_job_manager()
                    try:
                        self._refresh_job_tree(self._job_tree)
                    except Exception:
                        pass
                    paste_win.destroy()
                except Exception as e:
                    messagebox.showerror('错误', str(e))
            ttk.Button(paste_win, text='打印', command=do_print).pack(pady=8)
        except Exception as e:
            messagebox.showerror('错误', f'粘贴失败: {e}')

    def _reprint_selected(self):
        selected = self.stream_tree.selection()
        if not selected:
            messagebox.showwarning('警告', '请先选择一条弹幕')
            return
        success = 0
        last_jid = None
        errors = 0
        for item in selected:
            vals = self.stream_tree.item(item, 'values')
            if len(vals) < 5:
                errors += 1
                continue
            permanent_id = vals[1] if vals[1] not in ('', '-') else '0'
            unique_id = vals[2] if vals[2] not in ('', '-') else 'manual'
            display_name = vals[3] if vals[3] else 'manual'
            message = vals[4]
            timestamp = vals[0] if vals[0] else datetime.now().isoformat(sep=' ', timespec='seconds')
            if str(message).startswith('🔢 '):
                message = str(message)[2:]
            rendered = self._compose_print_rendered(permanent_id, unique_id, display_name, timestamp, message)
            try:
                size_str = f"{self.width_var.get().strip()}x{self.height_var.get().strip()}"
                trace_id = self._new_trace_id()
                last_jid = db.add_print_job(
                    permanent_id, unique_id, display_name, timestamp, message, rendered, ("" if bool(self.use_default_printer_var.get()) else self.printer_cb.get()), size_str,
                    raw_message=message, rule_hit="manual_reprint", trace_id=trace_id,
                )
                success += 1
            except Exception:
                errors += 1
        if success:
            msg = f'已加入打印队列 {success} 条'
            if last_jid is not None:
                msg += f' (最后JID:{last_jid})'
            if errors:
                msg += f'，失败 {errors} 条'
            messagebox.showinfo('成功', msg)
            self.open_job_manager()
            try:
                self._refresh_job_tree(self._job_tree)
            except Exception:
                pass
        else:
            messagebox.showerror('错误', '补充打印失败，请检查选择项和打印配置')

    def _get_selected_unique_ids(self):
        sel = self.stream_tree.selection()
        if not sel:
            return []
        unique_ids = []
        seen = set()
        for item in sel:
            vals = self.stream_tree.item(item, 'values')
            uid = vals[2] if len(vals) > 2 else ""
            uid = str(uid).strip()
            if not uid or uid in ("-", "manual") or uid in seen:
                continue
            seen.add(uid)
            unique_ids.append(uid)
        return unique_ids

    def _blacklist_selected(self):
        unique_ids = self._get_selected_unique_ids()
        if not unique_ids:
            messagebox.showwarning('警告', '请先选择有效的客户ID')
            return
        if not messagebox.askyesno('确认', f'确认拉黑并释放 {len(unique_ids)} 个用户ID?'):
            return
        blacklisted = 0
        released = 0
        for unique_id in unique_ids:
            try:
                db.add_blacklist(unique_id)
                blacklisted += 1
                freed = db.delete_user(unique_id)
                if freed is not None:
                    released += 1
                self.user_cache.pop(unique_id, None)
                self.guest_message_counter.pop(unique_id, None)
                self.today_guest_rank.pop(unique_id, None)
                self.user_msg_timestamps.pop(unique_id, None)
            except Exception:
                pass
        self._set_status(f"已拉黑并释放: 用户{blacklisted}，释放ID{released}")
        self._sync_local_permanent_ids_backup("blacklist_release")
        _sync_ok, sync_msg = self._sync_deleted_permanent_ids_to_server()
        if sync_msg:
            self._set_status(str(sync_msg))
        msg = f'已拉黑 {blacklisted} 个用户，释放永久ID {released} 个'
        if sync_msg:
            msg += f'\n\n服务器同步: {sync_msg}'
        messagebox.showinfo('完成', msg)
        if hasattr(self, '_pid_tree') and hasattr(self, '_pid_win') and self._pid_win.winfo_exists():
            self._refresh_pid_tree(self._pid_tree)

    def _release_selected(self):
        unique_ids = self._get_selected_unique_ids()
        if not unique_ids:
            messagebox.showwarning('警告', '请先选择有效的客户ID')
            return
        if not messagebox.askyesno('确认', f'确认仅释放 {len(unique_ids)} 个用户的永久ID（不拉黑）?'):
            return
        released = 0
        for unique_id in unique_ids:
            try:
                freed = db.delete_user(unique_id)
                if freed is not None:
                    released += 1
                self.user_cache.pop(unique_id, None)
            except Exception:
                pass
        self._set_status(f"已释放永久ID: {released}")
        self._sync_local_permanent_ids_backup("release_only")
        _sync_ok, sync_msg = self._sync_deleted_permanent_ids_to_server()
        if sync_msg:
            self._set_status(str(sync_msg))
        msg = f'已释放永久ID {released} 个（未拉黑）'
        if sync_msg:
            msg += f'\n\n服务器同步: {sync_msg}'
        messagebox.showinfo('完成', msg)
        if hasattr(self, '_pid_tree') and hasattr(self, '_pid_win') and self._pid_win.winfo_exists():
            self._refresh_pid_tree(self._pid_tree)

    def _copy_pid_selection(self, tree):
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], 'values')
        text = '\t'.join(str(v) for v in vals)
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo('复制', '已复制所选用户信息到剪贴板')
        except Exception:
            pass

    def _select_all_pid(self, tree):
        rows = tree.get_children()
        if not rows:
            return "break"
        tree.selection_set(rows)
        tree.focus(rows[-1])
        tree.see(rows[-1])
        return "break"

    def _paste_pid_from_clipboard(self, tree):
        try:
            text = self.root.clipboard_get().strip()
        except Exception as e:
            messagebox.showerror('错误', f'读取剪贴板失败: {e}')
            return "break"
        if not text:
            messagebox.showwarning('提示', '剪贴板为空')
            return "break"
        rows = []
        for line in text.splitlines():
            if not line.strip():
                continue
            # support tab/comma separated rows: permanent_id, display_name, unique_id
            if '\t' in line:
                cols = [c.strip() for c in line.split('\t')]
            else:
                cols = [c.strip() for c in line.split(',')]
            if len(cols) < 2:
                continue
            rows.append({
                'permanent_id': cols[0],
                'display_name': cols[1] if len(cols) >= 2 else '',
                'unique_id': cols[2] if len(cols) >= 3 else '',
            })
        rows = [r for r in rows if str(r.get('permanent_id', '')).strip() and str(r.get('unique_id', '')).strip()]
        if not rows:
            messagebox.showwarning('提示', '未识别到可导入的数据（需含 permanent_id 与 unique_id）')
            return "break"
        try:
            result = db.import_permanent_ids(rows)
            self.user_cache = {}
            self._sync_local_permanent_ids_backup("paste_import")
            self._refresh_pid_tree(tree)
            imported = result.get('imported', 0)
            skipped = result.get('skipped', 0)
            conflicts = result.get('conflicts', [])
            msg = f"粘贴导入完成：成功 {imported}，跳过 {skipped}"
            _sync_ok, sync_msg = self._sync_imported_permanent_ids_to_server()
            if sync_msg:
                msg += f"\n\n服务器同步: {sync_msg}"
            if conflicts:
                msg += "\n\n冲突信息:\n" + "\n".join(conflicts[:10])
                if len(conflicts) > 10:
                    msg += f"\n...其余 {len(conflicts)-10} 条已省略"
            messagebox.showinfo('导入结果', msg)
        except Exception as e:
            messagebox.showerror('错误', f'粘贴导入失败: {e}')
        return "break"

    def _pid_search(self, query: str):
        q = (query or '').strip().lower()
        for r in self._pid_tree.get_children():
            self._pid_tree.delete(r)
        users = self._get_pid_users_cached()
        for unique_id, display_name, permanent_id in users:
            if not q or q in str(unique_id).lower() or q in (display_name or '').lower() or q in str(permanent_id):
                self._pid_tree.insert('', tk.END, values=(permanent_id, unique_id, display_name))

    def _ensure_today_counters(self):
        today = datetime.now().date().isoformat()
        if today != self.today_date:
            self.today_date = today
            self.today_guest_rank = {}
            self.today_guest_counter = 0
            self.guest_message_counter = {}

    def _next_guest_metrics(self, unique_id: str):
        self._ensure_today_counters()
        if unique_id not in self.today_guest_rank:
            self.today_guest_counter += 1
            self.today_guest_rank[unique_id] = self.today_guest_counter
        rank = self.today_guest_rank[unique_id]
        self.guest_message_counter[unique_id] = self.guest_message_counter.get(unique_id, 0) + 1
        msg_count = self.guest_message_counter[unique_id]
        return msg_count, rank

    def _check_lock_order(self, unique_id: str, content: str, should_print: bool) -> tuple[bool, str]:
        if not self.lock_order_mode:
            return should_print, ""
        now_ts = time.time()
        if now_ts > self._lock_order_open_until:
            self._lock_order_open_until = now_ts + self.lock_order_window_seconds
            self._lock_order_winner_key = None
        if not should_print:
            return False, ""
        key = f"{unique_id}:{content}"
        if self._lock_order_winner_key is None:
            self._lock_order_winner_key = key
            return True, "抢单中签"
        if self._lock_order_winner_key == key:
            return True, "抢单中签"
        return False, "锁单未中"

    def _refresh_peak_mode(self):
        now_ts = time.time()
        if (now_ts - float(getattr(self, "_last_peak_eval_ts", 0.0))) < 1.0:
            return
        self._last_peak_eval_ts = now_ts
        pending = 0
        try:
            pending = len(db.list_print_jobs("pending"))
        except Exception:
            pending = int(getattr(self, "_last_peak_pending", 0))
        self._last_peak_pending = int(pending)
        if pending >= int(self.peak_critical_threshold):
            self.peak_mode = "critical"
        elif pending >= int(self.peak_warn_threshold):
            self.peak_mode = "warn"
        else:
            self.peak_mode = "normal"

    def _apply_peak_protection(self, unique_id: str, content: str, should_print: bool, rule_hit: str) -> tuple[bool, str]:
        """Return (should_print, tag). tag used for UI visibility."""
        if not should_print:
            return False, ""
        mode = str(getattr(self, "peak_mode", "normal"))
        if mode == "normal":
            return True, ""
        # warn: de-dup frequent repeats in short window
        if mode == "warn":
            try:
                win = max(1, int(self.peak_duplicate_window_seconds))
                if db.has_recent_job_duplicate(unique_id, content, win):
                    self.peak_merge_counter += 1
                    return False, "高峰合并"
            except Exception:
                pass
            return True, "高峰警戒"
        # critical: only keep high-priority prints (numeric / keyword), drop low-priority jobs
        if mode == "critical":
            hit = str(rule_hit or "")
            if not (hit == "numeric" or hit.startswith("keyword:")):
                self.peak_drop_counter += 1
                return False, "高峰降级丢弃"
            return True, "高峰关键优先"
        return should_print, ""

    def _dedupe_print_in_reconnect_window(self, unique_id: str, content: str, now_ts: float) -> bool:
        """True means skip printing duplicate message in reconnect window."""
        try:
            win_sec = max(0, int(self.reconnect_print_window_seconds))
        except Exception:
            win_sec = 20
        if win_sec <= 0:
            return False
        key = f"{str(unique_id).strip()}|{str(content).strip()}"
        if not key.strip("|"):
            return False
        # Durable dedupe: if same message already exists in recent
        # pending/processing/printed jobs, skip enqueue/print.
        try:
            if db.has_recent_job_duplicate(unique_id, content, win_sec):
                self.recent_print_keys[key] = float(now_ts)
                return True
        except Exception:
            pass
        cutoff = float(now_ts) - float(win_sec)
        try:
            stale = [k for k, ts in self.recent_print_keys.items() if float(ts) < cutoff]
            for k in stale:
                self.recent_print_keys.pop(k, None)
        except Exception:
            pass
        last_ts = self.recent_print_keys.get(key)
        if last_ts is not None and (float(now_ts) - float(last_ts)) <= float(win_sec):
            return True
        self.recent_print_keys[key] = float(now_ts)
        return False

    def _new_trace_id(self) -> str:
        return datetime.now().strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:12]

    def _match_print_rule(self, content: str) -> tuple[bool, str]:
        text = (content or "").strip()
        if not text:
            return False, ""
        if len(text) < self.print_min_len or len(text) > self.print_max_len:
            return False, ""
        is_numeric = text.isdigit()
        if is_numeric and self.auto_print_numeric:
            return True, "numeric"
        if self.keyword_print_enabled and self.keyword_print_list:
            for kw in self.keyword_print_list:
                if kw and kw in text:
                    return True, f"keyword:{kw}"
        return False, ""

    def _add_stream_row(self, when_ts, permanent_id, unique_id, display_name, message_text, guest_msg_count='-', today_guest_rank='-', source_room='-'):
        t = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(when_ts))
        try:
            text_msg = str(message_text or "")
            if text_msg.startswith("[") and "]" in text_msg:
                pass
            try:
                self.dashboard_recent_comments.append({
                    "time": t,
                    "room": str(source_room or "-"),
                    "uid": str(unique_id),
                    "name": str(display_name),
                    "content": text_msg,
                    "pid": str(permanent_id),
                })
            except Exception:
                pass
            row_id = self.stream_tree.insert(
                '',
                tk.END,
                values=(t, permanent_id, unique_id, display_name, message_text, guest_msg_count, today_guest_rank),
            )
            # keep last visible
            children = self.stream_tree.get_children()
            if children:
                self.stream_tree.see(children[-1])
            return row_id
        except Exception:
            return None

    def start_listen(self):
        url = self.room_url_var.get().strip()
        if not url:
            messagebox.showwarning("错误", "请输入直播网址")
            return
        listen_mode = self._get_listen_source_mode()
        if listen_mode == "relay" and (not self._license_server_url() or not self._license_key()):
            messagebox.showwarning("提示", "服务器中转监听必须先填写授权码")
            return
        if not self._ensure_license_active(show_dialog=True):
            return
        self._save_settings()
        unique_ids = self._parse_room_inputs(url)
        if not unique_ids:
            messagebox.showwarning("错误", "无法解析主播ID，请使用 @username / username / https://www.tiktok.com/@username，支持逗号分隔")
            return

        self._stop_event.clear()
        self.listening = True
        try:
            self.connect_btn.config(state=tk.DISABLED)
            self.stop_btn.config(state=tk.NORMAL)
        except Exception:
            pass
        new_count = 0
        for uid in unique_ids:
            if uid in self.listener_workers:
                continue
            block_until = 0.0 if listen_mode == "relay" else float(self._sign_rate_limited_until.get(uid, 0.0))
            if block_until > time.time():
                left_min = max(1, int((block_until - time.time()) / 60))
                self._set_status(f"签名限流冷却中 @{uid}，剩余约 {left_min} 分钟")
                continue
            self._start_relay_listener_for_uid(uid) if listen_mode == "relay" else self._start_listener_for_uid(uid)
            new_count += 1
        active = ", ".join("@"+u for u in self.listener_workers.keys())
        self._set_status(f"监听中: {active}")
        self._audit("start_listen", ",".join(unique_ids))
        self._track_business_event("listen_start", {"input_count": len(unique_ids), "active_rooms": list(self.listener_workers.keys())})
        self._sync_local_permanent_ids_backup("start_listen")
        if new_count > 0:
            messagebox.showinfo("开始", f"新增监听 {new_count} 个直播间")
        else:
            messagebox.showinfo("提示", "这些直播间已在监听中")

    def _poll(self):
        next_delay_ms = 200
        try:
            self._refresh_peak_mode()
            # Adaptive batching: process more when backlog grows, but keep one UI slice bounded.
            backlog = 0
            try:
                backlog = int(self.queue.qsize())
            except Exception:
                backlog = 0
            target_batch = 10
            if backlog > 300:
                target_batch = 100
            elif backlog > 120:
                target_batch = 60
            elif backlog > 40:
                target_batch = 30
            target_batch = max(10, min(POLL_MAX_BATCH, int(target_batch)))
            deadline = time.time() + (float(POLL_MAX_SLICE_MS) / 1000.0)
            batch = []
            for _ in range(target_batch):
                try:
                    msg = self.queue.get_nowait()
                    batch.append(msg)
                except queue.Empty:
                    break
                if time.time() >= deadline:
                    break
            try:
                left = int(self.queue.qsize())
            except Exception:
                left = 0
            if left > 200:
                next_delay_ms = 40
            elif left > 80:
                next_delay_ms = 80
            elif left > 20:
                next_delay_ms = 120

            for item in batch:
                if len(item) >= 7:
                    unique_id, name, content, timestamp, allow_auto_print, source_room, source_tag = item
                elif len(item) >= 6:
                    unique_id, name, content, timestamp, allow_auto_print, source_room = item
                    source_tag = "main"
                elif len(item) >= 5:
                    unique_id, name, content, timestamp, allow_auto_print = item
                    source_room = "-"
                    source_tag = "main"
                else:
                    unique_id, name, content, timestamp = item
                    allow_auto_print = True
                    source_room = "-"
                    source_tag = "main"
                if unique_id is None:
                    if isinstance(content, str) and content.startswith("__LIVE_END__::"):
                        ended_uid = content.split("::", 1)[1].strip()
                        if source_tag == "analysis":
                            self._stop_listener_for_uid(ended_uid, reason=f"分析直播已结束: @{ended_uid}", worker_map=self.analysis_listener_workers, source_tag="analysis")
                        else:
                            self._stop_listener_for_uid(ended_uid, reason=f"直播已结束: @{ended_uid}", worker_map=self.listener_workers, source_tag="main")
                        try:
                            when_ts = datetime.fromisoformat(timestamp).timestamp()
                        except Exception:
                            when_ts = time.time()
                        self._add_stream_row(when_ts, '-', ended_uid, '系统', f'检测到直播结束，已自动停止监听 @{ended_uid}', '-', '-', source_room)
                        continue
                    try:
                        when_ts = datetime.fromisoformat(timestamp).timestamp()
                    except Exception:
                        when_ts = time.time()
                    self._add_stream_row(when_ts, '-', '-', '系统', content, '-', '-', source_room)
                    continue

                guest_msg_count, today_guest_rank = self._next_guest_metrics(unique_id)
                try:
                    when_ts = datetime.fromisoformat(timestamp).timestamp()
                except Exception:
                    when_ts = time.time()
                text_lower = (content or "").lower()
                name_lower = (name or "").lower()
                white_hit = any((kw.lower() in text_lower) or (kw.lower() in name_lower) for kw in self.whitelist_keywords)

                # automatic blacklist by keyword/rate (unless whitelisted)
                if not white_hit and self.blacklist_keywords:
                    if any((kw.lower() in text_lower) or (kw.lower() in name_lower) for kw in self.blacklist_keywords):
                        db.add_blacklist(unique_id)
                if not white_hit and self.auto_blacklist_rate_limit > 0:
                    dq = self.user_msg_timestamps[unique_id]
                    now_ts = time.time()
                    dq.append(now_ts)
                    while dq and (now_ts - dq[0]) > 10:
                        dq.popleft()
                    if len(dq) >= self.auto_blacklist_rate_limit:
                        db.add_blacklist(unique_id)

                # check blacklist
                if db.is_blacklisted(unique_id):
                    msg = f"[黑名单] {content}"
                    self._record_overlap_event(source_room, unique_id, when_ts, source_tag=source_tag)
                    self._record_overlap_message_event(source_room, unique_id, name, msg, when_ts, source_tag=source_tag)
                    self._add_stream_row(when_ts, 0, unique_id, f"({name})", msg, guest_msg_count, today_guest_rank, source_room)
                    self._enqueue_cloud_event(source_room, unique_id, 0, name, msg, timestamp, False)
                    continue

                # detect pure numeric
                is_numeric = content.isdigit()
                rule_match, rule_hit = self._match_print_rule(content)
                should_print = rule_match and allow_auto_print
                should_print, lock_tag = self._check_lock_order(unique_id, content, should_print)
                if should_print and self._dedupe_print_in_reconnect_window(unique_id, content, when_ts):
                    should_print = False
                should_print, peak_tag = self._apply_peak_protection(unique_id, content, should_print, rule_hit)

                # Use cache to reduce DB queries.
                # Only create a permanent ID when needed (numeric flow).
                if unique_id not in self.user_cache:
                    try:
                        existing = db.get_user_by_unique_id(unique_id)
                        if existing:
                            permanent_id, stored_name = existing
                            resolved_name = name or stored_name
                            self.user_cache[unique_id] = (permanent_id, resolved_name)
                        elif is_numeric:
                            remote_pid = self._resolve_remote_permanent_id(unique_id, name)
                            if remote_pid:
                                permanent_id = remote_pid
                            else:
                                _, permanent_id = db.get_or_create_user(unique_id, name)
                            self.user_cache[unique_id] = (permanent_id, name)
                            self._sync_local_permanent_ids_backup("create_pid")
                        else:
                            permanent_id = ''
                            self.user_cache[unique_id] = ('', name)
                    except Exception:
                        permanent_id = ''
                        self.user_cache[unique_id] = ('', name)
                else:
                    permanent_id, cache_name = self.user_cache[unique_id]
                    if not name and cache_name:
                        name = cache_name

                # If the user first spoke non-numeric and now sends numeric, assign PID now.
                if is_numeric and not permanent_id:
                    try:
                        remote_pid = self._resolve_remote_permanent_id(unique_id, name)
                        if remote_pid:
                            permanent_id = remote_pid
                        else:
                            _, permanent_id = db.get_or_create_user(unique_id, name)
                        self.user_cache[unique_id] = (permanent_id, name)
                        self._sync_local_permanent_ids_backup("create_pid")
                    except Exception:
                        pass

                # ADD TO STREAM WINDOW
                try:
                    self._record_overlap_event(source_room, unique_id, when_ts, source_tag=source_tag)
                    self._record_overlap_message_event(source_room, unique_id, name, content, when_ts, source_tag=source_tag)
                    msg_display = f"🔢 {content}" if is_numeric else content
                    if lock_tag:
                        msg_display = f"[{lock_tag}] {msg_display}"
                    if peak_tag:
                        msg_display = f"[{peak_tag}] {msg_display}"
                    self._add_stream_row(when_ts, permanent_id, unique_id, name, msg_display, guest_msg_count, today_guest_rank, source_room)
                    self._enqueue_cloud_event(source_room, unique_id, permanent_id, name, content, timestamp, should_print)
                    self._archive_stream_overflow()
                    if hasattr(self, "_overlap_win") and self._overlap_win.winfo_exists():
                        now_ts = time.time()
                        if now_ts - self._last_overlap_refresh_ts > 2.0:
                            self._last_overlap_refresh_ts = now_ts
                            self._refresh_overlap_view()
                except Exception:
                    pass
                if self._feishu_should_send_comment(content, rule_hit):
                    self._enqueue_feishu_event(
                        "弹幕推送",
                        f"房间: @{source_room}\n客户: {name} ({unique_id})\n永久ID: {permanent_id}\n规则: {rule_hit or '-'}\n内容: {content}",
                    )

                # Match print rules -> persist print job
                if should_print:
                    rendered = self._compose_print_rendered(
                        permanent_id, unique_id, name, timestamp, content,
                        extra_vars={
                            "guest_msg_count": guest_msg_count,
                            "today_guest_rank": today_guest_rank,
                            "source_room": source_room,
                            "lock_tag": lock_tag or "",
                        },
                    )
                    try:
                        size_str = f"{self.width_var.get().strip()}x{self.height_var.get().strip()}"
                        trace_id = self._new_trace_id()
                        jid = db.add_print_job(
                            permanent_id, unique_id, name, timestamp, content, rendered, ("" if bool(self.use_default_printer_var.get()) else self.printer_cb.get()), size_str,
                            raw_message=content, rule_hit=rule_hit, trace_id=trace_id,
                        )
                        self._audit("print_job_enqueued", f"jid={jid}|trace={trace_id}|uid={unique_id}|rule={rule_hit}")
                    except Exception:
                        os.makedirs(DATA_DIR, exist_ok=True)
                        safe_name = f"{permanent_id}_{int(time.time())}.txt"
                        path = os.path.join(DATA_DIR, safe_name)
                        try:
                            rendered_text = str(rendered or "")
                            if rendered_text.startswith(CANVAS_PRINT_MARKER):
                                rendered_text = rendered_text[len(CANVAS_PRINT_MARKER):].lstrip("\n")
                            printer_utils.print_to_file(rendered_text, path)
                        except Exception:
                            pass
        except Exception:
            pass
        finally:
            self.root.after(max(30, int(next_delay_ms)), self._poll)

    def export_comments(self):
        """Export comments from stream_tree to a text file."""
        if not self.stream_tree.get_children():
            messagebox.showinfo("导出", "当前无弹幕可导出")
            return
        path = filedialog.asksaveasfilename(defaultextension='.txt', filetypes=[('Text files','*.txt')], title='保存弹幕为')
        if not path:
            return
        try:
            with open(path, 'w', encoding='utf-8') as f:
                for item in self.stream_tree.get_children():
                    vals = self.stream_tree.item(item, 'values')
                    # vals = (time, permanent_id, unique_id, display_name, message)
                    line = f"{vals[0]} | {vals[3]}: {vals[4]}\n" if len(vals) >= 5 else '\n'
                    f.write(line)
            messagebox.showinfo('导出', f'已保存到 {path}')
        except Exception as e:
            messagebox.showerror('错误', str(e))

    def export_permanent_ids(self):
        """Export permanent ID mapping to XLSX/CSV."""
        users = db.list_users()
        if not users:
            messagebox.showinfo('导出', '无永久编号数据')
            return
        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx'), ('CSV files', '*.csv')],
            title='保存永久编号为'
        )
        if not path:
            return
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext == '.xlsx':
                if Workbook is None:
                    messagebox.showerror('错误', '未安装 openpyxl，无法导出 XLSX。请执行: pip install openpyxl')
                    return
                wb = Workbook()
                ws = wb.active
                ws.title = "permanent_ids"
                # A列=永久ID, B列=昵称, C列=客户ID
                ws.append(['permanent_id', 'display_name', 'unique_id'])
                for unique_id, display_name, permanent_id in users:
                    ws.append([int(permanent_id), str(display_name or ''), str(unique_id or '')])
                wb.save(path)
            else:
                with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                    w = csv.writer(f)
                    # A列=永久ID, B列=昵称, C列=客户ID
                    w.writerow(['permanent_id', 'display_name', 'unique_id'])
                    for unique_id, display_name, permanent_id in users:
                        w.writerow([permanent_id, (display_name or '').replace('\n', ' '), unique_id])
            messagebox.showinfo('导出', f'已保存到 {path}')
        except Exception as e:
            messagebox.showerror('错误', str(e))

    def import_permanent_ids(self):
        """Import permanent IDs from XLSX/CSV."""
        if not self._require_admin():
            return
        path = filedialog.askopenfilename(
            filetypes=[('Excel files', '*.xlsx'), ('CSV files', '*.csv')],
            title='选择永久编号文件(XLSX/CSV)'
        )
        if not path:
            return
        try:
            rows = []
            ext = os.path.splitext(path)[1].lower()
            if ext == '.xlsx':
                if load_workbook is None:
                    messagebox.showerror('错误', '未安装 openpyxl，无法导入 XLSX。请执行: pip install openpyxl')
                    return
                wb = load_workbook(path, data_only=True)
                ws = wb.active
                header = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    vals = [("" if v is None else str(v).strip()) for v in row]
                    if i == 0:
                        header = [x.lower() for x in vals]
                        continue
                    if not any(vals):
                        continue
                    # A列永久ID B列昵称 C列客户ID（优先），同时兼容有表头
                    col_a = vals[0] if len(vals) > 0 else ""
                    col_b = vals[1] if len(vals) > 1 else ""
                    col_c = vals[2] if len(vals) > 2 else ""
                    row_map = {}
                    if header:
                        for idx, h in enumerate(header):
                            if h:
                                row_map[h] = vals[idx] if idx < len(vals) else ""
                    pid = (row_map.get('permanent_id') or row_map.get('永久id') or row_map.get('永久编号') or col_a).strip()
                    display_name = (row_map.get('display_name') or row_map.get('昵称') or row_map.get('name') or col_b).strip()
                    unique_id = (row_map.get('unique_id') or row_map.get('客户id') or row_map.get('customer_id') or col_c).strip()
                    if not unique_id and display_name and any(ch in display_name for ch in ('@', '_', '.')):
                        unique_id = display_name
                        display_name = ""
                    rows.append({'permanent_id': pid, 'unique_id': unique_id, 'display_name': display_name})
            else:
                with open(path, 'r', encoding='utf-8-sig', newline='') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        pid = (row.get('permanent_id') or row.get('永久ID') or row.get('永久编号') or '').strip()
                        # preferred: B列昵称, C列客户ID
                        display_name = (row.get('display_name') or row.get('昵称') or row.get('name') or '').strip()
                        unique_id = (row.get('unique_id') or row.get('客户ID') or row.get('customer_id') or '').strip()
                        # backward compatible: older exported format B=unique_id, C=display_name
                        if not unique_id and display_name and any(ch in display_name for ch in ('@', '_', '.')):
                            unique_id = display_name
                            display_name = (row.get('display_name') or '').strip()
                        rows.append({
                            'permanent_id': pid,
                            'unique_id': unique_id,
                            'display_name': display_name,
                        })
            rows = [r for r in rows if str(r.get('permanent_id', '')).strip() and str(r.get('unique_id', '')).strip()]
            if not rows:
                messagebox.showwarning('导入', '文件为空或没有有效数据')
                return
            result = db.import_permanent_ids(rows)
            self.user_cache = {}
            self._sync_local_permanent_ids_backup("file_import")
            if hasattr(self, '_pid_tree') and hasattr(self, '_pid_win') and self._pid_win.winfo_exists():
                self._refresh_pid_tree(self._pid_tree)
            imported = result.get('imported', 0)
            skipped = result.get('skipped', 0)
            conflicts = result.get('conflicts', [])
            msg = f"导入完成：成功 {imported}，跳过 {skipped}"
            _sync_ok, sync_msg = self._sync_imported_permanent_ids_to_server()
            if sync_msg:
                msg += f"\n\n服务器同步: {sync_msg}"
            if conflicts:
                msg += "\n\n冲突信息:\n" + "\n".join(conflicts[:20])
                if len(conflicts) > 20:
                    msg += f"\n...其余 {len(conflicts)-20} 条已省略"
            messagebox.showinfo('导入结果', msg)
        except Exception as e:
            messagebox.showerror('错误', f'导入失败: {e}')

    def _refresh_pid_tree(self, tree):
        for r in tree.get_children():
            tree.delete(r)
        users = db.list_users()
        try:
            self._pid_users_cache = list(users)
            self._pid_cache_ts = time.time()
        except Exception:
            pass
        for unique_id, display_name, permanent_id in users:
            tree.insert('', tk.END, values=(permanent_id, unique_id, display_name))

    def _get_pid_users_cached(self):
        try:
            if self._pid_users_cache and (time.time() - float(self._pid_cache_ts)) <= 20.0:
                return list(self._pid_users_cache)
        except Exception:
            pass
        users = db.list_users()
        try:
            self._pid_users_cache = list(users)
            self._pid_cache_ts = time.time()
        except Exception:
            pass
        return users

    def open_job_manager(self):
        if hasattr(self, '_job_win') and self._job_win.winfo_exists():
            self._job_win.lift()
            try:
                self._refresh_job_tree(self._job_tree)
            except Exception:
                pass
            return
        win = tk.Toplevel(self.root)
        win.title('打印任务管理')
        filter_frm = ttk.Frame(win)
        filter_frm.pack(fill=tk.X, padx=6, pady=6)
        ttk.Label(filter_frm, text="状态筛选").pack(side=tk.LEFT)
        status_filter_var = tk.StringVar(value="all")
        ttk.Combobox(filter_frm, textvariable=status_filter_var, values=["all", "pending", "processing", "printed", "failed"], width=12, state="readonly").pack(side=tk.LEFT, padx=4)
        ttk.Label(filter_frm, text="关键词").pack(side=tk.LEFT, padx=(8, 2))
        kw_filter_var = tk.StringVar(value="")
        ttk.Entry(filter_frm, textvariable=kw_filter_var, width=24).pack(side=tk.LEFT, padx=2)
        auto_refresh_var = tk.IntVar(value=1)
        ttk.Checkbutton(filter_frm, text="自动刷新", variable=auto_refresh_var).pack(side=tk.LEFT, padx=8)
        cols = ('id', 'trace_id', 'permanent_id', 'unique_id', 'display_name', 'time', 'status', 'printer', 'size', 'rule_hit', 'fail_reason')
        tree = ttk.Treeview(win, columns=cols, show='headings', selectmode='extended')
        for c, l in zip(cols, ('JID','Trace','永久ID','客户ID','昵称','时间','状态','打印机','尺寸','命中规则','失败原因')):
            tree.heading(c, text=l)
        tree.column('trace_id', width=160)
        tree.column('fail_reason', width=220)
        tree.column('rule_hit', width=120)
        tree.pack(fill=tk.BOTH, expand=True)

        btn_frm = ttk.Frame(win)
        btn_frm.pack(fill=tk.X)
        ttk.Button(btn_frm, text='刷新', command=lambda: self._refresh_job_tree(tree, status_filter_var.get(), kw_filter_var.get())).pack(side=tk.LEFT)
        ttk.Button(btn_frm, text='重试所选', command=lambda: self._job_retry_selected(tree)).pack(side=tk.LEFT)
        ttk.Button(btn_frm, text='批量删除所选', command=lambda: self._job_delete_selected(tree)).pack(side=tk.LEFT)
        ttk.Button(btn_frm, text='清空已完成/失败', command=lambda: self._job_clear_finished_failed(tree)).pack(side=tk.LEFT)

        tree.bind('<Control-c>', lambda e: self._copy_job_selection(tree))
        self._job_tree = tree
        self._job_win = win
        self._refresh_job_tree(tree, status_filter_var.get(), kw_filter_var.get())
        status_filter_var.trace_add("write", lambda *_: self._debounce_call("job_filter_refresh", 120, lambda: self._refresh_job_tree(tree, status_filter_var.get(), kw_filter_var.get())))
        kw_filter_var.trace_add("write", lambda *_: self._debounce_call("job_filter_refresh", 180, lambda: self._refresh_job_tree(tree, status_filter_var.get(), kw_filter_var.get())))
        def _auto_refresh():
            if not (hasattr(self, "_job_win") and self._job_win and self._job_win.winfo_exists()):
                return
            if not bool(auto_refresh_var.get()):
                try:
                    win.after(1500, _auto_refresh)
                except Exception:
                    pass
                return
            try:
                self._refresh_job_tree(tree, status_filter_var.get(), kw_filter_var.get())
            except Exception:
                pass
            try:
                win.after(1500, _auto_refresh)
            except Exception:
                pass
        win.after(1500, _auto_refresh)

    def _refresh_job_tree(self, tree, status_filter: str = "all", kw_filter: str = ""):
        for r in tree.get_children():
            tree.delete(r)
        rows = db.list_print_jobs()
        sf = str(status_filter or "all").strip().lower()
        kw = str(kw_filter or "").strip().lower()
        for row in rows:
            jid, pid, uid, name, when, content, status, printer, size, created, rule_hit, fail_reason, trace_id = row
            if sf not in ("", "all") and str(status).lower() != sf:
                continue
            if kw:
                hay = f"{jid} {trace_id} {pid} {uid} {name} {when} {status} {printer} {size} {rule_hit} {fail_reason} {content}".lower()
                if kw not in hay:
                    continue
            tree.insert('', tk.END, values=(jid, trace_id, pid, uid, name, when, status, printer, size, rule_hit, fail_reason))

    def _job_retry_selected(self, tree):
        sel = tree.selection()
        if not sel:
            messagebox.showwarning('提示', '请选择任务')
            return
        jid = tree.item(sel[0], 'values')[0]
        db.reset_job_to_pending(jid)
        self._refresh_job_tree(tree)
        messagebox.showinfo('已重试', f'任务 {jid} 已重置为 pending')

    def _job_delete_selected(self, tree):
        sel = tree.selection()
        if not sel:
            messagebox.showwarning('提示', '请选择任务')
            return
        count = 0
        for item in sel:
            jid = tree.item(item, 'values')[0]
            try:
                db.delete_print_job(int(jid))
                count += 1
            except Exception:
                pass
        self._refresh_job_tree(tree)
        messagebox.showinfo('完成', f'已删除 {count} 条任务')

    def _job_clear_finished_failed(self, tree):
        if not messagebox.askyesno('确认', '确认清空所有“已完成/失败”任务吗？'):
            return
        try:
            deleted = db.delete_print_jobs_by_status(['printed', 'failed'])
            self._refresh_job_tree(tree)
            messagebox.showinfo('完成', f'已清空 {deleted} 条任务')
        except Exception as e:
            messagebox.showerror('错误', str(e))

    def _copy_job_selection(self, tree):
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], 'values')
        text = '\t'.join(str(v) for v in vals)
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo('复制', '已复制任务信息到剪贴板')
        except Exception:
            pass

    def print_preview_selected(self):
        # preview last saved print file if exists
        try:
            files = sorted([f for f in os.listdir(DATA_DIR) if f.endswith('.txt')])
            if not files:
                messagebox.showinfo('预览', '无打印文件')
                return
            path = os.path.join(DATA_DIR, files[-1])
            if os.name == 'nt':
                os.startfile(path)
            else:
                subprocess.call(['open' if os.uname().sysname == 'Darwin' else 'xdg-open', path])
        except Exception as e:
            messagebox.showerror('错误', str(e))

    def _pid_delete_and_blacklist(self, tree):
        sel = tree.selection()
        if not sel:
            messagebox.showwarning('提示', '请选择一个用户')
            return
        vals = tree.item(sel[0], 'values')
        pid, unique_id = vals[0], vals[1]
        confirm = messagebox.askyesno('确认', f'是否删除并拉黑 {unique_id} (ID={pid}) ?')
        if not confirm:
            return
        db.blacklist_and_remove(unique_id)
        self._sync_local_permanent_ids_backup("pid_delete_blacklist")
        _sync_ok, sync_msg = self._sync_deleted_permanent_ids_to_server()
        if sync_msg:
            self._set_status(str(sync_msg))
        self._refresh_pid_tree(tree)
        msg = f'已删除并拉黑 {unique_id} (ID={pid})'
        if sync_msg:
            msg += f'\n\n服务器同步: {sync_msg}'
        messagebox.showinfo('完成', msg)

    def _pid_delete_release(self, tree):
        sel = tree.selection()
        if not sel:
            messagebox.showwarning('提示', '请选择一个用户')
            return
        vals = tree.item(sel[0], 'values')
        pid, unique_id = vals[0], vals[1]
        confirm = messagebox.askyesno('确认', f'是否删除 {unique_id} 并释放 ID={pid} ?')
        if not confirm:
            return
        db.delete_user(unique_id)
        self._sync_local_permanent_ids_backup("pid_delete_release")
        _sync_ok, sync_msg = self._sync_deleted_permanent_ids_to_server()
        if sync_msg:
            self._set_status(str(sync_msg))
        self._refresh_pid_tree(tree)
        msg = f'已删除 {unique_id} 并释放 ID={pid}'
        if sync_msg:
            msg += f'\n\n服务器同步: {sync_msg}'
        messagebox.showinfo('完成', msg)

    def _pid_delete_all(self, tree):
        if not self._require_admin():
            return
        users = db.list_users()
        total = len(users)
        if total == 0:
            messagebox.showinfo('提示', '当前没有永久编号数据')
            return
        if not messagebox.askyesno('确认', f'确认删除全部 {total} 条永久编号吗？此操作不可撤销。'):
            return
        snapshot_path, snapshot_count = self._snapshot_local_permanent_ids_backup("before_local_clear_all")
        deleted = db.delete_all_users()
        self.user_cache.clear()
        self.guest_message_counter.clear()
        self.today_guest_rank.clear()
        self._sync_local_permanent_ids_backup("pid_delete_all")
        sync_msg = "服务器保留原有永久编号，未跟随本地执行全部清空"
        self._set_status(str(sync_msg))
        self._refresh_pid_tree(tree)
        msg = f'已删除 {deleted} 条永久编号'
        if sync_msg:
            msg += f'\n\n服务器同步: {sync_msg}'
        if snapshot_path and snapshot_count > 0:
            msg += f'\n\n本地快照: 已额外保存 {snapshot_count} 条清空前备份'
        messagebox.showinfo('完成', msg)

    def _print_dispatcher(self):
        """Claim pending jobs in time windows and dispatch them to worker queue."""
        while True:
            try:
                rows = db.fetch_pending_jobs_batch(
                    window_seconds=self.print_batch_window_seconds,
                    limit=self.print_batch_limit,
                )
                if not rows:
                    time.sleep(0.1)
                    continue
                for row in rows:
                    self.print_task_queue.put(row)
            except Exception:
                try:
                    # silent fail-safe
                    pass
                except Exception:
                    pass
                time.sleep(0.2)

    def _print_worker(self):
        """Concurrent worker that processes one claimed print job at a time."""
        while True:
            try:
                row = self.print_task_queue.get(timeout=0.5)
            except queue.Empty:
                continue
            try:
                # row: (id, permanent_id, unique_id, display_name, time, content, rendered, printer, printer_size, trace_id)
                if len(row) >= 10:
                    jid, pid, unique_id, display_name, when, content, rendered, printer, printer_size, trace_id = row
                else:
                    jid, pid, unique_id, display_name, when, content, rendered, printer, printer_size = row
                    trace_id = ""
                rendered_text = str(rendered or "")
                is_canvas_render = False
                font_scale = 1.0
                canvas_payload = None
                if rendered_text.startswith(CANVAS_PRINT_MARKER):
                    is_canvas_render = True
                    rendered_text = rendered_text[len(CANVAS_PRINT_MARKER):].lstrip("\n")
                cleaned_lines = []
                for ln in str(rendered_text).splitlines():
                    t = str(ln).strip()
                    if t.startswith(CANVAS_PAYLOAD_MARKER_PREFIX) and t.endswith("]]"):
                        try:
                            raw_payload = t[len(CANVAS_PAYLOAD_MARKER_PREFIX):-2].strip()
                            canvas_payload = json.loads(base64.b64decode(raw_payload.encode("ascii")).decode("utf-8"))
                        except Exception:
                            canvas_payload = None
                        continue
                    if t.startswith(FONT_SCALE_MARKER_PREFIX) and t.endswith("]]"):
                        try:
                            raw_val = t[len(FONT_SCALE_MARKER_PREFIX):-2].strip()
                            font_scale = max(0.35, min(3.0, float(raw_val)))
                        except Exception:
                            font_scale = 1.0
                        continue
                    cleaned_lines.append(str(ln))
                rendered_text = "\n".join(cleaned_lines).lstrip("\n")

                # write rendered content to file first
                os.makedirs(DATA_DIR, exist_ok=True)
                safe_name = f"{pid}_{int(time.time())}_{jid}.txt"
                path = os.path.join(DATA_DIR, safe_name)
                try:
                    printer_utils.print_to_file(rendered_text, path)
                except Exception as e:
                    db.mark_job_failed(jid, str(e))
                    continue

                # Attempt to send to printer if set, with limited retries
                success = False
                selected_printer = str(getattr(self, "selected_printer_cached", "") or "").strip()
                default_printer = printer_utils.get_default_printer()
                candidates = []
                use_default = bool(getattr(self, "use_default_printer_cached", True))
                if use_default:
                    ordered = [default_printer, str(printer or "").strip(), selected_printer]
                else:
                    ordered = [str(printer or "").strip(), selected_printer, default_printer]
                for p in ordered:
                    if p and p not in candidates:
                        candidates.append(p)

                used_printer = ""
                failure_details = []
                width_mm = None
                height_mm = None
                try:
                    if printer_size and "x" in str(printer_size).lower():
                        p = str(printer_size).lower().replace("mm", "").replace(" ", "").split("x")
                        if len(p) >= 2:
                            width_mm = int(float(p[0]))
                            height_mm = int(float(p[1]))
                except Exception:
                    width_mm = None
                    height_mm = None
                if width_mm is None or height_mm is None:
                    try:
                        width_mm = int(float(str(self.settings.get("custom_paper_width_mm", "40")).strip()))
                        height_mm = int(float(str(self.settings.get("custom_paper_height_mm", "30")).strip()))
                    except Exception:
                        width_mm, height_mm = 40, 30
                for target_printer in candidates:
                    attempts = 0
                    lock = self.printer_locks[target_printer]
                    with lock:
                        char_width_mm, line_height_mm, margin_mm = self._print_calibration()
                        while attempts < self.print_retry_limit and not success:
                            try:
                                success, detail = printer_utils.send_to_printer_debug(
                                    target_printer,
                                    path,
                                    width_mm=width_mm,
                                    height_mm=height_mm,
                                    canvas_mode=is_canvas_render,
                                    preformatted_mode=True,
                                    canvas_payload=canvas_payload,
                                    font_scale=font_scale,
                                    char_width_mm=char_width_mm,
                                    line_height_mm=line_height_mm,
                                    margin_mm=margin_mm,
                                )
                                if not success and detail:
                                    failure_details.append(f"[{target_printer}] {detail}")
                            except Exception:
                                success = False
                                failure_details.append(f"[{target_printer}] unexpected_exception")
                            attempts += 1
                            if not success:
                                time.sleep(0.2)
                    if success:
                        used_printer = target_printer
                        break

                if success:
                    db.mark_job_printed(jid)
                    if trace_id:
                        self._audit("print_job_printed", f"jid={jid}|trace={trace_id}|printer={used_printer}")
                else:
                    if candidates:
                        joined = " | ".join(failure_details)[:900]
                        fail_reason = f"send_to_printer_failed:{joined}" if joined else ("send_to_printer_failed:" + ",".join(candidates))
                    else:
                        fail_reason = "no_printer_selected"
                    db.mark_job_failed(jid, fail_reason)
                    if trace_id:
                        self._audit("print_job_failed", f"jid={jid}|trace={trace_id}|reason={fail_reason[:200]}")
                    self._enqueue_feishu_event(
                        "打印失败告警",
                        f"JID:{jid}\n客户:{display_name} ({unique_id})\n永久ID:{pid}\n原因:{fail_reason[:300]}",
                    )
            except Exception:
                try:
                    pass
                except Exception:
                    pass
            finally:
                try:
                    self.print_task_queue.task_done()
                except Exception:
                    pass



def main():
    root = tk.Tk()
    app = App(root)
    # If AUTO_TEST env set, enqueue some test messages after startup
    if os.environ.get('AUTO_TEST'):
        def do_test():
            ts = datetime.now().isoformat(sep=' ', timespec='seconds')
            # non-numeric
            app.queue.put(("user_abc", "Alice", "hello world", ts))
            # numeric (should create PID and enqueue print)
            app.queue.put(("user_123", "Bob", "12345", ts))
            # another numeric
            app.queue.put(("user_456", "Carol", "67890", ts))
        root.after(1000, do_test)
    root.mainloop()


def _write_startup_crash_log(exc: Exception):
    try:
        log_path = os.path.join(RUNTIME_DIR, "sen_nails_crash.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Startup Crash\n")
            f.write("".join(traceback.format_exception(type(exc), exc, exc.__traceback__)))
        return log_path
    except Exception:
        return ""


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        log_file = _write_startup_crash_log(e)
        try:
            r = tk.Tk()
            r.withdraw()
            msg = f"程序启动失败: {e}"
            if log_file:
                msg += f"\n已写入日志: {log_file}"
            messagebox.showerror("Sen Nails 启动失败", msg)
            r.destroy()
        except Exception:
            pass
